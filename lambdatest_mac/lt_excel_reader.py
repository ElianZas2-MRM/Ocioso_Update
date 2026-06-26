"""
lt_excel_reader.py
==================
Lee los archivos Excel de Osocio (data/Lead_information_Formulario_<Pais>_Main.xlsx)
y construye los pares (URL pública, URL secure) con los datos de cada fila.

NO importa nada del Proyecto Osocio. Lee los archivos directamente con openpyxl.

Estructura esperada del Excel de Osocio:
    Fila 1: encabezados (URL, Formulario, Modelo, Nombre, Apellido, Documento,
             Celular, Email, ...)
    Desde fila 2: datos

Columnas mapeadas:
    A = URL pública (landing page)
    B = URL del formulario (secure-developments.com/...)
    C en adelante = datos del lead (Modelo, Nombre, Apellido, etc.)

El lector devuelve una lista de LeadRow, donde cada LeadRow tiene:
    - public_url
    - secure_url
    - data: dict con los campos del lead tal como vienen del Excel
"""

import os
from dataclasses import dataclass, field
from typing import List, Dict, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("openpyxl es requerido. Instalar con: pip install openpyxl")


@dataclass
class LeadRow:
    index: int               # nro de fila (1-based desde fila 2)
    public_url: str          # columna A
    secure_url: str          # columna B
    data: Dict[str, str]     # resto de columnas, clave = encabezado

    def __str__(self):
        return f"LeadRow #{self.index}: {self.public_url}"


def _normalize_url(value) -> str:
    """Convierte el valor de celda a string URL limpio."""
    if value is None:
        return ""
    return str(value).strip()


def read_osocio_excel(excel_path: str, inherit_urls: bool = True) -> List[LeadRow]:
    """
    Lee un Excel de Osocio y retorna una lista de LeadRow.

    excel_path: ruta completa al .xlsx
    inherit_urls: si True, cuando URL o Formulario están vacíos en una fila,
                  hereda los valores de la fila anterior (comportamiento Osocio).

    Retorna lista de LeadRow con las filas que tengan al menos URL pública.
    Filas completamente vacías se omiten.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No se encontró el Excel: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Fila 1 = encabezados
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]

    result: List[LeadRow] = []
    last_public_url = ""
    last_secure_url = ""

    for row_idx, row in enumerate(rows[1:], start=2):
        # Verificar si la fila está completamente vacía
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        # URL pública (columna A = índice 0)
        public_url = _normalize_url(row[0] if len(row) > 0 else None)
        # URL secure (columna B = índice 1)
        secure_url = _normalize_url(row[1] if len(row) > 1 else None)

        # Herencia de URLs vacías (igual que Osocio)
        if inherit_urls:
            if not public_url and last_public_url:
                public_url = last_public_url
            if not secure_url and last_secure_url:
                secure_url = last_secure_url

        if public_url:
            last_public_url = public_url
        if secure_url:
            last_secure_url = secure_url

        # Si no hay URL pública después de herencia, saltar fila
        if not public_url:
            continue

        # Construir dict con el resto de columnas (desde C en adelante)
        data: Dict[str, str] = {}
        for col_idx, header in enumerate(headers[2:], start=2):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                data[header] = str(val).strip()
            else:
                data[header] = ""

        result.append(LeadRow(
            index=row_idx,
            public_url=public_url,
            secure_url=secure_url,
            data=data,
        ))

    wb.close()
    return result


def find_osocio_excels(data_dir: str) -> Dict[str, str]:
    """
    Busca los Excels de Osocio en la carpeta data/ y retorna un dict:
        { "Argentina": "/ruta/a/Lead_information_Formulario_Argentina_Main.xlsx", ... }
    """
    if not os.path.exists(data_dir):
        return {}

    result = {}
    for filename in os.listdir(data_dir):
        if not filename.endswith(".xlsx"):
            continue
        if "Lead_information_Formulario_" not in filename:
            continue
        # Extraer nombre del país: Lead_information_Formulario_<Pais>_Main.xlsx
        try:
            # Remover prefijo y sufijo
            nombre = filename.replace("Lead_information_Formulario_", "")
            nombre = nombre.replace("_Main.xlsx", "")
            if nombre:
                full_path = os.path.join(data_dir, filename)
                result[nombre] = full_path
        except Exception:
            continue

    return result


def get_lead_data_from_row(row: LeadRow) -> Dict[str, str]:
    """
    Retorna el dict de datos de la fila tal como está en el Excel.
    Útil para logging y para el Excel de resultados.
    """
    return {
        "public_url": row.public_url,
        "secure_url": row.secure_url,
        **row.data,
    }


if __name__ == "__main__":
    # Test rápido
    import sys
    if len(sys.argv) < 2:
        print("Uso: python lt_excel_reader.py <ruta_al_excel>")
        sys.exit(1)
    rows = read_osocio_excel(sys.argv[1])
    print(f"Filas encontradas: {len(rows)}")
    for r in rows:
        print(f"  #{r.index}: {r.public_url}")
        print(f"          {r.secure_url}")
        print(f"          datos: {r.data}")
