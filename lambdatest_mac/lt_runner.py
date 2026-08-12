"""
lt_runner.py
============
Orquestador LambdaTest Mac para Osocio Form Automation.

Replica EXACTAMENTE la lógica de Osocio (base_form_filler.py) trasladada
a LambdaTest Web Automation Mac + Chrome.

Qué es igual a Osocio:
  - Manejo de cookies GM (gb-legal-notification, .close-btn.js-close-icon)
  - Scroll dinámico de la landing (pre_scroll_for_dynamic_content)
  - Secuencia de capturas: landing_inicial, form_errores, form_completado, ty_page
  - Dependencias entre dropdowns con polling activo (8s timeout, igual que Osocio)
  - _handle_terms_checkboxes: busca por name known_names + required, JS check +
    click label + PointerEvent fallback, 4 intentos por checkbox
  - _mark_preferred_radios: agrupa por name, elige opción sí/renovar/yes
  - Fuzzy matching en selects (exacto 100, contiene 80, contenido 70, palabras 50)
  - Guardado en resultados/resultados_<Pais><N>.xlsx con mismas columnas de Osocio
  - Screenshots en resultados/screenshots_<Pais><N>/

Qué es diferente (mejoras para LambdaTest):
  - Driver: LambdaTest Remote Mac + Safari 1440x900
  - Capturas: driver.save_screenshot() DIRECTAMENTE desde LT
    → LambdaTest captura la imagen con el frame de Mac incluido
    → NO se sale al default_content para las capturas (eso da pantalla en blanco)
    → Para la landing se hace desde el contexto principal normalmente
  - Ingreso de texto: JS puro (execute_script + dispatchEvent) en vez de send_keys
"""

import json
import os
import random
import sys
import threading
import time
import traceback
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Dict, List, Optional

try:
    import requests as _requests
    _REQUESTS_OK = True
except ImportError:
    _REQUESTS_OK = False

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, StaleElementReferenceException, NoSuchElementException
)
from selenium.webdriver import Remote
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.remote.client_config import ClientConfig

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font as XLFont
except ImportError:
    raise ImportError("openpyxl requerido: pip install openpyxl")

from lt_excel_reader import read_osocio_excel, find_osocio_excels, LeadRow

# Aliases ID para formularios del estándar visid (coexistencia con forms actuales)
_VISID_ID_ALIASES: dict = {
    "firstname":               "name",
    "models":                  "model",
    "model_1":                 "model",
    "model_2":                 "model",
    "estimated-date-purchase": "estimated-day",
    "estimated-date":          "estimated-day",
    "estimated_date_purchase": "estimated-day",
    # gm_front / alianzas modernas
    "telephone":               "phone",
    "cellphone":               "phone",
    "ci":                      "document",
}

# ── Rutas base ────────────────────────────────────────────────────────────────
_THIS_DIR   = os.path.dirname(os.path.abspath(__file__))   # lambdatest_mac/
if getattr(sys, 'frozen', False):
    _OSOCIO_DIR = os.path.dirname(sys.executable)
else:
    _OSOCIO_DIR = os.path.dirname(_THIS_DIR)                   # Form_Automation_Project/
_DATA_DIR       = os.path.join(_OSOCIO_DIR, "data")
_RESULTADOS_DIR = os.path.join(_OSOCIO_DIR, "resultados_lambdatestmac")
_JSON_DIR       = os.path.join(_OSOCIO_DIR, "json")        # json/ de Osocio


# ══════════════════════════════════════════════════════════════════════════════
# OPCIONES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class LTRunOptions:
    pais: str
    excel_path: str
    credentials_file: str = ""
    build_name: str = ""
    platform: str = "mac"   # "mac" o "iphone"
    brasil_docs: dict = None  # {"cpf_rows": [...], "cep_rows": [...], "cnpj_rows": [...]}


# ══════════════════════════════════════════════════════════════════════════════
# CREDENCIALES
# ══════════════════════════════════════════════════════════════════════════════

#: Variables de entorno que se leen antes del archivo. LT_USERNAME/LT_ACCESS_KEY son los
#: nombres que usa la propia documentación de LambdaTest.
_ENV_USER_VARS = ("LT_USERNAME", "LAMBDATEST_USERNAME")
_ENV_KEY_VARS = ("LT_ACCESS_KEY", "LAMBDATEST_ACCESS_KEY")


def _first_env(names):
    for n in names:
        v = (os.environ.get(n) or "").strip()
        if v:
            return v
    return ""


def load_credentials(filepath: str = ""):
    """Credenciales de LambdaTest: primero variables de entorno, si no el archivo.

    Se prioriza el entorno para poder correr sin dejar la access key escrita en claro en
    el disco (y para CI). El archivo lambdatest_credentials.txt sigue funcionando igual
    que siempre como fallback, así que no rompe las instalaciones que ya lo usan.
    """
    if not filepath:
        env_user, env_key = _first_env(_ENV_USER_VARS), _first_env(_ENV_KEY_VARS)
        if env_user and env_key:
            return env_user, env_key

    if not filepath:
        for candidate in [
            os.path.join(_THIS_DIR, "lambdatest_credentials.txt"),
            os.path.join(_OSOCIO_DIR, "lambdatest_credentials.txt"),
        ]:
            if os.path.exists(candidate):
                filepath = candidate
                break

    if not os.path.exists(filepath or ""):
        raise FileNotFoundError(
            "No se encontraron credenciales de LambdaTest.\n"
            "Opción 1 (recomendada) — variables de entorno:\n"
            "    setx LT_USERNAME tu_usuario\n"
            "    setx LT_ACCESS_KEY tu_key\n"
            "Opción 2 — archivo lambdatest_credentials.txt con:\n"
            "    username=tu_usuario\n    access_key=tu_key"
        )

    username = access_key = None
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k, v = k.strip().lower(), v.strip()
            if k == "username":
                username = v
            elif k == "access_key":
                access_key = v

    if not username or not access_key:
        raise ValueError("El archivo debe tener 'username' y 'access_key'.")
    return username, access_key


# ══════════════════════════════════════════════════════════════════════════════
# DRIVER LAMBDATEST MAC
# ══════════════════════════════════════════════════════════════════════════════

LT_HUB = "https://hub.lambdatest.com/wd/hub"


def create_lt_driver(username: str, access_key: str,
                     build_name: str = "Osocio LT",
                     test_name: str = "Form Run",
                     platform: str = "mac") -> Remote:
    """
    Crea driver LambdaTest para Mac o iPhone.
    platform: "mac"    -> macOS Sonoma + Chrome 1920x1080
              "iphone" -> iPhone 14 + Safari (real device)
    """
    _cfg = ClientConfig(remote_server_addr=LT_HUB, username=username, password=access_key)

    if platform == "iphone":
        # Web Automation — Safari en iPhone real
        # Usar ChromeOptions con browserName Safari fuerza Web Automation
        # (no App Automation). LambdaTest asigna el dispositivo disponible
        # entre iPhone 15, 16, etc.
        options = ChromeOptions()
        options.set_capability("LT:Options", {
            "username":      username,
            "accessKey":     access_key,
            "build":         build_name,
            "name":          test_name,
            "platformName":  "iOS",
            "browserName":   "Safari",
            "deviceName": "iPhone 14 Pro Max",
            "realMobile":    True,
            "visual":        True,
            "video":         True,
            "console":       False,
            "network":       False,
            "w3c":           True,
        })
        driver = Remote(command_executor=LT_HUB, options=options, client_config=_cfg)
    else:
        # Mac + Safari
        from selenium.webdriver.safari.options import Options as SafariOptions
        options = SafariOptions()
        options.set_capability("LT:Options", {
            "username":      username,
            "accessKey":     access_key,
            "build":         build_name,
            "name":          test_name,
            "platformName":  "macOS Sonoma",
            "browserName":   "Safari",
            "browserVersion":"latest",
            "visual":        False,
            "video":         True,
            "console":       False,
            "network":       True,
            "w3c":           True,
            "resolution":    "1440x900",
        })
        driver = Remote(command_executor=LT_HUB, options=options, client_config=_cfg)
        driver.maximize_window()

    driver.implicitly_wait(0)
    return driver


def _safe_log(log: Callable) -> Callable:
    """
    Blinda el log de la corrida. Los mensajes usan Unicode (✓ ✗ ⚠ 🎲 ⛔ →) y el destino
    no siempre lo soporta: una consola Windows cp1252 tira UnicodeEncodeError, y el .exe
    compilado con console=False deja sys.stdout en None. Sin esto, el primer log mata la
    corrida ANTES de crear el driver (la sesión nunca llega a LambdaTest).
    Reintenta en ASCII y, si aun así falla, descarta el mensaje: loguear nunca puede
    cortar la ejecución.
    """
    def _l(msg=""):
        try:
            log(msg)
        except UnicodeEncodeError:
            try:
                log(str(msg).encode("ascii", "replace").decode("ascii"))
            except Exception:
                pass
        except Exception:
            pass
    return _l


def mark_lt_status(driver, passed: bool = True):
    try:
        driver.execute_script(f"lambda-status={'passed' if passed else 'failed'}")
    except Exception:
        pass




# ══════════════════════════════════════════════════════════════════════════════
# MANEJO DE COOKIES (igual que Osocio)
# ══════════════════════════════════════════════════════════════════════════════

def _handle_gm_cookie_popup(driver, log: Callable = print) -> bool:
    """Maneja el popup de cookies GM (gb-legal-notification). Igual que Osocio."""
    # Intento 1: botón dentro del Shadow DOM del web component
    try:
        clicked = driver.execute_script("""
            var host = document.querySelector('gb-legal-notification');
            if (!host) return false;
            var root = host.shadowRoot || host;
            var btn = root.querySelector('.close-btn, .js-close-icon, .silent-consent, button[aria-label*="close"], button[aria-label*="Close"]');
            if (!btn) {
                // Fallback: cualquier botón dentro del shadow root
                var btns = root.querySelectorAll('button');
                for (var i = 0; i < btns.length; i++) {
                    var t = (btns[i].textContent || '').toLowerCase();
                    var cl = btns[i].className || '';
                    if (t.indexOf('aceitar') >= 0 || t.indexOf('accept') >= 0 || t.indexOf('ok') >= 0
                        || cl.indexOf('close') >= 0 || cl.indexOf('accept') >= 0) {
                        btn = btns[i]; break;
                    }
                }
            }
            if (btn) { btn.click(); return true; }
            return false;
        """)
        if clicked:
            time.sleep(0.5)
            log("  ✓ Cookie popup GM cerrado (shadow DOM)")
            return True
    except Exception:
        pass

    # Intento 2: botones en DOM normal
    for selector in [
        ".close-btn.js-close-icon.silent-consent",
        ".js-close-icon",
        ".silent-consent",
        "button[onclick*='cookie']",
        "button[class*='cookie-accept']",
        "button[id*='cookie-accept']",
        ".cookie-accept",
        "#cookie-accept",
    ]:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            for el in elements:
                if el.is_displayed():
                    driver.execute_script("arguments[0].click();", el)
                    time.sleep(0.5)
                    log("  ✓ Cookie popup cerrado")
                    return True
        except Exception:
            continue
    return False


def _handle_cookie_popups(driver, log: Callable = print) -> bool:
    """Maneja todos los popups de cookies. Igual que Osocio."""
    return _handle_gm_cookie_popup(driver, log)


# ══════════════════════════════════════════════════════════════════════════════
# SCROLL DINÁMICO (igual que Osocio: pre_scroll_for_dynamic_content)
# ══════════════════════════════════════════════════════════════════════════════

def _pre_scroll_for_dynamic_content(driver):
    """Scroll progresivo para activar contenido dinámico. Igual que Osocio."""
    total_height = driver.execute_script("return document.body.parentNode.scrollHeight")
    viewport_height = driver.execute_script("return window.innerHeight")
    current_pos = 0
    scroll_step = viewport_height * 0.8

    while current_pos < total_height:
        driver.execute_script(f"window.scrollTo(0, {current_pos});")
        time.sleep(0.1)
        current_pos += scroll_step
        if current_pos > total_height * 3:
            break

    driver.execute_script("window.scrollTo(0, document.body.parentNode.scrollHeight);")
    time.sleep(0.2)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.2)


# ══════════════════════════════════════════════════════════════════════════════
# PLACEHOLDERS Y OPCIONES VÁLIDAS
# ══════════════════════════════════════════════════════════════════════════════

_PLACEHOLDER_KW = (
    "seleccione", "selecciona", "seleccionar", "seleccion",
    "selecione", "selecionar", "selecao", "escolha", "escolher",
    "elija", "elegir", "opcao", "opciones", "opcion",
    "porfavor", "favor", "obrigatorio", "required",
    "select", "choose", "please",
)


# Mapeo flexible de encabezados Excel → IDs HTML (compartido por el llenado y por la
# comparación "pedido en el Excel" vs "lo que realmente quedó en el form")
_HEADER_TO_IDS = {
    "nombre":   ["firstname", "name"],
    "nombres":  ["firstname", "name"],
    "apellido": ["lastname"],
    "apellidos":["lastname"],
    "documento":["ci", "rut", "cpf", "document", "cnpj", "dni"],
    "celular":  ["cellphone", "telephone", "telephone-mask", "phone"],
    "teléfono": ["cellphone", "telephone", "telephone-mask", "phone"],
    "telefono": ["cellphone", "telephone", "telephone-mask", "phone"],
    "email":    ["email"],
    "correo":   ["email"],
    "modelo":   ["models", "model"],
    # Campos de elección: el Excel que genera 'Generar Excels' trae varias columnas
    # equivalentes ("Fecha estimada", "Fecha estimada de compra", "Fecha de compra") y el
    # id real del form cambia según la versión (estimated-date-purchase / estimated-day).
    # Sin estos alias, el valor cargado a mano en la columna "de más" se ignoraba y el
    # campo terminaba sorteado.
    "fecha estimada":            ["estimated-date-purchase", "estimated-day", "estimated-date"],
    "fecha estimada de compra":  ["estimated-date-purchase", "estimated-day", "estimated-date"],
    "fecha de compra":           ["estimated-date-purchase", "estimated-day", "estimated-date"],
    "región":                    ["region", "state"],
    "region":                    ["region", "state"],
    "ciudad":                    ["city"],
    "concesionario":             ["dealer"],
    "evento":                    ["event", "event-name"],
    "tipo de documento":         ["document-type"],
    "tipo de documento (perú)":  ["document-type"],
    "color":                     ["color"],
    "kit":                       ["kit"],
    "seguro":                    ["insurance"],
    "versión":                   ["version"],
    "version":                   ["version"],
}


def _is_placeholder(text: str) -> bool:
    t = (text or "").lower().strip()
    if not t or "--" in t:
        return True
    return any(kw in t for kw in _PLACEHOLDER_KW)


def _get_valid_options(select_el) -> List:
    try:
        return [o for o in Select(select_el).options
                if not _is_placeholder(o.text) and not o.get_attribute("disabled")]
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
# IDS DINÁMICOS Y DEPENDENCIAS DE OSOCIO
# ══════════════════════════════════════════════════════════════════════════════

def _load_ids_dinamicos(pais: str) -> Dict[str, str]:
    """Lee json/ids_dinamicos.json de Osocio. Primera entrada por ID gana."""
    path = os.path.join(_JSON_DIR, "ids_dinamicos.json")
    result = {}
    if not os.path.exists(path):
        return result
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for entry in data.get("entries", []):
            eid   = str(entry.get("id")    or "").strip()
            valor = str(entry.get("valor") or "").strip()
            paises = entry.get("paises") or []
            if not eid or not valor:
                continue
            if not paises or pais in paises:
                if eid not in result:
                    result[eid] = valor
    except Exception:
        pass
    return result


def _load_field_dependencies() -> Dict[str, str]:
    """Dependencias hardcoded + ids_dinamicos.json. Retorna {padre: hijo}."""
    deps = {
        "region": "city",
        "city": "dealer",
        "models": "kits[]",
        "model": "kits[]",
    }
    path = os.path.join(_JSON_DIR, "ids_dinamicos.json")
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for dep in data.get("dependencies", []):
                padre = str(dep.get("padre") or "").strip()
                hijo  = str(dep.get("hijo")  or "").strip()
                if padre and hijo:
                    deps[padre] = hijo
        except Exception:
            pass
    return deps


# ══════════════════════════════════════════════════════════════════════════════
# INGRESO DE TEXTO VÍA JS (motor del Proyecto 2, más confiable en remoto)
# ══════════════════════════════════════════════════════════════════════════════

def _fill_text_js(driver, element, value: str):
    """Una sola llamada de red. Usa native setter para React/Angular + dispara eventos."""
    driver.execute_script(
        "var el = arguments[0], val = arguments[1];"
        "el.focus();"
        "try {"
        "  var proto = el.tagName === 'TEXTAREA' ? window.HTMLTextAreaElement.prototype : window.HTMLInputElement.prototype;"
        "  Object.getOwnPropertyDescriptor(proto, 'value').set.call(el, val);"
        "} catch(e) { el.value = val; }"
        "el.dispatchEvent(new Event('input',  {bubbles:true}));"
        "el.dispatchEvent(new Event('change', {bubbles:true}));"
        "el.dispatchEvent(new Event('blur',   {bubbles:true}));",
        element, str(value),
    )


# Campos con máscara de jquery-mask (CPF/CNPJ/CEP/teléfono): la máscara se reformatea en
# cada tecla, así que estos se siguen tipeando carácter a carácter. El resto va en un solo
# send_keys: cada send_keys es un round-trip HTTP a LambdaTest, y letra por letra un email
# de 34 chars tardaba ~23s contra un device real.
_MASKED_FIELD_HINTS = ("cpf", "cnpj", "cep", "telephone", "phone", "celular", "zip", "postal")


def _is_masked_field(element) -> bool:
    try:
        ident = ((element.get_attribute("id") or "") + " " +
                 (element.get_attribute("name") or "")).lower()
    except Exception:
        return True  # ante la duda, el modo lento (seguro)
    return any(h in ident for h in _MASKED_FIELD_HINTS)


def _send_keys_fast(element, value: str):
    """send_keys en 1 sola llamada, salvo campos enmascarados (char a char)."""
    if _is_masked_field(element):
        for char in str(value):
            try:
                element.send_keys(char)
            except Exception:
                pass
            time.sleep(0.005)
        return
    try:
        element.send_keys(str(value))
    except Exception:
        pass


def _fill_text_sendkeys(driver, element, value: str):
    """
    Re-ingresa el input vía send_keys real, en vez de setear .value por JS.
    Usado como método PRIMARIO en Mac/Safari: el fill por JS (value + dispatchEvent)
    puede mostrar el valor en el DOM sin que el estado interno de React lo registre
    ahí — send_keys pasa por el pipeline real de eventos de teclado y evita ese desync.
    """
    # Si el campo no queda realmente vacío, send_keys concatena sobre lo que había
    # ('ApellidoApellido', email duplicado) y el lead viaja corrupto → en ese caso se
    # setea el valor por JS en vez de tipear.
    if not _hard_clear_input(driver, element, log=lambda *_a, **_k: None):
        _set_input_value_js(driver, element, str(value))
        return
    try:
        driver.execute_script("arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", element)
    except Exception:
        pass
    _send_keys_fast(element, value)
    try:
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
            "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
            element,
        )
    except Exception:
        pass


def _fill_required_synthetic_lt(driver, log=print):
    """Rellena con valores sintéticos aleatorios los campos REQUERIDOS visibles que
    quedaron vacíos (ej. Rua, Número, Data de Nascimento). Una sola llamada JS (rápido en LT).
    Fecha → DD/MM/AAAA >18; número → dígitos; email → @mrm.com; resto → alfanumérico."""
    try:
        filled = driver.execute_script(r"""
            function ri(a,b){return Math.floor(Math.random()*(b-a+1))+a;}
            function rs(n){var c='abcdefghijklmnopqrstuvwxyz0123456789',s='';for(var i=0;i<n;i++)s+=c[ri(0,c.length-1)];return s;}
            function pad(n){return (n<10?'0':'')+n;}
            var out=[];
            var els=document.querySelectorAll('input:not([type=hidden]):not([type=submit]):not([type=button]):not([type=checkbox]):not([type=radio]):not([type=file]):not([type=reset]),textarea');
            for(var i=0;i<els.length;i++){
                var el=els[i];
                if(!el.offsetParent||el.disabled) continue;
                if((el.value||'').trim()) continue;
                var req=el.required||el.getAttribute('aria-required')==='true';
                if(!req){var id=el.id; if(id){var l=document.querySelector('label[for="'+((window.CSS&&CSS.escape)?CSS.escape(id):id)+'"]'); if(l&&(l.textContent||'').indexOf('*')>=0)req=true;} var p=el.closest('label'); if(p&&(p.textContent||'').indexOf('*')>=0)req=true;}
                if(!req) continue;
                var hints=((el.id||'')+' '+(el.name||'')+' '+(el.placeholder||'')+' '+(el.getAttribute('aria-label')||'')).toLowerCase();
                var type=(el.getAttribute('type')||'text').toLowerCase();
                var dateH=['nascimento','nacimiento','birth','dob','aniversario','fecha','data-de','cumple'];
                var isDate=type==='date'||dateH.some(function(h){return hints.indexOf(h)>=0;});
                var numH=['numero','número','number','quantidade','cantidad'];
                var val;
                if(isDate){var y=ri(1975,2003),m=ri(1,12),d=ri(1,28); val=(type==='date')?(y+'-'+pad(m)+'-'+pad(d)):(pad(d)+'/'+pad(m)+'/'+y);}
                else if(type==='email'){val='auto'+rs(6)+ri(10,99)+'@mrm.com';}
                else if(type==='number'||type==='tel'||numH.some(function(h){return hints.indexOf(h)>=0;})){val=''+ri(1,9999);}
                else{val=rs(8);}
                try{var proto=el.tagName==='TEXTAREA'?window.HTMLTextAreaElement.prototype:window.HTMLInputElement.prototype; el.removeAttribute('maxlength'); Object.getOwnPropertyDescriptor(proto,'value').set.call(el,val);}catch(e){el.value=val;}
                el.dispatchEvent(new Event('input',{bubbles:true}));
                el.dispatchEvent(new Event('change',{bubbles:true}));
                el.dispatchEvent(new Event('blur',{bubbles:true}));
                out.push((el.id||el.name||'?')+'='+val);
            }
            return out;
        """) or []
        for f in filled:
            log(f"  🧩 Requerido sin dato completado (sintético): {f}")
        return len(filled)
    except Exception as e:
        log(f"  ⚠ Auto-relleno sintético: {e}")
        return 0


def _fill_text_android(driver, element, value: str):
    """
    Ingreso de texto para Android real device vía send_keys real (teclado real)
    — igual que Mac/Safari. Más confiable que el fill por JS para que React/Angular
    registre el valor correctamente.
    """
    try:
        element.click()
    except Exception:
        try:
            driver.execute_script("arguments[0].click();", element)
        except Exception:
            pass
    # En Android el clear() falla en silencio con campos enmascarados o de React: si no
    # queda vacío, send_keys concatena y se envía el valor duplicado.
    if not _hard_clear_input(driver, element, log=lambda *_a, **_k: None):
        _set_input_value_js(driver, element, str(value))
        return
    _send_keys_fast(element, value)
    try:
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
            "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
            element,
        )
    except Exception:
        pass


def _fill_text_js_mobile(driver, element, value: str):
    """
    Ingreso de texto para Safari iOS.
    Usa native setter + touch events para que Safari móvil registre la interacción.
    """
    driver.execute_script(
        "const el = arguments[0], val = arguments[1];"
        "el.focus();"
        "try {"
        "  var proto = el.tagName === 'TEXTAREA' ? window.HTMLTextAreaElement.prototype : window.HTMLInputElement.prototype;"
        "  Object.getOwnPropertyDescriptor(proto, 'value').set.call(el, val);"
        "} catch(e) { el.value = val; }"
        "el.dispatchEvent(new Event('touchstart', {bubbles:true}));"
        "el.dispatchEvent(new Event('touchend',   {bubbles:true}));"
        "el.dispatchEvent(new Event('input',      {bubbles:true}));"
        "el.dispatchEvent(new Event('change',     {bubbles:true}));"
        "el.dispatchEvent(new Event('blur',       {bubbles:true}));",
        element, str(value),
    )


def _tap_and_type_native(driver, iframe_el, field_id: str,
                          value: str, log: Callable = print) -> bool:
    """
    Para iOS Safari cross-origin: toca el campo por coordenadas en pantalla,
    cambia a NATIVE_APP y escribe con send_keys nativo de XCUITest.

    1. Calcula la posición del campo dentro del iframe en pantalla
    2. Tap en esas coordenadas (abre el campo y el teclado)
    3. Cambia a NATIVE_APP
    4. Busca el TextField con foco y escribe
    5. Vuelve al WEBVIEW
    """
    try:
        # Obtener posición del campo en pantalla via JS
        pos = driver.execute_script("""
            const iframe = arguments[0];
            const fieldId = arguments[1];
            const iRect = iframe.getBoundingClientRect();
            // No podemos acceder al DOM del iframe, usamos posicion aproximada
            // basada en la posicion del iframe en pantalla
            return {
                iframeLeft:   iRect.left,
                iframeTop:    iRect.top,
                iframeWidth:  iRect.width,
                iframeHeight: iRect.height,
                viewportW:    window.innerWidth,
                viewportH:    window.innerHeight,
            };
        """, iframe_el, field_id)

        cx = pos['iframeLeft'] + pos['iframeWidth'] / 2
        # Y aproximada — el campo estará en algún punto del iframe visible
        cy = max(pos['iframeTop'] + 50, pos['viewportH'] / 2)

        log(f"  📱 tap_and_type '{field_id}' en ({cx:.0f}, {cy:.0f})")

        # Tap en coordenadas
        try:
            driver.execute_script("mobile: tap", {"x": cx, "y": cy})
        except Exception:
            from selenium.webdriver.common.action_chains import ActionChains
            from selenium.webdriver.common.actions.action_builder import ActionBuilder
            from selenium.webdriver.common.actions.pointer_input import PointerInput
            from selenium.webdriver.common.actions import interaction
            actions = ActionChains(driver)
            actions.w3c_actions = ActionBuilder(
                driver, mouse=PointerInput(interaction.POINTER_TOUCH, "touch"))
            actions.w3c_actions.pointer_action.move_to_location(int(cx), int(cy))
            actions.w3c_actions.pointer_action.click()
            actions.perform()
        time.sleep(1)

        # Cambiar a NATIVE_APP para send_keys nativo
        contexts = driver.contexts
        driver.switch_to.context("NATIVE_APP")

        # Buscar campo con foco
        focused = driver.find_elements(By.XPATH,
            '//XCUIElementTypeTextField[@focused="true"] | '
            '//XCUIElementTypeSecureTextField[@focused="true"]')
        if not focused:
            focused = driver.find_elements(By.XPATH,
                '//XCUIElementTypeTextField | //XCUIElementTypeTextView')

        if focused:
            focused[0].send_keys(str(value))
            log(f"  ✓ '{field_id}' escrito via NATIVE_APP send_keys")
        else:
            log(f"  ⚠ Sin campo nativo con foco para '{field_id}'")

        # Volver al WEBVIEW
        webviews = [c for c in contexts if c != "NATIVE_APP"]
        if webviews:
            driver.switch_to.context(webviews[0])
        time.sleep(0.3)
        return bool(focused)

    except Exception as e:
        log(f"  ✗ tap_and_type '{field_id}': {e}")
        # Asegurar que volvemos al WEBVIEW
        try:
            contexts = driver.contexts
            webviews = [c for c in contexts if c != "NATIVE_APP"]
            if webviews:
                driver.switch_to.context(webviews[0])
        except Exception:
            pass
        return False


def _scroll_to(driver, element, known_iframe=None):
    if not element:
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", element)
        time.sleep(0.1)
    except Exception:
        pass

    try:
        # Comprobar si estamos dentro de un iframe
        in_iframe = driver.execute_script("return window.self !== window.top;")
        if in_iframe:
            rect = driver.execute_script(
                "const r = arguments[0].getBoundingClientRect(); return {top: r.top, height: r.height};",
                element
            )
            element_top_in_iframe = rect["top"]
            element_height = rect["height"]

            # Cambiar temporalmente al contenido principal
            driver.switch_to.default_content()

            try:
                # Usar el iframe CONOCIDO del form (la landing puede tener varios iframes de
                # marketing y agarrar el primero visible scrollea al lugar equivocado).
                target_iframe = known_iframe
                if target_iframe is None:
                    iframes = driver.find_elements(By.TAG_NAME, "iframe")
                    for iframe in iframes:
                        if iframe.is_displayed():
                            target_iframe = iframe
                            break

                if target_iframe:
                    # Obtener la posición del iframe en el parent
                    iframe_rect = driver.execute_script(
                        "const r = arguments[0].getBoundingClientRect(); return {top: r.top, yOffset: window.pageYOffset};",
                        target_iframe
                    )
                    iframe_top = iframe_rect["top"]
                    parent_y_offset = iframe_rect["yOffset"]

                    viewport_height = driver.execute_script("return window.innerHeight;")
                    
                    target_y_on_parent = parent_y_offset + iframe_top + element_top_in_iframe
                    # Centrar el elemento en la pantalla
                    scroll_y = target_y_on_parent - (viewport_height / 2) + (element_height / 2)
                    scroll_y = max(0, scroll_y)

                    driver.execute_script(f"window.scrollTo(0, {scroll_y});")
                    time.sleep(0.15)
            finally:
                # Siempre volver al iframe
                if target_iframe:
                    driver.switch_to.frame(target_iframe)
    except Exception:
        pass


def _scroll_to_mobile(driver, element):
    """
    Scroll para Safari iOS.
    scrollIntoView puede no funcionar en iOS Safari — usamos
    window.scrollTo calculando la posición del elemento.
    """
    try:
        driver.execute_script("""
            const el = arguments[0];
            const rect = el.getBoundingClientRect();
            const scrollTop = window.pageYOffset || document.documentElement.scrollTop;
            const targetY = rect.top + scrollTop - (window.innerHeight / 2);
            window.scrollTo({ top: Math.max(0, targetY), behavior: 'smooth' });
        """, element)
        time.sleep(0.3)
    except Exception:
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", element
            )
        except Exception:
            pass


def _get_maxlength(element) -> Optional[int]:
    try:
        ml = element.get_attribute("maxlength")
        if ml:
            n = int(ml)
            if n > 0:
                return n
    except Exception:
        pass
    return None


# ══════════════════════════════════════════════════════════════════════════════
# SELECTS CON DEPENDENCIAS (igual que Osocio: polling activo)
# ══════════════════════════════════════════════════════════════════════════════

_DEP_TIMEOUT  = 6.0
_DEP_POLL     = 0.075
_DEP_RETRIES  = 2


_VALID_CEPS = [
    "01310100", "04538133", "20040020", "30112010",
    "40015970", "60060100", "80010010", "90010280",
    "69010060", "74805100",
]


def _refill_brasil_doc_sendkeys(driver, field_mapping: list, pais: str,
                                 brasil_doc_type: str, all_tracked: dict,
                                 log) -> bool:
    """
    Re-ingresa el doc Brasil (CPF/CNPJ/CEP) carácter a carácter via send_keys,
    usando el MISMO valor que ya estaba en el Excel (no genera uno nuevo).
    Las máscaras JS (imask, jQuery Mask, etc.) necesitan keystrokes reales.
    Retorna True si logró re-llenar al menos un campo.
    """
    if pais.lower() not in ("brasil", "brazil", "br"):
        return False

    doc_keywords = ("cpf", "cnpj", "cep", "zip", "postal", "document", "ci")
    rellenados = 0
    for fc in field_mapping:
        fname   = fc.get("name", "")
        fid_raw = fc.get("id", "")
        fids    = fid_raw if isinstance(fid_raw, list) else [fid_raw]
        for fid in fids:
            if not fid or not any(k in str(fid).lower() for k in doc_keywords):
                continue
            # Usar el valor que ya se ingresó (del Excel), sin generar uno nuevo
            valor = all_tracked.get(fname) or all_tracked.get(fid) or ""
            if not valor:
                continue
            try:
                els = driver.find_elements(By.ID, fid)
                if not els or not els[0].is_displayed():
                    continue
                el = els[0]
                # Limpiar y re-ingresar carácter a carácter (si no queda vacío, el
                # send_keys concatenaría y el documento viajaría duplicado)
                if not _hard_clear_input(driver, el, log=log):
                    log(f"  ⚠ No se pudo vaciar '{fid}' — se omite el re-ingreso")
                    continue
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", el
                )
                time.sleep(0.05)
                for char in str(valor):
                    el.send_keys(char)
                    time.sleep(0.005)
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
                    "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
                    el,
                )
                log(f"  ↺ Re-ingresado '{fid}' via send_keys (mismo valor): '{valor}'")
                rellenados += 1
            except Exception as e:
                log(f"  ✗ Error send_keys retry '{fid}': {e}")
    return rellenados > 0


def _tiene_thankyou_texto_2_0(driver) -> bool:
    """TY de forms 2.0 (/tools/forms): confirmación de envío.
    No basta el texto (el encabezado '...e entraremos em contato' daría falso
    positivo): sólo es TY si el botón Enviar ya NO está visible (form reemplazado)
    y aparece una frase fuerte de confirmación.
    """
    try:
        _cur = (driver.current_url or "").lower()
    except Exception:
        _cur = ""
    if "/tools/" not in _cur:
        return False
    try:
        _submit_visible = driver.execute_script("""
            var els = document.querySelectorAll("button, input[type='submit'], input[type='button']");
            for (var i=0;i<els.length;i++){
                var e=els[i];
                var t=((e.innerText||e.value||"")+"").trim().toLowerCase();
                if(e.offsetParent!==null && (t==='enviar' || t.indexOf('enviar')===0)){ return true; }
            }
            return false;
        """)
        if _submit_visible:
            return False
    except Exception:
        pass
    try:
        _txt = driver.execute_script(
            "return (document.body && document.body.innerText) ? document.body.innerText : '';"
        ) or ""
    except Exception:
        _txt = ""
    import unicodedata as _ud
    _txt = _txt.lower()
    _txt_noaccent = "".join(c for c in _ud.normalize("NFKD", _txt) if not _ud.combining(c))
    _markers = (
        "obrigado", "obrigada",
        "recebemos sua solicit", "recebemos seu contato",
        "sua solicitação foi recebida", "solicitação enviada",
        "recebido com sucesso", "recebida com sucesso",
        "enviado com sucesso", "cadastro realizado",
        "dados enviados", "mensagem enviada",
        "gracias por", "thank you for",
        "sucesso", "exito", "éxito", "success",
    )
    return any(m in _txt or m in _txt_noaccent for m in _markers)


def _investigate_ty_cta(driver, log=print):
    """Wrapper del módulo compartido utils/ty_cta. En LambdaTest (Mac/Android) NO se
    guardan capturas — el click al CTA y la landing quedan grabados en el video."""
    if _OSOCIO_DIR not in sys.path:
        sys.path.insert(0, _OSOCIO_DIR)
    from utils.ty_cta import investigate_ty_cta
    return investigate_ty_cta(driver, log=log, take_screenshot=False)


def _format_ty_cta(info):
    if _OSOCIO_DIR not in sys.path:
        sys.path.insert(0, _OSOCIO_DIR)
    from utils.ty_cta import format_ty_cta
    return format_ty_cta(info)


def _format_link_issue(info):
    if _OSOCIO_DIR not in sys.path:
        sys.path.insert(0, _OSOCIO_DIR)
    from utils.ty_cta import format_link_issue
    return format_link_issue(info)


def _describir_errores_visuales(driver) -> str:
    """Devuelve todos los errores visuales con el campo asociado: 'campo: mensaje, ...'"""
    try:
        errores = driver.find_elements(
            By.XPATH,
            "//*[contains(@class,'error') or contains(@style,'red') or contains(@class,'invalid')]"
        )
        errores_visibles = [e for e in errores if e.is_displayed() and e.text.strip()]
    except Exception:
        return ""
    if not errores_visibles:
        return ""
    partes = []
    vistos = set()
    for err_el in errores_visibles:
        texto = err_el.text.strip()
        if texto in vistos:
            continue
        vistos.add(texto)
        campo = ""
        try:
            campo = driver.execute_script("""
                var el = arguments[0];
                var parent = el.parentElement;
                for (var i = 0; i < 4 && parent; i++) {
                    var inp = parent.querySelector('input,select,textarea');
                    if (inp) { return inp.id || inp.name || inp.getAttribute('placeholder') || ''; }
                    var lbl = parent.querySelector('label');
                    if (lbl && lbl.textContent.trim()) { return lbl.textContent.trim(); }
                    parent = parent.parentElement;
                }
                return '';
            """, err_el) or ""
        except Exception:
            pass
        partes.append(f'{campo}: "{texto}"' if campo else f'"{texto}"')
    return ", ".join(partes)


def _fetch_4devs(acao: str, extra_params: dict = None, timeout: int = 8) -> str:
    import urllib.request, urllib.parse
    data = {"acao": acao}
    if extra_params:
        data.update(extra_params)
    body = urllib.parse.urlencode(data).encode()
    req = urllib.request.Request(
        "https://www.4devs.com.br/ferramentas_online.php",
        data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded",
                 "User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8").strip()


def _generate_valid_cpf() -> str:
    """Fallback: CPF con dígitos verificadores válidos (sin puntos)."""
    while True:
        d = [random.randint(0, 9) for _ in range(9)]
        if len(set(d)) == 1:
            continue
        r1 = sum(v * (10 - i) for i, v in enumerate(d)) % 11
        c1 = 0 if r1 < 2 else 11 - r1
        d2 = d + [c1]
        r2 = sum(v * (11 - i) for i, v in enumerate(d2)) % 11
        c2 = 0 if r2 < 2 else 11 - r2
        return "".join(map(str, d + [c1, c2]))


def _generate_valid_cnpj() -> str:
    """Fallback: CNPJ con puntuación (XX.XXX.XXX/XXXX-XX)."""
    b = [random.randint(0, 9) for _ in range(8)] + [0, 0, 0, 1]
    w1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    r1 = sum(v * w for v, w in zip(b, w1)) % 11
    c1 = 0 if r1 < 2 else 11 - r1
    w2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    r2 = sum(v * w for v, w in zip(b + [c1], w2)) % 11
    c2 = 0 if r2 < 2 else 11 - r2
    digits = "".join(map(str, b + [c1, c2]))
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"


def _cpf_checksum_ok(digits: str) -> bool:
    """Valida el dígito verificador real de un CPF de 11 dígitos."""
    if len(digits) != 11 or len(set(digits)) == 1 or not digits.isdigit():
        return False
    d = [int(c) for c in digits]
    r1 = sum(v * (10 - i) for i, v in enumerate(d[:9])) % 11
    c1 = 0 if r1 < 2 else 11 - r1
    if c1 != d[9]:
        return False
    r2 = sum(v * (11 - i) for i, v in enumerate(d[:10])) % 11
    c2 = 0 if r2 < 2 else 11 - r2
    return c2 == d[10]


def _cnpj_checksum_ok(digits: str) -> bool:
    """Valida el dígito verificador real de un CNPJ de 14 dígitos."""
    if len(digits) != 14 or len(set(digits)) == 1 or not digits.isdigit():
        return False
    d = [int(c) for c in digits]
    w1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    r1 = sum(v * w for v, w in zip(d[:12], w1)) % 11
    c1 = 0 if r1 < 2 else 11 - r1
    if c1 != d[12]:
        return False
    w2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    r2 = sum(v * w for v, w in zip(d[:12] + [c1], w2)) % 11
    c2 = 0 if r2 < 2 else 11 - r2
    return c2 == d[13]


def _generate_brazil_document(field_id: str) -> str:
    """
    Genera CPF (sin puntos), CNPJ (con puntos) o CEP (sin puntos) SIEMPRE vía la API
    real de 4devs — nunca con el generador local sintético (checksum válido pero no
    es un documento real, y el form puede rechazarlo). Si la API falla tras varios
    reintentos, devuelve "" — el campo queda sin llenar y el error de validación
    real queda registrado en el Excel de resultados, en vez de disfrazarlo con un
    doc falso.
    """
    import re as _re
    fid = field_id.lower()
    is_cep  = "cep" in fid or "zip" in fid or "postal" in fid
    is_cnpj = "cnpj" in fid

    if is_cep:
        for _attempt in range(5):
            try:
                html = _fetch_4devs("gerar_cep", {
                    "estado": "", "cidade": "São Paulo",
                    "bairro": "", "tipo_cep": "residencial",
                })
                m = _re.search(r'(\d{5}-\d{3})', html)
                if m:
                    return m.group(1).replace("-", "")
            except Exception:
                pass
            time.sleep(0.3)
        return ""

    if is_cnpj:
        for _attempt in range(5):
            try:
                raw = _fetch_4devs("gerar_cnpj", {"pontuacao": "S"})
                # Extraer el patrón exacto (evita ruido de la respuesta) y validar el
                # dígito verificador real — la respuesta a veces trae basura y "al menos
                # 14 dígitos en cualquier lado" podía armar un CNPJ inválido.
                m = _re.search(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})', raw or "")
                if m:
                    formatted = m.group(1)
                    digits = "".join(c for c in formatted if c.isdigit())
                    if _cnpj_checksum_ok(digits):
                        return formatted
            except Exception:
                pass
            time.sleep(0.3)
        return ""

    # CPF — solo dígitos
    for _attempt in range(5):
        try:
            raw = _fetch_4devs("gerar_cpf", {"pontuacao": "N"})
            m = _re.search(r'(\d{11})', raw or "")
            if m and _cpf_checksum_ok(m.group(1)):
                return m.group(1)
        except Exception:
            pass
        time.sleep(0.3)
    return ""


def _sanitize_peru_document(doc_type_value: str, raw_value: str) -> str:
    """Corrige el número de documento según las reglas de Perú."""
    import random as _random
    dt = (doc_type_value or "").lower()
    if "dni" in dt:
        required_len, no_leading_zero = 8, True
    elif "ruc" in dt:
        required_len, no_leading_zero = 11, False
    elif "pasaporte" in dt:
        # El pasaporte es ALFANUMÉRICO: se conservan las letras del valor original y se
        # completa con caracteres alfanuméricos, no con dígitos.
        alnum = "".join(c for c in str(raw_value or "") if c.isalnum()).upper()
        _pool = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
        while len(alnum) < 12:
            alnum += _random.choice(_pool)
        return alnum[:12]
    elif "carn" in dt or "extran" in dt:
        required_len, no_leading_zero = 12, False
    else:
        return raw_value

    digits = "".join(c for c in str(raw_value or "") if c.isdigit())
    while len(digits) < required_len:
        digits += str(_random.randint(0, 9))
    digits = digits[:required_len]
    if no_leading_zero and digits[0] == "0":
        digits = str(_random.randint(1, 9)) + digits[1:]
    return digits


_VALID_OPTS_JS = (
    "var e=document.getElementById(arguments[0]);"
    "if(!e||!e.getClientRects().length||e.disabled)return false;"
    "return Array.from(e.options).some(function(o){"
    "return o.value&&o.value!=='0'&&!o.disabled&&o.text&&o.text.trim()!=='';"
    "});"
)

def _wait_for_dropdown_options(driver, select_id: str,
                                timeout: float = _DEP_TIMEOUT) -> bool:
    """Polling activo con un solo JS call por tick (en vez de 4 round-trips)."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            if driver.execute_script(_VALID_OPTS_JS, select_id):
                return True
        except Exception:
            pass
        time.sleep(_DEP_POLL)
    return False


_SELECT_JS = r"""
var id=arguments[0], val=arguments[1], isEmpty=arguments[2];
var el=document.getElementById(id);
if(!el)return{ok:false,r:'notfound'};
var cs=getComputedStyle(el);
// getClientRects()==0 tambien cuando un ANCESTRO esta display:none (paso oculto del wizard);
// getComputedStyle(el).display del hijo NO refleja eso.
if(!el.getClientRects().length||cs.visibility==='hidden')return{ok:false,r:'hidden'};
el.removeAttribute('disabled');el.disabled=false;
el.scrollIntoView({block:'center',behavior:'instant'});
var bad=/^(selec|eligi|choos|escolh|--)/i;
var opts=Array.from(el.options).filter(function(o){
    return o.value&&o.value!=='0'&&!o.disabled&&o.text&&o.text.trim()!==''&&!bad.test(o.text.trim());
});
if(!opts.length)return{ok:false,r:'noopts'};
// Normalizar ambos lados antes de comparar: los forms escriben las opciones con &nbsp;
// y acentos ("1 mes", "Más de 6 meses") y el Excel con espacios normales, asi que la
// comparacion cruda fallaba y caia al random, pisando lo que pidio el usuario.
var norm=function(s){return String(s||'').replace(/ /g,' ').normalize('NFD')
    .replace(/[̀-ͯ]/g,'').toLowerCase().replace(/\s+/g,' ').trim();};
var target=null;
if(!isEmpty&&val){
    var vl=norm(val);
    for(var i=0;i<opts.length;i++){if(norm(opts[i].text)===vl){target=opts[i];break;}}
    if(!target)for(var i=0;i<opts.length;i++){if(norm(opts[i].text).indexOf(vl)>=0){target=opts[i];break;}}
    if(!target)for(var i=0;i<opts.length;i++){var ot=norm(opts[i].text);if(ot&&vl.indexOf(ot)>=0){target=opts[i];break;}}
    if(!target)for(var i=0;i<opts.length;i++){if(norm(opts[i].value)===vl){target=opts[i];break;}}
}
var matched=!!target;
if(!target)target=opts[Math.floor(Math.random()*opts.length)];
try{var s=Object.getOwnPropertyDescriptor(HTMLSelectElement.prototype,'value');if(s&&s.set)s.set.call(el,target.value);else el.value=target.value;}catch(e){el.value=target.value;}
el.dispatchEvent(new Event('focus',{bubbles:true}));
el.dispatchEvent(new Event('change',{bubbles:true,cancelable:true}));
el.dispatchEvent(new Event('input',{bubbles:true}));
el.dispatchEvent(new Event('blur',{bubbles:true}));
return{ok:true,text:target.text,matched:matched,wasRandom:!matched};
"""

def _scroll_to_bottom_parent_aware(driver, log: Callable = print):
    """
    Centra el botón Enviar (o el checkbox de términos) en el viewport en vez de saltar
    ciegamente al fondo de la página — en páginas cortas (form pegado al footer) un salto
    fijo se pasa de largo y deja el footer en pantalla, sin el form a la vista.
    Si no encuentra ningún ancla, hace un scroll incremental acotado (no directo al fondo)
    y, si aun así queda posicionado fuera del form, vuelve a centrarlo.
    """
    try:
        anchored = driver.execute_script(
            """
            var sels = arguments[0];
            var el = null;
            for (var i=0;i<sels.length;i++){
                var e = document.querySelector(sels[i]);
                if (e && e.offsetParent !== null) { el = e; break; }
            }
            if (!el) el = document.querySelector('input[name="terms"],input[type="checkbox"]');
            if (el) { el.scrollIntoView({block:'center', behavior:'instant'}); return true; }
            return false;
            """,
            _SUBMIT_SELECTORS,
        )
        if anchored:
            time.sleep(0.3)
            return
    except Exception as e:
        log(f"  ⚠ Error anclando scroll a Enviar/checkbox: {e}")

    # Fallback: sin ancla visible — scroll acotado (no directo al fondo/footer)
    try:
        from selenium.webdriver.common.action_chains import ActionChains
        ActionChains(driver).scroll_by_amount(0, 400).perform()
        time.sleep(0.3)
    except Exception as e:
        log(f"  ⚠ Error scroll bottom parent-aware: {e}")

    # Red de seguridad: si el form/checkbox/botón quedó fuera del viewport (footer), re-centrar
    try:
        driver.execute_script(
            """
            var f = document.querySelector('input[type="checkbox"], button.submit-button, form');
            if (!f) return;
            var r = f.getBoundingClientRect();
            var vh = window.innerHeight || document.documentElement.clientHeight;
            if (!(r.top < vh && r.bottom > 0)) f.scrollIntoView({block:'center', behavior:'instant'});
            """
        )
        time.sleep(0.2)
    except Exception:
        pass

def _scroll_to_element_parent_aware(driver, element, log: Callable = print):
    """
    Scrollea incrementalmente hacia abajo a medida que se completan los campos
    (evitando saltos a la parte superior en Safari Mac).
    """
    try:
        from selenium.webdriver.common.action_chains import ActionChains
        ActionChains(driver).scroll_by_amount(0, 90).perform()
        time.sleep(0.15)
    except Exception as e:
        log(f"  ⚠ Error scroll element parent-aware: {e}")

def _select_option(driver, select_id: str, value: str, field_name: str,
                   is_dependent: bool = False,
                   log: Callable = print,
                   _out: Optional[list] = None) -> bool:
    """Un solo JS call por intento: busca, valida, selecciona y despacha eventos."""
    if is_dependent:
        if not _wait_for_dropdown_options(driver, select_id):
            log(f"  ⏱ Timeout esperando opciones en '{field_name}' (id={select_id})")
            return False

    is_empty = not value or not value.strip() or _is_placeholder(value)

    for attempt in range(1, _DEP_RETRIES + 1):
        try:
            try:
                el_select = driver.find_element(By.ID, select_id)
                if el_select:
                    _scroll_to_element_parent_aware(driver, el_select, log)
            except Exception:
                pass
            res = driver.execute_script(_SELECT_JS, select_id, value or "", is_empty)
            if not res:
                return False
            r = res.get("r", "")
            if r == "noopts":
                if attempt < _DEP_RETRIES:
                    time.sleep(0.1 if not is_dependent else 0.3)
                    continue
                return False
            if r in ("notfound", "hidden"):
                return False
            if res.get("ok"):
                txt = res.get("text", value or "?")
                if not is_empty and not res.get("matched"):
                    # El Excel pidió un valor concreto y el dropdown no lo tiene: se eligió
                    # uno al azar. Se avisa acá además de quedar en "Datos vs Excel", para
                    # que se vea en el log por qué el resultado no coincide con lo pedido.
                    log(f"  ⚠ {field_name}: el Excel pide '{value}' pero el dropdown no "
                        f"ofrece esa opción → quedó '{txt}' (aleatorio)")
                else:
                    log(f"  {'🎲' if is_empty else '✓'} {field_name} = '{txt}'")
                if _out is not None:
                    _out.append(txt)
                return True
        except Exception as e:
            if attempt < _DEP_RETRIES:
                time.sleep(0.05)
            else:
                log(f"  ✗ Error select '{field_name}': {e}")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# LLENADO DE CAMPOS (lógica de Osocio + JS text input)
# ══════════════════════════════════════════════════════════════════════════════

_RADIO_GROUP_JS = r"""
// Busca el grupo de radios de un campo del mapping (por name o por id/prefijo de id) y
// selecciona la opción que pidió el Excel. Los forms 2.0 dibujan 'Fecha estimada'
// (estimated-date-purchase) como radios, no como <select>: sin esto el campo no se
// resolvía, se salteaba, y después _mark_preferred_radios elegía la primera opción.
var ids = arguments[0], val = arguments[1];
var norm = function (s) {
    return String(s || '').replace(/ /g, ' ').normalize('NFD')
        .replace(/[̀-ͯ]/g, '').toLowerCase().replace(/\s+/g, ' ').trim();
};
var vis = function (e) {
    return e && e.getClientRects().length && getComputedStyle(e).visibility !== 'hidden';
};

var radios = [], seen = [];
for (var n = 0; n < ids.length; n++) {
    var nm = ids[n];
    if (!nm) continue;
    var sel = 'input[type="radio"][name="' + nm + '"],'
            + 'input[type="radio"][id="' + nm + '"],'
            + 'input[type="radio"][id^="' + nm + '-"]';
    var found;
    try { found = document.querySelectorAll(sel); } catch (e) { continue; }
    for (var k = 0; k < found.length; k++) {
        if (seen.indexOf(found[k]) < 0) { seen.push(found[k]); radios.push(found[k]); }
    }
}
radios = radios.filter(vis);
if (!radios.length) return {r: 'noradios'};

// Texto asociado a cada radio: value, title, aria-label y el label que lo describe.
var labelOf = function (r) {
    var out = [r.value, r.getAttribute('title'), r.getAttribute('aria-label'),
              r.getAttribute('data-dtm')];
    if (r.id) {
        var lb = document.querySelector('label[for="' + r.id + '"]');
        if (lb) out.push(lb.textContent);
    }
    var anc = r.closest ? r.closest('label') : null;
    if (anc) out.push(anc.textContent);
    return out.filter(function (t) { return t; }).map(norm);
};

var target = null, matched = false;
if (val) {
    var vl = norm(val);
    for (var i = 0; i < radios.length && !target; i++) {
        var c = labelOf(radios[i]);
        for (var j = 0; j < c.length; j++) { if (c[j] === vl) { target = radios[i]; break; } }
    }
    if (!target) {
        for (var i = 0; i < radios.length && !target; i++) {
            var c = labelOf(radios[i]);
            for (var j = 0; j < c.length; j++) {
                if (c[j].indexOf(vl) >= 0 || (c[j] && vl.indexOf(c[j]) >= 0)) { target = radios[i]; break; }
            }
        }
    }
    matched = !!target;
}
// Sin valor en el Excel: respetar lo que el form ya tenga marcado antes de tocar nada.
if (!target) {
    for (var i = 0; i < radios.length; i++) { if (radios[i].checked) { target = radios[i]; break; } }
    if (target) {
        var pre = labelOf(target);
        return {r: 'ok', ok: true, matched: false, kept: true, text: pre.length ? pre[0] : (target.value || '')};
    }
}
if (!target) {
    var enabled = radios.filter(function (r) { return !r.disabled; });
    if (!enabled.length) return {r: 'nonenabled'};
    target = enabled[Math.floor(Math.random() * enabled.length)];
}

target.scrollIntoView({block: 'center', behavior: 'instant'});
target.removeAttribute('disabled'); target.disabled = false;
try { target.click(); } catch (e) { target.checked = true; }
if (!target.checked) target.checked = true;
target.dispatchEvent(new Event('change', {bubbles: true}));
target.dispatchEvent(new Event('input', {bubbles: true}));
target.dispatchEvent(new Event('blur', {bubbles: true}));
var lbls = labelOf(target);
return {r: 'ok', ok: true, matched: matched, kept: false,
        text: target.getAttribute('title') || target.value || (lbls.length ? lbls[0] : '')};
"""


def _select_radio_group(driver, ids: List[str], value: str, field_name: str,
                        log: Callable = print):
    """Completa un campo del mapping que el form renderiza como radio group.

    Devuelve None si no hay radios para ese campo (el flujo sigue con select/input),
    o (ok: bool, texto_elegido: str). Respeta SIEMPRE el valor del Excel: sólo sortea
    cuando la columna viene vacía y el form no trae nada marcado.
    """
    try:
        res = driver.execute_script(_RADIO_GROUP_JS, [str(i) for i in ids if i], value or "")
    except Exception as e:
        log(f"  ✗ Error en radios de '{field_name}': {e}")
        return None
    if not res or res.get("r") == "noradios":
        return None
    if res.get("r") == "nonenabled":
        log(f"  ⚠ {field_name}: radios presentes pero todos deshabilitados")
        return (False, "")

    # El title del radio trae &nbsp; ("3\xa0meses"): se normaliza a espacio común para que
    # el Excel de resultados muestre el texto tal cual lo escribiría el usuario.
    txt = " ".join((res.get("text") or "").replace("\xa0", " ").split())
    if res.get("kept"):
        log(f"  ✓ {field_name} (radio) ya venía marcado: '{txt}' — se conserva")
    elif value and not res.get("matched"):
        log(f"  ⚠ {field_name}: el Excel pide '{value}' pero ningún radio coincide "
            f"→ quedó '{txt}' (aleatorio)")
    elif value:
        log(f"  ✓ {field_name} (radio) = '{txt}' (pedido en el Excel)")
    else:
        log(f"  🎲 {field_name} (radio) = '{txt}'")
    return (True, txt)


def _get_field_mapping_for_pais(pais: str) -> List[Dict]:
    """Carga el field_mapping efectivo de Osocio para el país."""
    try:
        _src_root = os.path.dirname(_THIS_DIR)
        if _src_root not in sys.path:
            sys.path.insert(0, _src_root)
        from utils.fixed_field_mapping_store import load_effective_country_form_config
        config = load_effective_country_form_config(pais)
        return config.get("field_mapping", [])
    except Exception as e:
        print(f"  ⚠ No se pudo cargar field_mapping de Osocio para {pais}: {e}")
        return []


def fill_form_fields(driver, lead: LeadRow, pais: str,
                     field_mapping: List[Dict],
                     dependencies: Dict[str, str],
                     ids_dinamicos: Dict[str, str],
                     tracked: Dict[str, str],
                     log: Callable = print,
                     is_mobile: bool = False,
                     is_android: bool = False,
                     iframe_el=None,
                     brasil_doc_type: str = "cpf",
                     cross_processed: Optional[set] = None):
    """
    Llena los campos del formulario.
    is_mobile=True: usa funciones adaptadas para Safari iOS.
    iframe_el: elemento iframe (para tap_and_type en iOS).
    """
    # Elegir funciones según plataforma
    if is_android:
        fill_text = _fill_text_android
    elif is_mobile:
        fill_text = _fill_text_js_mobile
    else:
        # Mac/Safari desktop: send_keys real desde el arranque (ver _fill_text_sendkeys)
        fill_text = _fill_text_sendkeys
    scroll_to  = _scroll_to_mobile     if is_mobile else _scroll_to

    child_to_parent = {v: k for k, v in dependencies.items()}
    lead_lower = {k.lower().strip(): v for k, v in lead.data.items()}

    HEADER_TO_IDS = _HEADER_TO_IDS

    def _get_value(field_name: str, field_id) -> str:
        ids = field_id if isinstance(field_id, list) else [field_id]
        # Por nombre del campo del mapping
        val = lead_lower.get(field_name.lower().strip(), "")
        if val and val.lower() not in ("", "none"):
            return val
        # Por ID HTML directo
        for fid in ids:
            val = lead_lower.get((fid or "").lower(), "")
            if val and val.lower() not in ("", "none"):
                return val
        # Por encabezados del Excel
        for header, mapped_ids in HEADER_TO_IDS.items():
            if any(mid in ids for mid in mapped_ids):
                val = lead_lower.get(header, "")
                if val and val.lower() not in ("", "none"):
                    return val
        return ""

    processed = set(cross_processed or [])

    # Resolver todos los IDs del mapping en un solo JS call (evita N round-trips a LambdaTest)
    _ids_batch = []
    for _fc in field_mapping:
        _fid = _fc.get("id", "")
        _ids = [f for f in (_fid if isinstance(_fid, list) else [_fid]) if f]
        # Agregar alias visid si existe (para coexistencia old/new forms)
        for _f in list(_ids):
            _alias = _VISID_ID_ALIASES.get(_f)
            if _alias and _alias not in _ids:
                _ids.append(_alias)
        _ids_batch.append(_ids)
    try:
        _batch_raw = driver.execute_script("""
            // getClientRects().length==0 cuando el campo O CUALQUIER ANCESTRO esta display:none
            // (paso oculto del wizard). getComputedStyle(campo).display NO detecta el ancestro:
            // por eso antes se "llenaban" los pasos 2/3 ocultos, el send_keys fallaba en silencio
            // y el campo quedaba vacio.
            var vis=function(e){return e&&e.getClientRects().length&&getComputedStyle(e).visibility!=='hidden';};
            var cfgs=arguments[0],res={};
            for(var i=0;i<cfgs.length;i++){
                var ids=cfgs[i];
                for(var j=0;j<ids.length;j++){
                    if(vis(document.getElementById(ids[j]))){res[i]=ids[j];break;}
                }
                if(res[i]===undefined){
                    for(var j=0;j<ids.length;j++){
                        var all=document.querySelectorAll('input[id^="'+ids[j]+'-"],select[id^="'+ids[j]+'-"],textarea[id^="'+ids[j]+'-"]');
                        for(var k=0;k<all.length;k++){
                            if(vis(all[k])){res[i]=all[k].id;break;}
                        }
                        if(res[i]!==undefined)break;
                    }
                }
            }
            return res;
        """, _ids_batch) or {}
        _batch = {int(k): v for k, v in _batch_raw.items() if v}
    except Exception:
        _batch = {}

    _text_fill_count = 0
    for i, fc in enumerate(field_mapping):
        fid_raw   = fc.get("id", "")
        fname     = fc.get("name", str(fid_raw))
        ftype     = fc.get("type", "text")
        resolved  = _batch.get(i)
        if not resolved:
            # Re-chequeo simple SIN scrollear: si el campo no está visible en el DOM actual,
            # se saltea. NO scroll-search agresivo: en forms multi-step (gm_forms) roaba toda
            # la página buscando campos de pasos siguientes y perdía los del paso actual
            # (ej. VIN/Modelo del paso 1 en clubemyev). El loop multi-step los llena cuando
            # aparecen tras 'Siguiente'.
            try:
                resolved = driver.execute_script("""
                    var vis=function(e){return e&&e.getClientRects().length&&getComputedStyle(e).visibility!=='hidden';};
                    var ids=arguments[0];
                    for(var j=0;j<ids.length;j++){
                        if(vis(document.getElementById(ids[j])))return ids[j];
                        var all=document.querySelectorAll('input[id^="'+ids[j]+'-"],select[id^="'+ids[j]+'-"],textarea[id^="'+ids[j]+'-"]');
                        for(var k=0;k<all.length;k++){if(vis(all[k]))return all[k].id;}
                    }
                    return null;
                """, _ids_batch[i])
            except Exception:
                resolved = None

        value = _get_value(fname, fid_raw)

        # Campo del mapping que el form dibuja como RADIOS en vez de <select> (típico de
        # 'Fecha estimada' / estimated-date-purchase en los forms 2.0). Se prueba SIEMPRE
        # antes que la rama select: el id suele resolverse igual al del radio individual
        # ('estimated-date-purchase-1'), con lo que _select_option fallaba en silencio, el
        # campo quedaba sin llenar y después _mark_preferred_radios ponía la 1ª opción,
        # pisando lo que pidió el Excel. Si no hay radios, _select_radio_group devuelve
        # None y el flujo sigue normalmente por select/input.
        if ftype == "select":
            radio_res = _select_radio_group(driver, _ids_batch[i], value, fname, log=log)
            if radio_res is not None:
                ok, chosen = radio_res
                if ok:
                    tracked[fname] = chosen
                    for _rid in _ids_batch[i]:
                        processed.add(_rid)
                        if cross_processed is not None:
                            cross_processed.add(_rid)
                continue

        if not resolved or resolved in processed:
            continue

        # ── Texto ────────────────────────────────────────────────────────────
        if ftype == "text":
            if resolved in ids_dinamicos:
                value = ids_dinamicos[resolved]
            if not value:
                processed.add(resolved)
                continue
            # Brasil: generar CPF/CNPJ/CEP según brasil_doc_type y maxlength del campo
            if pais.lower() in ("brasil", "brazil", "br") and any(x in resolved.lower() for x in ("cpf", "cnpj", "cep", "zip", "postal", "document", "ci")):
                # Si el usuario configuró un tipo específico, usarlo; si no, autodetectar por field_id
                _forced_type = brasil_doc_type  # "cpf" | "cnpj" | "cep"
                _is_cnpj = (_forced_type == "cnpj") or ("cnpj" in resolved.lower())
                _is_cep  = (_forced_type == "cep")  or ("cep" in resolved.lower() or "zip" in resolved.lower() or "postal" in resolved.lower())
                # Leer maxlength del campo antes de generar para decidir formato
                _field_maxlen = None
                try:
                    _tmp_els = driver.find_elements(By.ID, resolved)
                    if _tmp_els:
                        _field_maxlen = _get_maxlength(_tmp_els[0])
                except Exception:
                    pass
                try:
                    _clean = str(int(float(str(value)))) if value not in (None, "") else ""
                except (ValueError, TypeError):
                    _clean = str(value or "")
                _digits = "".join(c for c in _clean if c.isdigit())
                _needs = 14 if _is_cnpj else (8 if _is_cep else 11)
                if _digits:
                    value = _digits
                if len(_digits) != _needs:
                    if _is_cnpj:
                        # Sin puntuación si maxlength <= 14
                        pontuacao = "N" if (_field_maxlen and _field_maxlen <= 14) else "S"
                        raw = _fetch_4devs("gerar_cnpj", {"pontuacao": pontuacao})
                        generated = "".join(c for c in raw if c.isdigit()) if pontuacao == "N" else raw
                    elif _is_cep:
                        generated = _generate_brazil_document("cep")
                    else:
                        generated = _generate_brazil_document(resolved)
                    if generated:
                        value = generated

            # Perú: sanitizar número de documento según tipo seleccionado.
            # En los forms visid / gm_front el 'ci' del mapping se resuelve al alias
            # 'document', así que comparar sólo contra "ci" dejaba esos forms sin sanitizar.
            if resolved in ("ci", "document") and pais.lower() in ("peru", "pe"):
                try:
                    doc_el = driver.find_elements(By.ID, "document-type")
                    if doc_el:
                        doc_type_val = Select(doc_el[0]).first_selected_option.text
                        value = _sanitize_peru_document(doc_type_val, value)
                except Exception:
                    pass
            try:
                el = driver.execute_script(
                    "var e=document.getElementById(arguments[0]);"
                    "if(!e||e.disabled||!e.getClientRects().length"
                    "||getComputedStyle(e).visibility==='hidden')return null;"
                    "e.scrollIntoView({block:'center',behavior:'instant'});"
                    "return e;",
                    resolved
                )
                if not el:
                    if is_mobile and iframe_el:
                        _tap_and_type_native(driver, iframe_el, resolved, value, log)
                        tracked[fname] = value
                        processed.add(resolved)
                    continue
                _scroll_to_element_parent_aware(driver, el, log)
                ml = _get_maxlength(el)
                if ml and len(value) > ml:
                    value = value[:ml]
                fill_text(driver, el, value)
                tracked[fname] = value
                log(f"  ✓ {fname} ({resolved}) = '{value}'")
                processed.add(resolved)
                if cross_processed is not None:
                    cross_processed.add(resolved)
                _text_fill_count += 1
                if _text_fill_count == 2:
                    # Scroll al fondo tras los 2 primeros campos — solo en forms comprar-carro
                    try:
                        if "comprar-carro" in driver.current_url:
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    except Exception:
                        pass
            except Exception as e:
                log(f"  ✗ Error texto '{fname}': {e}")

        # ── Select ───────────────────────────────────────────────────────────
        elif ftype == "select":
            is_dep = resolved in child_to_parent
            _sel_result: list = []
            ok = _select_option(driver, resolved, value, fname,
                                 is_dependent=is_dep, log=log,
                                 _out=_sel_result)
            if ok:
                tracked[fname] = _sel_result[0] if _sel_result else value
                processed.add(resolved)
                if cross_processed is not None:
                    cross_processed.add(resolved)
                if is_dep:
                    time.sleep(0.2)  # pausa para que arranque AJAX del dependiente

    # ── IDs dinámicos no mapeados ─────────────────────────────────────────────
    for dyn_id, dyn_value in ids_dinamicos.items():
        if dyn_id in processed:
            continue
        try:
            el = driver.execute_script(
                "var e=document.getElementById(arguments[0]);"
                "if(!e||e.disabled||!e.getClientRects().length"
                "||getComputedStyle(e).visibility==='hidden')return null;"
                "e.scrollIntoView({block:'center',behavior:'instant'});"
                "return e;",
                dyn_id
            )
            if not el:
                continue
            _scroll_to_element_parent_aware(driver, el, log)
            tag = el.tag_name.lower()
            if tag == "select":
                _select_option(driver, dyn_id, dyn_value, dyn_id, log=log)
            elif tag in ("input", "textarea"):
                ml = _get_maxlength(el)
                v = dyn_value[:ml] if ml and len(dyn_value) > ml else dyn_value
                fill_text(driver, el, v)
                log(f"  📌 {dyn_id} = '{v}' (ID dinámico)")
            processed.add(dyn_id)
        except Exception:
            continue

    # Scroll al fondo al terminar de llenar para revelar checkboxes/botón Enviar en Mac/Safari
    try:
        if not is_android:
            _scroll_to_bottom_parent_aware(driver, log)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════════════
# TÉRMINOS Y CHECKBOXES (copiado exacto de Osocio)
# ══════════════════════════════════════════════════════════════════════════════

_KNOWN_CHECKBOX_NAMES = {
    "terms", "terms-contact", "terms_contact", "termscontact",
    "terms-platform", "terms_platform", "termsplatform",
    "accept-terms", "accept_terms", "privacy", "privacy_policy",
}

_CHECKBOX_PRIORITY = {
    "terms": 3,
    "terms-platform": 2, "terms_platform": 2, "termsplatform": 2,
    "terms-contact": 1, "terms_contact": 1, "termscontact": 1,
}


def _prepare_checkbox(driver, cb):
    """Elimina disabled y aria-disabled. Igual que Osocio."""
    try:
        driver.execute_script("""
            const cb = arguments[0];
            if (cb.hasAttribute('disabled')) cb.removeAttribute('disabled');
            cb.disabled = false;
            cb.setAttribute('aria-disabled', 'false');
            cb.classList.remove('is-invalid');
            if (cb.tabIndex === -1) cb.tabIndex = 0;
        """, cb)
    except Exception:
        pass


def _set_checkbox_js(driver, cb) -> bool:
    """
    Marca checkbox priorizando label.click() para React/custom checkboxes ocultos.
    Safari + React ignora cb.checked = true; label.click() es la única interacción
    que React procesa como evento nativo real.
    """
    try:
        result = driver.execute_script("""
            const cb = arguments[0];
            if (cb.checked) return 'already';
            // Prioridad: click en el label (React lo procesa como evento nativo)
            let label = null;
            try {
                if (cb.id) label = document.querySelector('label[for="' + CSS.escape(cb.id) + '"]');
            } catch(e) {
                if (cb.id) label = document.querySelector('label[for="' + cb.id + '"]');
            }
            if (!label) label = cb.closest('label');
            if (!label) {
                const sib = cb.nextElementSibling;
                if (sib && sib.tagName && sib.tagName.toLowerCase() !== 'input') label = sib;
            }
            if (label) {
                label.click();
                if (cb.checked) return 'label';
            }
            // Fallback: native property setter + MouseEvent (mejor que .checked = true directo)
            const n = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'checked');
            if (n && n.set) n.set.call(cb, true);
            else cb.checked = true;
            cb.setAttribute('checked', 'checked');
            cb.dispatchEvent(new MouseEvent('click',  {bubbles:true, cancelable:true}));
            cb.dispatchEvent(new Event('change', {bubbles:true}));
            return cb.checked ? 'setter' : 'failed';
        """, cb)
        return result not in (None, False, 'failed')
    except Exception:
        return False


def _checkbox_in_dom_js(driver, el) -> bool:
    """
    True si el checkbox tiene layout propio (opacity:0 OK) o tiene un label visible.
    Cubre el patrón custom: input display:none + label visible clickeable.
    """
    try:
        return bool(driver.execute_script("""
            var e=arguments[0];
            if(!e.isConnected)return false;
            var cs=getComputedStyle(e);
            var ownLayout=cs.display!=='none'&&cs.visibility!=='hidden'
                &&(e.offsetWidth>0||e.offsetHeight>0||e.getClientRects().length>0);
            if(ownLayout)return true;
            // Fallback: label visible (checkbox puede ser display:none con label clickeable)
            var lbl=null;
            if(e.id)lbl=document.querySelector('label[for="'+e.id+'"]');
            if(!lbl)lbl=e.closest('label');
            if(!lbl){var s=e.nextElementSibling;if(s&&s.tagName!=='INPUT')lbl=s;}
            if(!lbl)return false;
            var ls=getComputedStyle(lbl);
            return ls.display!=='none'&&ls.visibility!=='hidden'
                &&(lbl.offsetWidth>0||lbl.offsetHeight>0);
        """, el))
    except Exception:
        return False


def _find_click_target(driver, element):
    """Busca el mejor elemento clickeable (label, sibling). Igual que Osocio."""
    try:
        target = driver.execute_script("""
            const elem = arguments[0];
            if (elem.nextElementSibling &&
                elem.nextElementSibling.tagName.toLowerCase() !== 'input') {
                return elem.nextElementSibling;
            }
            const parentLabel = elem.closest('label');
            if (parentLabel) return parentLabel;
            if (elem.id) {
                const forLabel = document.querySelector('label[for="' + elem.id + '"]');
                if (forLabel) return forLabel;
            }
            return elem;
        """, element)
        return target
    except Exception:
        return element


def _click_stable(driver, element) -> bool:
    """Click con fallbacks: click normal → JS click → PointerEvent → ActionChains."""
    if element is None:
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    except Exception:
        pass

    for _ in range(2):
        try:
            element.click()
            return True
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", element)
                return True
            except Exception:
                try:
                    driver.execute_script("""
                        const el = arguments[0];
                        if (typeof PointerEvent !== 'undefined') {
                            el.dispatchEvent(new PointerEvent('pointerdown', {bubbles:true}));
                            el.dispatchEvent(new PointerEvent('pointerup',   {bubbles:true}));
                        }
                        el.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
                        el.dispatchEvent(new MouseEvent('click',     {bubbles:true}));
                        el.dispatchEvent(new MouseEvent('mouseup',   {bubbles:true}));
                    """, element)
                    return True
                except Exception:
                    pass

    # Fallback ActionChains — más confiable en Safari Mac
    try:
        from selenium.webdriver.common.action_chains import ActionChains
        ActionChains(driver).move_to_element(element).click().perform()
        return True
    except Exception:
        pass

    return False


def _locate_checkbox(driver, candidate: Dict):
    """Localiza el checkbox en el DOM por id/name/data-dtm. Igual que Osocio."""
    cb_id       = candidate.get("id", "")
    name_orig   = candidate.get("name_original", "")
    lower_name  = candidate.get("name", "")
    data_dtm    = candidate.get("data_dtm", "")
    cand_value  = candidate.get("value", "")

    selectors = []
    if cb_id:
        selectors.append((By.ID, cb_id))
    if name_orig:
        selectors.append((By.CSS_SELECTOR, f"input[type='checkbox'][name=\"{name_orig}\"]"))
    if data_dtm:
        selectors.append((By.CSS_SELECTOR, f"input[type='checkbox'][data-dtm=\"{data_dtm}\"]"))
    selectors.append((By.CSS_SELECTOR, "input[type='checkbox']"))

    for by, sel in selectors:
        try:
            elements = driver.find_elements(by, sel)
        except Exception:
            continue
        for el in elements:
            try:
                if cb_id and (el.get_attribute("id") or "").strip() != cb_id:
                    continue
                if name_orig and (el.get_attribute("name") or "").strip() != name_orig:
                    continue
                if lower_name and (el.get_attribute("name") or "").strip().lower() != lower_name:
                    continue
                if data_dtm and (el.get_attribute("data-dtm") or "").strip() != data_dtm:
                    continue
                val_attr = (el.get_attribute("value") or "").strip()
                if cand_value and val_attr and cand_value != val_attr:
                    continue
                return el
            except StaleElementReferenceException:
                continue
    return None


def _find_label_for_checkbox(driver, cb):
    """Devuelve el label visible asociado al checkbox (para scroll y click real)."""
    try:
        return driver.execute_script("""
            var cb=arguments[0], lbl=null;
            if(cb.id) try{lbl=document.querySelector('label[for="'+CSS.escape(cb.id)+'"]');}catch(e){lbl=document.querySelector('label[for="'+cb.id+'"]');}
            if(!lbl) lbl=cb.closest('label');
            if(!lbl){var s=cb.nextElementSibling;if(s&&s.tagName&&s.tagName.toLowerCase()!=='input')lbl=s;}
            return lbl||null;
        """, cb)
    except Exception:
        return None


def _react_check_checkbox(driver, cb) -> bool:
    """
    Marca checkbox vía React fiber onChange — funciona aunque el input tenga disabled.
    React re-renderiza con state checked=true, sobrevive re-renders posteriores.
    """
    try:
        return bool(driver.execute_script("""
            var cb=arguments[0];
            var fk=null;
            var keys=Object.keys(cb);
            for(var i=0;i<keys.length;i++){
                var k=keys[i];
                if(k.startsWith('__reactFiber')||k.startsWith('__reactInternalInstance')||k.startsWith('_reactFiber')){fk=k;break;}
            }
            if(!fk)return false;
            var fiber=cb[fk];
            while(fiber){
                var props=fiber.memoizedProps;
                if(props&&typeof props.onChange==='function'){
                    var desc=Object.getOwnPropertyDescriptor(HTMLInputElement.prototype,'checked');
                    if(desc&&desc.set)desc.set.call(cb,true);else cb.checked=true;
                    try{props.onChange({target:cb,currentTarget:cb,type:'change',bubbles:true,cancelable:true,nativeEvent:{type:'change'}});}catch(e){}
                    return true;
                }
                fiber=fiber.return;
            }
            return false;
        """, cb))
    except Exception:
        return False


def _ensure_checkbox_selected(driver, candidate: Dict, log: Callable = print,
                              is_android: bool = False) -> bool:
    """
    Marca un checkbox.
    Mac/Safari: ActionChains → teclado (Space) → JS label.click() (2 iteraciones)
    Android/Chrome: JS click → native click (2 iteraciones)
    """
    lbl_name = candidate.get("label", "checkbox")
    max_iters = 2

    for attempt in range(max_iters):
        cb = _locate_checkbox(driver, candidate)
        if not cb:
            time.sleep(0.1 if is_android else 0.15)
            continue
        try:
            _prepare_checkbox(driver, cb)

            if cb.is_selected():
                log(f"  ℹ {lbl_name} ya estaba marcado")
                return True

            label_el = _find_label_for_checkbox(driver, cb)
            click_target = label_el or cb
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center',behavior:'instant'});",
                click_target
            )
            time.sleep(0.1 if is_android else 0.2)

            if is_android:
                # Android Chrome: JS primero (confiable), fallback native click
                if _set_checkbox_js(driver, cb):
                    time.sleep(0.2)
                    cb = _locate_checkbox(driver, candidate) or cb
                    if cb.is_selected():
                        log(f"  ✓ {lbl_name} marcado vía JS")
                        return True
                try:
                    click_target.click()
                    time.sleep(0.2)
                    cb = _locate_checkbox(driver, candidate) or cb
                    if cb.is_selected():
                        log(f"  ✓ {lbl_name} marcado vía click nativo")
                        return True
                except Exception:
                    pass
            else:
                from selenium.webdriver.common.action_chains import ActionChains
                # Intento 1: ActionChains sobre el label (Safari Mac lo procesa como evento real)
                try:
                    ActionChains(driver).move_to_element(click_target).click().perform()
                    time.sleep(0.35)
                    cb = _locate_checkbox(driver, candidate) or cb
                    if cb.is_selected():
                        log(f"  ✓ {lbl_name} marcado vía ActionChains")
                        return True
                except Exception:
                    pass

                # Intento 2: click real más preciso, sobre el ícono visual del checkbox en vez
                # del label completo — en labels largos (ej. texto de términos) el "centro" del
                # label puede caer lejos del ícono clickeable.
                try:
                    icon_el = driver.execute_script(
                        "var l=arguments[0];"
                        "return l.querySelector('.icon,svg,.content') || l;",
                        click_target,
                    )
                    ActionChains(driver).move_to_element(icon_el).click().perform()
                    time.sleep(0.3)
                    cb = _locate_checkbox(driver, candidate) or cb
                    if cb.is_selected():
                        log(f"  ✓ {lbl_name} marcado vía click preciso")
                        return True
                except Exception:
                    pass

                # Intento 3: JS label.click() (último recurso, best-effort)
                if _set_checkbox_js(driver, cb):
                    time.sleep(0.3)
                    cb = _locate_checkbox(driver, candidate) or cb
                    if cb.is_selected():
                        log(f"  ✓ {lbl_name} marcado vía JS")
                        return True

        except StaleElementReferenceException:
            time.sleep(0.1 if is_android else 0.15)
            continue
        except Exception as e:
            log(f"  ✗ Error {lbl_name}: {e}")
            break

    cb = _locate_checkbox(driver, candidate)
    return cb.is_selected() if cb else False


_CHECKBOX_SI = {"si", "sí", "yes", "true", "1", "x", "marcar", "on"}
_CHECKBOX_NO = {"no", "false", "0", "off", "desmarcar"}


def _checkbox_prefs(lead) -> Dict[str, bool]:
    """
    Preferencias de checkbox tomadas del Excel: una columna cuyo encabezado es el `name`
    (o el `id`) del checkbox, con valor SI/NO. Ej: columna "test-drive" = "NO".

    Solo se aplica si el form tiene un checkbox con ese name/id, así que una columna con
    SI/NO que no corresponda a ningún checkbox se ignora sola.
    """
    prefs: Dict[str, bool] = {}
    for header, value in (getattr(lead, "data", None) or {}).items():
        v = str(value or "").strip().lower()
        key = str(header or "").strip().lower()
        if not v or not key:
            continue
        if v in _CHECKBOX_SI:
            prefs[key] = True
        elif v in _CHECKBOX_NO:
            prefs[key] = False
    return prefs


_UNCHECK_JS = r"""
var id = arguments[0], name = arguments[1];
var cb = (id && document.getElementById(id)) ||
         (name && document.querySelector('input[type="checkbox"][name="' + name + '"]'));
if (!cb) return false;
if (!cb.checked) return true;
cb.checked = false;
cb.dispatchEvent(new Event('click',  {bubbles:true}));
cb.dispatchEvent(new Event('change', {bubbles:true}));
cb.checked = false;   // por si algún handler lo volvió a marcar
return !cb.checked;
"""


def _mark_required_checkboxes(driver, log: Callable = print, is_android: bool = False,
                              prefs: Optional[Dict[str, bool]] = None) -> int:
    """
    Marca los checkboxes de términos/privacy (por name conocido + required).

    `prefs` (del Excel, ver _checkbox_prefs) manda sobre el default: uno con NO se desmarca
    y no se vuelve a tocar; uno con SI se marca aunque no sea required ni conocido.
    """
    prefs = prefs or {}
    try:
        raw = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
    except Exception:
        return 0

    candidates = []
    for idx, cb in enumerate(raw):
        try:
            name_attr  = (cb.get_attribute("name")     or "").strip()
            lower_name = name_attr.lower()
            cb_id      = (cb.get_attribute("id")       or "").strip()
            required   = cb.get_attribute("required")
            data_dtm   = (cb.get_attribute("data-dtm") or "").strip()
            value_attr = (cb.get_attribute("value")    or "").strip()

            pref = prefs.get(lower_name)
            if pref is None:
                pref = prefs.get(cb_id.lower())
            if pref is False:
                try:
                    driver.execute_script(_UNCHECK_JS, cb_id, name_attr)
                except Exception:
                    pass
                log(f"  ⊘ {name_attr or cb_id} desmarcado (Excel = NO)")
                continue

            is_known = lower_name in _KNOWN_CHECKBOX_NAMES
            is_html_required = bool(required)
            if not is_known and not is_html_required and pref is not True:
                # Incluir si tiene layout DOM y no está ya marcado (igual que desktop)
                try:
                    if cb.is_selected() or not cb.is_enabled():
                        continue
                    if not _checkbox_in_dom_js(driver, cb):
                        continue
                except Exception:
                    continue

            priority = _CHECKBOX_PRIORITY.get(lower_name, 0)
            display  = name_attr or cb_id or data_dtm or "checkbox"
            candidates.append({
                "name": lower_name, "name_original": name_attr,
                "id": cb_id, "value": value_attr, "data_dtm": data_dtm,
                "priority": priority, "order": idx, "label": display,
            })
        except StaleElementReferenceException:
            continue

    if not candidates:
        log("  ℹ No se encontraron checkboxes requeridos")
        return 0

    candidates.sort(key=lambda c: (c["priority"], c["order"]))
    marked = 0
    for cand in candidates:
        if _ensure_checkbox_selected(driver, cand, log, is_android=is_android):
            marked += 1
        else:
            log(f"  ⚠ No se pudo marcar {cand.get('label','checkbox')}")

    log(f"  Checkboxes marcados: {marked}/{len(candidates)}")
    return marked


def _choose_radio_option(group: Dict) -> Optional[Dict]:
    """Elige la opción preferida de un radio group (sí/yes/renovar). Igual que Osocio."""
    options = group.get("options", [])
    if not options:
        return None
    # Si el grupo YA tiene una opción marcada, esa gana: puede venir del valor que el
    # usuario puso en el Excel (fill_form_fields → _select_radio_group) o del default del
    # form. Sin este chequeo el grupo se re-marcaba con la preferida/primera y la fecha
    # estimada pedida en el Excel terminaba reemplazada por otra.
    for opt in options:
        try:
            el = opt.get("element")
            if el is not None and el.is_selected():
                return opt
        except StaleElementReferenceException:
            continue
        except Exception:
            continue
    prefer = {"si", "sí", "yes", "renovar", "acepto", "acepta", "1", "true", "on"}
    for opt in options:
        if (opt.get("value") or "").lower() in prefer:
            return opt
        if (opt.get("label") or "").lower() in prefer:
            return opt
    return options[0]


def _mark_preferred_radios(driver, log: Callable = print) -> int:
    """Agrupa radios por name y selecciona opción preferida. Igual que Osocio."""
    try:
        radios = driver.find_elements(By.CSS_SELECTOR, "input[type='radio']")
    except Exception:
        return 0

    groups: Dict[str, Dict] = {}
    for idx, radio in enumerate(radios):
        try:
            name = (radio.get_attribute("name") or f"radio_group_{idx}").strip()
            key  = name.lower()
            grp  = groups.setdefault(key, {"name": key, "original_name": name,
                                            "options": [], "first_index": idx})
            grp["options"].append({
                "value": (radio.get_attribute("value") or "").strip(),
                "label": (radio.get_attribute("title") or radio.get_attribute("data-dtm")
                          or radio.get_attribute("id") or radio.get_attribute("value") or f"op_{idx}"),
                "index": idx,
                "element": radio,
            })
        except StaleElementReferenceException:
            continue

    marked = 0
    for grp in sorted(groups.values(), key=lambda g: g["first_index"]):
        opt = _choose_radio_option(grp)
        if not opt:
            continue
        try:
            radio = opt["element"]
            if radio.is_selected():
                marked += 1
                continue
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", radio)
            if _set_checkbox_js(driver, radio):
                marked += 1
                log(f"  ✓ Radio '{grp['original_name']}' = '{opt['value']}'")
            else:
                if _click_stable(driver, radio):
                    marked += 1
                    log(f"  ✓ Radio '{grp['original_name']}' = '{opt['value']}' (click)")
        except Exception as e:
            log(f"  ✗ Error radio '{grp['original_name']}': {e}")

    return marked


_REVALIDAR_JS = r"""
// Limpia los mensajes de error que quedaron pintados en campos que YA estan completos.
//
// El "click enviar vacio" dispara las validaciones antes de llenar. Despues llenamos por
// JS/send_keys, pero jquery-validation solo borra el mensaje cuando escucha keyup/focusout,
// asi que el cartelito rojo quedaba puesto aunque el campo fuera valido -> falso error en
// la captura. Solo se revalida lo que tiene valor: los vacios conservan su error (es real).
var revalidados = 0;
var campos = document.querySelectorAll('input, select, textarea');
for (var i = 0; i < campos.length; i++) {
    var el = campos[i];
    if (el.type === 'hidden' || el.disabled) continue;
    if (!el.getClientRects().length) continue;      // no visible (paso oculto)
    if (!el.value || !String(el.value).trim()) continue;   // vacio: el error es legitimo
    el.dispatchEvent(new Event('keyup',    {bubbles: true}));
    el.dispatchEvent(new Event('focusout', {bubbles: true}));
    el.dispatchEvent(new Event('blur',     {bubbles: true}));
    revalidados++;
}
// jquery-validation: pedirle que revalide campo por campo (borra el mensaje si ya pasa)
try {
    var $ = window.jQuery;
    if ($) {
        $('form').each(function () {
            var v = $(this).data('validator');
            if (!v) return;
            $(this).find('input, select, textarea').each(function () {
                if (this.type === 'hidden' || this.disabled) return;
                if (!this.getClientRects().length) return;
                if (!this.value || !String(this.value).trim()) return;
                try { v.element(this); } catch (e) {}
            });
        });
    }
} catch (e) {}
return revalidados;
"""


def _revalidar_campos_llenos(driver, log: Callable = print):
    """Saca los mensajes de error que quedaron colgados en campos que ya están completos."""
    try:
        n = driver.execute_script(_REVALIDAR_JS)
        if n:
            log(f"  ✓ {n} campos revalidados (limpieza de errores ya resueltos)")
    except Exception as e:
        log(f"  ⚠ Error revalidando campos: {e}")


def _handle_terms_checkboxes(driver, log: Callable = print, is_android: bool = False,
                             prefs: Optional[Dict[str, bool]] = None) -> bool:
    """
    Marca radios y checkboxes de términos.
    `prefs`: preferencias SI/NO por checkbox tomadas del Excel (ver _checkbox_prefs).
    """
    try:
        radios_marked = _mark_preferred_radios(driver, log)
    except Exception as e:
        log(f"  ⚠ Error radios: {e}")
        radios_marked = 0

    try:
        checkboxes_marked = _mark_required_checkboxes(driver, log, is_android=is_android,
                                                      prefs=prefs)
    except Exception as e:
        log(f"  ⚠ Error checkboxes: {e}")
        checkboxes_marked = 0

    log(f"  Radios: {radios_marked}, Checkboxes: {checkboxes_marked}")
    return (radios_marked + checkboxes_marked) > 0


def _ensure_terms_marked_before_submit(driver, log: Callable = print,
                                        prefs: Optional[Dict[str, bool]] = None):
    """
    Re-verifica checkboxes de términos justo antes de cada click submit.
    Safari puede deseleccionarlos tras una re-renderización del formulario.
    Los que el Excel marcó como NO se dejan como están (ver _checkbox_prefs).
    """
    prefs = prefs or {}
    try:
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);"
            "var f=document.querySelector('form,#contact-form,.gm-form,.form-wrapper');"
            "if(f)f.scrollTop=f.scrollHeight;"
        )
        time.sleep(0.2)
    except Exception:
        pass
    try:
        checkboxes = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
        for cb in checkboxes:
            try:
                name = (cb.get_attribute("name") or "").strip().lower()
                required = cb.get_attribute("required")
                cb_id_l = (cb.get_attribute("id") or "").strip().lower()
                if prefs.get(name) is False or prefs.get(cb_id_l) is False:
                    continue  # el Excel pidió NO marcarlo
                if name not in _KNOWN_CHECKBOX_NAMES and not required:
                    if not _checkbox_in_dom_js(driver, cb):
                        continue
                if not cb.is_selected():
                    log(f"  ⚠ Checkbox '{name or 'required'}' deseleccionado — re-marcando")
                    cb_id = (cb.get_attribute("id") or "").strip()
                    cand = {"name": name, "id": cb_id, "label": name or "terms"}
                    _ensure_checkbox_selected(driver, cand, log)
            except Exception:
                continue
    except Exception as e:
        log(f"  Error verificando checkboxes pre-submit: {e}")


def _motivo_corto(result_text: str) -> str:
    """Motivo del fallo en pocas palabras, para la columna 'Motivo' del Excel.
    Mismo criterio que base_form_filler._short_fail_reason."""
    t = (result_text or "").lower()
    if "no inserto" in t or "formulario ausente" in t:
        return "form no inserto"
    if "no coincide con el esperado" in t or "formulario incorrecto" in t:
        return "form inserto no coincide"
    if "iframe ausente" in t or "no encontrado" in t and "form" in t:
        return "form no encontrado"
    if "link issue" in t:
        return "link issue TYP"
    if "event_id" in t or "event id" in t:
        return "error event_id del servidor"
    if "error visual" in t:
        # El detalle del campo que falló ya viene en result_text
        _det = result_text.split("Error visual:", 1)[-1].strip()
        _det = _det.split("(2 intentos)")[0].strip(" |")
        return f"validación del form: {_det[:120]}" if _det else "validación del form"
    if "sin confirmación ty" in t or "sin confirmacion ty" in t or "ty page" in t:
        return "sin TY page (no se confirmó el envío)"
    if "botón de envío" in t or "boton de envio" in t:
        return "no se encontró el botón Enviar"
    if "404" in t:
        return "landing 404"
    if "error" in t:
        return "error al completar el form"
    return "error"


def _set_input_value_js(driver, el, value: str) -> None:
    """Setea el value usando el setter nativo del prototipo (necesario en React/Angular:
    asignar el.value directo no actualiza el estado interno del componente)."""
    try:
        driver.execute_script(
            "var e=arguments[0], v=arguments[1];"
            "var proto = e instanceof window.HTMLTextAreaElement"
            "    ? window.HTMLTextAreaElement.prototype : window.HTMLInputElement.prototype;"
            "var d = Object.getOwnPropertyDescriptor(proto, 'value');"
            "if (d && d.set) { d.set.call(e, v); } else { e.value = v; }"
            "e.dispatchEvent(new Event('input',{bubbles:true}));"
            "e.dispatchEvent(new Event('change',{bubbles:true}));",
            el, value,
        )
    except Exception:
        pass


def _hard_clear_input(driver, el, log: Callable = print) -> bool:
    """Vacía un input de verdad, y confirma que quedó vacío.

    `el.clear()` en Android/Chrome puede no vaciar nada y tampoco lanzar excepción
    (campos con máscara, o manejados por React que reponen el valor). Cuando eso pasa,
    el send_keys posterior CONCATENA y se termina enviando 'ApellidoApellido' o un email
    duplicado — el lead viaja con datos corruptos y falla la validación del form.
    Devuelve True sólo si el campo quedó realmente vacío.
    """
    def _current():
        try:
            return el.get_attribute("value") or ""
        except Exception:
            return ""

    for intento in range(3):
        if not _current():
            return True
        if intento == 0:
            try:
                el.clear()
            except Exception:
                pass
        elif intento == 1:
            _set_input_value_js(driver, el, "")
        else:
            # Último recurso: borrar con teclado, carácter por carácter desde el final
            try:
                el.click()
            except Exception:
                pass
            for _ in range(len(_current()) + 2):
                try:
                    el.send_keys(Keys.END)
                    el.send_keys(Keys.BACKSPACE)
                except Exception:
                    break
        time.sleep(0.05)

    quedo = _current()
    if quedo:
        log(f"  ⚠ No se pudo vaciar el campo (sigue con '{quedo}')")
        return False
    return True


def _ensure_fields_filled_before_submit(driver, field_mapping: List[Dict], all_tracked: Dict[str, str],
                                         log: Callable = print,
                                         is_mobile: bool = False, is_android: bool = False):
    """
    Re-ingresa inputs de texto carácter a carácter (send_keys real) justo antes de
    cada click submit. Safari puede mostrar el valor en el DOM (el.value con texto)
    sin que el estado interno de React/validación lo haya registrado nunca — el
    fill por JS (value + dispatchEvent) no alcanza a sincronizar ahí, aunque en
    Chrome sí funciona. Por eso NO se chequea el valor actual: se re-ingresa
    siempre, igual que ya se hacía solo para CPF/CNPJ/CEP en
    _refill_brasil_doc_sendkeys.
    """
    # Último valor conocido por nombre de campo (name del mapping), sin importar el paso.
    last_value_by_name: Dict[str, str] = {}
    for key, val in (all_tracked or {}).items():
        name = key.split("::", 1)[1] if "::" in key else key
        if val:
            last_value_by_name[name] = val

    for fc in field_mapping or []:
        fname = fc.get("name", "")
        expected = last_value_by_name.get(fname)
        if not expected:
            continue
        ftype = fc.get("type", "text")
        fid_raw = fc.get("id", "")
        ids = fid_raw if isinstance(fid_raw, list) else [fid_raw]
        for fid in ids:
            if not fid:
                continue
            try:
                els = driver.find_elements(By.ID, fid)
            except Exception:
                els = []
            if not els or not els[0].is_displayed():
                continue
            el = els[0]
            try:
                if ftype == "select":
                    current = None
                    try:
                        current = Select(el).first_selected_option.text.strip()
                    except Exception:
                        pass
                    if current and current.strip().lower() == str(expected).strip().lower():
                        break
                    _select_option(driver, fid, expected, fname, log=log)
                else:
                    if not _hard_clear_input(driver, el, log=log):
                        log(f"  ⚠ No se pudo vaciar '{fname}' ({fid}) — se omite el "
                            f"re-ingreso para no duplicar el valor")
                        break
                    time.sleep(0.05)
                    for char in str(expected):
                        el.send_keys(char)
                        time.sleep(0.005)
                    driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
                        "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
                        el,
                    )
                    # El send_keys puede concatenar si el campo tiene máscara o si React
                    # repuso el valor anterior: verificar y corregir por JS.
                    try:
                        _final = (el.get_attribute("value") or "")
                    except Exception:
                        _final = str(expected)
                    if _norm_cmp(_final) != _norm_cmp(expected) and _final:
                        _maxlen = _get_maxlength(el)
                        _target = str(expected)[:_maxlen] if _maxlen else str(expected)
                        if _norm_cmp(_final) != _norm_cmp(_target):
                            log(f"  ⚠ '{fname}' ({fid}) quedó '{_final}' tras el re-ingreso "
                                f"(se esperaba '{_target}') — corrigiendo por JS")
                            _set_input_value_js(driver, el, _target)
                    log(f"  ↺ Re-ingresado '{fname}' ({fid}) via send_keys: '{expected}'")
            except Exception as e:
                log(f"  ✗ Error re-ingresando '{fname}' ({fid}): {e}")
            break


_INFER_DATAKEY_MAP = {
    "firstname":  ("nombre", "name", "first", "fname", "nome", "given"),
    "lastname":   ("apellido", "lastname", "surname", "lname", "sobrenome", "family"),
    "email":      ("email", "correo", "mail", "e-mail", "emailaddress"),
    "phone":      ("telefono", "celular", "phone", "mobile", "cel", "tel", "fono", "movil",
                   "whatsapp", "celphone", "telefone"),
    "document":   ("documento", "doc", "cedula", "dni", "rut", "ci", "cpf",
                   "rfc", "identif", "passport", "pasaporte"),
    "model":      ("modelo", "model", "vehiculo", "veiculo", "auto", "car"),
}


def _normalize_for_infer(text):
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", (text or "").lower())
    ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
    return ascii_str.replace(" ", "").replace("-", "").replace("_", "")


def _infer_data_key(field_id, field_name, field_placeholder=""):
    tokens = [
        _normalize_for_infer(field_id),
        _normalize_for_infer(field_name),
        _normalize_for_infer(field_placeholder),
    ]
    for data_key, keywords in _INFER_DATAKEY_MAP.items():
        for token in tokens:
            if token and any(kw in token or token in kw for kw in keywords):
                return data_key
    return None


def _get_inferred_value(inferred_key: str, lead) -> str:
    if not lead or not inferred_key:
        return ""
    lead_lower = {k.lower().strip(): v for k, v in lead.data.items()}
    INFERRED_TO_HEADERS = {
        "firstname": ["firstname", "name", "nombre", "nombres", "fullname", "full_name"],
        "lastname":  ["lastname", "apellido", "apellidos", "last_name"],
        "email":     ["email", "correo", "e-mail"],
        "phone":     ["phone", "celular", "telefono", "cellphone", "telephone", "telephone-mask"],
        "document":  ["document", "documento", "cpf", "cnpj", "dni", "rut", "ci"],
        "model":     ["model", "modelo", "models"],
    }
    headers = INFERRED_TO_HEADERS.get(inferred_key, [inferred_key])
    for h in headers:
        val = lead_lower.get(h, "")
        if val and str(val).strip().lower() not in ("", "none"):
            return str(val)
    return str(lead_lower.get(inferred_key, ""))


# ══════════════════════════════════════════════════════════════════════════════
# SNAPSHOT REAL DEL FORM ANTES DE ENVIAR
# El tracking normal guarda lo que se *intentó* escribir en cada paso. Si después el
# form re-renderiza, un dropdown se resetea o el auto-relleno de campos no mapeados
# vuelve a randomizar un select (models, fecha estimada de compra, ...), lo que viaja
# en el lead deja de ser lo trackeado. Por eso, justo antes del click en Enviar se lee
# el DOM real y se pisa el tracking con el valor efectivo: el Excel de resultados debe
# reflejar SIEMPRE lo último que quedó en el form, que es lo que se compara contra la DB.
# ══════════════════════════════════════════════════════════════════════════════

_SNAPSHOT_JS = r"""
var out = [];
var els = document.querySelectorAll("input, select, textarea");
for (var i = 0; i < els.length; i++) {
    var e = els[i];
    var tag = e.tagName.toLowerCase();
    var type = (e.type || "").toLowerCase();
    if (tag === "input" && (type === "hidden" || type === "submit" ||
                            type === "button" || type === "reset" || type === "file")) continue;
    var id = e.id || e.getAttribute("name") || "";
    if (!id) continue;
    var value = "", text = "";
    if (tag === "select") {
        var sel = e.multiple
            ? [].filter.call(e.options, function(o){return o.selected;})
            : (e.selectedIndex >= 0 ? [e.options[e.selectedIndex]] : []);
        value = sel.map(function(o){return o.value;}).join(" | ");
        text  = sel.map(function(o){return (o.text||"").trim();}).join(" | ");
    } else if (type === "checkbox" || type === "radio") {
        if (!e.checked) continue;
        value = e.value || "on";
        // El 'value' de un radio suele ser un código ("1", "3m"). Lo legible —y lo que el
        // usuario escribe en el Excel— es el title o el texto del label: sin esto el
        // resultado mostraba el código y "Datos vs Excel" marcaba diferencias falsas.
        var lbl = e.getAttribute("title") || e.getAttribute("aria-label") || "";
        if (!lbl && e.id) {
            var lb = document.querySelector('label[for="' + e.id + '"]');
            if (lb) lbl = lb.textContent;
        }
        if (!lbl && e.closest) { var anc = e.closest("label"); if (anc) lbl = anc.textContent; }
        text = String(lbl || value).replace(/ /g, " ").replace(/\s+/g, " ").trim();
    } else {
        value = e.value || "";
        text  = value;
    }
    out.push({
        id: id,
        realId: e.id || "",
        name: e.getAttribute("name") || "",
        tag: tag,
        type: type,
        value: value,
        text: text,
        visible: !!(e.getClientRects().length && getComputedStyle(e).visibility !== "hidden")
    });
}
return out;
"""


def _snapshot_form_state(driver, log: Callable = print) -> Dict[str, Dict]:
    """Lee el estado REAL de todos los campos del form (un solo JS call).

    Devuelve {id_o_name: {...}}. Incluye campos no visibles (pasos anteriores de un
    wizard): siguen en el DOM y su value es lo que efectivamente se envía.
    """
    try:
        raw = driver.execute_script(_SNAPSHOT_JS) or []
    except Exception as e:
        log(f"  ⚠ No se pudo leer el estado final del form: {e}")
        return {}
    snap: Dict[str, Dict] = {}
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        for key in (entry.get("realId"), entry.get("name"), entry.get("id")):
            key = (key or "").strip()
            if not key:
                continue
            prev = snap.get(key)
            # Si hay duplicados de id/name, gana el que tenga valor (y entre esos, el visible)
            if prev is None:
                snap[key] = entry
            elif not (prev.get("text") or "").strip() and (entry.get("text") or "").strip():
                snap[key] = entry
            elif entry.get("visible") and not prev.get("visible") and (entry.get("text") or "").strip():
                snap[key] = entry
    return snap


def _final_value_of(entry: Dict) -> str:
    """Valor legible de una entrada del snapshot (texto de la option para selects)."""
    if not entry:
        return ""
    txt = (entry.get("text") or "").strip()
    if txt:
        return txt
    return (entry.get("value") or "").strip()


def _lead_value_for(lead, field_name: str, field_id) -> str:
    """Valor pedido en el Excel para un campo del mapping (misma lógica que
    fill_form_fields._get_value, expuesta a nivel de módulo para poder comparar
    'lo pedido' contra 'lo que realmente quedó')."""
    if lead is None:
        return ""
    lead_lower = {str(k).lower().strip(): v for k, v in (lead.data or {}).items()}
    ids = field_id if isinstance(field_id, list) else [field_id]
    val = lead_lower.get((field_name or "").lower().strip(), "")
    if val and str(val).lower() not in ("", "none"):
        return str(val)
    for fid in ids:
        val = lead_lower.get((fid or "").lower(), "")
        if val and str(val).lower() not in ("", "none"):
            return str(val)
    for header, mapped_ids in _HEADER_TO_IDS.items():
        if any(mid in ids for mid in mapped_ids):
            val = lead_lower.get(header, "")
            if val and str(val).lower() not in ("", "none"):
                return str(val)
    return ""


def _norm_cmp(text) -> str:
    """Normaliza para comparar valores (sin acentos/espacios/mayúsculas)."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(text or "").replace("\xa0", " ").lower().strip())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def _sync_tracked_with_dom(driver, field_mapping: List[Dict], lead,
                           all_tracked: Dict[str, str],
                           log: Callable = print) -> Dict:
    """Pisa el tracking con el estado REAL del form justo antes de enviar.

    - Cada `PasoN::campo` pasa a valer lo que el campo tiene en el DOM en este momento
      (si el campo sigue existiendo). Así, si un dropdown se re-eligió más tarde, queda
      el ÚLTIMO valor — el que viaja en el lead.
    - Los selects que nadie trackeó (no mapeados, elegidos al azar) se agregan como
      `Final::<id>` para que el random elegido también quede registrado.
    - Compara contra lo pedido en el Excel y devuelve las diferencias.

    Devuelve {"tracked", "diffs", "excel_mismatch", "excel_mismatch_present"}.
    """
    snap = _snapshot_form_state(driver, log=log)
    out = {
        "tracked": dict(all_tracked or {}),
        "diffs": [],
        "excel_mismatch": "",
        "excel_mismatch_present": False,
    }
    if not snap:
        return out

    # name del mapping → ids HTML posibles (+ alias visid)
    name_to_ids: Dict[str, List[str]] = {}
    id_to_name: Dict[str, str] = {}
    mapped_ids: set = set()
    for fc in field_mapping or []:
        fid_raw = fc.get("id", "")
        fname = fc.get("name", str(fid_raw))
        ids = [f for f in (fid_raw if isinstance(fid_raw, list) else [fid_raw]) if f]
        for f in list(ids):
            alias = _VISID_ID_ALIASES.get(f)
            if alias and alias not in ids:
                ids.append(alias)
        name_to_ids.setdefault(fname, []).extend(ids)
        for f in ids:
            mapped_ids.add(str(f))
            id_to_name.setdefault(str(f), fname)

    def _resolve_entry(names_or_ids: List[str]):
        """Primer campo del snapshot que matchee alguno de los ids (exacto o por prefijo
        'id-xxx', que es como los forms AEM/gm_front renombran los inputs)."""
        for cand in names_or_ids:
            cand = str(cand or "")
            if not cand:
                continue
            entry = snap.get(cand)
            if entry and _final_value_of(entry):
                return entry
        for cand in names_or_ids:
            cand = str(cand or "")
            if not cand:
                continue
            for key, entry in snap.items():
                if key.startswith(cand + "-") and _final_value_of(entry):
                    return entry
        return None

    covered_ids: set = set()
    tracked_keys_seen: set = set()

    for key in list(out["tracked"].keys()):
        raw_name = key.split("::", 1)[1] if "::" in key else key
        tracked_keys_seen.add(raw_name)
        candidates = list(name_to_ids.get(raw_name, []))
        # El tracking de AEM / model-from-url usa el id como clave, no el name
        if raw_name not in name_to_ids:
            candidates.append(raw_name)
        entry = _resolve_entry(candidates)
        if not entry:
            continue
        eid = entry.get("realId") or entry.get("name") or ""
        if eid:
            covered_ids.add(eid)
        real_value = _final_value_of(entry)
        if not real_value:
            continue
        old_value = str(out["tracked"].get(key, "")).strip()
        if _norm_cmp(old_value) != _norm_cmp(real_value):
            out["tracked"][key] = real_value
            out["diffs"].append(f"{raw_name}: '{old_value}' → '{real_value}'")
            log(f"  🔄 {raw_name}: el form quedó con '{real_value}' "
                f"(trackeado era '{old_value}') — se registra el valor real")

    # Campos con valor que nadie trackeó: selects randomizados por el auto-relleno de no
    # mapeados, o campos del mapping que se completaron por fuera de fill_form_fields.
    # Sin esto, lo que realmente viajó en el lead no queda en ningún lado.
    for key, entry in snap.items():
        eid = entry.get("realId") or entry.get("name") or ""
        if not eid or eid != key:
            continue
        if eid in covered_ids:
            continue
        # Solo selects: un input de texto sin trackear suele ser ruido de la landing
        # (buscador, newsletter), mientras que un select con valor es una elección real
        # que viaja en el lead.
        if entry.get("tag") != "select":
            continue
        value = _final_value_of(entry)
        if not value or _is_placeholder(value):
            continue
        label = id_to_name.get(eid, eid)
        if label in tracked_keys_seen:
            continue
        out["tracked"][f"Final::{label}"] = value
        log(f"  📎 '{label}' ({eid}) quedó en '{value}' sin estar trackeado "
            f"— registrado como Final::{label}")

    # ── Comparación contra lo pedido en el Excel ─────────────────────────────
    # Sólo campos de elección (selects): en los de texto el propio flujo transforma el
    # dato a propósito (CPF/CNPJ/CEP regenerados en Brasil, documento saneado en Perú,
    # recorte por maxlength) y compararlos daría avisos falsos todo el tiempo.
    mismatches = []
    for fc in field_mapping or []:
        if fc.get("type", "text") != "select":
            continue
        fid_raw = fc.get("id", "")
        fname = fc.get("name", str(fid_raw))
        expected = _lead_value_for(lead, fname, fid_raw)
        if not expected or _is_placeholder(expected):
            continue
        entry = _resolve_entry(name_to_ids.get(fname, []))
        if not entry:
            continue
        real_value = _final_value_of(entry)
        if not real_value:
            mismatches.append(f"{fname}: pedido '{expected}' pero quedó vacío")
            continue
        exp_n, real_n = _norm_cmp(expected), _norm_cmp(real_value)
        if exp_n == real_n or exp_n in real_n or real_n in exp_n:
            continue
        mismatches.append(f"{fname}: pedido '{expected}' → quedó '{real_value}'")

    if mismatches:
        out["excel_mismatch"] = " ; ".join(mismatches)
        out["excel_mismatch_present"] = True
        log(f"  ⚠ Datos distintos a los pedidos en el Excel: {out['excel_mismatch']}")
    else:
        out["excel_mismatch"] = "OK"

    return out


# Valores de ?model= que NO son modelos (son servicios/secciones): no deben registrarse
# como "modelo elegido". Mismo criterio que base_form_filler._NON_MODEL_URL_TOKENS.
_NON_MODEL_URL_TOKENS = (
    "posventa", "postventa", "pos venta", "onstar", "servicio", "service",
    "agendamiento", "seminuevos", "seminovos", "revision", "revisao",
    "testdrive", "test drive", "cotizacion", "cotizar", "contacto", "contato",
    "oferta", "raq", "suscripcion", "renovacion", "financiamiento", "repuestos",
    "accesorios", "acessorios", "acessilab", "garantia", "posventa gral", "gral",
)


def _record_model_from_url_if_missing_lt(driver, field_mapping: List[Dict],
                                          tracked: Dict[str, str],
                                          log: Callable = print,
                                          expected_form_url: str = "") -> None:
    """Si ningún select de Modelo quedó registrado, usar el ?model= de la URL del FORM.

    OJO: estando dentro del iframe, `driver.current_url` devuelve la URL de la LANDING,
    que no lleva ?model=. Por eso se mira primero la URL esperada del form (columna
    'Formulario' del Excel), que es la que trae el modelo parametrizado. Sin esto, los
    forms que reciben el modelo por URL (movs-visid, accesorios, ...) quedaban sin
    ningún modelo registrado en el Excel de resultados.
    """
    # tracked está indexado por NOMBRE del mapping ('Modelo'), no por id HTML
    _model_names = {"modelo", "model", "models"}
    for k, v in (tracked or {}).items():
        if str(k).strip().lower() in _model_names and str(v).strip():
            return

    from urllib.parse import urlsplit, parse_qs, unquote
    url_model = ""
    for _url in (expected_form_url or "", (driver.current_url or "")):
        if not _url:
            continue
        try:
            raw = parse_qs(urlsplit(_url).query).get("model", [""])[0]
            candidate = unquote(raw).replace("+", " ").strip()
        except Exception:
            continue
        if not candidate:
            continue
        norm = _norm_cmp(candidate)
        if any(_norm_cmp(tok) in norm for tok in _NON_MODEL_URL_TOKENS):
            continue  # es un servicio/sección, no un modelo
        url_model = candidate
        break

    if not url_model:
        return
    model_name = "Modelo"
    for fc in (field_mapping or []):
        fid = fc.get("id")
        fid = fid[0] if isinstance(fid, list) else fid
        if fid in ("models", "model"):
            model_name = fc.get("name") or model_name
            break
    tracked[model_name] = url_model
    log(f"🚗 Modelo (de la URL ?model=) = {url_model}")


def _auto_fill_unmapped_dropdowns_lt(driver, field_mapping: List[Dict],
                                      ids_dinamicos: Dict[str, str],
                                      lead=None,
                                      log: Callable = print,
                                      tracked: Dict[str, str] = None,
                                      filled_ids: set = None) -> bool:
    """
    Completa selects e inputs visibles que no están en el field_mapping.
    Réplica de base_form_filler._auto_fill_unmapped_dropdowns para Safari LambdaTest.

    filled_ids: IDs HTML reales que ya llenó fill_form_fields en esta fila. Es lo que
    evita que un select ya resuelto (p. ej. 'estimated-day', cuyo id real no coincide con
    el id del mapping 'estimated-date-purchase') se vuelva a randomizar y pise el valor
    del Excel.
    """
    mapped_ids = set()
    for fc in field_mapping or []:
        fid_raw = fc.get("id", "")
        fids = fid_raw if isinstance(fid_raw, list) else [fid_raw]
        for fid in fids:
            if fid:
                mapped_ids.add(str(fid))

    # Ids ya llenados por el mapping en esta fila (id real del form ≠ id de config,
    # ej. 'estimated-day' vs 'estimated-date-purchase') → no re-seleccionar y pisar.
    # OJO: tracked está indexado por NOMBRE de campo del mapping, no por id HTML; por eso
    # hace falta filled_ids (ids reales resueltos por fill_form_fields) para que el guard
    # sirva de algo.
    already_filled_ids = set(tracked.keys()) if isinstance(tracked, dict) else set()
    if filled_ids:
        already_filled_ids |= {str(x) for x in filled_ids}

    filled_any = False

    # ── Selects no mapeados ───────────────────────────────────────────────────
    try:
        select_elements = driver.find_elements(By.XPATH, "//select")
    except Exception as e:
        log(f"  Error buscando selects no mapeados: {e}")
        return False

    for sel_el in select_elements:
        sel_id = None
        try:
            sel_id = sel_el.get_attribute("id") or sel_el.get_attribute("name")
            if not sel_id:
                continue
            real_id = sel_el.get_attribute("id") or ""
            if real_id and real_id in mapped_ids:
                continue
            if (real_id and real_id in already_filled_ids) or (sel_id in already_filled_ids):
                continue
            if not sel_el.is_displayed() or not sel_el.is_enabled():
                continue
            if sel_el.get_attribute("multiple") is not None:
                continue

            valid_opts = _get_valid_options(sel_el)
            if not valid_opts:
                continue

            chosen = None
            from_excel = False
            # Prioridad 1: valor pedido en el Excel para ese campo. Un select puede quedar
            # fuera del field_mapping (form nuevo, id distinto) y aun así tener su columna
            # cargada en el Excel: antes se sorteaba igual y se perdía el dato del usuario.
            excel_val = _lead_value_for(lead, sel_id, [real_id, sel_id]) if lead else ""
            if excel_val and not _is_placeholder(excel_val):
                exp = _norm_cmp(excel_val)
                chosen = next((o for o in valid_opts if _norm_cmp(o.text) == exp), None)
                if chosen is None:
                    chosen = next((o for o in valid_opts
                                   if exp and exp in _norm_cmp(o.text)), None)
                if chosen is None:
                    chosen = next((o for o in valid_opts
                                   if _norm_cmp(o.get_attribute("value") or "") == exp), None)
                if chosen is None:
                    log(f"  ⚠ Select '{sel_id}': el Excel pide '{excel_val}' pero no hay "
                        f"esa opción → se elige una al azar")
                else:
                    from_excel = True

            if chosen is None and sel_id in ids_dinamicos:
                target_val = ids_dinamicos[sel_id]
                chosen = next((o for o in valid_opts if o.text.strip() == target_val), None)
            if chosen is None:
                chosen = random.choice(valid_opts)

            opt_val  = chosen.get_attribute("value") or chosen.text.strip()
            opt_text = chosen.text.strip()

            try:
                Select(sel_el).select_by_value(opt_val)
            except Exception:
                try:
                    Select(sel_el).select_by_visible_text(opt_text)
                except Exception:
                    pass

            driver.execute_script(
                "var el=arguments[0],val=arguments[1];"
                "try{Object.getOwnPropertyDescriptor(window.HTMLSelectElement.prototype,'value').set.call(el,val);}catch(e){}"
                "el.dispatchEvent(new Event('focus',{bubbles:true}));"
                "el.dispatchEvent(new Event('change',{bubbles:true,cancelable:true}));"
                "el.dispatchEvent(new Event('input',{bubbles:true,cancelable:true}));"
                "el.dispatchEvent(new Event('blur',{bubbles:true}));",
                sel_el, opt_val
            )
            time.sleep(0.1)
            log(f"  {'✓' if from_excel else '🎲'} Select no mapeado '{sel_id}' → '{opt_text}'"
                f"{' (pedido en el Excel)' if from_excel else ''}")
            # Registrar el random elegido: si no queda trackeado, el Excel de resultados
            # no muestra qué se envió realmente en ese campo.
            if isinstance(tracked, dict):
                tracked[real_id or sel_id] = opt_text
            filled_any = True
        except Exception as e:
            log(f"  Select no mapeado '{sel_id}' error: {e}")
            continue

    # ── Inputs/textareas no mapeados ──────────────────────────────────────────
    try:
        text_elements = driver.find_elements(
            By.XPATH,
            "//input[not(@type='hidden')][not(@type='submit')][not(@type='button')]"
            "[not(@type='checkbox')][not(@type='radio')][not(@type='file')][not(@type='reset')]"
            " | //textarea"
        )
    except Exception as e:
        log(f"  Error buscando inputs no mapeados: {e}")
        text_elements = []

    for element in text_elements:
        field_id = None
        try:
            field_id = element.get_attribute("id") or element.get_attribute("name")
            if not field_id or field_id in mapped_ids:
                continue
            if not element.is_displayed() or not element.is_enabled():
                continue
            current_val = (element.get_attribute("value") or "").strip()
            if current_val:
                continue

            fill_value = None
            source_label = ""

            # Prioridad 1: IDs dinámicos
            if field_id in ids_dinamicos:
                fill_value = ids_dinamicos.get(field_id, "")
                source_label = "ID dinámico"

            # Prioridad 2: inferir data_key y tomar valor del lead
            if not fill_value and lead:
                field_name = (
                    element.get_attribute("aria-label")
                    or element.get_attribute("placeholder")
                    or element.get_attribute("name")
                    or field_id
                )
                placeholder = element.get_attribute("placeholder") or ""
                inferred_key = _infer_data_key(field_id, field_name, placeholder)
                if inferred_key:
                    inferred_val = _get_inferred_value(inferred_key, lead)
                    if inferred_val:
                        fill_value = inferred_val
                        source_label = f"inferido→{inferred_key}"

            # Prioridad 3: probe numérico → alfa para inputs requeridos sin valor
            if not fill_value:
                cls = element.get_attribute("class") or ""
                has_error_class = any(c in cls for c in ("error", "invalid", "is-invalid", "ng-invalid"))
                is_req = bool(
                    element.get_attribute("required")
                    or element.get_attribute("aria-required") == "true"
                    or has_error_class
                )
                if not is_req:
                    continue

                input_type = (element.get_attribute("type") or "text").lower()
                inputmode = (element.get_attribute("inputmode") or "").lower()
                likely_num = inputmode in ("numeric", "tel", "decimal") or input_type == "tel"
                probes = ["12345678", "Carlos"] if likely_num else ["Carlos", "12345678"]

                ml = _get_maxlength(element)
                fill_value = probes[0]
                if ml and len(fill_value) > ml:
                    fill_value = fill_value[:ml]

                for probe in probes:
                    if ml and len(probe) > ml:
                        probe = probe[:ml]
                    _fill_text_js(driver, element, probe)
                    time.sleep(0.15)
                    cls_after = element.get_attribute("class") or ""
                    if not any(c in cls_after for c in ("error", "invalid", "is-invalid", "ng-invalid")):
                        fill_value = probe
                        source_label = "auto-probe"
                        break

                log(f"  📌 Input no mapeado '{field_id}' → '{fill_value}' ({source_label})")
                filled_any = True
                continue

            # Rellenar con valor dinámico o inferido respetando maxlength
            ml = _get_maxlength(element)
            if ml and len(fill_value) > ml:
                fill_value = fill_value[:ml]
            _fill_text_js(driver, element, fill_value)
            log(f"  📌 Input no mapeado '{field_id}' → '{fill_value}' ({source_label})")
            filled_any = True
        except Exception as e:
            log(f"  Input no mapeado '{field_id}' error: {e}")
            continue

    if not filled_any:
        log("  ℹ No se encontraron campos no mapeados para auto-completar")
    return filled_any


# ══════════════════════════════════════════════════════════════════════════════
# BOTONES SIGUIENTE / ENVIAR
# Lógica copiada de Osocio: _is_next_button_element, _find_next_button,
# _click_next_button, submit_and_verify_form
# ══════════════════════════════════════════════════════════════════════════════

_NEXT_KEYWORDS = (
    "siguiente", "seguinte", "continuar", "continuacao",
    "proximo", "próximo", "next",
)

_SUBMIT_KEYWORDS = (
    "enviar", "submit", "confirmar", "finalizar", "enviar dados",
)

# Selectores CSS para encontrar el botón Siguiente (igual que Osocio)
_NEXT_SELECTORS = [
    ".button.next.pulsate.stat-button-link",
    "button[class*='next']",
    "button[data-dtm*='next']",
    "button[type='button']",
    ".next-button",
]

# Selectores CSS para el botón Enviar (igual que Osocio submit_and_verify_form)
_SUBMIT_SELECTORS = [
    "button.submit-button.stat-button-link",
    "button.btn-visid-submit.stat-button-link",
    "button.submit-button",
    "button[type='submit']",
    "input[type='submit']",
    "button[class*='submit']",
    "button[type='button'][class*='submit']",
    # AEM/T3 (guide container): en Safari/Mac el botón real puede ser este widget
    # específico en vez de un button[type=submit] genérico (igual que base_form_filler.py).
    'button[name="guideContainer-rootPanel-guidebutton___jqName"]',
    'button[type="submit"][aria-label="Enviar"]',
    "button[id*='guidebutton']",
    "button[name*='guidebutton']",
    "[id*='guidebutton'] button",
]


def _normalize_btn_text(text: str) -> str:
    """Normaliza texto de botón para comparación. Igual que Osocio._normalize_text."""
    import unicodedata
    t = (text or "").lower().strip()
    # Quitar acentos
    nfkd = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return t


def _is_next_button_element(driver, element) -> bool:
    """
    True si el elemento parece ser un botón para avanzar al siguiente paso.
    Copiado exacto de Osocio: chequea text, data-dtm, class, aria-label, title.
    """
    try:
        if not element.is_displayed():
            return False
        parts = [
            element.text or "",
            element.get_attribute("data-dtm") or "",
            element.get_attribute("class") or "",
            element.get_attribute("aria-label") or "",
            element.get_attribute("title") or "",
        ]
        normalized = " ".join(_normalize_btn_text(p) for p in parts if p)
        return any(kw in normalized for kw in _NEXT_KEYWORDS)
    except Exception:
        return False


def _find_next_button(driver):
    """
    Devuelve el primer botón visible que parezca avanzar al siguiente paso.
    Copiado exacto de Osocio: mismos selectores CSS + _is_next_button_element.
    """
    for selector in _NEXT_SELECTORS:
        try:
            btns = driver.find_elements(By.CSS_SELECTOR, selector)
            for btn in btns:
                if _is_next_button_element(driver, btn):
                    return btn
        except Exception:
            continue
    return None


def _has_next_button(driver) -> bool:
    """True si hay botón Siguiente visible. Igual que Osocio._has_next_button."""
    try:
        return _find_next_button(driver) is not None
    except Exception:
        return False


def _click_next_button(driver, log: Callable = print) -> bool:
    """
    Hace clic en el botón Siguiente.
    Copiado de Osocio: scroll → JS click → fallback click normal.
    """
    try:
        btn = _find_next_button(driver)
        if not btn:
            log("  No se encontró botón 'Siguiente' visible")
            return False

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", btn)
        log("  ✓ Siguiente clickeado")
        return True

    except Exception:
        try:
            btn = _find_next_button(driver)
            if btn:
                btn.click()
                log("  ✓ Siguiente clickeado (fallback)")
                return True
        except Exception:
            pass
        log("  ✗ Error al clickear Siguiente")
        return False


_DOM_SIGNATURE_JS = r"""
var cfgs = arguments[0], sig = [];
for (var i = 0; i < cfgs.length; i++) {
    var ids = cfgs[i];
    for (var j = 0; j < ids.length; j++) {
        var e = document.getElementById(ids[j]);
        if (e && e.getClientRects().length && getComputedStyle(e).visibility !== 'hidden') {
            sig.push(ids[j]);
            break;
        }
    }
}
return sig;
"""


def _dom_signature_visible(driver, field_mapping: List[Dict]) -> tuple:
    """
    Firma del DOM: IDs del mapping que están visibles ahora, para detectar si el DOM
    cambió tras clickear "Siguiente".

    Un solo execute_script: antes hacía un find_elements + is_displayed() por ID (~26
    round-trips a LambdaTest, ~10s por llamada contra un device real), y se llama varias
    veces por cada transición de paso.
    """
    ids_batch = []
    for fc in field_mapping or []:
        fid_raw = fc.get("id", "")
        ids = fid_raw if isinstance(fid_raw, list) else [fid_raw]
        ids_batch.append([f for f in ids if f])
    if not ids_batch:
        return tuple()
    try:
        sig = driver.execute_script(_DOM_SIGNATURE_JS, ids_batch) or []
    except Exception:
        return tuple()
    return tuple(sorted(sig))


# ¿El form reaccionó al click en Enviar? (TY visible, form removido del DOM, botón en
# loading/disabled, o errores de validación pintados). Si nada de eso pasó, el click no llegó.
_SUBMIT_REACTED_JS = r"""
var ty = document.getElementById('thank-you');
if (ty && ty.getClientRects().length) return true;
if (document.querySelector('div.rp-wrapper')) return true;
var form = document.getElementById('formulario');
if (form && !form.getClientRects().length) return true;   // form oculto/removido tras enviar
var sb = document.querySelector('button.submit-button');
if (sb && (sb.disabled || /loading/.test(sb.className || ''))) return true;
return [].some.call(document.querySelectorAll('span.error, label.error'), function(e){
    return e.textContent.trim() !== '' && e.getClientRects().length;
});
"""


def _click_and_verify(driver, btn, log: Callable = print) -> None:
    """
    Click nativo y, si el form no reaccionó, re-click por JS.

    En LambdaTest el click nativo es un tap por coordenadas: en algunos forms (raq-eletricos)
    no aterriza en el botón y Selenium NO lanza excepción, así que el fallback por except
    nunca se activaba y el submit no se disparaba nunca. Sólo re-clickea si el form no
    reaccionó, así que no puede duplicar el envío.
    """
    try:
        btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", btn)
        return
    time.sleep(1.2)
    try:
        if driver.execute_script(_SUBMIT_REACTED_JS):
            return
    except Exception:
        return
    log("  ↺ El click nativo no disparó el submit — reintentando por JS")
    try:
        driver.execute_script("arguments[0].click();", btn)
    except Exception:
        pass


def _click_submit(driver, log: Callable = print, is_android: bool = False) -> bool:
    """
    Hace clic en el botón Enviar.
    Igual que Osocio submit_and_verify_form: mismos selectores + XPath por texto.
    Android usa timeout reducido (1s vs 3s) ya que Chrome es más rápido que Safari.
    """
    _timeout = 1.0 if is_android else 3

    # Mac/Safari: el botón Enviar suele quedar por debajo del fold y el click no
    # registra si no está en viewport. Revelamos el fondo del formulario primero
    # (equivalente a "scrollear un poco más abajo" que hacía el usuario a mano).
    if not is_android:
        try:
            _scroll_to_bottom_parent_aware(driver, log)
        except Exception:
            pass

    # Selectores CSS específicos de GM (igual que Osocio)
    for sel in _SUBMIT_SELECTORS:
        try:
            btn = WebDriverWait(driver, _timeout).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
            )
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", btn
            )
            time.sleep(0.3 if is_android else 0.5)
            _click_and_verify(driver, btn, log)
            log(f"  ✓ Enviar clickeado ({sel})")
            return True
        except TimeoutException:
            continue
        except Exception:
            continue

    # Fallback XPath por texto (igual que Osocio)
    try:
        for xpath in [
            "//button[contains(@class,'submit')][.//span[contains(normalize-space(.),'Enviar')] or contains(normalize-space(.),'Enviar')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÑ','abcdefghijklmnopqrstuvwxyzáéíóúñ'),'enviar')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'submit')]",
            "//button[contains(@class,'submit') and not(contains(@class,'next'))]",
        ]:
            try:
                btn = WebDriverWait(driver, _timeout).until(
                    EC.element_to_be_clickable((By.XPATH, xpath))
                )
                driver.execute_script("arguments[0].click();", btn)
                log(f"  ✓ Enviar clickeado (XPath)")
                return True
            except TimeoutException:
                continue
    except Exception:
        pass

    # Si no se encontró inicialmente, scrolleamos al fondo e intentamos una vez más
    log("  ✗ No se encontró botón de envío inicialmente, scrolleando abajo e intentando de nuevo...")
    try:
        _scroll_to_bottom_parent_aware(driver, log)
    except Exception:
        pass

    for sel in _SUBMIT_SELECTORS:
        try:
            btn = WebDriverWait(driver, _timeout).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
            )
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", btn
            )
            time.sleep(0.3 if is_android else 0.5)
            _click_and_verify(driver, btn, log)
            log(f"  ✓ Enviar clickeado ({sel}) tras scroll")
            return True
        except Exception:
            continue

    try:
        for xpath in [
            "//button[contains(@class,'submit')][.//span[contains(normalize-space(.),'Enviar')] or contains(normalize-space(.),'Enviar')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÑ','abcdefghijklmnopqrstuvwxyzáéíóúñ'),'enviar')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'submit')]",
            "//button[contains(@class,'submit') and not(contains(@class,'next'))]",
        ]:
            try:
                btn = WebDriverWait(driver, _timeout).until(
                    EC.element_to_be_clickable((By.XPATH, xpath))
                )
                driver.execute_script("arguments[0].click();", btn)
                log(f"  ✓ Enviar clickeado (XPath) tras scroll")
                return True
            except Exception:
                continue
    except Exception:
        pass

    # Último recurso (Mac/Safari): un scroll más, sin tocar el zoom — reducir el zoom
    # deforma cómo se ve/renderiza el form (no representativo) y nunca funcionó en la
    # práctica, así que se saca directamente en vez de seguir intentándolo en vano.
    if not is_android:
        try:
            _scroll_to_bottom_parent_aware(driver, log)
            time.sleep(0.4)
            for sel in _SUBMIT_SELECTORS:
                try:
                    btn = WebDriverWait(driver, _timeout).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.4)
                    _click_and_verify(driver, btn, log)
                    log(f"  ✓ Enviar clickeado ({sel}) tras scroll adicional")
                    return True
                except Exception:
                    continue
        except Exception:
            pass

    # Fallback genérico: barrer TODOS los botones/inputs submit visibles dentro de un <form>
    # y clickear el que no sea 'Siguiente' ni cierre/cookie. Ataca el caso de "no encuentra el
    # botón que sí está" cuando el markup no matchea ninguno de los selectores fijos.
    _send_words = ("enviar", "submit", "solicitar", "finalizar", "confirmar",
                   "quero", "cadastrar", "registrar", "receber")
    try:
        _cands = driver.find_elements(
            By.CSS_SELECTOR, "form button, form input[type='submit'], form [role='button']")
        if not _cands:
            _cands = driver.find_elements(
                By.CSS_SELECTOR, "button, input[type='submit'], [role='button']")
        for el in _cands:
            try:
                if not (el.is_displayed() and el.is_enabled()):
                    continue
                if _is_next_button_element(driver, el):
                    continue
                blob = _normalize_btn_text(" ".join(filter(None, [
                    el.text or "", el.get_attribute("class") or "",
                    el.get_attribute("aria-label") or "", el.get_attribute("id") or "",
                    el.get_attribute("type") or "",
                ])))
                if any(bad in blob for bad in ("cookie", "cerrar", "close", "fechar")):
                    continue
                _is_type_submit = (el.get_attribute("type") or "").lower() == "submit"
                if _is_type_submit or any(w in blob for w in _send_words):
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    time.sleep(0.3)
                    _click_and_verify(driver, el, log)
                    log("  ✓ Enviar clickeado (scan genérico del form)")
                    return True
            except Exception:
                continue
    except Exception:
        pass

    log("  ✗ No se encontró botón de envío")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# RESULTADOS (lógica de Osocio: mismo Excel, mismas columnas)
# ══════════════════════════════════════════════════════════════════════════════

def _get_run_number(pais: str, plat_clean: str, prefix: str = "resultados_") -> int:
    """Número incremental igual que Osocio."""
    import glob
    pattern = os.path.join(_RESULTADOS_DIR, f"{prefix}{pais}_{plat_clean}*.xlsx")
    matches = glob.glob(pattern)
    max_n = 0
    for m in matches:
        base = os.path.basename(m).replace(f"{prefix}{pais}_{plat_clean}", "").replace(".xlsx", "")
        if base.isdigit():
            max_n = max(max_n, int(base))
    return max_n + 1


def _fetch_lt_video_url(session_id: str, username: str, access_key: str) -> str:
    """Obtiene la URL del video grabado de la sesión vía LambdaTest API."""
    if not _REQUESTS_OK or not session_id:
        return ""
    try:
        resp = _requests.get(
            f"https://api.lambdatest.com/automation/api/v1/sessions/{session_id}",
            auth=(username, access_key),
            timeout=15,
        )
        data = resp.json().get("data", {})
        return data.get("video_url") or data.get("video") or ""
    except Exception:
        return ""


def _setup_results_excel(pais: str, source_excel_path: str, platform: str, build_name: str = "") -> tuple:
    """
    Copia el Excel de entrada y agrega columnas de resultado:
    Form Encontrado, Enviado, Resultado, Video LT
    """
    os.makedirs(_RESULTADOS_DIR, exist_ok=True)
    plat_clean = "Mac" if "mac" in platform.lower() else ("Android" if "android" in platform.lower() else ("Iphone" if "iphone" in platform.lower() else platform.capitalize()))
    prefix = "Automatizacion_" if "Automatización" in build_name else "resultados_"
    run_number   = _get_run_number(pais, plat_clean, prefix=prefix)
    results_path = os.path.join(_RESULTADOS_DIR, f"{prefix}{pais}_{plat_clean}{run_number}.xlsx")

    wb = load_workbook(source_excel_path)
    ws = wb.active

    required_cols = [
        "Resultado", "Formulario Inserto", "Formulario Completado",
        "TY Page", "TYP con CTA", "LINK ISSUE TYP", "Form URL esperada", "Form URL encontrada", "Form coincide",
        "Datos vs Excel", "Motivo", "Estado URL landing", "Estado URL form",
        "Video LT", "Dashboard LT",
    ]
    headers = [cell.value for cell in ws[1] if cell.value]
    for col_name in required_cols:
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers)).value = col_name

    wb.save(results_path)

    headers = [cell.value for cell in ws[1] if cell.value]
    col_idx = {name: headers.index(name) + 1
               for name in required_cols if name in headers}
    return wb, ws, run_number, results_path, col_idx


def _mark_red(cell):
    cell.fill = PatternFill(patternType="solid", fgColor="FF0000")
    cell.font = XLFont(color="FFFFFF", bold=True)

def _mark_green(cell):
    cell.fill = PatternFill(patternType="solid", fgColor="00B050")
    cell.font = XLFont(color="FFFFFF", bold=True)

def _mark_amber(cell):
    """Enviado, pero con datos distintos a los pedidos en el Excel."""
    cell.fill = PatternFill(patternType="solid", fgColor="FFC000")
    cell.font = XLFont(color="000000", bold=True)

def _write_row_result(ws, row_num: int, col_idx: Dict,
                      iframe_found: bool, submitted: bool,
                      result_text: str, tracked: Dict[str, str],
                      ty_confirmed: bool = False,
                      iframe_url_found: str = "",
                      iframe_url_expected: str = "",
                      video_url: str = "",
                      dashboard_url: str = "",
                      ty_cta: str = "",
                      link_issue: str = "", link_issue_present: bool = False,
                      datos_vs_excel: str = "", datos_mismatch: bool = False,
                      motivo: str = "", estado_url_landing: str = "-",
                      estado_url_form: str = "-",
                      form_inserto_estado: str = "", form_mismatch_enviado: bool = False):
    """Escribe resultado con las mismas columnas que el desktop.
    OJO: NO se tocan las columnas de entrada (Modelo/Nombre/etc. las setea el usuario). El
    modelo elegido queda en las columnas de tracking (PasoN::models), para comparar pedido vs
    completado."""
    resultado_cell = ws.cell(row=row_num, column=col_idx.get("Resultado", 1))
    resultado_cell.value = result_text
    if form_mismatch_enviado:
        # El lead SÍ viajó, pero por un form distinto al esperado: ni verde ni rojo.
        _mark_amber(resultado_cell)
    elif not submitted:
        _mark_red(resultado_cell)
    elif datos_mismatch:
        # Se envió, pero con datos distintos a los pedidos en el Excel → no es "todo OK"
        _mark_amber(resultado_cell)
    else:
        _mark_green(resultado_cell)

    # Motivo del fallo: por qué no salió el lead, sin tener que leer el Resultado entero
    if "Motivo" in col_idx:
        _mot_cell = ws.cell(row=row_num, column=col_idx["Motivo"])
        _mot_cell.value = motivo or "-"
        if motivo and motivo != "-":
            _mark_amber(_mot_cell) if form_mismatch_enviado else _mark_red(_mot_cell)

    # Estado HTTP de las URLs: distingue "el form falló" de "la URL nunca cargó"
    for _col, _val in (("Estado URL landing", estado_url_landing),
                       ("Estado URL form", estado_url_form)):
        if _col not in col_idx:
            continue
        _cell = ws.cell(row=row_num, column=col_idx[_col])
        _cell.value = _val or "-"
        _txt = (_val or "").upper()
        if _txt.startswith("200"):
            _mark_green(_cell)
        elif "REDIRIGE" in _txt:
            _mark_amber(_cell)
        elif _txt not in ("-", ""):
            _mark_red(_cell)

    if "Datos vs Excel" in col_idx:
        _dv_cell = ws.cell(row=row_num, column=col_idx["Datos vs Excel"])
        _dv_cell.value = datos_vs_excel or "-"
        if datos_mismatch:
            _mark_amber(_dv_cell)
        elif (datos_vs_excel or "").strip().upper() == "OK":
            _mark_green(_dv_cell)

    # Si la landing nunca cargó (404 / sin respuesta / redirige a otro lado), el form
    # jamás se buscó: decir "no coincide con el esperado" sería engañoso. Se reporta la
    # causa real, que es la URL.
    _landing_txt = (estado_url_landing or "").upper()
    _landing_falla = bool(
        _landing_txt and _landing_txt not in ("-", "")
        and (not _landing_txt.startswith("200") or "REDIRIGE" in _landing_txt)
    )

    # Formulario Inserto en TRES estados (no dos): inserto y correcto / inserto pero
    # otro distinto al esperado / no inserto (iframe sin src o inexistente).
    inserto_cell = ws.cell(row=row_num, column=col_idx.get("Formulario Inserto", 2))
    if _landing_falla and form_inserto_estado != "inserto_ok":
        inserto_cell.value = f"N/D — no se buscó el form: la landing no cargó ({estado_url_landing})"
        _mark_red(inserto_cell)
    elif form_inserto_estado == "inserto_otro":
        _sent = "se envió lead igualmente" if form_mismatch_enviado else "lead no enviado"
        inserto_cell.value = f"⚠ Form inserto NO coincide con el esperado, {_sent}"
        _mark_amber(inserto_cell)
    elif form_inserto_estado == "no_inserto" or not iframe_found:
        inserto_cell.value = "✗ Form NO inserto (iframe sin src / sin form en la landing)"
        _mark_red(inserto_cell)
    else:
        inserto_cell.value = "✓ Form inserto"
        _mark_green(inserto_cell)

    ws.cell(row=row_num, column=col_idx.get("Formulario Completado", 3)).value = (
        "✓ Completado" if (submitted or iframe_found) else "-"
    )
    ws.cell(row=row_num, column=col_idx.get("TY Page", 4)).value = (
        "✓ TY detectada" if ty_confirmed else "-"
    )
    if "TYP con CTA" in col_idx:
        ws.cell(row=row_num, column=col_idx["TYP con CTA"]).value = ty_cta or "-"
    if "LINK ISSUE TYP" in col_idx:
        _li_cell = ws.cell(row=row_num, column=col_idx["LINK ISSUE TYP"])
        _li_cell.value = link_issue or "-"
        if link_issue_present:
            _mark_red(_li_cell)
    if "Form URL esperada" in col_idx and iframe_url_expected:
        ws.cell(row=row_num, column=col_idx["Form URL esperada"]).value = iframe_url_expected
    if "Form URL encontrada" in col_idx:
        ws.cell(row=row_num, column=col_idx["Form URL encontrada"]).value = iframe_url_found or ""
    if "Form coincide" in col_idx and iframe_url_expected:
        def _norm(u):
            if not u: return ""
            u = u.strip().split("?")[0].split("#")[0]
            if u.endswith("/"): u = u[:-1]
            return u.lower()
        coincide_ok = _norm(iframe_url_found) == _norm(iframe_url_expected)
        coincide_cell = ws.cell(row=row_num, column=col_idx["Form coincide"])
        if not coincide_ok and _landing_falla:
            # La landing no cargó: no hay comparación posible, el form nunca se buscó.
            coincide_cell.value = f"N/D — la landing no cargó ({estado_url_landing})"
            _mark_red(coincide_cell)
        elif coincide_ok:
            coincide_cell.value = "PASS"
            _mark_green(coincide_cell)
        elif form_mismatch_enviado:
            # El form no era el esperado pero el lead salió: naranja, no rojo — hay dato
            # en la base, solo que entró por otro formulario.
            coincide_cell.value = "FAIL — form inserto no coincide, se envió lead igualmente"
            _mark_amber(coincide_cell)
        elif form_inserto_estado == "no_inserto" or not iframe_found:
            coincide_cell.value = "FAIL — form NO inserto (iframe sin src / sin form)"
            _mark_red(coincide_cell)
        else:
            coincide_cell.value = "FAIL — form inserto no coincide con el esperado"
            _mark_red(coincide_cell)
    if "Video LT" in col_idx and video_url:
        ws.cell(row=row_num, column=col_idx["Video LT"]).value = video_url
    if "Dashboard LT" in col_idx and dashboard_url:
        ws.cell(row=row_num, column=col_idx["Dashboard LT"]).value = dashboard_url

    if tracked:
        current_headers = [cell.value for cell in ws[1]]
        for field_name, field_value in tracked.items():
            # field_name ya viene prefijado como "PasoN::campo" desde _run_single_lead
            col_header = field_name if "::" in field_name else f"Paso1::{field_name}"
            if col_header not in current_headers:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col).value = col_header
                current_headers.append(col_header)
            col_num = current_headers.index(col_header) + 1
            ws.cell(row=row_num, column=col_num).value = str(field_value)


# ══════════════════════════════════════════════════════════════════════════════
# PROCESAMIENTO DE UN LEAD (secuencia igual que Osocio)
# ══════════════════════════════════════════════════════════════════════════════

_CTA_EMPTY_JS = r"""
var subSel = arguments[0];
function vis(e){ return e && e.getClientRects().length && !e.disabled; }

// ¿Hay algun campo obligatorio visible? Si NO, clickear "Enviar" con el form vacio no
// dispara validaciones: lo manda de una y genera un lead basura (ej. el Libro de
// Reclamaciones de Peru, donde todos los campos son voluntarios). Ahi no se toca el Enviar.
var hayObligatorios = false;
var req = document.querySelectorAll(
    'input[required], select[required], textarea[required], [aria-required="true"]');
for (var i = 0; i < req.length; i++) {
    if (req[i].type !== 'hidden' && !req[i].disabled && req[i].getClientRects().length) {
        hayObligatorios = true;
        break;
    }
}

// 1) Enviar visible (form de un solo paso) → ese es el CTA del paso actual
if (hayObligatorios) {
    for (var i=0; i<subSel.length; i++){
        var e = document.querySelector(subSel[i]);
        if (vis(e)){ e.scrollIntoView({block:'center',behavior:'instant'}); e.click(); return 'Enviar'; }
    }
}
// 2) Wizard: el Enviar vive en un paso oculto → el CTA del paso actual es Siguiente
var nx = [].filter.call(
    document.querySelectorAll("button.next, button[class*='next'], .next-button"), vis)[0];
if (nx){ nx.scrollIntoView({block:'center',behavior:'instant'}); nx.click(); return 'Siguiente'; }
return '';
"""


def _click_cta_empty(driver, log: Callable = print):
    """
    Click en el CTA del paso actual con el form vacío, para disparar las validaciones.

    Un solo JS call. Antes se usaba _click_submit, que barre 12 selectores con un
    WebDriverWait cada uno: en los wizards (RAQ Brasil) el Enviar vive en un paso oculto,
    así que ese barrido SIEMPRE falla y quemaba ~150s en Mac / ~50s en Android por lead.
    """
    try:
        cta = driver.execute_script(_CTA_EMPTY_JS, _SUBMIT_SELECTORS)
    except Exception as e:
        log(f"  ⚠ Error en click CTA vacío: {e}")
        return
    if cta:
        log(f"  ✓ CTA vacío clickeado ({cta}) — validaciones disparadas")
    else:
        log("  ℹ Sin CTA para el click vacío (o el form no tiene campos obligatorios)")


def _raq_brasil_gate(driver, lead: LeadRow, pais: str, log: Callable = print):
    """
    El formulario RAQ de Brasil muestra primero una pantalla de selección
    "Formulário / WhatsApp". Hay que clickear id="contact-by-form" para que aparezca
    el formulario real. La variante "raq-eletricos" NO tiene esa pantalla (va directo
    al form real), así que se excluye para no perder tiempo buscando un botón que
    nunca va a aparecer.

    Se llama tanto en el primer ingreso al iframe como tras recargar la landing en el
    reintento: si no, el form queda tapado por el chooser, no hay botón "Siguiente"
    visible, el wizard nunca avanza y el submit (que vive en el paso 3) queda oculto.
    """
    iframe_src_lower = (lead.secure_url or "").lower()
    landing_url_lower = (lead.public_url or "").lower()
    _is_eletricos = "eletricos" in iframe_src_lower or "eletricos" in landing_url_lower
    _is_brasil = pais.lower() in ("brasil", "brazil", "br")
    _has_raq = "raq" in iframe_src_lower or "raq" in landing_url_lower
    if not (_is_brasil and _has_raq and (not _is_eletricos) and (
        "/brasil/gm_forms/raq" in iframe_src_lower or "solicitar-contato" in landing_url_lower
    )):
        return
    log("  🇧🇷 Excepción RAQ Brasil: buscando botón 'contact-by-form'...")
    try:
        btn_form = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "contact-by-form"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_form)
        time.sleep(0.5)
        try:
            btn_form.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn_form)
        log("  ✓ Botón 'Formulário' clickeado — formulario RAQ visible")
        time.sleep(1.5)  # esperar que cargue el formulario real
    except Exception as e:
        log(f"  ⚠ No se encontró contact-by-form (RAQ Brasil): {e}")


def _finalize_result(result: Dict, lead: LeadRow) -> Dict:
    """Cierre común de una fila: estado del form inserto, motivo del fallo y estado HTTP.

    Se aplica SIEMPRE, también a los cortes tempranos (iframe ausente, error general):
    si un lead no salió, la fila del Excel tiene que decir por qué sin leer el log.
    """
    if result.get("form_url_mismatch"):
        _found_str = (result.get("iframe_url_found") or "").strip()
        # El lead viajó de verdad si se confirmó la TY, aunque el form fuera otro.
        _enviado = bool(result.get("ty_confirmed") or result.get("submitted"))
        result["form_mismatch_enviado"] = _enviado
        _lead_str = "se envió lead igualmente" if _enviado else "lead no enviado"
        if _found_str:
            # SÍ se encontró/insertó un form (p.ej. gm_front en Mac/Safari en vez del gm_forms
            # del Excel): es un form INCORRECTO/distinto al esperado, NO "no inserto".
            result["result_text"] = (
                f"[Error Form] Form inserto NO coincide con el esperado, {_lead_str} — "
                f"URL esperada: {lead.secure_url or '?'} | "
                f"URL encontrada: {_found_str}"
            )
            result["motivo"] = f"form inserto no coincide ({_lead_str})"
        else:
            # No se encontró ningún form GM en la landing (iframe sin src / inexistente).
            result["result_text"] = (
                f"[Error Form] FORM NO INSERTO (iframe sin src / sin form en la landing) — "
                f"URL esperada: {lead.secure_url or '?'} | "
                f"URL encontrada: ninguno | {_lead_str}"
            )
            result["motivo"] = "form no inserto"
        result["submitted"] = False

    # Issue en el CTA/link de la TY page → cuenta como fail (columna LINK ISSUE en rojo).
    if result.get("link_issue_present") and not result.get("form_url_mismatch"):
        _li = result.get("link_issue", "") or ""
        _sent = "se envió lead igualmente" if result.get("submitted") else "lead no enviado"
        result["result_text"] = f"[Error Form] LINK ISSUE TYP: {_li} | {_sent}"
        result["motivo"] = "link issue TYP"
        result["submitted"] = False

    # Motivo del fallo, siempre explícito
    if not result.get("submitted") and not result.get("motivo"):
        result["motivo"] = _motivo_corto(result.get("result_text", ""))

    # Si además hay un problema de estado HTTP, es la causa raíz más probable:
    # se antepone al motivo (una landing 404 explica cualquier "form no encontrado").
    _url_prob = result.get("url_status_problema") or ""
    if _url_prob and not result.get("submitted"):
        result["motivo"] = f"{_url_prob} | {result.get('motivo') or 'error'}"
        if "Estado URL:" not in (result.get("result_text") or ""):
            result["result_text"] = f"{result.get('result_text','')} | Estado URL: {_url_prob}"

    if result.get("submitted") and not result.get("motivo"):
        result["motivo"] = "-"

    return result


def _run_single_lead(driver, pais: str, lead: LeadRow,
                     field_mapping: List[Dict],
                     dependencies: Dict[str, str],
                     ids_dinamicos: Dict[str, str],
                     is_mobile: bool = False,
                     is_android: bool = False,
                     brasil_doc_type: str = "cpf",
                     screenshot_manager=None,
                     log: Callable = print):
    """Wrapper: corre el lead y aplica SIEMPRE el cierre común (motivo/estado URL),
    incluso si el cuerpo cortó temprano por un error."""
    result = _run_single_lead_impl(
        driver, pais, lead, field_mapping, dependencies, ids_dinamicos,
        is_mobile=is_mobile, is_android=is_android,
        brasil_doc_type=brasil_doc_type, screenshot_manager=screenshot_manager, log=log,
    )
    try:
        return _finalize_result(result, lead)
    except Exception as _fe:
        log(f"  ⚠ Error armando el motivo del resultado: {_fe}")
        return result


def _run_single_lead_impl(driver, pais: str, lead: LeadRow,
                     field_mapping: List[Dict],
                     dependencies: Dict[str, str],
                     ids_dinamicos: Dict[str, str],
                     is_mobile: bool = False,
                     is_android: bool = False,
                     brasil_doc_type: str = "cpf",
                     screenshot_manager=None,
                     log: Callable = print):
    """
    Retorna result_dict.
    Secuencia idéntica a Osocio:
    1. Navegar → cookies → scroll
    2. Encontrar iframe → posicionar → switch
    3. Esperar form → click enviar vacío
    4. Llenar campos (pasos con Siguiente si hay)
    5. _handle_terms_checkboxes
    6. Submit → esperar
    """
    tracked: Dict[str, str] = {}
    result = {
        "result_text": "", "tracked": tracked,
        "iframe_found": False, "submitted": False,
        "ty_confirmed": False, "iframe_url_found": "",
        "ty_cta": "", "link_issue": "", "link_issue_present": False,
        "datos_vs_excel": "-", "datos_mismatch": False,
        "estado_url_landing": "-", "estado_url_form": "-",
        "url_status_problema": "", "motivo": "",
        "form_inserto_estado": "no_inserto", "form_mismatch_enviado": False,
    }

    _cb_prefs = _checkbox_prefs(lead)
    if _cb_prefs:
        log(f"  ☑ Preferencias de checkbox (Excel): {_cb_prefs}")

    # ── 0. Estado HTTP de las URLs ────────────────────────────────────────────
    # Selenium no expone el status de la navegación: si la landing da 404 o el form
    # está caído (503), el lead falla sin decir por qué. Se consulta aparte para que
    # el Excel muestre la causa real.
    try:
        from utils.url_status import check_url_status, format_status_pair
        _st_landing = check_url_status(lead.public_url)
        _st_form = check_url_status(lead.secure_url) if lead.secure_url else {}
        result["estado_url_landing"] = _st_landing.get("label", "-")
        result["estado_url_form"] = _st_form.get("label", "-") if _st_form else "-"
        result["url_status_problema"] = format_status_pair(_st_landing, _st_form)
        log(f"  🌐 Estado URL landing: {result['estado_url_landing']}")
        if lead.secure_url:
            log(f"  🌐 Estado URL form:    {result['estado_url_form']}")
    except Exception as _use:
        log(f"  ⚠ No se pudo verificar el estado de las URLs: {_use}")

    try:
        # ── 1. Landing page ──────────────────────────────────────────────────
        log(f"  Navegando a {lead.public_url}")
        driver.get(lead.public_url)

        try:
            WebDriverWait(driver, 8).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass
        time.sleep(1)

        # Zoom 80% en forms comprar-carro para que todo el form sea visible
        if "comprar-carro" in lead.public_url:
            try:
                driver.execute_script("document.body.style.zoom='80%';")
            except Exception:
                pass

        if is_android:
            # Scroll único al fondo para activar lazy loading, sin volver al top.
            # El posicionamiento del iframe (más abajo) ya scrollea hasta él.
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.5)
            except Exception:
                pass
        else:
            _pre_scroll_for_dynamic_content(driver)
        _handle_cookie_popups(driver, log)

        # gm_front (React SPA): esperar a que los componentes monten tras el scroll
        if "gm_front" in lead.public_url.lower():
            time.sleep(3)

        # ── 2. Encontrar y posicionar iframe ─────────────────────────────────
        use_iframe = bool(lead.secure_url and lead.secure_url.strip())
        iframe_el = None

        if use_iframe:
            def _buscar_iframe():
                # Dos pasadas para no agarrar "el primero que aparezca": primero el iframe que
                # matchea EXACTO la URL esperada del Excel; si no, cualquier iframe GM
                # (gm_forms/gm_admin/gm_front). Evita tomar un iframe GM que no es el correcto
                # cuando la landing tiene más de uno.
                _exp = (lead.secure_url or "").strip()
                _iframes = driver.find_elements(By.TAG_NAME, "iframe")
                if _exp:
                    for iframe in _iframes:
                        try:
                            if _exp in (iframe.get_attribute("src") or ""):
                                return iframe
                        except Exception:
                            continue
                for iframe in _iframes:
                    try:
                        src = iframe.get_attribute("src") or ""
                        if "gm_forms" in src or "gm_admin" in src or "gm_front" in src:
                            return iframe
                    except Exception:
                        continue
                return None

            # Scroll INCREMENTAL: bajar de a un viewport y frenar APENAS aparece el iframe,
            # así el scroll queda posicionado en él (NO seguir hasta el footer, si no el
            # botón Enviar queda fuera de vista). Una sola pasada — sin reintentos (lento en LT).
            iframe_el = _buscar_iframe()
            if not iframe_el:
                try:
                    _h = driver.execute_script("return document.body.scrollHeight") or 3000
                    _step = int((driver.execute_script("return window.innerHeight") or 800) * 0.8)
                    _y = 0
                    while _y < _h:
                        _y += _step
                        driver.execute_script(f"window.scrollTo(0, {_y});")
                        time.sleep(0.4)
                        iframe_el = _buscar_iframe()
                        if iframe_el:
                            break
                except Exception:
                    pass

            if not iframe_el:
                # Último recurso: cualquier iframe visible con keyword GM en su src
                try:
                    for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
                        try:
                            src = iframe.get_attribute("src") or ""
                            if any(kw in src for kw in ("gm_forms", "gm_admin", "gm_front")):
                                iframe_el = iframe
                                log(f"  ⚠ Iframe fallback por keyword: {src[:80]}")
                                break
                        except Exception:
                            continue
                except Exception:
                    pass

            if not iframe_el and is_android:
                log("  ↺ Iframe no encontrado — recargando landing (Android)...")
                driver.get(lead.public_url)
                try:
                    WebDriverWait(driver, 8).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                except Exception:
                    pass
                time.sleep(1)
                _handle_cookie_popups(driver, log)

                def _android_find_iframe():
                    for _if in driver.find_elements(By.TAG_NAME, "iframe"):
                        try:
                            src = _if.get_attribute("src") or ""
                            if lead.secure_url and lead.secure_url.strip() in src:
                                return _if
                            if any(kw in src for kw in ("gm_forms", "gm_admin", "gm_front")):
                                return _if
                        except Exception:
                            continue
                    return None

                # Scroll escalonado: mitad → fondo (el form puede estar en cualquier posición)
                for _scroll_pct in (0.5, 1.0):
                    try:
                        driver.execute_script(
                            f"window.scrollTo(0, document.body.scrollHeight * {_scroll_pct});"
                        )
                        time.sleep(0.8)
                    except Exception:
                        pass
                    iframe_el = _android_find_iframe()
                    if iframe_el:
                        break

        if use_iframe and not iframe_el:
            result["result_text"] = "[Error] Formulario no encontrado — iframe ausente"
            log("  ✗ No se encontró iframe del formulario")
            return result

        _iframe_src = ""
        if iframe_el:
            try:
                _iframe_src = iframe_el.get_attribute("src") or ""
            except Exception:
                pass
            if not _iframe_src or _iframe_src.strip() == "" or _iframe_src.startswith("about:"):
                try:
                    driver.switch_to.frame(iframe_el)
                    _iframe_src = driver.execute_script("return window.location.href;") or ""
                    driver.switch_to.parent_frame()
                except Exception:
                    pass
            if _iframe_src and not _iframe_src.startswith("http"):
                try:
                    from urllib.parse import urljoin
                    _iframe_src = urljoin(driver.current_url, _iframe_src)
                except Exception:
                    pass

        result["iframe_found"] = bool(iframe_el) if use_iframe else True
        result["iframe_url_found"] = _iframe_src

        # Detectar si la URL del iframe encontrado coincide con la esperada
        _expected_url = (lead.secure_url or "").strip()
        
        # Comparación normalizada para mismatch
        def _norm(u):
            if not u: return ""
            u = u.strip().split("?")[0].split("#")[0]
            if u.endswith("/"): u = u[:-1]
            return u.lower()
            
        result["form_url_mismatch"] = bool(
            iframe_el and _expected_url and _norm(result["iframe_url_found"]) != _norm(_expected_url)
        )

        # Estado del form inserto, en tres niveles (no dos):
        #   inserto_ok    → hay iframe con src y coincide con el esperado
        #   inserto_otro  → hay form, pero es OTRO distinto al del Excel
        #   no_inserto    → no hay iframe, o su src está vacío (nada que cargar)
        if not iframe_el and use_iframe:
            result["form_inserto_estado"] = "no_inserto"
        elif use_iframe and not (_iframe_src or "").strip():
            result["form_inserto_estado"] = "no_inserto"
        elif result["form_url_mismatch"]:
            result["form_inserto_estado"] = "inserto_otro"
        else:
            result["form_inserto_estado"] = "inserto_ok"

        # Posicionar scroll igual que Osocio
        if iframe_el:
            try:
                loc = iframe_el.location
                vph = driver.execute_script("return window.innerHeight")
                elem_y = loc.get('y', 0)
                if elem_y > vph:
                    scroll_to = elem_y - (vph * 0.3)
                else:
                    scroll_to = max(elem_y - 100, 0)
                driver.execute_script(f"window.scrollTo(0, {scroll_to});")
                time.sleep(0.5)
            except Exception:
                pass
        else:
            # Form standalone (sin iframe): puede estar más abajo y renderizar recién al
            # scrollear (lazy). En Safari, si el campo no está en viewport no se detecta/llena
            # bien → recorrer toda la página y posicionar en el PRIMER campo visible del form.
            try:
                for _pct in (0.35, 0.7, 1.0):
                    driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight*{_pct});")
                    time.sleep(0.6)
                driver.execute_script("""
                    var sel='input:not([type=hidden]):not([type=submit]):not([type=button]):not([type=reset]),select,textarea';
                    for (var el of document.querySelectorAll(sel)){
                        if(el.offsetParent){ el.scrollIntoView({block:'center',behavior:'instant'}); return; }
                    }
                    window.scrollTo(0, 0);
                """)
                time.sleep(0.6)
            except Exception:
                pass

        # Sólo hay iframe si el form está inserto en una landing. En los forms sueltos
        # (gm_front standalone) este log salía igual y hacía creer que se había encontrado
        # un iframe que nunca existió.
        if iframe_el:
            log("  ✓ Iframe encontrado")
        if screenshot_manager:
            screenshot_manager.captura_landing_inicial()

        # ── 3. Entrar al iframe ──────────────────────────────────────────────
        if iframe_el:
            if is_mobile:
                # iOS Safari: switch_to.frame() falla por cross-origin.
                # Solución: scrollear la landing hasta que el iframe ocupe el
                # viewport y operar via NATIVE_APP + tap en coordenadas.
                log("  📱 iOS: usando scroll al iframe (no switch_to.frame)")
                try:
                    driver.execute_script("""
                        const iframe = arguments[0];
                        const rect   = iframe.getBoundingClientRect();
                        const scrollY = window.pageYOffset || document.documentElement.scrollTop;
                        const targetY = rect.top + scrollY - 10;
                        window.scrollTo({ top: Math.max(0, targetY), behavior: 'instant' });
                    """, iframe_el)
                    time.sleep(1.5)
                    log("  ✓ Iframe posicionado en viewport (iOS)")
                except Exception as e:
                    log(f"  ⚠ Error posicionando iframe: {e}")
            else:
                # Desktop: switch_to.frame() normal
                driver.switch_to.frame(iframe_el)
                try:
                    WebDriverWait(driver, 10).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                except Exception:
                    pass
                time.sleep(0.3)
                log("  ✓ Dentro del iframe")
        else:
            log("  ✓ Formulario embebido en documento principal (sin iframe)")

        _raq_brasil_gate(driver, lead, pais, log)

        # ── 4. Click en el CTA vacío (para activar validaciones antes de llenar) ─
        _click_cta_empty(driver, log)
        time.sleep(0.5)  # esperar a que JS muestre los errores de validación
        if screenshot_manager:
            screenshot_manager.captura_form_errores()

        # ── 5. Loop de pasos — LÓGICA EXACTA DE OSOCIO fill_form_fields_auto_step ──
        # Igual que Osocio: llenar → _has_next_button → _dom_signature → click →
        # esperar 1.5s → verificar 4 veces que el DOM cambió → repetir
        all_tracked: Dict[str, str] = {}
        max_iter = 15  # igual que auto_step_max_iterations de Osocio
        _xp: set = set()  # IDs ya llenados — se pasa entre iteraciones para no repetir

        # Para forms sin iframe (gm_front y similares React/SPA): esperar a que el
        # framework monte los componentes antes de escanear el DOM.
        if not use_iframe:
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: bool(
                        d.find_elements(By.CSS_SELECTOR,
                            "input:not([type='hidden']):not([type='submit']):not([type='button']), select"
                        )
                    )
                )
            except Exception:
                time.sleep(2)

        # T3/AEM (forms 2.0): IDs volátiles → se llena por keyword del <label> con el
        # motor compartido (utils/aem_fill). Misma lógica que desktop.
        from utils import aem_fill as _aem
        _is_aem_form = _aem.is_aem_adaptive_form(driver)
        _is_brasil_aem = str(pais).lower() in ("brasil", "brazil", "br")

        def _aem_form_data():
            _ll = {str(k).lower().strip(): v for k, v in lead.data.items()}
            def _pick(*keys):
                for k in keys:
                    v = _ll.get(k, "")
                    if v and str(v).lower() not in ("", "none"):
                        return str(v)
                return ""
            return {
                "firstname": _pick("nombre", "nombres", "nome"),
                "lastname":  _pick("apellido", "apellidos", "sobrenome"),
                "email":     _pick("email", "correo"),
                "phone":     _pick("celular", "telefono", "teléfono"),
                "document":  _pick("documento", "cpf", "cnpj"),
                "cpf":       _pick("cpf"),
                "cnpj":      _pick("cnpj"),
                "cep":       _pick("cep", "zip", "postal"),
                "vin":       _pick("vin", "chasis", "chassi", "chassis"),
                "comment":   _pick("comentario", "comentarios"),
                "model":     _pick("modelo"),
                "city":      _pick("ciudad", "cidade"),
                "dealer":    _pick("concesionario", "concessionaria"),
            }

        if _is_aem_form:
            log("🧩 [LT] Formulario AEM/T3 detectado → llenado por keyword (motor compartido)")

        for iteration in range(max_iter):
            log(f"\n  --- Iteración {iteration + 1} ---")

            # Esperar que haya campos en el DOM
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "input,select,textarea")
                    )
                )
            except TimeoutException:
                log("  No hay campos en esta iteración.")
                break

            # Llenar campos visibles del paso actual (saltando los ya llenados en pasos anteriores)
            step_tracked: Dict[str, str] = {}
            if _is_aem_form:
                # No usar _auto_fill_unmapped_dropdowns_lt en AEM: re-randomizaría el
                # select persona → Pessoa y ocultaría CNPJ/empresa que ya llenamos.
                _aem.fill_aem_form(
                    driver, _aem_form_data(), _is_brasil_aem,
                    gen_doc=_generate_brazil_document, log=log,
                    record=lambda k, v: step_tracked.__setitem__(k, v),
                    is_android=is_android,
                )
                _aem.mark_aem_terms(driver, log=log)
            else:
                fill_form_fields(
                    driver, lead, pais, field_mapping,
                    dependencies, ids_dinamicos, step_tracked, log=log,
                    is_mobile=is_mobile,
                    is_android=is_android,
                    iframe_el=iframe_el if is_mobile else None,
                    brasil_doc_type=brasil_doc_type,
                    cross_processed=_xp,
                )
                _auto_fill_unmapped_dropdowns_lt(driver, field_mapping, ids_dinamicos, lead=lead, log=log,
                                                 tracked=step_tracked, filled_ids=_xp)
                _record_model_from_url_if_missing_lt(driver, field_mapping, step_tracked, log=log,
                                                     expected_form_url=lead.secure_url or "")
                # Requeridos sin dato (Rua/Número/Data de Nascimento, etc.) → sintético random
                _fill_required_synthetic_lt(driver, log=log)
            paso_num = iteration + 1
            for _k, _v in step_tracked.items():
                all_tracked[f"Paso{paso_num}::{_k}"] = _v
            log(f"  {len(step_tracked)} campos llenados (Paso {paso_num}).")
            if screenshot_manager:
                screenshot_manager.captura_paso(paso_num)

            # ¿Hay botón Siguiente? Si no, llegamos al último paso → salir
            if not _has_next_button(driver):
                log("  No hay botón Siguiente — último paso alcanzado.")
                break

            # Tomar firma del DOM ANTES de clickear (igual que Osocio)
            sig_before = _dom_signature_visible(driver, field_mapping)

            # Clickear Siguiente (igual que Osocio: JS click + fallback)
            if not _click_next_button(driver, log):
                log("  No se pudo clickear Siguiente — saliendo del loop.")
                break

            # Esperar que el DOM cambie tras click en Siguiente
            time.sleep(0.5)
            stuck = True
            for _ in range(4):
                sig_after = _dom_signature_visible(driver, field_mapping)
                if sig_after != sig_before:
                    stuck = False
                    break
                time.sleep(0.25)

            if stuck:
                log("  ⚠ DOM no cambió tras Siguiente (validación fallida o form atascado).")
                break

            log(f"  ✓ Transición al siguiente paso detectada.")

        else:
            log(f"  ⚠ Se alcanzó el máximo de iteraciones ({max_iter}).")

        result["tracked"] = all_tracked

        # ── 6. Terms — scroll agresivo al fondo para revelar checkboxes ──────────
        try:
            driver.execute_script("""
                // Scroll de la ventana y todos los contenedores padres al fondo
                window.scrollTo(0, document.body.scrollHeight);
                document.documentElement.scrollTop = document.documentElement.scrollHeight;
                // Scroll de cualquier contenedor de formulario con overflow
                ['form','section','.form-wrapper','.gm-form','#contact-form','.container'].forEach(function(sel){
                    var el=document.querySelector(sel);
                    if(el){el.scrollTop=el.scrollHeight;}
                });
                // Buscar el label/área del checkbox de términos y scrollear a él
                var cb=document.querySelector('input[name="terms"],input[name="terms-contact"],input[type="checkbox"]');
                if(cb){
                    var lbl=null;
                    if(cb.id)try{lbl=document.querySelector('label[for="'+CSS.escape(cb.id)+'"]');}catch(e){}
                    if(!lbl)lbl=cb.closest('label');
                    if(!lbl){var s=cb.nextElementSibling;if(s&&s.tagName!=='INPUT')lbl=s;}
                    var target=lbl||cb;
                    // Scroll de todos los ancestros
                    var el=target;
                    while(el&&el!==document.body){
                        if(el.parentElement)el.parentElement.scrollTop=el.parentElement.scrollHeight;
                        el=el.parentElement;
                    }
                    target.scrollIntoView({block:'center',behavior:'instant'});
                }
            """)
            time.sleep(0.4)
        except Exception:
            pass
        _handle_terms_checkboxes(driver, log, is_android=is_android, prefs=_cb_prefs)
        # Antes de la captura: limpiar los errores que quedaron pintados en campos que ya
        # están completos (si no, la captura muestra falsos errores).
        _revalidar_campos_llenos(driver, log)
        if screenshot_manager:
            screenshot_manager.captura_form_completado()

        # ── 7. Submit (dentro del iframe) ────────────────────────────────────
        for _submit_attempt in range(1, 3):  # max 2 intentos totales
            # El re-ingreso vía send_keys es costoso (varios campos, char a char) y el
            # llenado normal ya funciona la mayoría de las veces — solo se hace en el
            # reintento (cuando ya hubo un submit fallido), no antes del primer intento.
            if _submit_attempt > 1:
                _ensure_fields_filled_before_submit(driver, field_mapping, all_tracked, log,
                                                     is_mobile=is_mobile, is_android=is_android)
            _ensure_terms_marked_before_submit(driver, log, prefs=_cb_prefs)

            # Foto del estado REAL del form justo antes de enviar: es lo que viaja en el
            # lead y lo que hay que comparar contra la base de datos.
            _sync = _sync_tracked_with_dom(driver, field_mapping, lead, all_tracked, log=log)
            all_tracked = _sync["tracked"]
            result["tracked"] = all_tracked
            result["datos_vs_excel"] = _sync["excel_mismatch"]
            result["datos_mismatch"] = _sync["excel_mismatch_present"]

            submitted_click = _click_submit(driver, log, is_android=is_android)

            if not submitted_click:
                result["result_text"] = "[Error] No se pudo encontrar el botón de envío"
                result["submitted"] = False
                break

            # Verificar div#thank-you con display:block o display: block
            ty_confirmed = False
            try:
                def _ty_visible(d):
                    try:
                        el = d.find_element(By.CSS_SELECTOR, "div#thank-you")
                        style = el.get_attribute("style") or ""
                        if "display: block" in style or "display:block" in style:
                            return True
                    except Exception:
                        pass
                    # Bolivia: el iframe navega a nueva URL con div.rp-wrapper
                    try:
                        d.find_element(By.CSS_SELECTOR, "div.rp-wrapper")
                        return True
                    except Exception:
                        pass
                    # Forms 2.0 (/tools/forms): la TY es un mensaje de agradecimiento.
                    try:
                        if _tiene_thankyou_texto_2_0(d):
                            return True
                    except Exception:
                        pass
                    return False
                WebDriverWait(driver, 15).until(_ty_visible)
                ty_confirmed = True
                log("  ✓ TY div detectado.")
                try:
                    _cta_info = _investigate_ty_cta(driver, log=log)
                    result["ty_cta"] = _format_ty_cta(_cta_info)
                    result["link_issue"] = _format_link_issue(_cta_info)
                    result["link_issue_present"] = bool(_cta_info.get("has_weird"))
                except Exception:
                    pass
                if is_android:
                    try:
                        ty_el = driver.find_element(By.CSS_SELECTOR, "div#thank-you, div.rp-wrapper")
                        driver.execute_script(
                            "arguments[0].scrollIntoView({block:'center',behavior:'smooth'});", ty_el
                        )
                        time.sleep(1.5)
                    except Exception:
                        pass
                else:
                    try:
                        driver.switch_to.default_content()
                        iframe_el = None
                        for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
                            try:
                                src = iframe.get_attribute("src") or ""
                                if any(kw in src for kw in ("gm_forms", "gm_admin", "gm_front")):
                                    iframe_el = iframe
                                    break
                            except Exception:
                                continue
                        if iframe_el:
                            driver.execute_script("arguments[0].scrollIntoView({block:'center',behavior:'smooth'});", iframe_el)
                        else:
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 3);")
                        time.sleep(6.0)  # Wait 6 seconds so the thank you page is visible in the video
                    except Exception:
                        pass
                if screenshot_manager:
                    screenshot_manager.captura_ty_page()
            except TimeoutException:
                time.sleep(0.5)

            if ty_confirmed:
                result["result_text"] = "✓ Lead enviado correctamente"
                result["submitted"] = True
                result["ty_confirmed"] = True
                break

            # Detectar error de Event ID ("Lo siento... ese error" aparece debajo del form)
            try:
                _page_src = driver.page_source or ""
                if "Lo siento" in _page_src or "ese error" in _page_src.lower():
                    result["result_text"] = "No se envió por Error de Event ID, revisar"
                    result["submitted"] = False
                    return result
            except Exception:
                pass

            # Sin TY: verificar errores visuales
            desc_errores = _describir_errores_visuales(driver)

            # Brasil intento 1 con error: re-ingresar doc via send_keys antes de recargar
            if _submit_attempt == 1 and pais.lower() in ("brasil", "brazil", "br") and desc_errores:
                log(f"  ⚠ Error visual Brasil (intento 1): {desc_errores}")
                log("  ↺ Reintentando doc Brasil via send_keys...")
                if _refill_brasil_doc_sendkeys(
                    driver, field_mapping, pais, brasil_doc_type, all_tracked, log
                ):
                    time.sleep(0.3)
                    continue  # intento 2 sin reload

            # Intento 1: rellenar campos con error class visible y reintentar sin reload
            # (no en AEM: re-randomizaría el select Pessoa/Empresa y pisaría el CNPJ/CPF
            # ya cargado según el Excel — ver _is_aem_form)
            if _submit_attempt == 1 and not _is_aem_form:
                # (lo que elija acá queda registrado igual: el snapshot pre-submit relee el DOM)
                _filled_post = _auto_fill_unmapped_dropdowns_lt(driver, field_mapping, ids_dinamicos, lead=lead, log=log,
                                                                filled_ids=_xp)
                if _filled_post:
                    log("  ↺ Campos con error rellenados — reintentando submit sin reload...")
                    _ensure_fields_filled_before_submit(driver, field_mapping, all_tracked, log,
                                                         is_mobile=is_mobile, is_android=is_android)
                    _ensure_terms_marked_before_submit(driver, log, prefs=_cb_prefs)
                    _sync = _sync_tracked_with_dom(driver, field_mapping, lead, all_tracked, log=log)
                    all_tracked = _sync["tracked"]
                    result["tracked"] = all_tracked
                    result["datos_vs_excel"] = _sync["excel_mismatch"]
                    result["datos_mismatch"] = _sync["excel_mismatch_present"]
                    if _click_submit(driver, log, is_android=is_android):
                        try:
                            WebDriverWait(driver, 15).until(_ty_visible)
                            ty_confirmed = True
                            try:
                                _cta_info = _investigate_ty_cta(driver, log=log)
                                result["ty_cta"] = _format_ty_cta(_cta_info)
                                result["link_issue"] = _format_link_issue(_cta_info)
                                result["link_issue_present"] = bool(_cta_info.get("has_weird"))
                            except Exception:
                                pass
                        except TimeoutException:
                            time.sleep(0.5)
                    if ty_confirmed:
                        result["result_text"] = "✓ Lead enviado correctamente"
                        result["submitted"] = True
                        result["ty_confirmed"] = True
                        break
                    desc_errores = _describir_errores_visuales(driver)

            # Intento 2: fallo definitivo
            if _submit_attempt >= 2:
                if desc_errores:
                    result["result_text"] = f"✗ Error visual: {desc_errores} (2 intentos)"
                else:
                    result["result_text"] = "~ Enviado sin confirmación TY Page (2 intentos)"
                result["submitted"] = False
                break

            # Intento 1 falló: para Android recarga landing; para desktop solo reubica el iframe
            motivo = desc_errores or "TY no detectada"
            if _submit_attempt < 2:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass

                if is_android:
                    log(f"  ↺ Intento 1 falló ({motivo!r}) — recargando landing (Android)...")
                    driver.get(lead.public_url)
                    try:
                        WebDriverWait(driver, 8).until(
                            lambda d: d.execute_script("return document.readyState") == "complete"
                        )
                    except Exception:
                        pass
                    time.sleep(0.8)
                    try:
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        time.sleep(0.5)
                    except Exception:
                        pass
                    _handle_cookie_popups(driver, log)
                    if "comprar-carro" in lead.public_url:
                        try:
                            driver.execute_script("document.body.style.zoom='80%';")
                        except Exception:
                            pass
                else:
                    log(f"  ↺ Intento 1 falló ({motivo!r}) — re-ubicando form sin recargar...")
                    time.sleep(0.5)

                # Re-encontrar iframe (el form se resetó solo dentro del iframe)
                iframe_el = None
                for _a in range(2):
                    for _if in driver.find_elements(By.TAG_NAME, "iframe"):
                        try:
                            src = _if.get_attribute("src") or ""
                            if lead.secure_url and lead.secure_url.strip() in src:
                                iframe_el = _if
                                break
                            if any(kw in src for kw in ("gm_forms", "gm_admin", "gm_front")):
                                iframe_el = _if
                                break
                        except Exception:
                            continue
                    if iframe_el:
                        break
                    time.sleep(0.5)
                if iframe_el:
                    driver.switch_to.frame(iframe_el)
                try:
                    WebDriverWait(driver, 5).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                except Exception:
                    pass
                # Tras recargar, el chooser "Formulário / WhatsApp" vuelve a tapar el form
                _raq_brasil_gate(driver, lead, pais, log)
                # Re-llenar campos — respetar el mismo motor que el paso 1 (AEM/T3 vs
                # genérico); usar fill_form_fields en un form AEM re-randomiza el select
                # Pessoa/Empresa y pisa el CNPJ/empresa ya cargado (ver _is_aem_form arriba).
                _step_tracked: Dict[str, str] = {}
                _xp_r: set = set()  # set fresco para el reintento (form recargado)
                for _iter in range(max_iter):
                    _iter_tracked: Dict[str, str] = {}
                    if _is_aem_form:
                        _aem.fill_aem_form(
                            driver, _aem_form_data(), _is_brasil_aem,
                            gen_doc=_generate_brazil_document, log=log,
                            record=lambda k, v: _iter_tracked.__setitem__(k, v),
                            is_android=is_android,
                        )
                        _aem.mark_aem_terms(driver, log=log)
                    else:
                        fill_form_fields(
                            driver, lead, pais, field_mapping,
                            dependencies, ids_dinamicos, _iter_tracked, log=log,
                            is_mobile=is_mobile,
                            is_android=is_android,
                            iframe_el=iframe_el if is_mobile else None,
                            brasil_doc_type=brasil_doc_type,
                            cross_processed=_xp_r,
                        )
                        _auto_fill_unmapped_dropdowns_lt(driver, field_mapping, ids_dinamicos, lead=lead, log=log,
                                                         tracked=_iter_tracked, filled_ids=_xp_r)
                        _record_model_from_url_if_missing_lt(driver, field_mapping, _iter_tracked, log=log,
                                                             expected_form_url=lead.secure_url or "")
                    _paso_r = _iter + 1
                    for _k, _v in _iter_tracked.items():
                        _step_tracked[f"Reintento{_paso_r}::{_k}"] = _v
                    if not _has_next_button(driver):
                        break
                    _sig_b = _dom_signature_visible(driver, field_mapping)
                    if not _click_next_button(driver, log):
                        break
                    time.sleep(0.5)
                    for _ in range(4):
                        if _dom_signature_visible(driver, field_mapping) != _sig_b:
                            break
                        time.sleep(0.25)
                all_tracked.update(_step_tracked)
                _handle_terms_checkboxes(driver, log, is_android=is_android, prefs=_cb_prefs)
                # continuar loop → intento 2 (con _ensure_terms al inicio del loop)

        # Volver al contexto principal para el próximo lead
        try:
            driver.switch_to.default_content()
        except Exception:
            pass

    except Exception as e:
        result["result_text"] = f"✗ Error: {e}"
        log(f"  ✗ Error: {e}")
        traceback.print_exc()
        try:
            driver.switch_to.default_content()
        except Exception:
            pass

    # El cierre (form inserto / motivo / estado URL) lo aplica _finalize_result desde
    # el wrapper _run_single_lead, para que también alcance a los cortes tempranos.
    return result


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def run_lt_batch(opts: LTRunOptions, log: Callable = print,
                 stop_event: Optional[threading.Event] = None) -> Dict:
    """
    Ejecuta todos los leads del Excel de Osocio en LambdaTest Mac.
    Guarda resultados en lambdatest_mac/resultados/ con la misma
    estructura que Osocio.
    """
    log = _safe_log(log)
    summary = {
        "pais": opts.pais, "results_excel": None,
        "session_id": None, "total": 0, "ok": 0, "failed": 0,
        "error": None, "video_url": "",
    }
    driver     = None
    session_id = None
    username   = ""
    access_key = ""
    video_url  = ""

    try:
        username, access_key = load_credentials(opts.credentials_file)
        log(f"✓ Credenciales OK: {username}")

        leads = read_osocio_excel(opts.excel_path)
        if not leads:
            raise ValueError(f"No hay filas en: {opts.excel_path}")
        log(f"  {len(leads)} leads encontrados.")
        summary["total"] = len(leads)

        build_name = opts.build_name or f"{opts.pais} {len(leads)} leads"

        field_mapping = _get_field_mapping_for_pais(opts.pais)
        log(f"  {len(field_mapping)} campos en el mapping de Osocio.")

        dependencies = _load_field_dependencies()
        ids_dinamicos = _load_ids_dinamicos(opts.pais)
        log(f"  IDs dinámicos: {list(ids_dinamicos.keys())}")

        wb, ws, run_number, results_path, col_idx = _setup_results_excel(
            opts.pais, opts.excel_path, opts.platform, build_name
        )
        summary["results_excel"] = results_path
        log(f"Resultados: {results_path} (Run #{run_number})")

        log(f"\nConectando a LambdaTest ({build_name})...")
        driver = create_lt_driver(
            username, access_key,
            build_name=build_name,
            test_name=f"{opts.pais} - {len(leads)} leads",
            platform=opts.platform,
        )
        session_id = driver.session_id
        summary["session_id"] = session_id
        log(f"✓ Session ID: {session_id}")
        log(f"  Dashboard: https://automation.lambdatest.com/logs/?testID={session_id}")

        for i, lead in enumerate(leads, start=1):
            if stop_event and stop_event.is_set():
                log("  ⛔ Ejecución detenida por el usuario.")
                break

            row_num = lead.index
            log(f"\n{'='*60}")
            log(f"  LEAD {i}/{len(leads)} — fila #{row_num}")
            log(f"  URL: {lead.public_url}")
            log(f"{'='*60}")

            # Determinar tipo de documento Brasil para esta fila
            _br_docs = opts.brasil_docs or {}
            _br_doc_type = "cpf"  # default
            if opts.pais.lower() in ("brasil", "brazil"):
                _form_idx = i  # número de form (1-based)
                if _form_idx in _br_docs.get("cnpj_rows", []):
                    _br_doc_type = "cnpj"
                elif _form_idx in _br_docs.get("cep_rows", []):
                    _br_doc_type = "cep"
                elif _form_idx in _br_docs.get("cpf_rows", []):
                    _br_doc_type = "cpf"

            result = _run_single_lead(
                driver, opts.pais, lead,
                field_mapping, dependencies, ids_dinamicos,
                is_mobile=(opts.platform == "iphone"),
                brasil_doc_type=_br_doc_type,
                log=log,
            )

            _write_row_result(
                ws, row_num, col_idx,
                iframe_found=result.get("iframe_found", False),
                submitted=result.get("submitted", False),
                result_text=result["result_text"],
                tracked=result.get("tracked", {}),
                ty_confirmed=result.get("ty_confirmed", False),
                iframe_url_found=result.get("iframe_url_found", ""),
                iframe_url_expected=lead.secure_url or "",
                ty_cta=result.get("ty_cta", ""),
                link_issue=result.get("link_issue", ""),
                link_issue_present=result.get("link_issue_present", False),
                datos_vs_excel=result.get("datos_vs_excel", "-"),
                datos_mismatch=result.get("datos_mismatch", False),
                motivo=result.get("motivo", "-"),
                estado_url_landing=result.get("estado_url_landing", "-"),
                estado_url_form=result.get("estado_url_form", "-"),
                form_inserto_estado=result.get("form_inserto_estado", ""),
                form_mismatch_enviado=result.get("form_mismatch_enviado", False),
            )
            try:
                wb.save(results_path)
            except Exception as e:
                log(f"  ⚠ Error guardando Excel: {e}")

            log(f"  → {result['result_text']}")
            if "✓" in result["result_text"]:
                summary["ok"] += 1
            else:
                summary["failed"] += 1

    except Exception as e:
        summary["error"] = str(e)
        log(f"✗ Error general: {e}")
        log(traceback.format_exc())
    finally:
        if driver:
            try:
                time.sleep(2)  # margen para que el último lead termine de procesar
            except Exception:
                pass
            try:
                mark_lt_status(driver, passed=(summary["failed"] == 0 and summary["error"] is None))
            except Exception as e:
                log(f"  ⚠ Error marcando estado LT: {e}")
            try:
                driver.quit()
                log("\n✓ Driver cerrado.")
            except Exception as e:
                log(f"  ⚠ Error cerrando driver: {e}")

            # Dashboard/video: independientes de que el driver haya cerrado bien —
            # solo dependen de session_id, que ya lo tenemos guardado localmente.
            if session_id:
                dashboard_url = f"https://automation.lambdatest.com/test?testID={session_id}"
                video_url = ""
                if username and access_key:
                    try:
                        video_url = _fetch_lt_video_url(session_id, username, access_key)
                    except Exception as e:
                        log(f"  ⚠ Error obteniendo video LT: {e}")
                summary["video_url"] = video_url
                if video_url:
                    log(f"  Video LT: {video_url}")
                log(f"  Dashboard LT: {dashboard_url}")
                if results_path:
                    try:
                        # Actualizar video/dashboard en todas las filas de datos
                        for _r in range(2, ws.max_row + 1):
                            if video_url and "Video LT" in col_idx:
                                ws.cell(row=_r, column=col_idx["Video LT"]).value = video_url
                            if "Dashboard LT" in col_idx:
                                ws.cell(row=_r, column=col_idx["Dashboard LT"]).value = dashboard_url
                        wb.save(results_path)
                    except Exception as e:
                        log(f"  ⚠ Error guardando Video/Dashboard LT en Excel: {e}")

    # Layout final: columnas de RESULTADO primero, datos de entrada al final. Se hace
    # recién acá porque durante la corrida los runners leen los datos del lead de la
    # misma hoja (URL en A, Formulario en B) y reordenar antes desalinearía la lectura.
    try:
        from utils.excel_layout import reordenar_archivo
        reordenar_archivo(results_path, log=log)
    except Exception as _re:
        log(f"  Aviso: no se pudo reordenar el Excel: {_re}")

    log(f"\n{'='*60}")
    log(f"RESUMEN: {summary['ok']}/{summary['total']} OK, {summary['failed']} con issues.")
    if summary.get("results_excel"):
        log(f"Excel: {summary['results_excel']}")
    if video_url:
        log(f"Video: {video_url}")
    log(f"{'='*60}")

    return summary
