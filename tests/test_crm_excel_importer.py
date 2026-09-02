"""Tests de utils.crm_excel_importer.CrmValidacionesImporter.

No usan el excel real del usuario (dato externo sensible, nunca se versiona): cada test
arma un workbook sintético con openpyxl en un archivo temporal.
"""
import datetime
import json
import os
import re

import openpyxl
import pytest

from utils.crm_excel_importer import CrmValidacionesImporter, construir_reporte, guardar_reporte


def _crear_excel_sintetico(tmp_path):
    ruta = os.path.join(tmp_path, "excel_sintetico.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja no-país: debe ser ignorada por completo.
    hoja_reportes = wb.create_sheet("GMSA | Reportes")
    hoja_reportes.append(["DASH", "https://example.com", "algo"])

    # Argentina: usa "MENSAJE DE ERROR2" (sin espacio) como variante de header.
    ar = wb.create_sheet("Argentina")
    ar.append(["CAMPO", "TIPO", "OBLIGATORIO", "VALIDACIONES", "MENSAJE DE ERROR",
               "MENSAJE DE ERROR2", "ÚLTIMO AJUSTE"])
    # Ya existe en el JSON con un regex_full que coincide con lo que deriva su prosa -> "ok".
    ar.append(["NOMBRE", "Rellenable", True, "De 2 a 50 caracteres.",
               "Ingresá tu nombre.", "Ingresá mínimo 2 letras.", "19 de Junio"])
    # Existe en el JSON pero regex_full vacío y obligatorio=True -> "incompletos", y su
    # prosa alcanza para derivar la regla, así que el merge se la completa.
    ar.append(["APELLIDO", "Rellenable", True, "De 2 a 50 caracteres.",
               "Ingresá tu apellido.", "Ingresá mínimo 2 letras.", None])
    # No existe en el JSON -> "faltantes", se debe crear en el merge.
    ar.append(["DNI", "Rellenable", True, "De 7 a 9 caracteres numéricos.",
               "Ingresá tu DNI.", "Ingresá mínimo 7 dígitos.", None])
    # Caso real que motivó la feature: la descripción dice "de 1 a 10" pero el mensaje de
    # error dice "10 caracteres", y el form rechaza cualquier cosa que no tenga 10 -> el
    # regex cargado quedó mal y esto debe caer en "conflictos".
    ar.append(["NÚMERO DE CONTRATO", "Rellenable", True,
               "De 1 a 10 caracteres numéricos. No permite iniciar con 0.",
               "Ingresá el número de contrato.",
               "Ingresá 10 caracteres que no comiencen con 0.", None])

    # Chile: usa "MENSAJE DE ERROR 2" (con espacio) como variante de header, y
    # "DESCRIPCIÓN" en vez de "VALIDACIONES" para la columna de prosa.
    cl = wb.create_sheet("Chile")
    cl.append(["CAMPO", "TIPO", "OBLIGATORIO", "DESCRIPCIÓN", "MENSAJE DE ERROR",
               "MENSAJE DE ERROR 2", "ÚLTIMO AJUSTE"])
    # ÚLTIMO AJUSTE con datetime real de Excel (varias hojas del CRM real lo traen así,
    # no como texto) — tiene que quedar como string parseado, nunca crudo.
    cl.append(["RUT", "Rellenable", True, "Posee un algoritmo de validación.",
               "Ingresá tu RUT.", None, datetime.datetime(2026, 6, 30, 0, 0)])

    wb.save(ruta)
    return ruta


# El regex que la prosa de NOMBRE deriva, calcado: si el JSON ya trae exactamente este,
# el campo va a "ok" y nada se recalcula.
REGEX_NOMBRE_DERIVADO = r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ ]{2,50}$"
# El regex viejo de NÚMERO DE CONTRATO, el que aceptaba 1 dígito y dejaba pasar "324".
REGEX_CONTRATO_VIEJO = r"^(?=(?:1|2|3|4|5|6|7|8|9))[0-9]{1,10}$"


def _reglas_json_argentina():
    return {
        "fields": {
            "NOMBRE": {
                "regex_full": REGEX_NOMBRE_DERIVADO,
                "descripcion": "NOMBRE",
            },
            "APELLIDO": {
                "regex_full": "",
                "descripcion": "APELLIDO",
            },
            "NÚMERO DE CONTRATO": {
                "regex_full": REGEX_CONTRATO_VIEJO,
                "descripcion": "NÚMERO DE CONTRATO",
            },
        }
    }


@pytest.fixture
def excel_sintetico(tmp_path):
    return _crear_excel_sintetico(str(tmp_path))


class TestDisponible:
    def test_archivo_inexistente_devuelve_false_con_mensaje_claro(self, tmp_path):
        importer = CrmValidacionesImporter(ruta_excel=str(tmp_path / "no_existe.xlsx"))
        disponible, mensaje = importer.disponible()
        assert disponible is False
        assert "no_existe.xlsx" in mensaje

    def test_archivo_existente_devuelve_true_con_ruta(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        disponible, ruta = importer.disponible()
        assert disponible is True
        assert ruta == excel_sintetico


class TestParsear:
    def test_ignora_hojas_no_pais_conocidas_sin_marcarlas_como_no_reconocidas(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        campos_por_pais = importer.parsear()
        assert "GMSA | Reportes" not in campos_por_pais
        # Es una hoja conocida-y-esperada (NON_COUNTRY_SHEETS), no una desconocida.
        assert "GMSA | Reportes" not in importer.hojas_no_reconocidas

    def test_hoja_con_nombre_de_pais_no_soportado_va_a_no_reconocidas(self, tmp_path):
        ruta = os.path.join(str(tmp_path), "excel_con_hoja_rara.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        rara = wb.create_sheet("Venezuela")
        rara.append(["CAMPO", "TIPO", "OBLIGATORIO"])
        rara.append(["NOMBRE", "Rellenable", True])
        wb.save(ruta)

        importer = CrmValidacionesImporter(ruta_excel=ruta)
        campos_por_pais = importer.parsear()

        assert "Venezuela" not in campos_por_pais
        assert "Venezuela" in importer.hojas_no_reconocidas

    def test_parsea_ambas_hojas_de_pais(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        campos_por_pais = importer.parsear()
        assert set(campos_por_pais.keys()) == {"Argentina", "Chile"}
        assert len(campos_por_pais["Argentina"]) == 4
        assert len(campos_por_pais["Chile"]) == 1

    def test_detecta_columna_mensaje_error_con_headers_no_uniformes(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        campos_por_pais = importer.parsear()

        nombre = next(c for c in campos_por_pais["Argentina"] if c["nombre_campo"] == "NOMBRE")
        assert nombre["mensajes_error"] == ["Ingresá tu nombre.", "Ingresá mínimo 2 letras."]

        rut = next(c for c in campos_por_pais["Chile"] if c["nombre_campo"] == "RUT")
        assert rut["mensajes_error"] == ["Ingresá tu RUT."]

    def test_ultimo_ajuste_datetime_se_convierte_a_string(self, excel_sintetico):
        """El excel real del CRM trae ÚLTIMO AJUSTE como datetime nativo de Excel en
        varias hojas (Chile, Colombia, Ecuador, Paraguay, Peru, Uruguay, Bolivia) — si se
        guarda crudo, json.dump revienta más adelante en construir_reporte/guardar_reporte
        ("Object of type datetime is not JSON serializable")."""
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        campos_por_pais = importer.parsear()

        rut = next(c for c in campos_por_pais["Chile"] if c["nombre_campo"] == "RUT")
        assert isinstance(rut["ultimo_ajuste"], str)
        assert rut["ultimo_ajuste"] == "2026-06-30"

    def test_cachea_el_parseo_no_relee_el_archivo(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        primera = importer.parsear()
        os.remove(excel_sintetico)
        segunda = importer.parsear()
        assert segunda is primera


class TestComparar:
    def test_campo_con_regex_ya_armada_va_a_ok_y_no_a_incompletos(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Argentina", _reglas_json_argentina())

        nombres_ok = {c["nombre_campo"] for c in comparacion["ok"]}
        assert "NOMBRE" in nombres_ok
        assert "NOMBRE" not in {c["nombre_campo"] for c in comparacion["incompletos"]}
        assert "NOMBRE" not in {c["nombre_campo"] for c in comparacion["faltantes"]}
        assert "NOMBRE" not in {c["nombre_campo"] for c in comparacion["conflictos"]}

    def test_campo_cuya_prosa_contradice_el_regex_cargado_va_a_conflictos(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Argentina", _reglas_json_argentina())

        conflictos = {c["nombre_campo"]: c for c in comparacion["conflictos"]}
        assert "NÚMERO DE CONTRATO" in conflictos
        contrato = conflictos["NÚMERO DE CONTRATO"]
        assert contrato["regex_actual"] == REGEX_CONTRATO_VIEJO
        assert contrato["regex_derivado"] == r"^[1-9][0-9]{9}$"
        # Lo que realmente se aplica es el ajuste de largo sobre la regla existente.
        assert contrato["regex_ajustado"] == r"^(?=(?:1|2|3|4|5|6|7|8|9))[0-9]{10}$"
        # El motivo cita el mensaje de error, que es el que ganó sobre la descripción.
        assert "10 caracteres" in contrato["motivo_derivacion"]

    def test_campo_existente_sin_regex_y_obligatorio_va_a_incompletos(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Argentina", _reglas_json_argentina())

        nombres_incompletos = {c["nombre_campo"] for c in comparacion["incompletos"]}
        assert "APELLIDO" in nombres_incompletos

    def test_campo_inexistente_va_a_faltantes(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Argentina", _reglas_json_argentina())

        nombres_faltantes = {c["nombre_campo"] for c in comparacion["faltantes"]}
        assert "DNI" in nombres_faltantes


class TestAplicarMerge:
    def test_no_toca_un_campo_cuyo_regex_ya_coincide_con_su_prosa(self, excel_sintetico):
        """Un regex que la prosa confirma queda intacto, y la entrada original no se muta."""
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        reglas = _reglas_json_argentina()
        nombre_original = dict(reglas["fields"]["NOMBRE"])

        comparacion = importer.comparar("Argentina", reglas)
        actualizado, _ = importer.aplicar_merge("Argentina", comparacion, reglas)

        assert actualizado["fields"]["NOMBRE"] == nombre_original
        # aplicar_merge trabaja sobre una copia: el dict de entrada no se muta.
        assert reglas["fields"]["NOMBRE"] == nombre_original

    def test_completa_el_regex_de_un_obligatorio_que_no_lo_tenia(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        reglas = _reglas_json_argentina()
        comparacion = importer.comparar("Argentina", reglas)

        actualizado, cambios = importer.aplicar_merge("Argentina", comparacion, reglas)

        apellido = actualizado["fields"]["APELLIDO"]
        assert apellido["regex_full"] == REGEX_NOMBRE_DERIVADO
        assert apellido["origen"] == "regex_derivado_prosa"
        assert apellido["motivo_derivacion"]
        assert "pendiente_regex" not in apellido
        assert cambios["completados"] == 1

    def test_conflicto_corrige_el_largo_y_conserva_el_resto_de_la_regla(self, excel_sintetico):
        """El caso que motivó la feature: la descripción decía 1-10 y el form pide 10.

        Se ajusta el cuantificador y NADA más: el lookahead que impide empezar con 0 sigue
        ahí, aunque la prosa no vuelva a mencionarlo."""
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        reglas = _reglas_json_argentina()
        comparacion = importer.comparar("Argentina", reglas)

        actualizado, cambios = importer.aplicar_merge("Argentina", comparacion, reglas)

        contrato = actualizado["fields"]["NÚMERO DE CONTRATO"]
        assert contrato["regex_full"] == r"^(?=(?:1|2|3|4|5|6|7|8|9))[0-9]{10}$"
        assert contrato["regex_full_previo"] == REGEX_CONTRATO_VIEJO
        assert contrato["origen"] == "largo_ajustado_prosa"
        assert cambios["recalculados"] == 1
        # El valor que fallaba en producción ya no pasa; uno de 10 dígitos sí.
        assert re.fullmatch(contrato["regex_full"], "5372819044")
        assert not re.fullmatch(contrato["regex_full"], "324")
        assert not re.fullmatch(contrato["regex_full"], "0372819044")

    def test_conflicto_conserva_la_clase_de_caracteres_al_ajustar_el_largo(self, excel_sintetico):
        """La regla admite espacios y puntuación; la prosa de NOMBRE sólo habla de "letras"
        y de un largo 2-50. Se corrige el largo y la clase queda intacta: derivar el regex
        entero desde la prosa dejaría un comentario sin espacios."""
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        reglas = {"fields": {"NOMBRE": {"regex_full": r"^[a-zA-ZñÑ .,;:!?]{5,500}$"}}}
        comparacion = importer.comparar("Argentina", reglas)

        actualizado, cambios = importer.aplicar_merge("Argentina", comparacion, reglas)

        assert actualizado["fields"]["NOMBRE"]["regex_full"] == r"^[a-zA-ZñÑ .,;:!?]{5,50}$"
        assert cambios["recalculados"] == 1
        # Lo que importa: la puntuación y el espacio siguen permitidos.
        assert re.fullmatch(actualizado["fields"]["NOMBRE"]["regex_full"], "Hola, me interesa.")

    def test_conflicto_sin_largo_aplicable_no_toca_la_regla(self, excel_sintetico):
        """Cuando el largo cargado ya cae dentro de lo que pide la prosa no hay nada que
        corregir, pero el conflicto igual queda registrado para revisarlo a mano."""
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        regla = r"^[a-zA-ZñÑ .,;:!?]{2,50}$"
        reglas = {"fields": {"NOMBRE": {"regex_full": regla}}}
        comparacion = importer.comparar("Argentina", reglas)

        actualizado, cambios = importer.aplicar_merge("Argentina", comparacion, reglas)

        assert actualizado["fields"]["NOMBRE"]["regex_full"] == regla
        assert "regex_full_previo" not in actualizado["fields"]["NOMBRE"]
        assert cambios["recalculados"] == 0
        assert "NOMBRE" in {c["nombre_campo"] for c in comparacion["conflictos"]}

    def test_agrega_solo_los_campos_faltantes(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        reglas = _reglas_json_argentina()
        comparacion = importer.comparar("Argentina", reglas)

        actualizado, cambios = importer.aplicar_merge("Argentina", comparacion, reglas)

        assert cambios["agregados"] == 1
        assert "DNI" in actualizado["fields"]
        nuevo = actualizado["fields"]["DNI"]
        # Su prosa alcanza para derivar la regla, así que nace con regex en vez de pendiente.
        assert nuevo["regex_full"] == r"^[0-9]{7,9}$"
        assert nuevo["origen"] == "regex_derivado_prosa"
        assert "pendiente_regex" not in nuevo
        assert nuevo["obligatorio_excel"] is True
        assert nuevo["mensajes_error_excel"] == ["Ingresá tu DNI.", "Ingresá mínimo 7 dígitos."]

    def test_faltante_sin_prosa_util_nace_pendiente_de_revision(self, excel_sintetico):
        """Sin con qué derivar, el campo se crea igual pero marcado para revisión manual:
        el módulo no inventa una regla."""
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Chile", {"fields": {}})

        actualizado, cambios = importer.aplicar_merge("Chile", comparacion, {"fields": {}})

        rut = actualizado["fields"]["RUT"]  # "Posee un algoritmo de validación."
        assert rut["regex_full"] == ""
        assert rut["pendiente_regex"] is True
        assert rut["origen"] == "crm_excel_import"
        assert cambios["agregados"] == 1

    def test_no_cambia_nada_si_no_hay_faltantes_ni_conflictos(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        reglas_chile = {"fields": {"RUT": {"regex_full": "^[0-9kK.-]+$"}}}
        comparacion = importer.comparar("Chile", reglas_chile)

        actualizado, cambios = importer.aplicar_merge("Chile", comparacion, reglas_chile)

        assert cambios == {"agregados": 0, "completados": 0, "recalculados": 0}
        assert actualizado["fields"]["RUT"]["regex_full"] == "^[0-9kK.-]+$"


class TestReporte:
    def test_construir_reporte_tiene_el_shape_documentado(self, excel_sintetico):
        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Argentina", _reglas_json_argentina())

        reporte = construir_reporte("Argentina", "Argentina", comparacion)

        assert reporte["pais"] == "Argentina"
        assert reporte["hoja_excel"] == "Argentina"
        assert "fecha_importacion" in reporte
        assert reporte["resumen"] == {"faltantes": 1, "incompletos": 1,
                                      "conflictos": 1, "ok": 1}
        assert reporte["conflictos"] == comparacion["conflictos"]
        assert reporte["faltantes"] == comparacion["faltantes"]
        assert reporte["incompletos"] == comparacion["incompletos"]
        assert reporte["ok"] == comparacion["ok"]

    def test_guardar_reporte_no_revienta_con_ultimo_ajuste_datetime(self, excel_sintetico, tmp_path, monkeypatch):
        json_dir = str(tmp_path / "json")
        monkeypatch.setattr("utils.crm_excel_importer.JSON_DIR", json_dir)

        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Chile", {"fields": {}})
        reporte = construir_reporte("Chile", "Chile", comparacion)

        ruta_guardada = guardar_reporte("Chile", reporte)  # no debe lanzar TypeError

        with open(ruta_guardada, "r", encoding="utf-8") as fh:
            contenido = json.load(fh)
        assert contenido["faltantes"][0]["ultimo_ajuste"] == "2026-06-30"

    def test_guardar_reporte_escribe_json_dir_con_nombre_por_pais(self, excel_sintetico, tmp_path, monkeypatch):
        json_dir = str(tmp_path / "json")
        monkeypatch.setattr("utils.crm_excel_importer.JSON_DIR", json_dir)

        importer = CrmValidacionesImporter(ruta_excel=excel_sintetico)
        comparacion = importer.comparar("Argentina", _reglas_json_argentina())
        reporte = construir_reporte("Argentina", "Argentina", comparacion)

        ruta_guardada = guardar_reporte("Argentina", reporte)

        assert ruta_guardada == os.path.join(json_dir, "crm_import_report_argentina.json")
        with open(ruta_guardada, "r", encoding="utf-8") as fh:
            contenido = json.load(fh)
        assert contenido["pais"] == "Argentina"
        assert contenido["resumen"]["faltantes"] == 1
