"""Tests de utils.muestreo_gm_forms.MuestreadorGmForms.

No usan el excel real del usuario (dato externo sensible, nunca se versiona): cada test
arma un workbook sintético con openpyxl en un archivo temporal, con la misma forma que
'GM Forms - 2026.xlsx' (fila de título del país, fila de headers, filas de datos).
"""
import os

import openpyxl
import pytest

from utils.muestreo_gm_forms import MuestreadorGmForms


def _crear_excel_sintetico(tmp_path, filas_argentina=None):
    ruta = os.path.join(tmp_path, "gm_forms_sintetico.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ar = wb.create_sheet("ARGENTINA")
    ar.append(["ARGENTINA", None, None, "Última revisión:"])
    ar.append(["SEGMENTO", "ESTADO", "URL LIVE", "URL SECURE"])
    for fila in (filas_argentina or []):
        ar.append(fila)

    # Hoja que no corresponde a ningún país soportado: debe ignorarse.
    otra = wb.create_sheet("RESUMEN GLOBAL")
    otra.append(["algo", "que", "no", "es", "un", "pais"])

    wb.save(ruta)
    return ruta


FILAS_MIXTAS = [
    ["VEHICLE SHOPPER", "✔️ ON", "https://a.com/1", "https://secure.com/1"],
    ["SIN STOCK", "❌ OFF", "https://a.com/2", "https://secure.com/2"],
    ["NUEVO", "🆕 PRIORIDAD", "https://a.com/3", "https://secure.com/3"],
    ["OTRO", "✔️ ON", "https://a.com/4", "https://secure.com/4"],
    ["OTRO MAS", "❌ OFF", "https://a.com/5", "https://secure.com/5"],
]

# Fila basura real vista en 'GM Forms - 2026.xlsx' (Uruguay/Bolivia): el estado quedó
# corrido a la columna SEGMENTO y el resto de la fila está vacío — sin URL no sirve para
# verificar nada, aunque el texto no contenga "OFF".
FILAS_CON_BASURA = FILAS_MIXTAS + [["✔️ ON", None, None, None]]


class TestMuestrearHoja:
    def test_nunca_incluye_filas_con_estado_off(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=10, seed=1)

        libro = openpyxl.load_workbook(ruta, data_only=True)
        header, muestra = muestreador._muestrear_hoja(libro["ARGENTINA"])

        estados_muestreados = {fila[1] for fila in muestra}
        assert "❌ OFF" not in estados_muestreados
        assert len(muestra) == 3  # las 2 "OFF" quedan afuera de las 5 filas

    def test_respeta_n_por_pais(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=2, seed=1)

        libro = openpyxl.load_workbook(ruta, data_only=True)
        _, muestra = muestreador._muestrear_hoja(libro["ARGENTINA"])

        assert len(muestra) == 2

    def test_toma_todas_si_hay_menos_activas_que_n(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=50, seed=1)

        libro = openpyxl.load_workbook(ruta, data_only=True)
        _, muestra = muestreador._muestrear_hoja(libro["ARGENTINA"])

        assert len(muestra) == 3

    def test_preserva_el_header_completo(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=1, seed=1)

        libro = openpyxl.load_workbook(ruta, data_only=True)
        header, _ = muestreador._muestrear_hoja(libro["ARGENTINA"])

        assert header == ["SEGMENTO", "ESTADO", "URL LIVE", "URL SECURE"]

    def test_descarta_filas_sin_ninguna_url_aunque_no_diga_off(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_CON_BASURA)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=None, seed=1)

        libro = openpyxl.load_workbook(ruta, data_only=True)
        _, muestra = muestreador._muestrear_hoja(libro["ARGENTINA"])

        assert len(muestra) == 3  # las 4 activas de FILAS_MIXTAS menos la basura sin URL
        assert all(fila[2] or fila[3] for fila in muestra)

    def test_n_por_pais_none_toma_todas_las_activas_sin_muestrear(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=None, seed=1)

        libro = openpyxl.load_workbook(ruta, data_only=True)
        _, muestra = muestreador._muestrear_hoja(libro["ARGENTINA"])

        assert len(muestra) == 3
        assert all(fila[1] != "❌ OFF" for fila in muestra)

    def test_misma_seed_da_la_misma_muestra(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)

        libro1 = openpyxl.load_workbook(ruta, data_only=True)
        _, muestra1 = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=2, seed=42)._muestrear_hoja(libro1["ARGENTINA"])

        libro2 = openpyxl.load_workbook(ruta, data_only=True)
        _, muestra2 = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=2, seed=42)._muestrear_hoja(libro2["ARGENTINA"])

        assert muestra1 == muestra2


class TestGenerarExcelMuestra:
    def test_genera_archivo_con_solo_hojas_de_pais(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=2, seed=1)
        salida = str(tmp_path / "muestra.xlsx")

        ruta_generada = muestreador.generar_excel_muestra(salida=salida)

        assert ruta_generada == salida
        wb = openpyxl.load_workbook(salida)
        assert wb.sheetnames == ["ARGENTINA"]
        assert "RESUMEN GLOBAL" not in wb.sheetnames

    def test_hoja_de_salida_tiene_header_mas_filas_muestreadas(self, tmp_path):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=2, seed=1)
        salida = str(tmp_path / "muestra.xlsx")

        muestreador.generar_excel_muestra(salida=salida)

        wb = openpyxl.load_workbook(salida)
        ws = wb["ARGENTINA"]
        filas = list(ws.iter_rows(values_only=True))
        assert filas[0] == ("SEGMENTO", "ESTADO", "URL LIVE", "URL SECURE")
        assert len(filas) == 1 + 2  # header + n_por_pais

    def test_archivo_inexistente_levanta_error_claro(self, tmp_path):
        muestreador = MuestreadorGmForms(ruta_excel=str(tmp_path / "no_existe.xlsx"))
        with pytest.raises(FileNotFoundError, match="no_existe.xlsx"):
            muestreador.generar_excel_muestra(salida=str(tmp_path / "salida.xlsx"))

    def test_sin_salida_explicita_guarda_en_results_dir(self, tmp_path, monkeypatch):
        ruta = _crear_excel_sintetico(str(tmp_path), FILAS_MIXTAS)
        results_dir = str(tmp_path / "resultados")
        monkeypatch.setattr("utils.muestreo_gm_forms.RESULTS_DIR", results_dir)

        muestreador = MuestreadorGmForms(ruta_excel=ruta, n_por_pais=1, seed=1)
        ruta_generada = muestreador.generar_excel_muestra()

        assert ruta_generada.startswith(results_dir)
        assert os.path.exists(ruta_generada)
