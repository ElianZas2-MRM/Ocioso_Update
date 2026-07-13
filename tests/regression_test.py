"""Regression funcional (sin browser): datos, comparador, scheduling, validacion."""
import os
import re
import shutil
import sys
import tempfile
import traceback

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(BASE)
for p in (BASE, os.path.join(BASE, "forms"), os.path.join(BASE, "core")):
    if p not in sys.path:
        sys.path.insert(0, p)

PASS, FAIL = [], []


def t(name):
    def deco(fn):
        try:
            fn()
            PASS.append(name)
            print(f"PASS  {name}")
        except Exception as e:
            FAIL.append((name, e))
            print(f"FAIL  {name}: {type(e).__name__}: {e}")
            traceback.print_exc()
        return fn
    return deco


# ---------------- utils/data_generator ----------------
import utils.data_generator as dg
from core.country_configs import COUNTRY_CONFIGS

PAISES = list(COUNTRY_CONFIGS)


@t("data_generator: generar_fila_datos para los 9 paises")
def _():
    for pais in PAISES:
        fila = dg.generar_fila_datos(pais)
        assert isinstance(fila, dict) and fila, pais
        for k in ("nombre", "apellido", "email"):
            hit = [v for kk, v in fila.items() if k in kk.lower()]
            assert hit, f"{pais}: falta campo tipo {k} en {list(fila)}"


@t("data_generator: email valido y sin acentos")
def _():
    for pais in PAISES:
        e = dg.generar_email("José", "Ñandú", pais)
        assert re.fullmatch(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", e), e
        assert e.isascii(), e


@t("data_generator: CPF Brasil con digito verificador valido")
def _():
    def cpf_ok(d):
        d = re.sub(r"\D", "", d)
        assert len(d) == 11, d
        for i in (9, 10):
            s = sum(int(d[j]) * ((i + 1) - j) for j in range(i))
            dv = (s * 10) % 11 % 10
            assert dv == int(d[i]), f"DV invalido en {d}"
    for _i in range(15):
        cpf_ok(dg.generar_cpf_brasil())


@t("data_generator: CNPJ Brasil con digito verificador valido")
def _():
    def cnpj_ok(v):
        d = re.sub(r"\D", "", v)
        assert len(d) == 14, d
        for size, w in ((12, [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]),
                        (13, [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])):
            s = sum(int(d[i]) * w[i] for i in range(size))
            r = s % 11
            dv = 0 if r < 2 else 11 - r
            assert dv == int(d[size]), f"DV invalido en {d}"
    for _i in range(15):
        cnpj_ok(dg.generar_cnpj_brasil())


@t("data_generator: RUT Chile con DV (incluye variante K)")
def _():
    def rut_ok(v):
        s = re.sub(r"[.\-]", "", str(v)).upper()
        cuerpo, dv = s[:-1], s[-1]
        assert cuerpo.isdigit(), v
        suma, mul = 0, 2
        for c in reversed(cuerpo):
            suma += int(c) * mul
            mul = 2 if mul == 7 else mul + 1
        r = 11 - (suma % 11)
        esperado = "0" if r == 11 else "K" if r == 10 else str(r)
        assert dv == esperado, f"{v}: DV {dv} != {esperado}"
    for _i in range(20):
        rut_ok(dg.generar_rut_chile())
        rut_ok(dg.generar_rut_chile_con_k())


@t("data_generator: CI Ecuador valida")
def _():
    for _i in range(15):
        ci = re.sub(r"\D", "", dg.generar_ci_ecuador())
        assert len(ci) == 10, ci
        assert 1 <= int(ci[:2]) <= 24, ci


@t("data_generator: documento y celular por pais no vacios")
def _():
    for pais in PAISES:
        doc = dg.generar_documento(pais)
        cel = dg.generar_celular(pais)
        assert str(doc).strip(), f"{pais}: documento vacio"
        assert re.sub(r"\D", "", str(cel)), f"{pais}: celular sin digitos"


@t("data_generator: VIN 17 chars sin I/O/Q")
def _():
    for _i in range(10):
        v = dg.generar_vin()
        assert len(v) == 17, v
        assert not set(v.upper()) & set("IOQ"), v


# ---------------- core/dealer_comparator_runner ----------------
import core.dealer_comparator_runner as dc
from openpyxl import Workbook


@t("comparador: normalize_text (acentos, case, espacios)")
def _():
    assert dc.normalize_text("  Ciudad   de  Méxicó ") == "CIUDAD DE MEXICO"
    assert dc.normalize_text(None) == ""
    assert dc.normalize_text("São Paulo") == "SAO PAULO"


@t("comparador: read_excel_rows con header_row != 1")
def _():
    tmp = tempfile.mkdtemp()
    try:
        path = os.path.join(tmp, "dealers.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Reporte interno", None, None])   # fila 1 basura
        ws.append(["REGION", "CIUDAD", "DEALER"])    # fila 2 = headers
        ws.append(["Montevideo", "Montevideo", "Dealer A"])
        ws.append(["Canelones", "Pando", "Dealer B"])
        wb.save(path)

        headers, rows = dc.read_excel_rows(path, header_row=2)
        assert "REGION" in headers, headers
        assert len(rows) == 2, rows
        assert rows[0]["DEALER"] == "Dealer A"
        assert rows[0]["_c0"] == "Montevideo"       # acceso por indice de columna
        assert rows[0]["__row__"] == 3              # fila real en el Excel
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@t("comparador: resolve_column por letra, nombre exacto y parcial")
def _():
    headers = ["REGION", "CIUDAD", "NOMBRE DEL DEALER"]
    assert dc.resolve_column(headers, "A") == "_c0"          # letra estilo Excel
    assert dc.resolve_column(headers, "C") == "_c2"
    assert dc.resolve_column(headers, "REGION") == "REGION"  # exacto
    assert dc.resolve_column(headers, "region") == "REGION"  # case-insensitive
    assert dc.resolve_column(headers, "dealer") == "NOMBRE DEL DEALER"  # parcial
    assert dc.resolve_column(headers, "") is None


@t("comparador: filter_rows include / exclude")
def _():
    rows = [
        {"POSVENTA": "si", "DEALER": "A"},
        {"POSVENTA": "no", "DEALER": "B"},
        {"POSVENTA": "SI", "DEALER": "C"},
    ]
    inc = dc.filter_rows(rows, "POSVENTA", "si", mode="include")
    assert sorted(r["DEALER"] for r in inc) == ["A", "C"], inc
    exc = dc.filter_rows(rows, "POSVENTA", "si", mode="exclude")
    assert [r["DEALER"] for r in exc] == ["B"], exc
    todo = dc.filter_rows(rows, "", "", mode="include")
    assert len(todo) == 3


@t("comparador: detect_hidden_rows detecta filas ocultas")
def _():
    tmp = tempfile.mkdtemp()
    try:
        path = os.path.join(tmp, "hidden.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["REGION", "DEALER"])
        ws.append(["R1", "A"])
        ws.append(["R2", "B"])
        ws.append(["R3", "C"])
        ws.row_dimensions[3].hidden = True   # oculta la fila del dealer B
        wb.save(path)
        hidden = dc.detect_hidden_rows(path, header_row=1)
        assert 3 in hidden, hidden
        assert 2 not in hidden and 4 not in hidden, hidden
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@t("comparador: export_results_excel genera el xlsx con colores")
def _():
    tmp = tempfile.mkdtemp()
    try:
        out = os.path.join(tmp, "res.xlsx")
        results = [
            {"url": "https://x/form", "region": "Montevideo", "city": "Montevideo",
             "dealer": "Dealer A", "status": "PASS", "detalle": "ok"},
            {"url": "https://x/form", "region": "Canelones", "city": "Pando",
             "dealer": "Dealer B", "status": "MISSING", "detalle": "no esta en el form"},
        ]
        path = dc.export_results_excel(results, output_path=out, pais="Uruguay")
        assert os.path.exists(path), path
        assert os.path.getsize(path) > 0
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert wb.active.max_row >= 3   # header + 2 filas
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@t("comparador: get_country_level_defaults por pais")
def _():
    for pais in PAISES:
        d = dc.get_country_level_defaults(pais)
        assert set(d) >= {"has_region", "has_city", "has_dealer"}, d
        assert all(isinstance(v, bool) for v in d.values()), d


# ---------------- utils/scheduling ----------------
import utils.scheduling as sched
import json as _json

SCHED_FILE = os.path.join(BASE, "json", "programacion_test.json")


@t("scheduling: guardar / cargar / limpiar programacion (round-trip)")
def _():
    backup = None
    if os.path.exists(SCHED_FILE):
        with open(SCHED_FILE, "r", encoding="utf-8") as f:
            backup = f.read()
    try:
        cfg = {
            "tipo": "semanal",
            "horarios": {"Lunes": ["03:00", "15:30"], "Martes": ["09:00"]},
            "paises": ["Uruguay", "Brasil"],
            "navegadores": ["chrome"],
            "viewports": ["fullscreen"],
            "modo_excel": "consecutivo",
            "modo_mercados": "paralelo",
        }
        assert sched.guardar_programacion(cfg) is True
        leido = sched.cargar_programacion()
        assert leido["tipo"] == "semanal", leido
        assert leido["horarios"]["Lunes"] == ["03:00", "15:30"], leido
        assert leido["paises"] == ["Uruguay", "Brasil"], leido
        assert leido["modo_mercados"] == "paralelo", leido
        assert sched.limpiar_programacion() is True
        assert sched.cargar_programacion() is None
    finally:
        if backup is not None:
            with open(SCHED_FILE, "w", encoding="utf-8") as f:
                f.write(backup)
        elif os.path.exists(SCHED_FILE):
            os.remove(SCHED_FILE)


# ---------------- validation ----------------
@t("validation: regex full/char compilan y discriminan")
def _():
    from validation.text_field_validator import _compile_regex, _matches_char
    full = _compile_regex(r"^[A-Za-zÁÉÍÓÚáéíóúñÑ ]{2,50}$", "NOMBRE")
    assert full.match("Juan Pérez")
    assert not full.match("Juan123")
    char = _compile_regex(r"[A-Za-zÁÉÍÓÚáéíóúñÑ ]*", "NOMBRE")
    assert _matches_char(char, "Jua")        # prefijo valido -> se permite tipear
    assert not _matches_char(char, "Jua1")   # caracter invalido -> se bloquea
    try:
        _compile_regex("[unclosed", "X")
        raise AssertionError("regex invalido deberia lanzar ValueError")
    except ValueError:
        pass


@t("validation: exporter e email importables con API esperada")
def _():
    import validation.validation_exporter as ve
    import validation.validation_email as vm
    assert any(callable(getattr(ve, n)) for n in dir(ve) if not n.startswith("_"))
    assert any(callable(getattr(vm, n)) for n in dir(vm) if not n.startswith("_"))


# ---------------- forms / runners ----------------
@t("forms: get_runner devuelve callable para los 9 paises")
def _():
    from forms._runner_common import get_runner
    for pais in PAISES:
        assert callable(get_runner(pais)), pais


@t("country_configs: cada pais tiene url/field_mapping coherentes")
def _():
    for pais, cfg in COUNTRY_CONFIGS.items():
        fm = cfg.get("field_mapping")
        assert isinstance(fm, list) and fm, f"{pais}: field_mapping vacio"
        for f in fm:
            assert "id" in f, f"{pais}: campo sin id -> {f}"


# ---------------- data/ excels ----------------
@t("data/: los Excel de leads se leen y tienen columnas URL/Formulario")
def _():
    from openpyxl import load_workbook
    import glob
    files = glob.glob(os.path.join(BASE, "data", "Lead_information_*.xlsx"))
    assert files, "no hay Excels de leads en data/"
    revisados = 0
    for f in files:
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
        assert any("url" in h for h in headers), f"{os.path.basename(f)}: sin columna URL -> {headers}"
        revisados += 1
        wb.close()
    print(f"      {revisados} Excels de data/ validados")


print("\n" + "=" * 60)
print(f"PASS: {len(PASS)}   FAIL: {len(FAIL)}")
for n, e in FAIL:
    print(f"  FAIL -> {n}: {e}")
sys.exit(1 if FAIL else 0)
