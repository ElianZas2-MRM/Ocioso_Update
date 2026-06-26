import os
import sys

import pandas as pd
from openpyxl.styles import Font, PatternFill


def _get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _ensure_results_dir():
    results_dir = os.path.join(_get_base_dir(), "resultados")
    os.makedirs(results_dir, exist_ok=True)
    return results_dir


def _build_incremental_output_path(results_dir):
    base_name = "Resultado_validaciones"
    index = 1
    while True:
        candidate = os.path.join(results_dir, f"{base_name}{index}.xlsx")
        if not os.path.exists(candidate):
            return candidate
        index += 1


def _normalize_descripcion_column(df):
    if "descripcion" not in df.columns and "field_name" in df.columns:
        df["descripcion"] = df["field_name"]
    return df


def _exclude_no_baseline_rows(df):
    """Omite filas NO_BASELINE para no reportarlas como error en el Excel final."""
    if df.empty or "resultado" not in df.columns:
        return df

    normalized = df["resultado"].fillna("").astype(str).str.strip().str.upper()
    return df[normalized != "NO_BASELINE"].copy()


def _reorder_columns(df, preferred_columns):
    existing_preferred = [col for col in preferred_columns if col in df.columns]
    remaining = [col for col in df.columns if col not in existing_preferred]
    return df[existing_preferred + remaining]


def _sort_errors_first(df):
    if "resultado" not in df.columns or df.empty:
        return df

    normalized = df["resultado"].fillna("").astype(str).str.strip().str.upper()
    priority = normalized.map({"ERROR": 2, "NO_BASELINE": 1}).fillna(0).astype(int)
    sorted_df = df.assign(_result_priority=priority).sort_values(
        by=["_result_priority"], ascending=False, kind="stable"
    )
    return sorted_df.drop(columns=["_result_priority"])


def _build_error_focus_column(df):
    if df.empty or "resultado" not in df.columns:
        return df

    if "error_esperado" in df.columns and "error_real" in df.columns:
        # Hoja de errores UI
        where = []
        for _, row in df.iterrows():
            resultado = str(row.get("resultado") or "").strip().upper()
            if resultado == "NO_BASELINE":
                where.append("Sin mensaje esperado")
                continue

            if resultado != "ERROR":
                where.append("")
                continue

            expected = str(row.get("error_esperado") or "").strip()
            real = str(row.get("error_real") or "").strip()
            detail = str(row.get("detalle") or "").strip()
            if expected != real:
                where.append("Mensaje UI distinto")
            elif detail:
                where.append("Fallo tecnico UI")
            else:
                where.append("Validacion UI")

        enhanced = df.copy()
        enhanced.insert(enhanced.columns.get_loc("resultado") + 1, "donde_falla", where)
        return enhanced

    if "esperado" in df.columns and "real" in df.columns:
        # Hoja detalle por caracter
        where = []
        for _, row in df.iterrows():
            resultado = str(row.get("resultado") or "").strip().upper()
            if resultado != "ERROR":
                where.append("")
                continue

            esperado = str(row.get("esperado") or "").strip().upper()
            real = str(row.get("real") or "").strip().upper()
            if esperado != real:
                where.append("Ingreso de caracter")
            else:
                where.append("Comparacion de resultado")

        enhanced = df.copy()
        enhanced.insert(enhanced.columns.get_loc("resultado") + 1, "donde_falla", where)
        return enhanced

    return df


def _format_worksheet_layout(worksheet):
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_cells in worksheet.columns:
        first = col_cells[0]
        col_letter = first.column_letter
        max_len = len(str(first.value or ""))
        for cell in col_cells[1:]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)

        adjusted = min(max(max_len + 2, 12), 60)
        worksheet.column_dimensions[col_letter].width = adjusted


def _apply_error_highlight(worksheet):
    """Pinta filas con resultado=ERROR/NO_BASELINE y enfatiza la celda de resultado."""
    if worksheet.max_row < 2:
        return

    header = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(worksheet[1], start=1) if cell.value}
    result_col_idx = header.get("resultado")
    if not result_col_idx:
        return

    # Excel theme equivalents aproximados:
    # - Red Accent 2 (base): C0504D
    # - Red Accent 2 40% Lighter: D99694
    row_error_fill = PatternFill(fill_type="solid", fgColor="00D99694")
    result_error_fill = PatternFill(fill_type="solid", fgColor="00C0504D")
    result_error_font = Font(color="00FFFFFF", bold=True)

    row_baseline_fill = PatternFill(fill_type="solid", fgColor="00FCE4D6")
    result_baseline_fill = PatternFill(fill_type="solid", fgColor="00ED7D31")
    result_baseline_font = Font(color="00000000", bold=True)

    for row_idx in range(2, worksheet.max_row + 1):
        result_value = str(worksheet.cell(row=row_idx, column=result_col_idx).value or "").strip().upper()
        if result_value == "ERROR":
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = row_error_fill

            result_cell = worksheet.cell(row=row_idx, column=result_col_idx)
            result_cell.fill = result_error_fill
            result_cell.font = result_error_font
            continue

        if result_value == "NO_BASELINE":
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = row_baseline_fill

            result_cell = worksheet.cell(row=row_idx, column=result_col_idx)
            result_cell.fill = result_baseline_fill
            result_cell.font = result_baseline_font


def export_validation_results(validation_result, output_path=None):
    """Exporta el detalle por caracter y un resumen a un archivo Excel."""
    rows = list(validation_result.get("rows") or [])
    fields = list(validation_result.get("fields") or [])
    error_rows = list(validation_result.get("error_rows") or [])
    unmapped_rows = list(validation_result.get("unmapped_ids") or [])
    summary = dict(validation_result.get("summary") or {})

    if output_path is None:
        output_path = _build_incremental_output_path(_ensure_results_dir())

    detail_df = _normalize_descripcion_column(pd.DataFrame(rows))
    fields_df = _normalize_descripcion_column(pd.DataFrame(fields))
    errors_df = _normalize_descripcion_column(pd.DataFrame(error_rows))
    unmapped_df = pd.DataFrame(unmapped_rows)

    if not unmapped_df.empty:
        dedupe_cols = [col for col in ("landing_url", "form_url", "element_id") if col in unmapped_df.columns]
        if dedupe_cols:
            unmapped_df = unmapped_df.drop_duplicates(subset=dedupe_cols, keep="first")

    errors_df = _exclude_no_baseline_rows(errors_df)

    # El resumen exportado debe reflejar el mismo criterio del reporte filtrado.
    summary["ui_error_tests"] = int(len(errors_df))
    if "resultado" in errors_df.columns:
        ui_errors = int((errors_df["resultado"].fillna("").astype(str).str.strip().str.upper() == "ERROR").sum())
    else:
        ui_errors = 0
    summary["ui_errors"] = ui_errors
    summary["ui_no_baseline"] = 0
    summary["ui_ok"] = int(len(errors_df) - ui_errors)
    summary["ids_no_mapeados"] = int(len(unmapped_df))
    summary_df = pd.DataFrame([summary])

    detail_df = _sort_errors_first(_build_error_focus_column(detail_df))
    errors_df = _sort_errors_first(_build_error_focus_column(errors_df))

    detail_df = _reorder_columns(
        detail_df,
        [
            "element_id",
            "descripcion",
            "step",
            "landing_url",
            "form_url",
            "teclado_mobile",
            "teclado_mobile_ok",
            "char_index",
            "char",
            "esperado",
            "real",
            "resultado",
            "donde_falla",
            "valor_final",
            "regex_ok",
            "regex_full",
            "regex_char",
            "test_text",
            "url",
            "browser",
            "viewport",
            "timestamp",
        ],
    )

    fields_df = _reorder_columns(
        fields_df,
        [
            "element_id",
            "descripcion",
            "step",
            "landing_url",
            "form_url",
            "teclado_mobile",
            "teclado_mobile_ok",
            "regex_ok",
            "final_value",
            "errores",
            "tests_ui",
            "errores_ui",
        ],
    )

    errors_df = _reorder_columns(
        errors_df,
        [
            "campo",
            "input",
            "regex_disparada",
            "error_esperado",
            "error_real",
            "resultado",
            "donde_falla",
            "regla_principal",
            "detalle",
            "trigger",
            "trigger_inicial",
            "fallback_cta",
            "error_selector",
            "element_id",
            "descripcion",
            "field_name",
            "step",
            "landing_url",
            "form_url",
            "url",
            "browser",
            "viewport",
            "timestamp",
        ],
    )

    unmapped_df = _reorder_columns(
        unmapped_df,
        [
            "element_id",
            "tag",
            "name",
            "input_type",
            "placeholder",
            "motivo",
            "step",
            "landing_url",
            "form_url",
            "url",
            "browser",
            "viewport",
            "timestamp",
        ],
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail_df.to_excel(writer, sheet_name="detalle", index=False)
        fields_df.to_excel(writer, sheet_name="campos", index=False)
        errors_df.to_excel(writer, sheet_name="errores_ui", index=False)
        unmapped_df.to_excel(writer, sheet_name="ids_no_testeados", index=False)
        summary_df.to_excel(writer, sheet_name="resumen", index=False)

        _format_worksheet_layout(writer.book["detalle"])
        _format_worksheet_layout(writer.book["campos"])
        _format_worksheet_layout(writer.book["errores_ui"])
        _format_worksheet_layout(writer.book["ids_no_testeados"])
        _format_worksheet_layout(writer.book["resumen"])

        _apply_error_highlight(writer.book["detalle"])
        _apply_error_highlight(writer.book["errores_ui"])

    return output_path