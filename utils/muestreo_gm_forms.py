"""
muestreo_gm_forms.py — Arma un excel chico con N URLs de formulario activas al azar por
país, tomadas de 'GM Forms - 2026.xlsx' (mapa completo de URLs del CRM, externo, nunca se
versiona). El excel resultante conserva el mismo shape (headers y hojas por país) que ya
lee el Chequeo Masivo existente (core/massive_check_runner.py::run_massive_check), así que
alcanza con apuntarlo desde ese botón para correr la detección + llenado reales contra una
muestra en vez de esperar a una corrida completa para notar que un form dejó de detectar
campos.
"""
import datetime
import os
import random
import re

import openpyxl

from utils.paths import BASE_DIR, RESULTS_DIR

GM_FORMS_EXCEL_FILENAME = "GM Forms - 2026.xlsx"

# Mismos 9 países que interface.ids_dinamicos_ui.MAPEO_PAISES — duplicado deliberado para
# no acoplar utils/ a interface/ (ver la misma decisión en utils/crm_excel_importer.py).
PAISES_SOPORTADOS = (
    "Argentina", "Bolivia", "Brasil", "Chile", "Colombia",
    "Ecuador", "Paraguay", "Peru", "Uruguay",
)

_ACENTOS = str.maketrans("ÁÉÍÓÚáéíóúÑñ", "AEIOUaeiouNn")


def _normalizar(texto):
    sin_acentos = str(texto or "").translate(_ACENTOS)
    return re.sub(r"\s+", " ", sin_acentos.strip()).upper()


class MuestreadorGmForms:
    """Lee 'GM Forms - 2026.xlsx' y produce un excel chico del mismo shape con filas al
    azar por país, filtrando las que están dadas de baja (ESTADO '❌ OFF').

    n_por_pais=None desactiva el muestreo: toma TODAS las filas activas de cada país,
    sin recorte ni azar (para una verificación completa en vez de una muestra)."""

    def __init__(self, ruta_excel=None, n_por_pais=3, seed=None):
        self.ruta_excel = ruta_excel or os.path.join(BASE_DIR, GM_FORMS_EXCEL_FILENAME)
        self.n_por_pais = n_por_pais
        self._random = random.Random(seed)

    def disponible(self):
        if os.path.exists(self.ruta_excel):
            return True, self.ruta_excel
        mensaje = (
            f"No se encontró el archivo '{os.path.basename(self.ruta_excel)}'. "
            f"Colocalo en la carpeta del proyecto (al lado de run.py, o del .exe "
            f"en la versión compilada): {self.ruta_excel}"
        )
        return False, mensaje

    def _es_hoja_pais(self, nombre_hoja):
        normalizado = _normalizar(nombre_hoja)
        return any(_normalizar(p) == normalizado for p in PAISES_SOPORTADOS)

    def _muestrear_hoja(self, ws):
        """Busca la fila de headers (la que tiene 'ESTADO', porque 'GM Forms - 2026.xlsx'
        trae una fila de título del país arriba del header real) y devuelve
        (header_completo, filas_muestreadas) filtrando ESTADO '❌ OFF' y filas sin
        ninguna URL. Esto último importa: se vieron filas basura reales (Uruguay/Bolivia
        en 'GM Forms - 2026.xlsx') con el estado corrido a la columna SEGMENTO y el resto
        vacío — no dicen "OFF" pero tampoco sirven para verificar nada sin URL."""
        header = None
        idx_estado = None
        idx_urls = []
        activas = []

        for fila in ws.iter_rows(values_only=True):
            if header is None:
                candidato_idx = next(
                    (i for i, v in enumerate(fila) if _normalizar(v) == "ESTADO"), None
                )
                if candidato_idx is not None:
                    header = list(fila)
                    idx_estado = candidato_idx
                    idx_urls = [
                        i for i, v in enumerate(fila)
                        if _normalizar(v) in ("URL LIVE", "URL SECURE")
                    ]
                continue

            if all(v is None for v in fila):
                continue
            estado = _normalizar(fila[idx_estado]) if idx_estado < len(fila) else ""
            if "OFF" in estado:
                continue
            if idx_urls and not any(fila[i] for i in idx_urls if i < len(fila)):
                continue
            activas.append(list(fila))

        if header is None:
            return [], []

        if self.n_por_pais is None:
            return header, activas

        cantidad = min(self.n_por_pais, len(activas))
        muestra = self._random.sample(activas, cantidad) if cantidad else []
        return header, muestra

    def generar_excel_muestra(self, salida=None):
        disponible, info = self.disponible()
        if not disponible:
            raise FileNotFoundError(info)

        libro = openpyxl.load_workbook(self.ruta_excel, data_only=True)
        try:
            libro_salida = openpyxl.Workbook()
            libro_salida.remove(libro_salida.active)
            for nombre_hoja in libro.sheetnames:
                if not self._es_hoja_pais(nombre_hoja):
                    continue
                header, muestra = self._muestrear_hoja(libro[nombre_hoja])
                if not header:
                    continue
                hoja_salida = libro_salida.create_sheet(nombre_hoja)
                hoja_salida.append(header)
                for fila in muestra:
                    hoja_salida.append(fila)
        finally:
            libro.close()

        if salida is None:
            os.makedirs(RESULTS_DIR, exist_ok=True)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            salida = os.path.join(RESULTS_DIR, f"muestra_gm_forms_{timestamp}.xlsx")
        libro_salida.save(salida)
        return salida
