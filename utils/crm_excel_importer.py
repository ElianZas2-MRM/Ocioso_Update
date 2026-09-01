"""
crm_excel_importer.py — Cruza el excel externo de validaciones del CRM contra
field_validation_rules_<pais>.json y detecta campos que el excel documenta pero que el
JSON todavía no tiene mapeados (el gap que hoy se pierde por detección DOM incompleta).

La columna de validaciones del excel es prosa libre, no regex: este módulo NUNCA
autogenera regex ni pisa una entrada ya existente en el JSON, aunque el excel diga que
un campo es obligatorio y el JSON no tenga ninguna regla real — eso queda marcado como
"incompleto" para revisión manual, no se escribe solo.
"""
import datetime
import json
import os
import re

import openpyxl

from utils.paths import BASE_DIR, JSON_DIR

CRM_EXCEL_FILENAME = "CRM - GMSA - Validaciones Formularios (1).xlsx"

# Duplicado deliberado de los países de interface.ids_dinamicos_ui.MAPEO_PAISES: importar
# desde interface/ crearía un ciclo (field_validation_ui importa este módulo). Si se agrega
# un país nuevo a la app, hay que sumarlo acá también.
PAISES_SOPORTADOS = (
    "Argentina", "Bolivia", "Brasil", "Chile", "Colombia",
    "Ecuador", "Paraguay", "Peru", "Uruguay",
)

_OBLIGATORIO_TRUE = {"TRUE", "VERDADERO", "SI", "SÍ", "X"}
_OBLIGATORIO_FALSE = {"FALSE", "FALSO", "NO"}

_ACENTOS = str.maketrans("ÁÉÍÓÚáéíóúÑñ", "AEIOUaeiouNn")


def _normalizar_nombre_campo(texto):
    """Upper + strip + colapsar espacios — SIN tocar acentos: las keys de
    field_validation_rules_<pais>.json['fields'] los conservan, así que el cruce con la
    columna CAMPO del excel tiene que hacerse tal cual."""
    return re.sub(r"\s+", " ", str(texto or "").strip()).upper()


def _normalizar_header(texto):
    """Para MATCHEAR nombres de columna (no nombres de campo): acá sí conviene ignorar
    acentos, porque el mismo header aparece como 'ÚLTIMO AJUSTE' o 'ULTIMA ACTUALIZACION'
    según la hoja."""
    sin_acentos = str(texto or "").translate(_ACENTOS)
    return re.sub(r"\s+", " ", sin_acentos.strip()).upper()


def _parsear_ultimo_ajuste(valor):
    """openpyxl devuelve datetime.datetime nativo para celdas con formato de fecha (varias
    hojas del CRM real: Chile, Colombia, Ecuador, Paraguay, Peru, Uruguay, Bolivia) — hay
    que convertirlo a string acá, en el borde de parseo, para que nada río abajo
    (comparar/aplicar_merge/construir_reporte) tenga que preocuparse por tipos no-JSON."""
    if isinstance(valor, (datetime.datetime, datetime.date)):
        return valor.date().isoformat() if isinstance(valor, datetime.datetime) else valor.isoformat()
    return str(valor).strip() if valor else None


def _parsear_obligatorio(valor):
    if isinstance(valor, bool):
        return valor
    texto = _normalizar_header(valor)
    if texto in _OBLIGATORIO_TRUE:
        return True
    if texto in _OBLIGATORIO_FALSE:
        return False
    return None


class CrmValidacionesImporter:
    """Lee el excel del CRM y lo compara contra las reglas de validación ya configuradas.

    No escribe nada a disco: el caller (la UI de Validación de Campos) es el dueño de
    persistir field_validation_rules_<pais>.json vía sus propias funciones de carga y
    guardado, así no hay dos puntos distintos escribiendo el mismo archivo.
    """

    NON_COUNTRY_SHEETS = {"GMSA | Reportes", "Aclaraciones | Campo Celular"}

    def __init__(self, ruta_excel=None):
        self.ruta_excel = ruta_excel or os.path.join(BASE_DIR, CRM_EXCEL_FILENAME)
        self.hojas_no_reconocidas = []
        self._campos_por_pais = None

    def disponible(self):
        if os.path.exists(self.ruta_excel):
            return True, self.ruta_excel
        mensaje = (
            f"No se encontró el archivo '{os.path.basename(self.ruta_excel)}'. "
            f"Colocalo en la carpeta del proyecto (al lado de run.py, o del .exe "
            f"en la versión compilada): {self.ruta_excel}"
        )
        return False, mensaje

    def parsear(self):
        if self._campos_por_pais is not None:
            return self._campos_por_pais

        disponible, info = self.disponible()
        if not disponible:
            raise FileNotFoundError(info)

        libro = openpyxl.load_workbook(self.ruta_excel, read_only=True, data_only=True)
        try:
            paises_normalizados = {p.upper(): p for p in PAISES_SOPORTADOS}
            campos_por_pais = {}
            self.hojas_no_reconocidas = []

            for nombre_hoja in libro.sheetnames:
                if nombre_hoja in self.NON_COUNTRY_SHEETS:
                    continue
                pais = paises_normalizados.get(nombre_hoja.strip().upper())
                if not pais:
                    self.hojas_no_reconocidas.append(nombre_hoja)
                    continue
                campos_por_pais[pais] = self._parsear_hoja(libro[nombre_hoja])

            self._campos_por_pais = campos_por_pais
            return campos_por_pais
        finally:
            libro.close()

    def _parsear_hoja(self, ws):
        columnas = None
        campos = []
        for fila in ws.iter_rows(values_only=True):
            if columnas is None:
                candidata = self._detectar_columnas(fila)
                if candidata.get("campo") is not None:
                    columnas = candidata
                continue
            nombre_campo = fila[columnas["campo"]] if columnas["campo"] < len(fila) else None
            nombre_campo = _normalizar_nombre_campo(nombre_campo)
            if not nombre_campo:
                continue

            def _valor(idx):
                return fila[idx] if idx is not None and idx < len(fila) else None

            mensajes_error = []
            for idx in columnas.get("mensajes_error", []):
                texto = _valor(idx)
                if texto and str(texto).strip():
                    mensajes_error.append(str(texto).strip())

            campos.append({
                "nombre_campo": nombre_campo,
                "tipo": str(_valor(columnas.get("tipo")) or "").strip(),
                "obligatorio": _parsear_obligatorio(_valor(columnas.get("obligatorio"))),
                "descripcion": str(_valor(columnas.get("descripcion")) or "").strip(),
                "mensajes_error": mensajes_error,
                "ultimo_ajuste": _parsear_ultimo_ajuste(_valor(columnas.get("ultimo_ajuste"))),
            })
        return campos

    def _detectar_columnas(self, header_row):
        """None si la fila no parece un header (no tiene 'CAMPO') — así _parsear_hoja
        sigue buscando en las filas siguientes en vez de asumir que la primera fila
        siempre es el header."""
        columnas = {"campo": None, "tipo": None, "obligatorio": None,
                    "descripcion": None, "mensajes_error": [], "ultimo_ajuste": None}
        for idx, valor in enumerate(header_row):
            header = _normalizar_header(valor)
            if not header:
                continue
            if header == "CAMPO":
                columnas["campo"] = idx
            elif header == "TIPO":
                columnas["tipo"] = idx
            elif header == "OBLIGATORIO":
                columnas["obligatorio"] = idx
            elif header in ("DESCRIPCION", "VALIDACIONES"):
                columnas["descripcion"] = idx
            elif header.startswith("MENSAJE DE ERROR"):
                columnas["mensajes_error"].append(idx)
            elif header.startswith("ULTIMO AJUSTE") or header.startswith("ULTIMA ACTUALIZACION"):
                columnas["ultimo_ajuste"] = idx
        if columnas["campo"] is None:
            return columnas
        return columnas

    def comparar(self, pais, reglas_json):
        campos_por_pais = self.parsear()
        campos_excel = campos_por_pais.get(pais, [])
        fields = (reglas_json or {}).get("fields") or {}

        faltantes, incompletos, ok = [], [], []
        for campo in campos_excel:
            entry = fields.get(campo["nombre_campo"])
            if entry is None:
                faltantes.append(campo)
                continue
            tiene_regex = bool(str(entry.get("regex_full") or "").strip())
            if tiene_regex:
                ok.append({**campo, "tiene_regex": True})
            elif campo.get("obligatorio"):
                incompletos.append({**campo, "tiene_regex": False})
            else:
                ok.append({**campo, "tiene_regex": False})

        return {"faltantes": faltantes, "incompletos": incompletos, "ok": ok}

    def aplicar_merge(self, pais, comparacion, reglas_json):
        import copy
        actualizado = copy.deepcopy(reglas_json or {"fields": {}})
        fields = actualizado.setdefault("fields", {})

        agregados = 0
        fecha_importacion = datetime.datetime.now().isoformat()
        for campo in comparacion.get("faltantes", []):
            nombre = campo["nombre_campo"]
            if nombre in fields:
                # Ya se agregó en una corrida previa de "todos los países" — no duplicar.
                continue
            fields[nombre] = {
                "descripcion": campo.get("descripcion", ""),
                "campo": "",
                "element_id": "",
                "regex_full": "",
                "regex_char": "",
                "test_text": "",
                "dropdown": False,
                "dropdown_error_message": "",
                "dependencies": [],
                "paises": [pais],
                "teclado_mobile": False,
                "rules": {},
                "error_messages": {},
                "error_message_patterns": [],
                "error_config": {},
                "error_priority": [],
                "obligatorio_excel": bool(campo.get("obligatorio")),
                "mensajes_error_excel": list(campo.get("mensajes_error") or []),
                "pendiente_regex": True,
                "origen": "crm_excel_import",
                "fecha_importacion": fecha_importacion,
            }
            agregados += 1

        return actualizado, agregados


def construir_reporte(pais, hoja_excel, comparacion):
    """Reporte de auditoría de una corrida — de solo lectura, no se mezcla con el flujo
    de asignación de valores de la pestaña Campos Detectados."""
    return {
        "pais": pais,
        "fecha_importacion": datetime.datetime.now().isoformat(),
        "hoja_excel": hoja_excel,
        "resumen": {
            "faltantes": len(comparacion.get("faltantes", [])),
            "incompletos": len(comparacion.get("incompletos", [])),
            "ok": len(comparacion.get("ok", [])),
        },
        "faltantes": comparacion.get("faltantes", []),
        "incompletos": comparacion.get("incompletos", []),
        "ok": comparacion.get("ok", []),
    }


def guardar_reporte(pais, reporte):
    """Escritura atómica (tmp + os.replace), mismo patrón que
    utils/fixed_field_mapping_store.py, para no dejar el archivo a medio escribir si el
    proceso se corta a mitad de un guardado."""
    os.makedirs(JSON_DIR, exist_ok=True)
    nombre_archivo = f"crm_import_report_{_normalizar_nombre_campo(pais).lower()}.json"
    ruta_final = os.path.join(JSON_DIR, nombre_archivo)
    ruta_tmp = ruta_final + ".tmp"
    with open(ruta_tmp, "w", encoding="utf-8") as fh:
        json.dump(reporte, fh, indent=2, ensure_ascii=False)
    os.replace(ruta_tmp, ruta_final)
    return ruta_final
