"""
excel_layout.py
===============
Reordena las columnas del Excel de resultados para dejar PRIMERO las columnas de
resultado y al final los datos de entrada (los que el usuario cargó).

Por qué como post-proceso y no al construir la hoja: los runners leen los datos del
lead de la misma hoja en la que escriben (URL en la columna A, Formulario en la B,
y los campos desde `data_start_index`). Insertar columnas al principio durante la
corrida desalinearía esa lectura. Así que la corrida trabaja con el layout de
siempre y recién al terminar se reordena el archivo ya guardado.

Conserva valores, colores, fuente, formato de número, alineación y ancho de columna.
"""

import os
from copy import copy
from typing import List

# Orden en que se quieren las columnas de resultado (las que existan en la hoja).
# El tracking por paso (PasoN::campo, Final::campo, Reintento::campo) se detecta
# aparte y va detrás de estas, antes de los datos de entrada.
ORDEN_RESULTADO = [
    "Resultado",
    "Motivo",
    "Formulario Inserto",
    "Form coincide",
    "Datos vs Excel",
    "Estado URL landing",
    "Estado URL form",
    "Formulario Completado",
    "TY Page",
    "TYP con CTA",
    "LINK ISSUE TYP",
    "Form URL esperada",
    "Form URL encontrada",
    "Video LT",
    "Dashboard LT",
]

_PREFIJOS_TRACKING = ("Paso", "Final::", "Reintento", "ID::")


def _es_columna_tracking(nombre: str) -> bool:
    n = str(nombre or "")
    return "::" in n and n.startswith(_PREFIJOS_TRACKING)


def calcular_nuevo_orden(headers: List) -> List[int]:
    """Devuelve la lista de índices (1-based) de la hoja en el nuevo orden:
    resultado → tracking por paso → datos de entrada."""
    resultado, tracking, entrada = [], [], []
    usados = set()

    for nombre in ORDEN_RESULTADO:
        for idx, h in enumerate(headers, start=1):
            if idx in usados:
                continue
            if str(h or "").strip() == nombre:
                resultado.append(idx)
                usados.add(idx)
                break

    for idx, h in enumerate(headers, start=1):
        if idx in usados:
            continue
        if _es_columna_tracking(h):
            tracking.append(idx)
        else:
            entrada.append(idx)
        usados.add(idx)

    return resultado + tracking + entrada


def reordenar_hoja(ws) -> bool:
    """Reordena las columnas de la hoja in-place. True si cambió algo."""
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col < 2 or max_row < 1:
        return False

    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    nuevo_orden = calcular_nuevo_orden(headers)
    if nuevo_orden == list(range(1, max_col + 1)):
        return False  # ya está en el orden deseado

    # Snapshot de todas las celdas (valor + formato) antes de sobrescribir
    snapshot = []
    for c in range(1, max_col + 1):
        col = []
        for r in range(1, max_row + 1):
            cell = ws.cell(row=r, column=c)
            col.append({
                "value": cell.value,
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
            })
        snapshot.append(col)

    from openpyxl.utils import get_column_letter
    anchos = {}
    for c in range(1, max_col + 1):
        letra = get_column_letter(c)
        dim = ws.column_dimensions.get(letra)
        if dim is not None and dim.width:
            anchos[c] = dim.width

    for destino, origen in enumerate(nuevo_orden, start=1):
        col = snapshot[origen - 1]
        for r, datos in enumerate(col, start=1):
            cell = ws.cell(row=r, column=destino)
            cell.value = datos["value"]
            cell.fill = datos["fill"]
            cell.font = datos["font"]
            cell.number_format = datos["number_format"]
            cell.alignment = datos["alignment"]
            cell.border = datos["border"]
        ancho = anchos.get(origen)
        if ancho:
            ws.column_dimensions[get_column_letter(destino)].width = ancho

    return True


def reordenar_archivo(path: str, log=print) -> bool:
    """Abre el xlsx de resultados, deja las columnas de resultado primero y lo guarda.

    Silencioso ante errores: es cosmético, nunca debe hacer perder los resultados.
    """
    try:
        if not path or not os.path.exists(path):
            return False
        from openpyxl import load_workbook
        wb = load_workbook(path)
        cambio = False
        for ws in wb.worksheets:
            if reordenar_hoja(ws):
                cambio = True
        if cambio:
            wb.save(path)
            log("Excel de resultados reordenado: columnas de resultado primero")
        return cambio
    except Exception as e:
        try:
            log(f"Aviso: no se pudo reordenar las columnas del Excel: {e}")
        except Exception:
            pass
        return False
