"""
lt_android_runner.py
--------------------
Runner LambdaTest para Android (Samsung Galaxy S24 + Chrome).

Reutiliza la lógica de llenado de formularios de lambdatest_mac/lt_runner.py
(fill_form_fields, _run_single_lead, helpers de cookies/checkboxes, etc.).
Solo define lo que es Android-específico: driver con capacidades Android y
directorio de resultados propio.
"""

import os
import sys
import glob
import time
import traceback
import threading
from dataclasses import dataclass, field
from typing import Callable, Dict, List, Optional

from openpyxl import load_workbook
from selenium.webdriver import ChromeOptions
from selenium.webdriver.remote.webdriver import WebDriver as Remote
from selenium.webdriver.remote.client_config import ClientConfig

_ANDROID_DEVICES = {
    "Galaxy S23":  "13",
    "Galaxy S22":  "12",
    "Galaxy S21":  "11",
    "Galaxy S24":  "14",
    "Galaxy A54":  "13",
}
_DEFAULT_ANDROID_DEVICE = "Galaxy S23"

# ── Rutas base ────────────────────────────────────────────────────────────────
_THIS_DIR   = os.path.dirname(os.path.abspath(__file__))   # lambdatest_android/
if getattr(sys, 'frozen', False):
    _OSOCIO_DIR = os.path.dirname(sys.executable)
else:
    _OSOCIO_DIR = os.path.dirname(_THIS_DIR)                   # raíz del proyecto
_MAC_DIR    = os.path.join(os.path.dirname(_THIS_DIR), "lambdatest_mac")
_DATA_DIR   = os.path.join(_OSOCIO_DIR, "data")
_ANDROID_RESULTADOS_DIR = os.path.join(_OSOCIO_DIR, "resultados_lambdatest_android")

# Exponer lambdatest_mac en el path para importar desde lt_runner y lt_excel_reader
if _MAC_DIR not in sys.path:
    sys.path.insert(0, _MAC_DIR)
_src_root = os.path.dirname(_THIS_DIR)
if _src_root not in sys.path:
    sys.path.insert(0, _src_root)

# Importar toda la lógica reutilizable desde lambdatest_mac
from lt_runner import (  # type: ignore[import]
    load_credentials,
    mark_lt_status,
    LT_HUB,
    _fetch_lt_video_url,
    _load_field_dependencies,
    _load_ids_dinamicos,
    _get_field_mapping_for_pais,
    fill_form_fields,
    _run_single_lead,
    _write_row_result,
)
from lt_excel_reader import read_osocio_excel  # type: ignore[import]
from lt_screenshots import LTScreenshotManager  # type: ignore[import]


# ══════════════════════════════════════════════════════════════════════════════
# OPCIONES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class LTAndroidRunOptions:
    pais: str
    excel_path: str
    credentials_file: str = ""
    build_name: str = ""
    device_name: str = _DEFAULT_ANDROID_DEVICE
    with_screenshots: bool = False
    brasil_docs: dict = None


# ══════════════════════════════════════════════════════════════════════════════
# DRIVER LAMBDATEST ANDROID
# ══════════════════════════════════════════════════════════════════════════════

def create_lt_android_driver(username: str, access_key: str,
                              build_name: str = "Osocio LT Android",
                              test_name: str = "Form Run Android",
                              device_name: str = _DEFAULT_ANDROID_DEVICE) -> Remote:
    """
    Driver LambdaTest Web Automation para Android real device (Chrome).
    Nombre del device sin prefijo "Samsung" — formato del catálogo de LambdaTest.
    """
    _cfg = ClientConfig(remote_server_addr=LT_HUB, username=username, password=access_key)
    platform_version = _ANDROID_DEVICES.get(device_name, "13")
    options = ChromeOptions()
    options.set_capability("LT:Options", {
        "username":        username,
        "accessKey":       access_key,
        "build":           build_name,
        "name":            test_name,
        "platformName":    "Android",
        "browserName":     "Chrome",
        "deviceName":      device_name,
        "platformVersion": platform_version,
        "realMobile":      True,
        "visual":          True,
        "video":           True,
        "console":         False,
        "network":         False,
        "w3c":             True,
    })
    driver = Remote(command_executor=LT_HUB, options=options, client_config=_cfg)
    driver.implicitly_wait(0)
    return driver


# ══════════════════════════════════════════════════════════════════════════════
# RESULTADOS EXCEL (directorio propio para Android)
# ══════════════════════════════════════════════════════════════════════════════

def _get_android_run_number(pais: str, prefix: str = "resultados_") -> int:
    pattern = os.path.join(_ANDROID_RESULTADOS_DIR, f"{prefix}{pais}_Android*.xlsx")
    matches = glob.glob(pattern)
    max_n = 0
    for m in matches:
        base = os.path.basename(m).replace(f"{prefix}{pais}_Android", "").replace(".xlsx", "")
        if base.isdigit():
            max_n = max(max_n, int(base))
    return max_n + 1


def _setup_android_results_excel(pais: str, source_excel_path: str, build_name: str = "") -> tuple:
    os.makedirs(_ANDROID_RESULTADOS_DIR, exist_ok=True)
    prefix = "Automatizacion_" if "Automatización" in build_name else "resultados_"
    run_number   = _get_android_run_number(pais, prefix=prefix)
    results_path = os.path.join(_ANDROID_RESULTADOS_DIR, f"{prefix}{pais}_Android{run_number}.xlsx")

    wb = load_workbook(source_excel_path)
    ws = wb.active

    required_cols = [
        "Resultado", "Formulario Inserto", "Formulario Completado",
        "TY Page", "Form URL esperada", "Form URL encontrada", "Form coincide",
        "Video LT", "Dashboard LT",
    ]
    headers = [cell.value for cell in ws[1] if cell.value]
    for col_name in required_cols:
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers)).value = col_name

    wb.save(results_path)

    headers = [cell.value for cell in ws[1] if cell.value]
    col_idx = {name: headers.index(name) + 1
               for name in required_cols if name in headers}
    return wb, ws, run_number, results_path, col_idx


# ══════════════════════════════════════════════════════════════════════════════
# BATCH RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def run_lt_android_batch(opts: LTAndroidRunOptions, log: Callable = print,
                          stop_event: Optional[threading.Event] = None) -> Dict:
    """
    Ejecuta todos los leads del Excel en LambdaTest Android (Samsung + Chrome).
    Misma lógica que run_lt_batch de lambdatest_mac pero con driver Android
    y resultados en resultados_lambdatest_android/.
    """
    summary = {
        "pais": opts.pais, "results_excel": None,
        "session_id": None, "total": 0, "ok": 0, "failed": 0,
        "error": None, "video_url": "",
    }
    driver     = None
    session_id = None
    username   = ""
    access_key = ""
    video_url  = ""
    results_path = None
    wb = ws = col_idx = None

    try:
        username, access_key = load_credentials(opts.credentials_file)
        log(f"✓ Credenciales OK: {username}")

        leads = read_osocio_excel(opts.excel_path)
        if not leads:
            raise ValueError(f"No hay filas en: {opts.excel_path}")
        log(f"  {len(leads)} leads encontrados.")
        summary["total"] = len(leads)

        build_name = opts.build_name or f"Osocio LT Android — {opts.pais} ({opts.device_name})"

        field_mapping = _get_field_mapping_for_pais(opts.pais)
        log(f"  {len(field_mapping)} campos en el mapping de Osocio.")

        dependencies  = _load_field_dependencies()
        ids_dinamicos = _load_ids_dinamicos(opts.pais)
        log(f"  IDs dinámicos: {list(ids_dinamicos.keys())}")

        wb, ws, run_number, results_path, col_idx = _setup_android_results_excel(
            opts.pais, opts.excel_path
        )
        summary["results_excel"] = results_path
        log(f"Resultados: {results_path} (Run #{run_number})")

        log(f"\nConectando a LambdaTest Android ({build_name})...")
        log(f"  Dispositivo: {opts.device_name}")
        driver = create_lt_android_driver(
            username, access_key,
            build_name=build_name,
            test_name=f"{opts.pais} Android - {len(leads)} leads",
            device_name=opts.device_name,
        )
        session_id = driver.session_id
        summary["session_id"] = session_id
        log(f"✓ Session ID: {session_id}")
        log(f"  Dashboard: https://automation.lambdatest.com/logs/?testID={session_id}")

        screenshots_dir = os.path.join(_ANDROID_RESULTADOS_DIR,
                                       f"screenshots_{opts.pais}{run_number}")
        _all_screenshot_managers = []

        for i, lead in enumerate(leads, start=1):
            if stop_event and stop_event.is_set():
                log("  ⛔ Ejecución detenida por el usuario.")
                break

            row_num = lead.index
            log(f"\n{'='*60}")
            log(f"  LEAD {i}/{len(leads)} — fila #{row_num}")
            log(f"  URL: {lead.public_url}")
            log(f"{'='*60}")

            _br_docs    = opts.brasil_docs or {}
            _br_doc_type = "cpf"
            if opts.pais.lower() in ("brasil", "brazil"):
                _form_idx = i
                if _form_idx in _br_docs.get("cnpj_rows", []):
                    _br_doc_type = "cnpj"
                elif _form_idx in _br_docs.get("cep_rows", []):
                    _br_doc_type = "cep"
                elif _form_idx in _br_docs.get("cpf_rows", []):
                    _br_doc_type = "cpf"

            sm = None
            if opts.with_screenshots:
                sm = LTScreenshotManager(
                    driver=driver,
                    screenshots_dir=screenshots_dir,
                    lead_num=i,
                    session_id=session_id,
                    username=username,
                    access_key=access_key,
                    iframe_src=lead.secure_url or "",
                    log=log,
                    landing_url=lead.public_url or "",
                )
                _all_screenshot_managers.append(sm)

            result = _run_single_lead(
                driver, opts.pais, lead,
                field_mapping, dependencies, ids_dinamicos,
                is_mobile=False,   # Android Chrome soporta switch_to.frame() normal
                is_android=True,   # click antes de fill + scroll al TY
                brasil_doc_type=_br_doc_type,
                screenshot_manager=sm,
                log=log,
            )

            _write_row_result(
                ws, row_num, col_idx,
                iframe_found=result.get("iframe_found", False),
                submitted=result.get("submitted", False),
                result_text=result["result_text"],
                tracked=result.get("tracked", {}),
                ty_confirmed=result.get("ty_confirmed", False),
                iframe_url_found=result.get("iframe_url_found", ""),
                iframe_url_expected=lead.secure_url or "",
            )
            try:
                wb.save(results_path)
            except Exception as e:
                log(f"  ⚠ Error guardando Excel: {e}")

            log(f"  → {result['result_text']}")
            if "✓" in result["result_text"]:
                summary["ok"] += 1
            else:
                summary["failed"] += 1

    except Exception as e:
        summary["error"] = str(e)
        log(f"✗ Error general: {e}")
        log(traceback.format_exc())
    finally:
        if driver:
            try:
                time.sleep(2)
                mark_lt_status(driver, passed=(summary["failed"] == 0 and summary["error"] is None))
                driver.quit()
                log("\n✓ Driver cerrado.")

                if _all_screenshot_managers:
                    for _sm in _all_screenshot_managers:
                        _sm.descargar_con_frame_mac()

                if session_id and username and access_key:
                    video_url = _fetch_lt_video_url(session_id, username, access_key)
                    dashboard_url = f"https://automation.lambdatest.com/test?testID={session_id}"
                    summary["video_url"] = video_url
                    if video_url:
                        log(f"  Video LT: {video_url}")
                    log(f"  Dashboard LT: {dashboard_url}")
                    if results_path and ws and wb:
                        try:
                            for _r in range(2, ws.max_row + 1):
                                if video_url and "Video LT" in col_idx:
                                    ws.cell(row=_r, column=col_idx["Video LT"]).value = video_url
                                if "Dashboard LT" in col_idx:
                                    ws.cell(row=_r, column=col_idx["Dashboard LT"]).value = dashboard_url
                            wb.save(results_path)
                        except Exception:
                            pass
            except Exception:
                pass

    log(f"\n{'='*60}")
    log(f"RESUMEN: {summary['ok']}/{summary['total']} OK, {summary['failed']} con issues.")
    if summary.get("results_excel"):
        log(f"Excel: {summary['results_excel']}")
    if video_url:
        log(f"Video: {video_url}")
    log(f"{'='*60}")

    return summary
