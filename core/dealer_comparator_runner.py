"""
dealer_comparator_runner.py — Lógica de negocio de la pestaña "Comparador Dealers".

Lee un Excel de dealers esperados (fila de encabezado y columnas configurables porque
el Excel que mandan varía de país a país), navega región→ciudad→dealer en un formulario
real via Selenium (reusando BrowserManager), compara contra el Excel filtrado y arma un
reporte PASS/FAIL/EXTRA/DUPLICADO/NOTA con export a Excel (openpyxl) y capturas opcionales
(ScreenshotManager, con banner de URL) — una carpeta propia por formulario, sin ZIP.
"""
import os
import re
import time
import unicodedata
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException

from core.browser_manager import BrowserManager
from core.country_configs import COUNTRY_CONFIGS
from core.screenshot_manager import ScreenshotManager

try:
    from utils.paths import RESULTS_DIR
except Exception:
    RESULTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "resultados")


DEFAULT_SELECT_IDS = {"region": "region", "city": "city", "dealer": "dealer"}

# El <select> de modelos aparece con el id en plural o en singular segun el form.
# Se prueba el id pedido primero y luego el resto como fallback.
MODEL_FIELD_ID_ALIASES = ("models", "model")

_PASS_FILL = PatternFill(fill_type="solid", fgColor="00C6EFCE")
_FAIL_FILL = PatternFill(fill_type="solid", fgColor="00FFC7CE")
_EXTRA_FILL = PatternFill(fill_type="solid", fgColor="00FFEB9C")
_MISSING_FILL = PatternFill(fill_type="solid", fgColor="00D9D2E9")
_DUPLICATE_FILL = PatternFill(fill_type="solid", fgColor="00FFCC99")
_NOTA_FILL = PatternFill(fill_type="solid", fgColor="00BDD7EE")
_OCULTO_FILL = PatternFill(fill_type="solid", fgColor="00FFB84D")  # naranja: relación con filas OCULTAS del Excel
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="007D4E9F")
_HEADER_FONT = Font(color="00FFFFFF", bold=True)


class StopRequested(Exception):
    """Señal interna para cortar la corrida cuando el usuario aprieta Detener."""


# ──────────────────────────────────────────────────────────────────────────────
# Config por país
# ──────────────────────────────────────────────────────────────────────────────
def get_country_level_defaults(country_name):
    """A partir de core/country_configs.py, infiere qué niveles (región/ciudad/dealer)
    tiene mapeados el país, para usar como default editable en la config persistida."""
    config = COUNTRY_CONFIGS.get(country_name, {})
    ids = {fm.get("id") for fm in config.get("field_mapping", []) if isinstance(fm.get("id"), str)}
    return {
        "has_region": "region" in ids,
        "has_city": "city" in ids,
        "has_dealer": "dealer" in ids,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Normalización de texto (acentos / mayúsculas), igual criterio que el resto de la app
# ──────────────────────────────────────────────────────────────────────────────
def normalize_text(value):
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", text)


_PLACEHOLDER_KEYWORDS = (
    "SELECCION", "SELECIONE", "SELECIONAR", "ESCOLHA", "ESCOLHER",
    "ELIJA", "ELEGIR", "OPCION", "OPCIONES", "SELECT", "CHOOSE", "REQUIRED",
)


def _is_placeholder(text):
    normalized = normalize_text(text)
    if not normalized:
        return True
    return any(k in normalized for k in _PLACEHOLDER_KEYWORDS)


# ──────────────────────────────────────────────────────────────────────────────
# Matcheo tolerante de nombres de dealer: apóstrofes/comillas (D'hola vs Dhola o
# "D AMICO"), guiones y guiones largos, paréntesis ("- RPM" vs "(RPM)"), sufijo
# parentético numérico ("(1000km)") y diferencias de espaciado.
# ──────────────────────────────────────────────────────────────────────────────
_TRAILING_PAREN_RE = re.compile(r"\(\s*[^()]*\d[^()]*\)\s*$")  # "(1000KM)" al final
_PUNCT_STRIP_RE = re.compile(r"[\'’‘´`\.\,\-–—()\"]+")


def _canonical_dealer_name(text):
    """Forma canónica para el matcheo tolerante: normaliza mayúsculas/acentos, saca el
    sufijo parentético numérico final (ej. '(1000km)'), toda la puntuación (apóstrofes,
    comillas, puntos, comas, guiones, guiones largos, paréntesis) y TODOS los espacios.
    Dos nombres que coinciden solo tras esta limpieza son 'el mismo dealer con
    diferencias menores' (ej. D'AMICO vs D AMICO, '- RPM' vs '(RPM)') — cualquier
    diferencia de CONTENIDO (letras/números de más o de menos) NO matchea."""
    t = normalize_text(text)
    t = _TRAILING_PAREN_RE.sub("", t).strip()
    t = _PUNCT_STRIP_RE.sub("", t)
    return re.sub(r"\s+", "", t)


def match_dealer_names(excel_name, form_name):
    """Compara el nombre del Excel contra el texto real del form.
    Devuelve (match: bool, disclaimer: str|None):
      - Coincidencia exacta (normalize_text) → (True, None)
      - Coincidencia solo tras sacar apóstrofes/comillas o un sufijo parentético numérico
        final → (True, "disclaimer explicando la diferencia menor")
      - Cualquier otra diferencia (más o menos contenido) → (False, None)
    """
    excel_norm = normalize_text(excel_name)
    form_norm = normalize_text(form_name)
    if not excel_norm or not form_norm:
        return False, None
    if excel_norm == form_norm:
        return True, None
    excel_canon = _canonical_dealer_name(excel_name)
    form_canon = _canonical_dealer_name(form_name)
    if excel_canon and excel_canon == form_canon:
        disclaimer = (
            f"Nombre difiere en detalles menores (mayúsculas/caracteres especiales/sufijo) "
            f"— Excel: '{str(excel_name).strip()}' | Form: '{str(form_name).strip()}'"
        )
        return True, disclaimer
    return False, None


# ──────────────────────────────────────────────────────────────────────────────
# Lectura de Excel con fila de encabezado configurable
# ──────────────────────────────────────────────────────────────────────────────
def read_excel_rows(file_path, header_row=1, sheet_name=None):
    """Lee un Excel con la fila de encabezado en `header_row` (1-indexado, como lo ve el usuario en Excel).

    Devuelve (headers, rows). Cada row es un dict con:
      - clave = texto del encabezado (si existe y no está vacío)
      - clave = "_c0", "_c1", ... (índice de columna, para resolver por letra A/B/K/...)
      - "__row__" = número de fila real en el Excel (1-indexado)
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = wb[sheet_name] if sheet_name else wb.worksheets[0]
        all_rows = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()

    if header_row < 1 or header_row > len(all_rows):
        raise ValueError(f"La fila de encabezado ({header_row}) está fuera de rango del Excel.")

    header_cells = all_rows[header_row - 1]
    headers = [str(h).strip() if h is not None else "" for h in header_cells]

    rows = []
    for i in range(header_row, len(all_rows)):
        raw = all_rows[i]
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {"__row__": i + 1}
        for idx, header in enumerate(headers):
            val = raw[idx] if idx < len(raw) else None
            val = "" if val is None else str(val).strip()
            row[f"_c{idx}"] = val
            if header:
                row[header] = val
        rows.append(row)

    return headers, rows


def detect_hidden_rows(file_path, header_row=1, sheet_name=None):
    """Devuelve un dict {nº de fila (1-indexado): 'filtrada' | 'oculta'} con las filas de
    datos que NO se ven en el Excel. Distingue:
      - 'filtrada': escondida por un AutoFilter activo (con criterios) — el Excel viene
        pre-filtrado; la fila reaparece si sacás el filtro.
      - 'oculta': escondida a mano (fila oculta manualmente), independiente de cualquier filtro.
    En Excel ambas quedan hidden=True; se diferencian porque las filtradas caen dentro del
    rango del AutoFilter que tiene criterios aplicados. Solo informativo (no cambia qué se
    compara). read_only no expone row_dimensions, por eso se abre en modo normal."""
    result = {}
    try:
        from openpyxl.utils import range_boundaries
        wb = load_workbook(file_path, data_only=True, read_only=False)
        try:
            sheet = wb[sheet_name] if sheet_name else wb.worksheets[0]

            # ¿Hay un AutoFilter con criterios activos? (filas escondidas por filtro)
            af = getattr(sheet, "auto_filter", None)
            af_ref = getattr(af, "ref", None) if af else None
            af_activo = bool(af_ref) and bool(getattr(af, "filterColumn", None))
            fr_min = fr_max = None
            if af_activo:
                try:
                    _c1, fr_min, _c2, fr_max = range_boundaries(af_ref)
                except Exception:
                    af_activo = False

            for row_num, dim in sheet.row_dimensions.items():
                if row_num <= header_row or not getattr(dim, "hidden", False):
                    continue
                if af_activo and fr_min is not None and fr_min <= row_num <= fr_max:
                    result[row_num] = "filtrada"
                else:
                    result[row_num] = "oculta"
        finally:
            wb.close()
    except Exception:
        pass
    return result


def detect_hidden_columns(file_path, header_row=1, sheet_name=None):
    """Devuelve [{'column': 'K', 'header': 'VENTA 0KM'}, ...] de columnas OCULTAS en el
    Excel — igual criterio que detect_hidden_rows pero para columnas. Solo informativo:
    no cambia qué se compara, pero avisa porque una columna oculta puede pasar
    desapercibida para quien preparó el Excel."""
    hidden = []
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
        try:
            sheet = wb[sheet_name] if sheet_name else wb.worksheets[0]
            header_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            header_vals = header_cells[0] if header_cells else []
            for col_letter, dim in sheet.column_dimensions.items():
                if not getattr(dim, "hidden", False):
                    continue
                try:
                    idx = column_index_from_string(col_letter) - 1
                    header_name = str(header_vals[idx]).strip() if (0 <= idx < len(header_vals) and header_vals[idx]) else ""
                except Exception:
                    header_name = ""
                hidden.append({"column": col_letter, "header": header_name})
        finally:
            wb.close()
    except Exception:
        pass
    return sorted(hidden, key=lambda h: h["column"])


def resolve_column(headers, input_name):
    """Resuelve un nombre de columna ingresado por el usuario a la clave real del row dict.
    Soporta: letra de columna estilo Excel (A, B, K...), nombre exacto, o coincidencia parcial
    (sin distinguir mayúsculas/acentos), igual criterio que el bookmarklet original."""
    input_name = (input_name or "").strip()
    if not input_name:
        return None

    if re.fullmatch(r"[A-Za-z]{1,3}", input_name) and input_name not in headers:
        try:
            idx = column_index_from_string(input_name.upper()) - 1
            if idx >= 0:
                return f"_c{idx}"
        except Exception:
            pass

    if input_name in headers:
        return input_name

    normalized_input = normalize_text(input_name)
    for h in headers:
        if h and normalize_text(h) == normalized_input:
            return h
    for h in headers:
        if h and (normalized_input in normalize_text(h) or normalize_text(h) in normalized_input):
            return h
    return None


def filter_rows(rows, filter_key, filter_value, mode="include"):
    """Filtra filas por columna+valor. mode='include' deja solo las que matchean el valor,
    'exclude' saca las que matchean (para el caso de "chequear todo MENOS X")."""
    if not filter_key or filter_value in (None, ""):
        return list(rows)
    target = normalize_text(filter_value)

    def _matches(row):
        return normalize_text(row.get(filter_key, "")) == target

    if mode == "exclude":
        return [r for r in rows if not _matches(r)]
    return [r for r in rows if _matches(r)]


# ──────────────────────────────────────────────────────────────────────────────
# Navegación de selects dependientes (región → ciudad → dealer) en el form real
# ──────────────────────────────────────────────────────────────────────────────
def _get_select(driver, element_id, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        if (element.tag_name or "").lower() != "select":
            return None
        return element
    except Exception:
        return None


def _valid_options(select_element):
    try:
        return [o for o in select_element.find_elements(By.TAG_NAME, "option") if o.get_attribute("value")]
    except Exception:
        return []


def _find_option_by_text(select_element, text):
    """Busca la <option> cuyo texto matchea: exacto primero, y si no, solo tolerando
    diferencias menores (apóstrofes/comillas, sufijo numérico entre paréntesis) vía
    match_dealer_names — NO hace substring genérico: un nombre con contenido de más o de
    menos (ej. otro apellido agregado) NO debe considerarse el mismo dealer.
    Devuelve (option, disclaimer_o_None) o (None, None) si no hay match."""
    target = normalize_text(text)
    if not target:
        return None, None
    options = _valid_options(select_element)
    for opt in options:
        if normalize_text(opt.text) == target:
            return opt, None
    for opt in options:
        match, disclaimer = match_dealer_names(text, opt.text)
        if match:
            return opt, disclaimer
    return None, None


def _select_option(driver, select_element, option_element):
    Select(select_element).select_by_value(option_element.get_attribute("value"))
    driver.execute_script(
        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('input', {bubbles:true}));",
        select_element,
    )


def _select_by_text_robust(driver, select_id, target_text, wait_for_option=False, timeout=None):
    """
    Selecciona la <option> cuyo texto matchea target_text en el <select> id=select_id,
    de forma resistente a stale elements y a selects dependientes que se repueblan
    tarde (city tras region, dealer tras city). En cada intento re-busca el <select>
    y sus <option> FRESCOS — así un re-render del form (que invalida referencias
    viejas) no rompe la selección ni genera un "no encontrado" falso.

    wait_for_option=True: reintenta hasta `timeout` esperando a que la opción aparezca
    (el select dependiente todavía puede estar cargando las opciones de la nueva región).
    Devuelve (found: bool, option_text: str|None, disclaimer: str|None) — disclaimer no-None
    si matcheó tolerando una diferencia menor de nombre (ver match_dealer_names).
    """
    if timeout is None:
        # Selects dependientes (city/dealer) pueden poblar la opción buscada con retraso
        # (ej. la 3ra opción de dealer en algunos forms carga tarde). Presupuesto amplio:
        # el poll corta APENAS aparece, así que sólo se agota en dealers realmente ausentes.
        timeout = DEPENDENT_SELECT_TIMEOUT if wait_for_option else 0.4
    deadline = time.time() + timeout
    while True:
        try:
            el = _get_select(driver, select_id, timeout=0.5)
            if el is not None:
                opt, disclaimer = _find_option_by_text(el, target_text)
                if opt is not None:
                    text = opt.text.strip()
                    try:
                        _select_option(driver, el, opt)
                    except Exception as sel_err:
                        # La opción EXISTE (dealer presente en el form) pero está
                        # deshabilitada / no seleccionable: a los fines de la comparación
                        # (¿está el dealer en el form?) cuenta como encontrado.
                        if "disabled" in str(sel_err).lower():
                            return True, text, disclaimer
                        raise
                    return True, text, disclaimer
        except StaleElementReferenceException:
            pass
        if time.time() >= deadline:
            return False, None, None
        time.sleep(0.1)


def _read_valid_option_texts(driver, select_id, wait_nonempty=True, timeout=None):
    """Devuelve la lista de textos de <option> válidas (no placeholder) del select,
    leídas de forma stale-safe (re-busca el elemento y re-lee si algo se pone stale
    a mitad de la iteración — el form re-renderiza los selects).

    wait_nonempty=True: espera hasta `timeout` a que el select tenga opciones (el select
    dependiente se puebla async tras cambiar el padre); corta apenas aparecen."""
    if timeout is None:
        timeout = DEPENDENT_SELECT_TIMEOUT if wait_nonempty else 0.4
    deadline = time.time() + timeout
    while True:
        try:
            el = _get_select(driver, select_id, timeout=0.5)
            if el is not None:
                texts = []
                for o in el.find_elements(By.TAG_NAME, "option"):
                    if not o.get_attribute("value"):
                        continue
                    t = (o.text or "").strip()
                    if t and not _is_placeholder(t):
                        texts.append(t)
                if texts or not wait_nonempty:
                    return texts
        except StaleElementReferenceException:
            pass
        if time.time() >= deadline:
            return []
        time.sleep(0.1)


def _selected_dealer_option_attr(driver, dealer_id, attr):
    """Lee un atributo (ej. data-bac) de la <option> de dealer actualmente seleccionada,
    re-buscándola fresca (stale-safe)."""
    for _ in range(3):
        try:
            el = _get_select(driver, dealer_id, timeout=0.5)
            if el is None:
                return ""
            return (Select(el).first_selected_option.get_attribute(attr) or "").strip()
        except StaleElementReferenceException:
            time.sleep(0.1)
    return ""


FAST_TIMEOUT = 0.6  # primer intento: rápido, alcanza en la gran mayoría de los forms
MAX_TIMEOUT = 1.2    # segundo intento (solo si el primero no alcanzó): tope máximo
DEPENDENT_SELECT_TIMEOUT = 5.0  # espera máxima a que aparezca la opción buscada en un
                                 # select dependiente (city/dealer); corta apenas la halla
                      # (ej. el select de ciudad se puebla más lento en departamentos con
                      # muchas opciones, como Montevideo)


def _wait_options_loaded_once(driver, element_id, timeout, poll=0.1):
    deadline = time.time() + timeout
    element = None
    while time.time() < deadline:
        element = _get_select(driver, element_id, timeout=0.3)
        if element is not None and any(not _is_placeholder(o.text) for o in _valid_options(element)):
            return element
        time.sleep(poll)
    return element


def _wait_options_loaded(driver, element_id):
    """Espera a que un <select> hijo tenga opciones válidas (no placeholder) tras disparar el padre.
    Dos intentos: rápido (FAST_TIMEOUT) y, solo si no alcanzó, uno más largo hasta MAX_TIMEOUT total."""
    element = _wait_options_loaded_once(driver, element_id, FAST_TIMEOUT)
    if element is not None and any(not _is_placeholder(o.text) for o in _valid_options(element)):
        return element
    remaining = max(0.0, MAX_TIMEOUT - FAST_TIMEOUT)
    if remaining <= 0:
        return element
    return _wait_options_loaded_once(driver, element_id, remaining)


def _read_form_field_value(driver, field_id):
    """Lee el valor actual de un campo del form por id: texto de la opción elegida si es
    <select>, o el atributo value si es un input/textarea."""
    try:
        el = driver.find_element(By.ID, field_id)
    except Exception:
        return None
    try:
        tag = (el.tag_name or "").lower()
    except Exception:
        return None
    if tag == "select":
        try:
            return (Select(el).first_selected_option.text or "").strip()
        except Exception:
            return None
    try:
        return (el.get_attribute("value") or "").strip()
    except Exception:
        return None


def resolve_model_field_id(driver, field_id=None):
    """Devuelve el id real del <select> de modelos en el form actual, tolerando la variante
    singular/plural ("models" / "model"): prueba el id pedido y luego los alias conocidos.
    Devuelve None si ninguna variante existe — ahí no hay selector de modelo que usar."""
    candidates = [field_id] if field_id else []
    candidates += [alias for alias in MODEL_FIELD_ID_ALIASES if alias != field_id]
    for idx, candidate in enumerate(candidates):
        # El primer candidato es el esperado y merece la espera normal; los alias son
        # sólo una variante de nombre y se prueban con timeout corto para no encadenar
        # esperas largas cuando el form directamente no tiene selector de modelo.
        if _get_select(driver, candidate, timeout=10 if idx == 0 else 2) is not None:
            return candidate
    return None


def _get_model_select(driver, field_id):
    """Resuelve el <select> de modelos por id o alias. None si no está en el form."""
    resolved = resolve_model_field_id(driver, field_id)
    return _get_select(driver, resolved, timeout=2) if resolved else None


def list_model_options(driver, field_id):
    """Devuelve los textos de todas las opciones válidas (no placeholder) del <select> de
    modelos, tal como están hoy en el form real. Usado para el modo 'Todos los modelos'."""
    el = _get_model_select(driver, field_id)
    if el is None:
        return []
    return [o.text.strip() for o in _valid_options(el) if not _is_placeholder(o.text)]


def _select_model(driver, field_id, model_text):
    el = _get_model_select(driver, field_id)
    if el is None:
        return False
    opt, _disclaimer = _find_option_by_text(el, model_text)
    if opt is None:
        return False
    _select_option(driver, el, opt)
    return True


_LEVEL_RETRY_PAUSE = 0.6


def _select_is_effectively_empty(driver, select_id):
    """True si el <select> no tiene NINGUNA opción real todavía — sólo el placeholder
    ("Seleccione"/"Selecione"/etc., que no cuenta como opción) o directamente no se lo
    pudo encontrar. Es la señal de que esa parte de la página no terminó de cargar (no
    sólo que la opción buscada tarda), y amerita una recarga en vez de un reintento liviano."""
    el = _get_select(driver, select_id, timeout=1)
    if el is None:
        return True
    return not any(not _is_placeholder(o.text) for o in _valid_options(el))


def _select_level_with_retries(driver, select_id, target_text, log, level_label,
                                retrigger_fn=None, reload_and_replay_fn=None, skip=False):
    """Selecciona una opción por texto, con UN reintento ante un "no encontrado" antes de
    darlo por real. Contra el form real (probado con LambdaTest y sitios de producción
    lentos/con red inestable) un dropdown dependiente a veces no termina de repoblarse en
    el primer intento — sobre todo justo después de cambiar de modelo o de nivel padre — y
    sin reintentar eso se reportaba como FAIL aunque el dealer/ciudad/región sí estuviera:
    la misma corrida, contra los mismos datos, daba resultados distintos cada vez.

    El reintento elige entre dos caminos según el estado del dropdown:
      - Si quedó EFECTIVAMENTE VACÍO (sin ninguna opción real, ver
        _select_is_effectively_empty): esa página no cargó bien esa parte, y confiar en
        que un simple re-disparo del evento la arregle es optimista. Se recarga la
        página entera y se repite la navegación hasta este punto (reload_and_replay_fn),
        que es más confiable — vuelve a un estado conocido en vez de intentar reparar uno
        roto.
      - Si el dropdown SÍ tiene opciones (sólo no la buscada, o la está por recibir):
        reintento liviano — re-dispara el 'change' del nivel padre (retrigger_fn).

    `skip=True` (modo exclusión con "no encontrado" esperado) evita reintentar de más.
    Devuelve (found, text, disclaimer)."""
    found, text, disclaimer = _select_by_text_robust(driver, select_id, target_text, wait_for_option=True)
    if found or skip:
        return found, text, disclaimer

    log(f"  {level_label} '{target_text}' no encontrado, reintentando...", "warn")
    if _select_is_effectively_empty(driver, select_id) and reload_and_replay_fn:
        log(f"  {level_label}: el dropdown está vacío/no cargó — recargando la página...", "warn")
        reload_and_replay_fn()
    elif retrigger_fn:
        retrigger_fn()
    time.sleep(_LEVEL_RETRY_PAUSE)
    return _select_by_text_robust(driver, select_id, target_text, wait_for_option=True)


def _check_stop(stop_flag):
    if stop_flag is not None and stop_flag.is_set():
        raise StopRequested()


def _wait_document_ready(driver, timeout=30):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def _find_form_iframe(driver, expected_form_url):
    expected_form_url = (expected_form_url or "").strip()
    if not expected_form_url:
        return None
    try:
        for iframe in driver.find_elements(By.TAG_NAME, "iframe"):
            src = iframe.get_attribute("src") or ""
            if expected_form_url in src:
                return iframe
    except Exception:
        pass
    return None


def _pre_scroll_for_dynamic_content(driver):
    """Dispara eventos de scroll para activar IntersectionObserver/lazy-loading de iframes
    embebidos en la landing (mismo criterio que 'Envío de Leads')."""
    try:
        total_height = driver.execute_script("return document.body.parentNode.scrollHeight") or 0
        viewport_height = driver.execute_script("return window.innerHeight") or 800
    except Exception:
        return
    scroll_step = max(viewport_height * 0.8, 800)
    _scroll_js = (
        "window.scrollTo(0, arguments[0]);"
        "window.dispatchEvent(new Event('scroll', {bubbles:true,cancelable:false}));"
        "document.dispatchEvent(new Event('scroll', {bubbles:true}));"
    )
    current_position = 0
    while current_position < total_height:
        try:
            driver.execute_script(_scroll_js, current_position)
        except Exception:
            break
        time.sleep(0.15)
        current_position += scroll_step
    try:
        driver.execute_script(_scroll_js, 999999)
        time.sleep(0.3)
        driver.execute_script(_scroll_js, 0)
        time.sleep(0.3)
    except Exception:
        pass


def _locate_form_iframe(driver, form_url, wait_seconds=0):
    """Encuentra el iframe del form DENTRO de la landing (mismo criterio que 'Envío de Leads'):
    primero busca el que matchea la URL esperada (con espera activa si wait_seconds>0, para
    lazy-loading justo tras navegar); si no lo encuentra, cae a 'cualquier iframe visible'
    (el form pudo redirigir a otro dominio/URL). Nunca navega directo a form_url."""
    form_url = (form_url or "").strip()
    iframe = None
    if form_url:
        if wait_seconds > 0:
            try:
                iframe = WebDriverWait(driver, wait_seconds).until(
                    lambda d: next(
                        (f for f in d.find_elements(By.TAG_NAME, "iframe")
                         if form_url in (f.get_attribute("src") or "")),
                        None
                    )
                )
            except Exception:
                iframe = None
        else:
            iframe = _find_form_iframe(driver, form_url)
    if iframe is None:
        try:
            for f in driver.find_elements(By.TAG_NAME, "iframe"):
                try:
                    if f.is_displayed():
                        iframe = f
                        break
                except Exception:
                    continue
        except Exception:
            pass
    return iframe


def handle_cookie_popups(driver):
    """Cierra popups de cookies/legales (mismo criterio que BaseFormFiller.handle_cookie_popups,
    incluye el caso puntual de GM/Chevrolet vía <gb-legal-notification>)."""
    try:
        popups = driver.find_elements(By.TAG_NAME, "gb-legal-notification")
        if popups:
            close_buttons = driver.find_elements(By.CSS_SELECTOR, ".close-btn.js-close-icon.silent-consent")
            for btn in close_buttons:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(1)
                    return True
    except Exception:
        pass

    selectors = [
        "button[onclick*='cookie']",
        "button[class*='cookie-accept']",
        "button[id*='cookie-accept']",
        ".cookie-accept",
        "#cookie-accept",
        ".js-close-icon",
        ".silent-consent",
        ".close-btn",
    ]
    for selector in selectors:
        try:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                if element.is_displayed():
                    driver.execute_script("arguments[0].click();", element)
                    time.sleep(1)
                    return True
        except Exception:
            continue
    return False


def open_target(driver, url_mode, landing_url="", form_url="", page_ready_timeout=30):
    """Abre el form real, igual criterio que 'Envío de Leads':
    - url_mode='landing_form': abre la landing y hace el chequeo DENTRO de ella — espera
      activamente a que aparezca el iframe del form (con reintento + pre-scroll para
      lazy-loading) y cambia de contexto a ese iframe. Si no hay iframe reconocible, el form
      vive directo en la landing (sin iframe) y se sigue trabajando ahí. Nunca se navega
      directo a form_url: sería saltarse la landing, que es justo lo que se quiere chequear.
    - url_mode='solo_forms': abre form_url directamente (sin landing) — para cuando la URL
      configurada es la del formulario suelto (secure/stage/revise).
    En ambos casos cierra popups de cookies antes y después de entrar al form.
    Devuelve True si terminó sobre el contexto del form (o su iframe)."""
    def _same_url(a, b):
        return (a or "").strip().rstrip("/").lower() == (b or "").strip().rstrip("/").lower()

    # Form suelto: modo explícito, sin landing, o la misma URL repetida en las dos columnas.
    # Este último caso es el de los forms .../gm_frontend/chevrolet/t3/<pais>/form/<slug>,
    # que son una página entera; si se tratara como landing, el fallback "cualquier iframe
    # visible" podría meterse en un iframe de tracking y no encontrar ningún campo.
    if url_mode == "solo_forms" or not landing_url or _same_url(landing_url, form_url):
        driver.get(form_url or landing_url)
        _wait_document_ready(driver, page_ready_timeout)
        handle_cookie_popups(driver)
        return True

    driver.get(landing_url)
    _wait_document_ready(driver, page_ready_timeout)
    handle_cookie_popups(driver)

    iframe = _locate_form_iframe(driver, form_url, wait_seconds=8)
    if iframe is None:
        # Iframe embebido con lazy-loading: activar con scroll y reintentar una vez más
        # antes de asumir que el form no vive en un iframe.
        _pre_scroll_for_dynamic_content(driver)
        handle_cookie_popups(driver)
        iframe = _locate_form_iframe(driver, form_url, wait_seconds=8)

    if iframe is not None:
        driver.switch_to.frame(iframe)
        handle_cookie_popups(driver)

    return True


# ──────────────────────────────────────────────────────────────────────────────
# Avance de pasos en formularios multi-paso (igual criterio que Envío de Leads):
# clickea Siguiente/Next hasta que aparezcan los dropdowns region/city/dealer.
# ──────────────────────────────────────────────────────────────────────────────
_NEXT_KEYWORDS = (
    "siguiente", "seguinte", "continuar", "continuacao", "continue",
    "proximo", "next", "avanzar", "avancar", "seguir",
)
_NEXT_SELECTORS = [
    "button.btn-steps-submit",
    ".button.next.pulsate.stat-button-link",
    "button[class*='next']",
    "button[data-dtm*='next']",
    "a[class*='next']",
    ".next-button",
    "button[type='button']",
    "button[type='submit']",
    "input[type='submit']",
    "input[type='button']",
]


def _looks_like_next(el):
    try:
        if not el.is_displayed() or not el.is_enabled():
            return False
        parts = [
            el.text or "",
            el.get_attribute("data-dtm") or "",
            el.get_attribute("class") or "",
            el.get_attribute("aria-label") or "",
            el.get_attribute("title") or "",
            el.get_attribute("value") or "",
        ]
        blob = normalize_text(" ".join(p for p in parts if p))
        return any(normalize_text(k) in blob for k in _NEXT_KEYWORDS)
    except Exception:
        return False


def _find_next_button(driver):
    """Primer botón/enlace visible que parezca avanzar de paso (Siguiente/Next/Continuar)."""
    for sel in _NEXT_SELECTORS:
        try:
            for btn in driver.find_elements(By.CSS_SELECTOR, sel):
                if _looks_like_next(btn):
                    return btn
        except Exception:
            continue
    return None


def _select_visible(driver, element_id):
    if not element_id:
        return False
    try:
        for el in driver.find_elements(By.ID, element_id):
            if (el.tag_name or "").lower() == "select" and el.is_displayed():
                return True
    except Exception:
        pass
    return False


def _anchor_level_id(level_ids, has_region, has_city, has_dealer):
    """Nivel 'ancla' = el más alto que el usuario seleccionó (region → city → dealer). Es el
    punto donde arranca la cascada de comparación, así que el avance de pasos apunta a ÉL."""
    if has_region:
        return level_ids.get("region")
    if has_city:
        return level_ids.get("city")
    if has_dealer:
        return level_ids.get("dealer")
    return None


def _any_level_select_present(driver, level_ids, has_region, has_city, has_dealer):
    """True cuando el <select> del nivel ANCLA (el más alto seleccionado) está VISIBLE — ahí
    empieza la comparación. Se apunta al ancla, no a 'cualquiera', para no frenar en un paso
    donde, por ejemplo, ya se ve 'dealer' pero todavía no 'region' que el usuario también pidió.
    Fallback: si el ancla no está pero sí otro nivel activo, también sirve (forms atípicos)."""
    anchor = _anchor_level_id(level_ids, has_region, has_city, has_dealer)
    if _select_visible(driver, anchor):
        return True
    # Fallback tolerante: cualquier nivel activo visible (forms que no siguen el orden típico)
    for eid, activo in (
        (level_ids.get("region"), has_region),
        (level_ids.get("city"), has_city),
        (level_ids.get("dealer"), has_dealer),
    ):
        if activo and eid != anchor and _select_visible(driver, eid):
            return True
    return False


def _step_signature(driver):
    """Firma del paso actual: ids/names de los campos visibles. Sirve para detectar si el
    formulario realmente avanzó tras apretar Siguiente (si no cambia, el paso está gateado)."""
    sig = []
    try:
        for el in driver.find_elements(By.CSS_SELECTOR, "input, select, textarea"):
            try:
                if not el.is_displayed():
                    continue
                if (el.get_attribute("type") or "") == "hidden":
                    continue
                sig.append((el.get_attribute("id") or el.get_attribute("name") or "", el.tag_name))
            except Exception:
                continue
    except Exception:
        pass
    return tuple(sorted(sig))


def _dummy_for_input(el):
    """Valor sintético razonable para un input de texto, según su tipo/name/placeholder."""
    kind = " ".join([
        (el.get_attribute("type") or ""), (el.get_attribute("name") or ""),
        (el.get_attribute("id") or ""), (el.get_attribute("placeholder") or ""),
        (el.get_attribute("inputmode") or ""),
    ]).lower()
    if "mail" in kind or (el.get_attribute("type") or "").lower() == "email":
        return "prueba.qa@mrm.com"
    if any(k in kind for k in ("tel", "phone", "celular", "telefono", "telefone", "movil")):
        return "1122334455"
    if any(k in kind for k in ("dni", "documento", "cpf", "cuit", "rut", "cedula", "number", "numeric", "cep", "cod")):
        return "30111222"
    if any(k in kind for k in ("date", "fecha", "data")):
        return ""  # las fechas suelen ser selects/datepicker; no forzamos texto
    return "Prueba"


def _fill_step_for_advance(driver, level_ids, log=None):
    """Completa los campos VISIBLES del paso actual con valores sintéticos (sin tocar los
    selects region/city/dealer que queremos alcanzar), para poder pasar de paso en forms
    gateados por campos requeridos — mismo espíritu que Envío de Leads pero acotado a lo
    mínimo para navegar hasta los dropdowns de dealer. No envía el formulario."""
    log = log or (lambda *_a, **_k: None)
    protegidos = {level_ids.get("region"), level_ids.get("city"), level_ids.get("dealer")}
    # 1) Selects visibles (no protegidos, sin selección válida): primera opción válida
    try:
        for sel in driver.find_elements(By.TAG_NAME, "select"):
            try:
                if not sel.is_displayed() or not sel.is_enabled():
                    continue
                if (sel.get_attribute("id") or sel.get_attribute("name")) in protegidos:
                    continue
                if sel.get_attribute("multiple") is not None:
                    continue
                # ¿ya tiene una opción real elegida?
                cur = ""
                try:
                    cur = (Select(sel).first_selected_option.text or "").strip()
                except Exception:
                    cur = ""
                if cur and not _is_placeholder(cur):
                    continue
                opcion = next((o for o in _valid_options(sel)
                               if o.text.strip() and not _is_placeholder(o.text)), None)
                if opcion is not None:
                    _select_option(driver, sel, opcion)
            except Exception:
                continue
    except Exception:
        pass
    # 2) Inputs de texto y textareas visibles y vacíos
    try:
        for el in driver.find_elements(By.CSS_SELECTOR, "input, textarea"):
            try:
                if not el.is_displayed() or not el.is_enabled():
                    continue
                tag = (el.tag_name or "").lower()
                itype = (el.get_attribute("type") or "text").lower() if tag == "input" else "textarea"
                if itype in ("hidden", "submit", "button", "reset", "file", "image", "range", "color"):
                    continue
                if itype in ("checkbox", "radio"):
                    continue
                if (el.get_attribute("value") or "").strip():
                    continue
                val = _dummy_for_input(el) if itype != "textarea" else "Prueba"
                if not val:
                    continue
                driver.execute_script(
                    "arguments[0].focus();"
                    "var s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value')||"
                    "Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype,'value');"
                    "try{s.set.call(arguments[0], arguments[1]);}catch(e){arguments[0].value=arguments[1];}"
                    "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                    "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
                    "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
                    el, val,
                )
            except Exception:
                continue
    except Exception:
        pass
    # 3) Checkboxes/radios visibles sin marcar (términos, etc.)
    try:
        for el in driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox'], input[type='radio']"):
            try:
                if not el.is_displayed() or not el.is_enabled() or el.is_selected():
                    continue
                try:
                    el.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", el)
            except Exception:
                continue
    except Exception:
        pass


def advance_to_selects(driver, level_ids=None, has_region=True, has_city=True, has_dealer=True,
                        log_cb=None, stop_flag=None, max_steps=12):
    """Avanza los pasos del formulario hasta que alguno de los dropdowns activos
    (region/city/dealer) esté visible. En forms de 1 paso los encuentra al toque; en
    multi-paso completa los campos requeridos de cada paso (sin tocar region/city/dealer)
    y aprieta Siguiente, igual que Envío de Leads. Detecta si el form no avanza (paso
    gateado) para no quedarse en loop. Devuelve True si llegó a los dropdowns."""
    level_ids = level_ids or DEFAULT_SELECT_IDS
    log = log_cb or (lambda *_a, **_k: None)
    sin_progreso = 0
    for step in range(max_steps + 1):
        _check_stop(stop_flag)
        if _any_level_select_present(driver, level_ids, has_region, has_city, has_dealer):
            if step:
                log(f"  Dropdowns de dealer encontrados tras avanzar {step} paso(s).", "ok")
            return True
        sig_antes = _step_signature(driver)
        # Completar el paso actual para destrabar el botón Siguiente
        _fill_step_for_advance(driver, level_ids, log=log)
        time.sleep(0.4)
        btn = _find_next_button(driver)
        if not btn:
            log("  No se ven los dropdowns region/city/dealer y no hay botón 'Siguiente'.", "warn")
            return False
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
            time.sleep(0.2)
            try:
                btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
            log(f"  Formulario multi-paso: paso {step + 1} completado, avanzando...", "info")
            _wait_document_ready(driver, 15)
            time.sleep(1.2)
        except StopRequested:
            raise
        except Exception as e:  # noqa: BLE001
            log(f"  No se pudo avanzar de paso: {e}", "warn")
            return False
        # ¿avanzó? (cambió el conjunto de campos visibles)
        if _step_signature(driver) == sig_antes:
            sin_progreso += 1
            if sin_progreso >= 2:
                log("  El formulario no avanza (posible campo requerido no reconocido). Se corta el avance.", "warn")
                return _any_level_select_present(driver, level_ids, has_region, has_city, has_dealer)
        else:
            sin_progreso = 0
    log(f"  No aparecieron los dropdowns tras {max_steps} pasos.", "warn")
    return False


# ──────────────────────────────────────────────────────────────────────────────
# Comparación principal: recorre las filas esperadas del Excel contra el form real
# ──────────────────────────────────────────────────────────────────────────────
def compare_dealers(
    driver,
    rows,
    column_map,
    level_ids=None,
    has_region=True,
    has_city=True,
    chk_bac=True,
    extra_validations=None,
    field_checks=None,
    model_field_id=None,
    models=None,
    log_cb=None,
    progress_cb=None,
    stop_flag=None,
    screenshot_cb=None,
    expect_absent=False,
    has_dealer=True,
    reload_cb=None,
):
    """
    rows: filas del Excel ya filtradas (dicts, ver read_excel_rows/filter_rows).
    column_map: {"region": <col_key>, "city": <col_key>, "dealer": <col_key>, "bac": <col_key>}
    level_ids: ids de los <select> del form, default DEFAULT_SELECT_IDS (region/city/dealer).
    reload_cb: callable sin argumentos que deja al driver de vuelta en la página del form,
               recién navegada (mismo criterio que el `open_target` + `advance_to_selects`
               que hace el caller antes de llamar a compare_dealers). Si se pasa, se usa
               como recuperación cuando un dropdown queda efectivamente vacío (sin ninguna
               opción real, sólo el placeholder) — más confiable que reintentar sobre una
               página que puede haber quedado en un estado roto. Sin reload_cb, ese caso
               cae al reintento liviano normal (re-disparar el nivel padre).
    extra_validations: lista de {"attr": "lat", "label": "Latitud", "column": <col_key>}
                       — compara data-<attr> del <option> de dealer contra la columna del Excel.
    field_checks: lista de {"column": <col_key>, "field_id": "id-del-campo"} — compara el valor
                  actual de cualquier campo del form (input o select, por id) contra una columna
                  del Excel. Se corre por fila, independientemente de si el dealer fue encontrado.
    model_field_id + models: si se pasan ambos (models no vacía), antes de cada tanda de filas se
                  selecciona ese modelo en el <select> de modelos (por id) y se repite la comparación
                  completa para cada modelo de la lista — útil cuando el listado de dealers cambia
                  según el modelo elegido (ej. modelos eléctricos con dealers limitados).
    Los tiempos de espera de los selects dependientes son fijos e internos (FAST_TIMEOUT/MAX_TIMEOUT,
    con reintento automático) — no son configurables desde afuera.

    Devuelve una lista de dicts de resultado: status, region, city, dealer, modelo, bac_excel, bac_form,
    fails (lista de strings), fila (número de fila en el Excel original).
    """
    level_ids = level_ids or DEFAULT_SELECT_IDS
    extra_validations = extra_validations or []
    field_checks = field_checks or []
    _log_raw = log_cb or (lambda *_a, **_k: None)
    def log(msg, *a, **k):
        # Nunca debe poder tirar una excepción: un log_cb que falla (ej. encoding de
        # consola en un caracter especial) no puede convertir un resultado válido en FAIL.
        try:
            _log_raw(msg, *a, **k)
        except Exception:
            pass

    region_key = column_map.get("region")
    city_key = column_map.get("city")
    dealer_key = column_map.get("dealer") if has_dealer else None
    bac_key = column_map.get("bac")

    results = []
    models_to_run = list(models) if (model_field_id and models) else [None]
    total = len(rows) * len(models_to_run)
    counter = 0

    for model_text in models_to_run:
        if model_text:
            _check_stop(stop_flag)
            log(f"\n=== MODELO: {model_text} ===", "info")
            if not _select_model(driver, model_field_id, model_text):
                log(f"  Modelo '{model_text}' no encontrado en el form, se omite", "warn")
                counter += len(rows)
                continue
            # Cambiar de modelo suele repoblar TODA la cascada región→ciudad→dealer (algunos
            # forms limitan qué dealers venden qué modelo). Un sleep fijo corto alcanzaba para
            # el primer modelo (ya viene con la cascada cargada) pero no siempre para el
            # siguiente: se medía en pruebas reales contra el form real que, sin esperar a que
            # el nivel ancla tenga opciones de verdad, la primera tanda de filas del modelo
            # nuevo arrancaba a comparar contra un select todavía vacío/a medio poblar y daba
            # FAIL falso. Se espera activamente (con tope) a que el nivel ancla tenga opciones.
            anchor_id = level_ids.get("region") if has_region else (
                level_ids.get("city") if has_city else level_ids.get("dealer"))
            if anchor_id:
                _wait_options_loaded_once(driver, anchor_id, timeout=DEPENDENT_SELECT_TIMEOUT)
            time.sleep(0.3)

        # Estado de navegación: qué región/ciudad están seleccionadas ahora mismo en el
        # form. Re-seleccionar el MISMO valor rompe algunos forms (limpian el select hijo
        # y no lo repueblan porque "no cambió") → dealers presentes daban FAIL falso. Por
        # eso sólo se re-navega región/ciudad cuando cambian respecto a la fila anterior;
        # si son iguales, el select de dealer ya tiene la lista cargada y se elige directo.
        cur_region = None
        cur_city = None

        for row in rows:
            counter += 1
            _check_stop(stop_flag)

            region_text = row.get(region_key, "") if has_region else ""
            city_text = row.get(city_key, "") if has_city else ""
            dealer_text = row.get(dealer_key, "") if (has_dealer and dealer_key) else ""

            if progress_cb:
                # Etiqueta explícita con la combinación que se está revisando, para que el
                # modal muestre Región / Ciudad / Dealer (los niveles activos).
                _partes = []
                if has_region and region_text:
                    _partes.append(f"Región: {region_text}")
                if has_city and city_text:
                    _partes.append(f"Ciudad: {city_text}")
                if has_dealer and dealer_text:
                    _partes.append(f"Dealer: {dealer_text}")
                label = " · ".join(_partes) or f"Fila {row.get('__row__')}"
                progress_cb(counter, total, f"Modelo {model_text} · {label}" if model_text else label)

            bac_excel = row.get(bac_key, "") if chk_bac and bac_key else ""

            fails = []
            region_found = not has_region
            city_found = not has_city
            dealer_found = False
            bac_ok = None
            extra_results = {}
            name_disclaimer = None
            detalle_info = ""

            try:
                # Orden SIEMPRE: región → ciudad → dealer (si no hay región, arranca en ciudad).
                # Cada nivel usa _select_by_text_robust: re-busca el <select> fresco en cada
                # intento (stale-safe ante re-render del form) y espera a que la opción aparezca
                # en los selects dependientes (city tras region, dealer tras city). Sólo se
                # (re)selecciona un nivel si su valor cambió respecto a la fila anterior.
                if has_region and region_text:
                    if region_text == cur_region:
                        region_found = True  # ya seleccionada de una fila previa
                    elif _get_select(driver, level_ids["region"], timeout=2) is None:
                        fails.append(f"Select de región '{level_ids['region']}' no encontrado en el form")
                        cur_region = None
                        cur_city = None
                    else:
                        _reload_for_region = None
                        if reload_cb:
                            def _reload_for_region():
                                reload_cb()
                                if model_text:
                                    _select_model(driver, model_field_id, model_text)
                                    time.sleep(0.4)  # asentar la cascada tras re-seleccionar el modelo
                        region_found, _, _ = _select_level_with_retries(
                            driver, level_ids["region"], region_text, log, "Región",
                            reload_and_replay_fn=_reload_for_region, skip=expect_absent)
                        cur_city = None  # cambió la región → el select de ciudad se repuebla
                        if region_found:
                            cur_region = region_text
                            log(f"  Región: {region_text}", "ok")
                        else:
                            cur_region = None
                            fails.append(f"Región '{region_text}' no encontrada")
                            log(f"  Región no encontrada: {region_text}", "warn")

                if has_city and city_text and (region_found or not has_region):
                    if city_text == cur_city:
                        city_found = True  # ya seleccionada (misma región y ciudad que la fila previa)
                    else:
                        _retrigger_region = (
                            (lambda: _select_by_text_robust(driver, level_ids["region"], region_text, wait_for_option=True))
                            if (has_region and region_text) else None
                        )
                        _reload_for_city = None
                        if reload_cb:
                            def _reload_for_city():
                                reload_cb()
                                if model_text:
                                    _select_model(driver, model_field_id, model_text)
                                    time.sleep(0.4)
                                if has_region and region_text:
                                    _select_by_text_robust(driver, level_ids["region"], region_text, wait_for_option=True)
                                    time.sleep(0.4)  # asentar la cascada tras re-seleccionar región
                        city_found, _, _ = _select_level_with_retries(
                            driver, level_ids["city"], city_text, log, "Ciudad",
                            retrigger_fn=_retrigger_region, reload_and_replay_fn=_reload_for_city,
                            skip=expect_absent)
                        if city_found:
                            cur_city = city_text
                            log(f"  Ciudad: {city_text}", "ok")
                        else:
                            cur_city = None
                            fails.append(f"Ciudad '{city_text}' no encontrada")
                            log(f"  Ciudad no encontrada: {city_text}", "warn")

                ready_for_dealer = (region_found or not has_region) and (city_found or not has_city)
                if has_dealer and dealer_text and ready_for_dealer:
                    if has_city and city_text:
                        _retrigger_parent = (
                            lambda: _select_by_text_robust(driver, level_ids["city"], city_text, wait_for_option=True))
                    elif has_region and region_text:
                        _retrigger_parent = (
                            lambda: _select_by_text_robust(driver, level_ids["region"], region_text, wait_for_option=True))
                    else:
                        _retrigger_parent = None
                    _reload_for_dealer = None
                    if reload_cb:
                        def _reload_for_dealer():
                            reload_cb()
                            if model_text:
                                _select_model(driver, model_field_id, model_text)
                                time.sleep(0.4)
                            if has_region and region_text:
                                _select_by_text_robust(driver, level_ids["region"], region_text, wait_for_option=True)
                                time.sleep(0.4)
                            if has_city and city_text:
                                _select_by_text_robust(driver, level_ids["city"], city_text, wait_for_option=True)
                                time.sleep(0.4)  # asentar la cascada tras re-seleccionar ciudad, antes del retry final de dealer
                    dealer_found, _, name_disclaimer = _select_level_with_retries(
                        driver, level_ids["dealer"], dealer_text, log, "Dealer",
                        retrigger_fn=_retrigger_parent, reload_and_replay_fn=_reload_for_dealer,
                        skip=expect_absent)
                    if dealer_found:
                        log(f"  Dealer: {dealer_text}"
                            + (f"  ⚠ {name_disclaimer}" if name_disclaimer else ""), "ok")
                        time.sleep(0.2)
                    elif not expect_absent:
                        fails.append(f"Dealer '{dealer_text}' no encontrado")
                        log(f"  Dealer no encontrado: {dealer_text}", "warn")

                if expect_absent:
                    # Modo EXCLUIR: este dealer NO debería estar en el form. Sólo importa
                    # su presencia — se descartan los fails de región/ciudad (si no están,
                    # el dealer tampoco puede estar → correctamente ausente). detalle_info
                    # explica en QUÉ nivel se confirmó la ausencia (misma lógica que usa
                    # la captura de evidencia para elegir qué dropdown desplegar).
                    fails = []
                    if dealer_found:
                        fails.append(f"Dealer '{dealer_text}' ESTÁ en el form pero NO debería (excluido)")
                        log(f"  ✗ '{dealer_text}' presente cuando debería estar ausente", "warn")
                    else:
                        if has_region and region_text and not region_found:
                            detalle_info = (f"Correctamente ausente: la región '{region_text}' no está en el "
                                            f"form, por lo tanto el dealer tampoco puede aparecer")
                        elif has_city and city_text and not city_found:
                            detalle_info = (f"Correctamente ausente: la ciudad '{city_text}' no está en el "
                                            f"form, por lo tanto el dealer tampoco puede aparecer")
                        else:
                            detalle_info = f"Correctamente ausente: '{dealer_text}' no aparece en el dropdown de dealers"
                        log(f"  ✓ '{dealer_text}' correctamente ausente del form", "ok")
                elif dealer_found:
                    if chk_bac and bac_key and bac_excel:
                        bac_form = _selected_dealer_option_attr(driver, level_ids["dealer"], "data-bac")
                        bac_ok = normalize_text(bac_excel) == normalize_text(bac_form)
                        if not bac_ok:
                            fails.append(f"BAC no coincide (excel='{bac_excel}' form='{bac_form}')")
                    for extra in extra_validations:
                        excel_val = row.get(extra["column"], "")
                        if not excel_val:
                            continue
                        form_val = _selected_dealer_option_attr(driver, level_ids["dealer"], f"data-{extra['attr']}")
                        ok = normalize_text(excel_val) == normalize_text(form_val)
                        extra_results[extra["label"]] = {"ok": ok, "excel": excel_val, "form": form_val}
                        if not ok:
                            fails.append(f"{extra['label']} no coincide (excel='{excel_val}' form='{form_val}')")

                if not expect_absent:
                    for check in field_checks:
                        excel_val = row.get(check["column"], "")
                        if not excel_val:
                            continue
                        form_val = _read_form_field_value(driver, check["field_id"]) or ""
                        ok = normalize_text(excel_val) == normalize_text(form_val)
                        extra_results[check["field_id"]] = {"ok": ok, "excel": excel_val, "form": form_val}
                        if not ok:
                            fails.append(
                                f"Campo '{check['field_id']}' no coincide (excel='{excel_val}' form='{form_val}')"
                            )

                status = "PASS" if not fails else "FAIL"

            except StopRequested:
                raise
            except Exception as e:  # noqa: BLE001
                status = "FAIL"
                fails.append(f"Error inesperado: {e}")
                log(f"  Error: {e}", "err")

            result = {
                "status": status,
                "modelo": model_text or "",
                "region": region_text,
                "city": city_text,
                "dealer": dealer_text,
                "bac_excel": bac_excel,
                "bac_ok": bac_ok,
                "extra_results": extra_results,
                "fails": fails,
                "fila": row.get("__row__"),
                "name_disclaimer": name_disclaimer if (dealer_found and not expect_absent) else None,
                "detalle_info": detalle_info,
                # En qué nivel cortó (o no) la navegación — lo usa la captura de evidencia
                # de exclusión para desplegar el dropdown correcto (región/ciudad/dealer).
                "region_found": bool(region_found),
                "city_found": bool(city_found),
                "dealer_found": bool(dealer_found),
            }

            if screenshot_cb is not None:
                try:
                    screenshot_cb(result)
                except StopRequested:
                    raise
                except Exception as e:  # noqa: BLE001
                    log(f"  Error tomando captura: {e}", "err")

            results.append(result)

    return results


def find_extra_dealers(
    driver,
    rows,
    column_map,
    level_ids=None,
    has_region=True,
    has_city=True,
    log_cb=None,
    progress_cb=None,
    stop_flag=None,
    has_dealer=True,
    hidden_rows_data=None,
    model_field_id=None,
    models=None,
):
    """
    Recorre las combinaciones del nivel superior presentes en el Excel y reporta:
    - EXTRA: opciones que aparecen en el form pero no están en la lista esperada del Excel.
    - DUPLICADO: opciones que aparecen más de una vez como <option> en el form.
    - OCULTO: opciones del form que SÓLO están declaradas en filas que NO se ven (ocultas o filtradas) del Excel
      (hidden_rows_data) — están mal porque la fila está oculta; van en naranja.
    Este chequeo se realiza de forma jerárquica (Región -> Ciudad -> Dealer) según los niveles activos,
    buscando en conjunto y reportando la combinación exacta de forma contextual.
    """
    level_ids = level_ids or DEFAULT_SELECT_IDS
    _log_raw = log_cb or (lambda *_a, **_k: None)
    def log(msg, *a, **k):
        # Nunca debe poder tirar una excepción: un log_cb que falla (ej. encoding de
        # consola en un caracter especial) no puede convertir un resultado válido en FAIL.
        try:
            _log_raw(msg, *a, **k)
        except Exception:
            pass

    region_key = column_map.get("region")
    city_key = column_map.get("city")
    dealer_key = column_map.get("dealer")

    def _build_expected_maps(source_rows):
        """Indexa (regiones, ciudades_por_región, dealers_por_región_ciudad, canon_dealers,
        orig_region, orig_city) a partir de un set de filas del Excel."""
        regions = set()
        cities_by_region = {}
        dealers_by_rc = {}
        dealer_canon_by_rc = {}
        orig_reg = {}
        orig_city = {}
        for row in source_rows or []:
            reg_val = str(row.get(region_key, "")).strip() if (has_region and region_key) else ""
            city_val = str(row.get(city_key, "")).strip() if (has_city and city_key) else ""
            dealer_val = str(row.get(dealer_key, "")).strip() if (has_dealer and dealer_key) else ""

            norm_reg = normalize_text(reg_val)
            norm_city = normalize_text(city_val)

            if has_region and reg_val:
                regions.add(norm_reg)
                orig_reg[norm_reg] = reg_val
            if has_city and city_val:
                cities_by_region.setdefault(norm_reg, set()).add(norm_city)
                orig_city[norm_city] = city_val
            if has_dealer and dealer_val:
                dealers_by_rc.setdefault((norm_reg, norm_city), set()).add(normalize_text(dealer_val))
                dealer_canon_by_rc.setdefault((norm_reg, norm_city), {})[
                    _canonical_dealer_name(dealer_val)
                ] = dealer_val
        return regions, cities_by_region, dealers_by_rc, dealer_canon_by_rc, orig_reg, orig_city

    # Filas visibles (lo esperado real) vs filas OCULTAS del Excel (lo que NO debería contar
    # pero, si aparece en el form, hay que señalarlo en naranja como OCULTO)
    (expected_regions, expected_cities_by_region, expected_dealers_by_region_city,
     expected_dealer_canon_by_region_city, orig_region_by_norm, orig_city_by_norm) = _build_expected_maps(rows)
    (hidden_regions, hidden_cities_by_region, hidden_dealers_by_region_city,
     _hidden_canon, hidden_orig_reg, hidden_orig_city) = _build_expected_maps(hidden_rows_data)
    # Para poder navegar también las combinaciones que sólo existen en filas ocultas
    orig_region_by_norm = {**hidden_orig_reg, **orig_region_by_norm}
    orig_city_by_norm = {**hidden_orig_city, **orig_city_by_norm}

    def _classify_dealer_options(option_texts, key):
        """Separa las opciones vistas en el form en (duplicadas, extras, notas, ocultos)
        contra lo esperado para (norm_region, norm_city). Tolera diferencias menores de
        nombre (ver match_dealer_names) → 'notas'. Lo declarado sólo en filas OCULTAS del
        Excel → 'ocultos' (mal: la fila está oculta), no 'extras'."""
        expected_norm = expected_dealers_by_region_city.get(key, set())
        expected_canon = expected_dealer_canon_by_region_city.get(key, {})
        hidden_norm = hidden_dealers_by_region_city.get(key, set())
        seen_in_form = {}
        extras, notas, ocultos = [], [], []
        for text in option_texts:
            norm = normalize_text(text)
            seen_in_form.setdefault(norm, []).append(text)
            if norm in expected_norm:
                continue
            if norm in hidden_norm:
                ocultos.append(text)
                continue
            canon = _canonical_dealer_name(text)
            if canon in expected_canon:
                notas.append((text, expected_canon[canon]))
            else:
                extras.append(text)
        duplicated = [(texts[0], len(texts)) for texts in seen_in_form.values() if len(texts) > 1]
        return duplicated, extras, notas, ocultos

    extra_results = []

    models_to_run = list(models) if (model_field_id and models) else [None]

    for model_text in models_to_run:
        if model_text:
            _check_stop(stop_flag)
            log(f"\n=== MODELO: {model_text} ===", "info")
            if not _select_model(driver, model_field_id, model_text):
                log(f"  Modelo '{model_text}' no encontrado en el form, se omite (extras/duplicados)", "warn")
                continue
            anchor_id = level_ids.get("region") if has_region else (
                level_ids.get("city") if has_city else level_ids.get("dealer"))
            if anchor_id:
                _wait_options_loaded_once(driver, anchor_id, timeout=DEPENDENT_SELECT_TIMEOUT)
            time.sleep(0.3)

        _model_start_idx = len(extra_results)

        # --------------------------------------------------------------------------
        # FASE A: Chequeo de Región (sólo si has_region es True)
        # --------------------------------------------------------------------------
        regions_to_check = [""]
        if has_region:
            log("Chequeando EXTRAS y DUPLICADOS en Regiones...", "info")
            option_texts = _read_valid_option_texts(driver, level_ids["region"])
            seen_in_form = {}
            for text in option_texts:
                norm = normalize_text(text)
                seen_in_form.setdefault(norm, []).append(text)
                if norm in expected_regions:
                    pass
                elif norm in hidden_regions:
                    log(f"  OCULTO REGION: {text} (declarada solo en fila(s) que no se ven (ocultas/filtradas) del Excel)", "warn")
                    extra_results.append({
                        "status": "OCULTO",
                        "region": text,
                        "city": "",
                        "dealer": "",
                        "fails": ["Región presente en el form pero declarada SOLO en fila(s) que NO se ven (ocultas o filtradas) del Excel"],
                    })
                else:
                    log(f"  EXTRA REGION: {text}", "warn")
                    extra_results.append({
                        "status": "EXTRA",
                        "region": text,
                        "city": "",
                        "dealer": "",
                        "fails": ["Región extra (no declarada en el Excel)"],
                    })
            for norm, texts in seen_in_form.items():
                if len(texts) > 1:
                    log(f"  DUPLICADA REGION: {texts[0]} x{len(texts)}", "warn")
                    extra_results.append({
                        "status": "DUPLICADO",
                        "region": texts[0],
                        "city": "",
                        "dealer": "",
                        "fails": [f"Región aparece {len(texts)} veces en el dropdown"],
                    })

            # Regiones a recorrer: las declaradas (visibles u ocultas) que existen en el form —
            # las ocultas también se navegan para poder clasificar sus ciudades/dealers.
            regions_to_check = [norm for norm in (expected_regions | hidden_regions) if norm in seen_in_form]

        # --------------------------------------------------------------------------
        # FASE B: Chequeo de Ciudad (sólo si has_city es True)
        # --------------------------------------------------------------------------
        cities_to_check = []  # lista de (norm_reg, norm_city)
    
        if has_city:
            if has_region:
                log("Chequeando EXTRAS y DUPLICADOS en Ciudades por Región...", "info")
                for norm_reg in regions_to_check:
                    _check_stop(stop_flag)
                    reg_orig_text = orig_region_by_norm[norm_reg]
                    if progress_cb:
                        progress_cb(1, 1, f"Región: {reg_orig_text}")

                    # Seleccionar la región para ver sus ciudades
                    ok, _, _ = _select_by_text_robust(driver, level_ids["region"], reg_orig_text)
                    if not ok:
                        continue
                    time.sleep(0.6)  # Pausa estable para repoblar ciudades

                    option_texts = _read_valid_option_texts(driver, level_ids["city"])
                    expected_cities = expected_cities_by_region.get(norm_reg, set())
                    hidden_cities = hidden_cities_by_region.get(norm_reg, set())
                    seen_in_form = {}
                    for text in option_texts:
                        norm = normalize_text(text)
                        seen_in_form.setdefault(norm, []).append(text)
                        if norm in expected_cities:
                            pass
                        elif norm in hidden_cities:
                            log(f"  OCULTO CIUDAD: {text} en {reg_orig_text} (solo en fila(s) que no se ven: ocultas/filtradas)", "warn")
                            extra_results.append({
                                "status": "OCULTO",
                                "region": reg_orig_text,
                                "city": text,
                                "dealer": "",
                                "fails": [f"Ciudad presente en el form para '{reg_orig_text}' pero declarada SOLO en fila(s) que NO se ven (ocultas o filtradas) del Excel"],
                            })
                        else:
                            log(f"  EXTRA CIUDAD: {text} (en Región: {reg_orig_text})", "warn")
                            extra_results.append({
                                "status": "EXTRA",
                                "region": reg_orig_text,
                                "city": text,
                                "dealer": "",
                                "fails": [f"Ciudad extra para la región '{reg_orig_text}'"],
                            })
                    for norm, texts in seen_in_form.items():
                        if len(texts) > 1:
                            log(f"  DUPLICADA CIUDAD: {texts[0]} x{len(texts)} (en Región: {reg_orig_text})", "warn")
                            extra_results.append({
                                "status": "DUPLICADO",
                                "region": reg_orig_text,
                                "city": texts[0],
                                "dealer": "",
                                "fails": [f"Ciudad aparece {len(texts)} veces en el dropdown de la región '{reg_orig_text}'"],
                            })

                    for norm_city in (expected_cities | hidden_cities):
                        if norm_city in seen_in_form:
                            cities_to_check.append((norm_reg, norm_city))
            else:
                # Sin Región, las ciudades son globales
                log("Chequeando EXTRAS y DUPLICADOS en Ciudades...", "info")
                option_texts = _read_valid_option_texts(driver, level_ids["city"])
                expected_cities = expected_cities_by_region.get("", set())
                hidden_cities = hidden_cities_by_region.get("", set())
                seen_in_form = {}
                for text in option_texts:
                    norm = normalize_text(text)
                    seen_in_form.setdefault(norm, []).append(text)
                    if norm in expected_cities:
                        pass
                    elif norm in hidden_cities:
                        log(f"  OCULTO CIUDAD: {text} (solo en fila(s) que no se ven: ocultas/filtradas)", "warn")
                        extra_results.append({
                            "status": "OCULTO",
                            "region": "",
                            "city": text,
                            "dealer": "",
                            "fails": ["Ciudad presente en el form pero declarada SOLO en fila(s) que NO se ven (ocultas o filtradas) del Excel"],
                        })
                    else:
                        log(f"  EXTRA CIUDAD: {text}", "warn")
                        extra_results.append({
                            "status": "EXTRA",
                            "region": "",
                            "city": text,
                            "dealer": "",
                            "fails": ["Ciudad extra (no declarada en el Excel)"],
                        })
                for norm, texts in seen_in_form.items():
                    if len(texts) > 1:
                        log(f"  DUPLICADA CIUDAD: {texts[0]} x{len(texts)}", "warn")
                        extra_results.append({
                            "status": "DUPLICADO",
                            "region": "",
                            "city": texts[0],
                            "dealer": "",
                            "fails": [f"Ciudad aparece {len(texts)} veces en el dropdown"],
                        })
                for norm_city in (expected_cities | hidden_cities):
                    if norm_city in seen_in_form:
                        cities_to_check.append(("", norm_city))

        # --------------------------------------------------------------------------
        # FASE C: Chequeo de Dealer (sólo si has_dealer es True)
        # --------------------------------------------------------------------------
        def _append_dealer_classification(option_texts, key, reg_orig_text, city_orig_text):
            duplicated, extras, notas, ocultos = _classify_dealer_options(option_texts, key)
            for text in extras:
                log(f"  EXTRA DEALER: {text} ({reg_orig_text} / {city_orig_text})", "warn")
                extra_results.append({
                    "status": "EXTRA", "region": reg_orig_text, "city": city_orig_text, "dealer": text,
                    "fails": [f"Dealer extra para '{reg_orig_text} / {city_orig_text}'".strip(" /")],
                })
            for text in ocultos:
                log(f"  OCULTO DEALER: {text} ({reg_orig_text} / {city_orig_text}) — solo en fila(s) que no se ven (ocultas/filtradas)", "warn")
                extra_results.append({
                    "status": "OCULTO", "region": reg_orig_text, "city": city_orig_text, "dealer": text,
                    "fails": [f"Dealer presente en el form para '{reg_orig_text} / {city_orig_text}' pero declarado SOLO en fila(s) que NO se ven (ocultas o filtradas) del Excel".strip(" /")],
                })
            for text, count in duplicated:
                log(f"  DUPLICADO DEALER: {text} x{count} ({reg_orig_text} / {city_orig_text})", "warn")
                extra_results.append({
                    "status": "DUPLICADO", "region": reg_orig_text, "city": city_orig_text, "dealer": text,
                    "fails": [f"Dealer aparece {count} veces en el dropdown de '{reg_orig_text} / {city_orig_text}'".strip(" /")],
                })
            for text, excel_orig in notas:
                log(f"  NOTA DEALER: '{text}' matchea '{excel_orig}' con diferencias menores "
                    f"({reg_orig_text} / {city_orig_text})", "info")
                extra_results.append({
                    "status": "NOTA", "region": reg_orig_text, "city": city_orig_text, "dealer": excel_orig,
                    "fails": [], "name_disclaimer": match_dealer_names(excel_orig, text)[1],
                })

        if has_dealer:
            log("Chequeando EXTRAS y DUPLICADOS en Dealers...", "info")
            if has_region and has_city:
                for norm_reg, norm_city in cities_to_check:
                    _check_stop(stop_flag)
                    reg_orig_text = orig_region_by_norm[norm_reg]
                    city_orig_text = orig_city_by_norm[norm_city]
                    if progress_cb:
                        progress_cb(1, 1, f"{reg_orig_text} / {city_orig_text}")

                    ok, _, _ = _select_by_text_robust(driver, level_ids["region"], reg_orig_text)
                    if not ok:
                        continue
                    ok, _, _ = _select_by_text_robust(driver, level_ids["city"], city_orig_text, wait_for_option=True)
                    if not ok:
                        continue
                    time.sleep(0.6)  # Pausa estable para repoblar dealers

                    option_texts = _read_valid_option_texts(driver, level_ids["dealer"])
                    _append_dealer_classification(option_texts, (norm_reg, norm_city), reg_orig_text, city_orig_text)
            elif not has_region and has_city:
                for _, norm_city in cities_to_check:
                    _check_stop(stop_flag)
                    city_orig_text = orig_city_by_norm[norm_city]
                    if progress_cb:
                        progress_cb(1, 1, f"Ciudad: {city_orig_text}")

                    ok, _, _ = _select_by_text_robust(driver, level_ids["city"], city_orig_text)
                    if not ok:
                        continue
                    time.sleep(0.6)

                    option_texts = _read_valid_option_texts(driver, level_ids["dealer"])
                    _append_dealer_classification(option_texts, ("", norm_city), "", city_orig_text)
            elif has_region and not has_city:
                for norm_reg in regions_to_check:
                    _check_stop(stop_flag)
                    reg_orig_text = orig_region_by_norm[norm_reg]
                    if progress_cb:
                        progress_cb(1, 1, f"Región: {reg_orig_text}")

                    ok, _, _ = _select_by_text_robust(driver, level_ids["region"], reg_orig_text)
                    if not ok:
                        continue
                    time.sleep(0.6)

                    option_texts = _read_valid_option_texts(driver, level_ids["dealer"])
                    _append_dealer_classification(option_texts, (norm_reg, ""), reg_orig_text, "")
            else:
                # Sin región ni ciudad, dealers globales
                option_texts = _read_valid_option_texts(driver, level_ids["dealer"])
                _append_dealer_classification(option_texts, ("", ""), "", "")


        for _er in extra_results[_model_start_idx:]:
            _er.setdefault("modelo", model_text or "")

    return extra_results


# ──────────────────────────────────────────────────────────────────────────────
# Captura de pantalla con banner de URL + ZIP
# ──────────────────────────────────────────────────────────────────────────────
def _ensure_window_covers_content(driver):
    """Agranda la ventana SOLO si hay overflow horizontal real (el contenido es más ancho
    que el viewport y quedaría recortado en la captura). Compara contra el viewport interior
    (clientWidth) y no contra el ancho de la ventana: un layout fluido siempre llena el 100%
    del viewport y NO debe disparar el agrandado (antes esto hacía crecer la ventana +60px
    en cada captura, en loop). Tope 2200px."""
    try:
        doc_w, client_w = driver.execute_script(
            "return [Math.max(document.body.scrollWidth || 0, document.documentElement.scrollWidth || 0),"
            " document.documentElement.clientWidth || 0];"
        )
        if not doc_w or not client_w or doc_w <= client_w + 5:
            return  # sin overflow horizontal: no tocar la ventana
        size = driver.get_window_size()
        extra = int(doc_w) - int(client_w)
        new_w = min(size.get("width", 1366) + extra + 20, 2200)
        if new_w > size.get("width", 0):
            driver.set_window_size(new_w, size.get("height", 1080))
    except Exception:
        pass


def banner_lines_for_result(result, has_region=True, has_city=True, has_dealer=True):
    """Arma las líneas extra del banner para una captura del Comparador: qué combinación
    se revisó (región/ciudad/dealer según niveles activos) y el estado PASS/FAIL."""
    partes = []
    if has_region and result.get("region"):
        partes.append(f"Región: {result['region']}")
    if has_city and result.get("city"):
        partes.append(f"Ciudad: {result['city']}")
    if has_dealer and result.get("dealer"):
        partes.append(f"Dealer: {result['dealer']}")
    if result.get("modelo"):
        partes.append(f"Modelo: {result['modelo']}")
    lines = []
    if partes:
        lines.append((f"Revisado:        {' · '.join(partes)}", (220, 220, 100)))
    status = result.get("status", "")
    color = (80, 200, 80) if status == "PASS" else (220, 80, 80)
    if result.get("fails"):
        detalle = " — " + " | ".join(result.get("fails", []))
    elif result.get("detalle_info"):
        detalle = " — " + result["detalle_info"]
    else:
        detalle = ""
    lines.append((f"Estado:          {status}{detalle}"[:220], color))
    return lines


def capture_result_screenshot(driver, screenshot_dir, filename, form_url="", landing_url="",
                              extra_lines=None):
    """Toma una captura de página completa con las URLs del form y landing pegadas en un banner
    superior (reusa ScreenshotManager). extra_lines: líneas extra del banner, ej. las de
    banner_lines_for_result (revisado + estado PASS/FAIL)."""
    os.makedirs(screenshot_dir, exist_ok=True)
    _ensure_window_covers_content(driver)
    manager = ScreenshotManager(driver, screenshot_dir)
    # Pasar ambas URLs para que el banner muestre las dos (si aplica)
    if landing_url:
        manager.url_landing = landing_url
    manager.url_form_esperado = form_url
    manager.extra_lines = list(extra_lines or [])
    manager.take_full_page_screenshot(filename)
    manager._add_url_banner(os.path.join(screenshot_dir, filename))
    return os.path.join(screenshot_dir, filename)


# ──────────────────────────────────────────────────────────────────────────────
# Evidencia de dropdown desplegado DENTRO del form (para verificar exclusión).
# El popup nativo del <select> es UI del sistema operativo y Selenium no puede
# fotografiarlo, pero sí se puede convertir temporalmente el <select> en una lista
# abierta (atributo size = cantidad de opciones): el dropdown queda desplegado en su
# lugar real dentro del formulario mostrando TODAS las opciones, se captura, y se
# restaura. Si el buscado aparece, se resalta en amarillo.
# ──────────────────────────────────────────────────────────────────────────────
_EXPAND_SELECT_JS = """
var id = arguments[0], searched = (arguments[1] || '').trim().toLowerCase();
var el = document.getElementById(id);
if (!el || !el.options) return -1;
el.setAttribute('data-ocioso-size-prev', el.getAttribute('size') || '');
el.size = Math.max(2, el.options.length);
el.style.height = 'auto';
el.style.maxHeight = 'none';
el.style.overflow = 'visible';
el.style.zIndex = '999999';
el.style.position = 'relative';
// Que ningún contenedor recorte la lista desplegada
var p = el.parentElement, hops = 0;
while (p && p !== document.body && hops < 6) {
    if (getComputedStyle(p).overflow !== 'visible') p.style.overflow = 'visible';
    p.style.maxHeight = 'none';
    p = p.parentElement; hops++;
}
for (var i = 0; i < el.options.length; i++) {
    var o = el.options[i];
    if (searched && (o.text || '').trim().toLowerCase() === searched) {
        o.style.background = '#ffd873';
        o.style.color = '#000000';
        o.style.fontWeight = 'bold';
        o.setAttribute('data-ocioso-marked', '1');
    }
}
try { el.scrollIntoView({block: 'center'}); } catch (e) {}
return el.options.length;
"""

_RESTORE_SELECT_JS = """
var id = arguments[0];
var el = document.getElementById(id);
if (!el) return;
var prev = el.getAttribute('data-ocioso-size-prev');
if (prev) el.setAttribute('size', prev); else el.removeAttribute('size');
el.removeAttribute('data-ocioso-size-prev');
el.style.height = ''; el.style.maxHeight = ''; el.style.overflow = '';
el.style.zIndex = ''; el.style.position = '';
for (var i = 0; i < el.options.length; i++) {
    var o = el.options[i];
    if (o.getAttribute('data-ocioso-marked')) {
        o.style.background = ''; o.style.color = ''; o.style.fontWeight = '';
        o.removeAttribute('data-ocioso-marked');
    }
}
"""


def evidence_level_for_result(result, has_region=True, has_city=True, has_dealer=True):
    """Decide en QUÉ dropdown desplegar la evidencia de exclusión: el primer nivel de la
    cadena región→ciudad→dealer que NO se encontró en el form (si la región no está, la
    ciudad y el dealer tampoco pueden estar — la evidencia es el dropdown de regiones).
    Si todo se encontró (dealer presente = FAIL de exclusión), la evidencia es el dropdown
    de dealer con el buscado resaltado. Sólo considera los niveles activos.
    Devuelve (nivel_label, select_key, texto_buscado, found_bool)."""
    if has_region and result.get("region") and not result.get("region_found"):
        return "Región", "region", result.get("region", ""), False
    if has_city and result.get("city") and not result.get("city_found"):
        return "Ciudad", "city", result.get("city", ""), False
    if has_dealer:
        return "Dealer", "dealer", result.get("dealer", ""), bool(result.get("dealer_found"))
    if has_city:
        return "Ciudad", "city", result.get("city", ""), bool(result.get("city_found"))
    return "Región", "region", result.get("region", ""), bool(result.get("region_found"))


_RESIZE_IFRAME_JS = """
var iframe = arguments[0], newHeight = arguments[1];
iframe.setAttribute('data-ocioso-h-prev', iframe.style.height || '');
iframe.style.height = newHeight + 'px';
iframe.style.maxHeight = 'none';
iframe.setAttribute('scrolling', 'no');
"""

_RESTORE_IFRAME_JS = """
var iframe = arguments[0];
iframe.style.height = iframe.getAttribute('data-ocioso-h-prev') || '';
iframe.style.maxHeight = '';
iframe.removeAttribute('data-ocioso-h-prev');
iframe.removeAttribute('scrolling');
"""


def capture_dropdown_evidence(driver, select_id, screenshot_dir, filename, title="", searched_text="",
                               found=None, form_url="", landing_url="", extra_lines=None):
    """Captura el formulario con el <select> indicado DESPLEGADO en su lugar: convierte
    temporalmente el dropdown en lista abierta (atributo size) para que se vean TODAS sus
    opciones dentro del form real — el popup nativo del <select> no es fotografiable. Si el
    texto buscado aparece entre las opciones queda resaltado. Como el form suele vivir en un
    iframe de altura fija que recortaría la lista desplegada, también se agranda el iframe
    desde la página padre a la altura del contenido expandido, se captura la página completa
    desde el documento principal, y se restaura todo (select, iframe y contexto)."""
    os.makedirs(screenshot_dir, exist_ok=True)
    expanded = False
    iframe_resized = None
    try:
        try:
            n = driver.execute_script(_EXPAND_SELECT_JS, select_id, searched_text or "")
            expanded = isinstance(n, (int, float)) and n >= 0
        except Exception:
            expanded = False
        time.sleep(0.3)  # dejar que el layout se acomode con la lista abierta

        # Altura real del documento del form con la lista abierta (contexto actual = form)
        try:
            inner_height = driver.execute_script(
                "return Math.max(document.documentElement.scrollHeight || 0,"
                " document.body ? document.body.scrollHeight : 0);"
            ) or 0
            driver.execute_script("window.scrollTo(0, 0);")
        except Exception:
            inner_height = 0

        # Si el form está embebido en un iframe, agrandarlo desde la página padre para que
        # la lista desplegada no quede recortada por la altura fija del iframe.
        try:
            driver.switch_to.default_content()
            iframe = _locate_form_iframe(driver, form_url)
        except Exception:
            iframe = None
        if iframe is not None and inner_height:
            try:
                driver.execute_script(_RESIZE_IFRAME_JS, iframe, int(inner_height) + 30)
                iframe_resized = iframe
                time.sleep(0.3)
            except Exception:
                iframe_resized = None
        elif iframe is None:
            # No hay iframe reconocible: el form está en la página principal, seguimos ahí.
            pass

        _ensure_window_covers_content(driver)
        manager = ScreenshotManager(driver, screenshot_dir)
        if landing_url:
            manager.url_landing = landing_url
        manager.url_form_esperado = form_url
        manager.extra_lines = list(extra_lines or [])
        manager.take_full_page_screenshot(filename)
        manager._add_url_banner(os.path.join(screenshot_dir, filename))
    finally:
        # Restaurar iframe (estamos en el documento principal desde el resize)
        if iframe_resized is not None:
            try:
                driver.execute_script(_RESTORE_IFRAME_JS, iframe_resized)
            except Exception:
                pass
        # Volver al contexto del form para restaurar el select y que la próxima fila navegue
        try:
            iframe = _locate_form_iframe(driver, form_url)
            if iframe is not None:
                driver.switch_to.frame(iframe)
        except Exception:
            pass
        if expanded:
            try:
                driver.execute_script(_RESTORE_SELECT_JS, select_id)
            except Exception:
                pass
    return os.path.join(screenshot_dir, filename)


def zip_screenshots(screenshot_paths, output_zip_path):
    os.makedirs(os.path.dirname(output_zip_path), exist_ok=True)
    with zipfile.ZipFile(output_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in screenshot_paths:
            if os.path.exists(path):
                zf.write(path, arcname=os.path.basename(path))
    return output_zip_path


# ──────────────────────────────────────────────────────────────────────────────
# Export a Excel (PASS verde / FAIL rojo / EXTRA amarillo / MISSING violeta)
# ──────────────────────────────────────────────────────────────────────────────
_RESULTS_HEADERS = ["Estado", "URL Landing", "URL Form", "Modelo", "Región", "Ciudad", "Dealer",
                    "BAC Excel", "BAC OK", "Detalle", "Nota Nombre", "Fila Excel"]

_RESULTS_COL_WIDTHS = {1: 10, 2: 45, 3: 45, 4: 16, 5: 22, 6: 22, 7: 30, 8: 14, 9: 10, 10: 50, 11: 55, 12: 10}

_STATUS_FILL_BY_NAME = {
    "PASS": _PASS_FILL, "FAIL": _FAIL_FILL, "EXTRA": _EXTRA_FILL,
    "MISSING": _MISSING_FILL, "DUPLICADO": _DUPLICATE_FILL, "NOTA": _NOTA_FILL,
    "OCULTO": _OCULTO_FILL,
}

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_name(name, taken):
    """Nombre de hoja Excel válido (sin \\/*?:[], máx 31 caracteres) y sin colisión con
    los ya usados en `taken` (se le agrega un sufijo numérico si hace falta)."""
    base = _INVALID_SHEET_CHARS.sub(" ", str(name or "").strip()) or "Modelo"
    base = re.sub(r"\s+", " ", base).strip()[:31] or "Modelo"
    candidate = base
    n = 2
    while candidate.lower() in taken:
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(candidate.lower())
    return candidate


def _write_results_sheet(wb, sheet_name, sheet_results, thin_border, header_border):
    """Escribe una hoja de resultados (mismo formato que antes, factorizado para poder
    repetirlo una vez por modelo cuando el form tiene selector de Modelo)."""
    from openpyxl.styles import Alignment

    ws = wb.create_sheet(sheet_name)
    ws.append(_RESULTS_HEADERS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in sheet_results:
        ws.append([
            r.get("status", ""),
            r.get("url_landing", ""),
            r.get("url_form", ""),
            r.get("modelo", ""),
            r.get("region", ""),
            r.get("city", ""),
            r.get("dealer", ""),
            r.get("bac_excel", ""),
            {True: "OK", False: "MISMATCH", None: ""}.get(r.get("bac_ok")),
            " | ".join(r.get("fails", [])) if r.get("fails") else (r.get("detalle_info") or ""),
            r.get("name_disclaimer", "") or "",
            r.get("fila", ""),
        ])
        row_idx = ws.max_row
        fill = _STATUS_FILL_BY_NAME.get(r.get("status"))
        # Todo lo relacionado con filas OCULTAS del Excel va en naranja, pise lo que pise
        if r.get("hidden_in_excel"):
            fill = _OCULTO_FILL
        for cell in ws[row_idx]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
        # Nota de nombre presente en fila PASS: resaltar solo esa celda para no
        # confundir con FAIL, ya que el dealer sí se encontró.
        if r.get("name_disclaimer") and r.get("status") == "PASS" and not r.get("hidden_in_excel"):
            ws.cell(row=row_idx, column=11).fill = _NOTA_FILL

    for col_idx, width in _RESULTS_COL_WIDTHS.items():
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def export_results_excel(results, output_path=None, pais="", hidden_rows=None, hidden_columns=None):
    """hidden_rows / hidden_columns: filas y columnas OCULTAS del Excel de dealers de ORIGEN
    (detect_hidden_rows / detect_hidden_columns). Se listan en el Resumen, y todo resultado
    relacionado con filas ocultas se pinta de NARANJA (estado OCULTO o flag hidden_in_excel).

    Si `results` trae más de un Modelo distinto (form con selector de Modelo, comparado
    contra varios), se separa en UNA HOJA POR MODELO en vez de una sola hoja mezclada —
    así no hay que filtrar la columna Modelo a mano para ver el detalle de cada uno. Los
    forms SIN selector de Modelo (o comparados contra uno solo) siguen igual que antes:
    una única hoja "Resultados"."""
    if output_path is None:
        results_dir = os.path.join(RESULTS_DIR, "dealer_comparator")
        os.makedirs(results_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(results_dir, f"dealer_comparator_{pais}_{stamp}.xlsx")

    from openpyxl.styles import Alignment, Border, Side

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="medium", color="7D4E9F"),
    )

    # Modelos distintos presentes (orden de primera aparición). Los resultados de
    # find_extra_dealers no traen "modelo" (todavía no es multi-modelo) — quedan con "".
    modelos_orden = []
    for r in results:
        m = (r.get("modelo") or "").strip()
        if m and m not in modelos_orden:
            modelos_orden.append(m)
    multi_modelo = len(modelos_orden) > 1

    wb = Workbook()
    wb.remove(wb.active)  # se crean las hojas explícitamente abajo, en el orden deseado

    if multi_modelo:
        taken_names = set()
        for modelo in modelos_orden:
            sheet_results = [r for r in results if (r.get("modelo") or "").strip() == modelo]
            sheet_name = _safe_sheet_name(modelo, taken_names)
            _write_results_sheet(wb, sheet_name, sheet_results, thin_border, header_border)
        # Resultados sin modelo asociado (ej. EXTRA/DUPLICADO de find_extra_dealers, que
        # no depende de qué modelo esté seleccionado): hoja aparte, no se pierden ni se
        # mezclan arbitrariamente con el primer modelo.
        sin_modelo = [r for r in results if not (r.get("modelo") or "").strip()]
        if sin_modelo:
            sheet_name = _safe_sheet_name("Extras", taken_names)
            _write_results_sheet(wb, sheet_name, sin_modelo, thin_border, header_border)
    else:
        _write_results_sheet(wb, "Resultados", results, thin_border, header_border)

    # ── Hoja Resumen ──
    total = len(results)
    counts = {"PASS": 0, "FAIL": 0, "EXTRA": 0, "MISSING": 0, "DUPLICADO": 0, "NOTA": 0, "OCULTO": 0}
    for r in results:
        counts[r.get("status", "FAIL")] = counts.get(r.get("status", "FAIL"), 0) + 1
    disclaimer_count = sum(1 for r in results if r.get("name_disclaimer"))

    summary_ws = wb.create_sheet("Resumen", 0)  # primera hoja: es lo primero que se quiere ver
    summary_data = [
        ["País", pais],
        ["Fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total chequeados", total],
        ["", ""],
        ["🟢 PASS", counts.get("PASS", 0)],
        ["🔴 FAIL", counts.get("FAIL", 0)],
        ["🟡 EXTRA (en el form, no declarado en el Excel)", counts.get("EXTRA", 0)],
        ["🔵 DUPLICADO (repetido en el dropdown del form)", counts.get("DUPLICADO", 0)],
        ["🟠 OCULTO (en el form, pero declarado solo en filas que no se ven (ocultas/filtradas))", counts.get("OCULTO", 0)],
        ["🟣 MISSING", counts.get("MISSING", 0)],
        ["🔷 NOTA (dealer del form no está en el Excel, solo difiere en nombre)", counts.get("NOTA", 0)],
        ["⚠ PASS con nombre distinto en detalles menores", disclaimer_count],
    ]

    # Desglose por modelo: mismo resumen PASS/FAIL/etc. pero por modelo, para ver de un
    # vistazo si algún modelo en particular tiene más problemas que otro — sin esto había
    # que abrir cada pestaña y contar a mano.
    if multi_modelo:
        summary_data.append(["", ""])
        summary_data.append(["Desglose por Modelo", ""])
        for modelo in modelos_orden:
            m_results = [r for r in results if (r.get("modelo") or "").strip() == modelo]
            m_counts = {}
            for r in m_results:
                m_counts[r.get("status", "FAIL")] = m_counts.get(r.get("status", "FAIL"), 0) + 1
            resumen_txt = " · ".join(
                f"{k}={v}" for k, v in m_counts.items() if v
            ) or "(sin resultados)"
            summary_data.append([f"  {modelo}", f"{len(m_results)} chequeados — {resumen_txt}"])

    # Agregar URLs únicas usadas
    urls_form = sorted(set(r.get("url_form", "") for r in results if r.get("url_form")))
    urls_landing = sorted(set(r.get("url_landing", "") for r in results if r.get("url_landing")))
    if urls_form or urls_landing:
        summary_data.append(["", ""])
        summary_data.append(["URLs procesadas", ""])
        for url in urls_landing:
            summary_data.append(["  Landing", url])
        for url in urls_form:
            summary_data.append(["  Form", url])

    # Avisos de calidad de datos del Excel de dealers (no afectan la comparación en sí).
    # hidden_rows es un dict {nº fila: 'filtrada'|'oculta'} (o, por compatibilidad, un set).
    if hidden_rows:
        if isinstance(hidden_rows, dict):
            filtradas = sorted(r for r, tipo in hidden_rows.items() if tipo == "filtrada")
            ocultas = sorted(r for r, tipo in hidden_rows.items() if tipo != "filtrada")
        else:
            filtradas, ocultas = [], sorted(hidden_rows)
        summary_data.append(["", ""])
        summary_data.append(["⚠ Filas del Excel que no se ven", len(filtradas) + len(ocultas)])
        if filtradas:
            summary_data.append(["  🔎 Filtradas por AutoFilter", len(filtradas)])
            summary_data.append(["     Número de fila(s)", ", ".join(str(r) for r in filtradas)])
        if ocultas:
            summary_data.append(["  🙈 Ocultas a mano", len(ocultas)])
            summary_data.append(["     Número de fila(s)", ", ".join(str(r) for r in ocultas)])
    if hidden_columns:
        summary_data.append(["", ""])
        summary_data.append(["⚠ Columnas OCULTAS en el Excel de dealers", len(hidden_columns)])
        for hc in hidden_columns:
            label = f"  Columna {hc['column']}" + (f" ({hc['header']})" if hc.get("header") else "")
            summary_data.append([label, ""])

    for row_data in summary_data:
        summary_ws.append(row_data)

    # Formato básico de la hoja resumen
    summary_ws.column_dimensions["A"].width = 30
    summary_ws.column_dimensions["B"].width = 60
    for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    # Encabezados en negrita
    for cell in [summary_ws.cell(1, 1), summary_ws.cell(1, 2),
                 summary_ws.cell(4, 1), summary_ws.cell(4, 2)]:
        cell.font = Font(bold=True)

    wb.save(output_path)
    return output_path
