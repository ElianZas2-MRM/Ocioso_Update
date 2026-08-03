"""
base_form_filler.py — Motor central de llenado de formularios (todos los países).
Maneja navegación multi-paso, dropdowns con dependencias, campos de texto, checkboxes,
generación de documentos brasileños, captura de screenshots y escritura de resultados a Excel.
"""
import os
import sys
import time
import re
import shutil
import unicodedata
import random
import json
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Parchear print global para evitar caídas por UnicodeEncodeError en consolas Windows (CP1252)
import builtins
_original_print = builtins.print
def _safe_global_print(*args, **kwargs):
    try:
        _original_print(*args, **kwargs)
    except UnicodeEncodeError:
        try:
            encoding = sys.stdout.encoding or "utf-8"
            safe_args = []
            for arg in args:
                s = str(arg)
                safe_args.append(s.encode(encoding, errors="replace").decode(encoding))
            _original_print(*safe_args, **kwargs)
        except Exception:
            pass
    except Exception:
        pass
builtins.print = _safe_global_print

# Importar directamente
from browser_manager import BrowserManager
from screenshot_manager import ScreenshotManager

from utils.field_id_aliases import VISID_ID_ALIASES

try:
    from utils.popup_logger import popup_log, log_runtime
except Exception:
    def popup_log(title, message, level="ERROR"):
        print(f"[{level}] {title}: {message}")

    def log_runtime(message, level="INFO"):
        print(f"[{level}] {message}")

try:
    from utils.fixed_field_mapping_store import load_effective_country_form_config
except Exception:
    def load_effective_country_form_config(country_name, fallback_config=None):
        fallback_country_fields = dict((fallback_config or {}).get('country_fields', {}) or {})
        return {
            'pais': str(country_name or '').strip(),
            'excel_file': str((fallback_config or {}).get('excel_file') or '').strip(),
            'data_start_index': int((fallback_config or {}).get('data_start_index', 2)),
            'field_mapping': list((fallback_config or {}).get('field_mapping') or []),
            'country_fields': fallback_country_fields,
        }

# Callback global para input manual — se setea desde la UI antes de ejecutar
_global_manual_input_callback = None


def set_global_manual_input_callback(callback):
    global _global_manual_input_callback
    _global_manual_input_callback = callback


class BaseFormFiller:
    """Clase base unificada para todos los formularios de países"""

    _EVENT_ID_PATTERNS = (
        "ocurrió un inconveniente al realizar el envío del formulario",
        "ocorreu um problema ao enviar o formulário neste momento",
        "ocorreu um problema ao enviar o formulario neste momento",
        "lo sentimos, ocurrió",
        "lo siento, ocurrió",
    )

    # Marcadores de URL de los formularios GM hechos por DEV (NO los AEM): el iframe del
    # formulario SIEMPRE tiene un src que contiene uno de estos fragmentos. Cuando una landing
    # tiene varios iframes hay que priorizar SIEMPRE el que matchee estos marcadores, y recién
    # como último recurso caer a cualquier iframe visible.
    # "gm_front" cubre tambien "gm_frontend" (los forms nuevos
    # .../gm_frontend/chevrolet/t3/<pais>/form/<slug>), porque el match es por substring.
    GM_FORM_URL_MARKERS = ("gm_forms", "gm_formns", "gm_front", "gm_admin")

    @staticmethod
    def _iframe_src_of(iframe):
        """src del iframe, tolerante a stale/errores (devuelve "")."""
        try:
            return (iframe.get_attribute("src") or "")
        except Exception:
            return ""

    def _is_gm_form_src(self, src):
        s = (src or "").lower()
        return any(m in s for m in self.GM_FORM_URL_MARKERS)

    def _pick_gm_iframe(self, iframes, expected_url=None):
        """
        Elige el iframe correcto entre varios candidatos.
        Prioridad:
          1. iframe GM (src contiene gm_forms/gm_front/gm_admin) que además matchee expected_url.
          2. cualquier iframe GM visible.
          3. cualquier iframe GM (aunque no esté 'displayed' todavía — lazy).
          4. último recurso: primer iframe visible (marca mismatch aguas arriba).
        Devuelve (iframe|None, es_gm: bool).
        """
        exp = (expected_url or "").strip().lower()
        gm_visible, gm_any, first_visible = [], [], None
        for f in (iframes or []):
            src = self._iframe_src_of(f)
            try:
                disp = f.is_displayed()
            except Exception:
                disp = False
            if disp and first_visible is None:
                first_visible = f
            if self._is_gm_form_src(src):
                gm_any.append((f, src))
                if disp:
                    gm_visible.append((f, src))
        # 1. GM que matchea expected
        if exp:
            for bucket in (gm_visible, gm_any):
                for f, src in bucket:
                    if exp in src.lower():
                        return f, True
        # 2/3. cualquier GM (visible primero)
        if gm_visible:
            return gm_visible[0][0], True
        if gm_any:
            return gm_any[0][0], True
        # 4. último recurso
        return first_visible, False

    # Aliases ID para formularios del estándar visid (coexistencia con forms actuales)
    # Si el ID del mapping no se encuentra en el DOM, se prueba el alias visid.
    # La tabla vive en utils/field_id_aliases.py: la comparte la Validación de Campos.
    _VISID_ID_ALIASES: dict = VISID_ID_ALIASES

    # Adobe AEM Adaptive Form (Guide) — términos / checkbox
    _GUIDE_CHECKBOX_CONTAINER_ID = "guideContainer-rootPanel-guidecheckbox___guide-item"
    _GUIDE_CHECKBOX_INPUT_ID_PREFIX = "guideContainer-rootPanel-guidecheckbox___"
    # El id ...___guide-item es el contenedor; el <button type="submit"> usa ...___widget
    _GUIDE_SUBMIT_CONTAINER_ID = "guideContainer-rootPanel-guidebutton___guide-item"
    _GUIDE_SUBMIT_WIDGET_ID_PREFIX = "guideContainer-rootPanel-guidebutton___"

    DATA_COLUMNS = [
        "model", "firstname", "lastname", "document", "phone", "email",
        "region", "city", "dealer", "purchase_date", "doc_type",
        "patent", "mileage", "acquisition_year",
        "advisor_code", "event", "kit", "chassis", "color", "insurance", "comment"
    ]
    
    def __init__(self, config):
        """
        Configura el form filler base
        
        Args:
            config (dict): Configuración específica del país
        """
        self.config = config
        try:
            from utils.paths import BASE_DIR, DATA_DIR, RESULTS_DIR
            self.BASE_DIR = BASE_DIR
            self.DATA_DIR = DATA_DIR
            self.RESULTADOS_DIR = RESULTS_DIR
        except ImportError:
            if getattr(sys, 'frozen', False):
                self.BASE_DIR = os.path.dirname(sys.executable)
            else:
                self.BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            self.DATA_DIR = os.path.join(self.BASE_DIR, "data")
            self.RESULTADOS_DIR = os.path.join(self.BASE_DIR, "resultados")
        
        os.makedirs(self.DATA_DIR, exist_ok=True)
        os.makedirs(self.RESULTADOS_DIR, exist_ok=True)
        
        # Configurar paths específicos del país
        self.EXCEL_PATH = os.path.join(self.DATA_DIR, config['excel_file'])
        _b = str(config.get('browser', '') or '').strip().lower()
        _dev = {"chrome": "Chrome", "firefox": "Firefox", "edge": "Edge"}.get(_b, _b.capitalize() if _b else "")
        _dev_sfx = f"_{_dev}" if _dev else ""
        prefix = "Automatizacion_" if config.get('is_scheduled') else "resultados_"
        prefix_ss = "Automatizacion_screenshots_" if config.get('is_scheduled') else "screenshots_"
        self.SCREENSHOT_BASENAME = f"{prefix_ss}{config['pais']}{_dev_sfx}"
        self.RESULTADOS_BASENAME = f"{prefix}{config['pais']}{_dev_sfx}"
        
        # Inicializar en setup_directories_and_files
        self.SCREENSHOT_DIR = None
        self.RESULTADOS_PATH = None
        self.RUN_NUMBER = None
        # Resumen por formulario de la última corrida (lo consume la UI para el modal detallado).
        self.run_summary = {"ok": 0, "fail": 0, "total": 0, "fail_rows": []}
        self.driver = None
        self.screenshot_manager = None
        
        effective_country_config = load_effective_country_form_config(config.get('pais'), fallback_config=config)

        # Campos específicos del país
        self.country_fields = effective_country_config.get('country_fields', config.get('country_fields', {}))
        self.field_mapping = effective_country_config.get('field_mapping', config.get('field_mapping', []))
        self.data_columns = config.get('data_columns', list(self.DATA_COLUMNS))
        self.data_start_index = effective_country_config.get('data_start_index', config.get('data_start_index', 2))
        self.effective_data_keys = self._build_effective_data_keys()
        # Preferencias SI/NO de checkbox del lead actual (se recarga por fila del Excel)
        self.checkbox_prefs = {}
        # Cargar timeouts desde config_global.json (con defaults si no existe)
        _global_timeouts = {}
        try:
            import json as _json
            _cfg_path = os.path.join(self.BASE_DIR, "json", "config_global.json")
            with open(_cfg_path, "r", encoding="utf-8") as _f:
                _global_timeouts = _json.load(_f).get("timeouts", {})
        except Exception:
            pass

        # Defaults auto-step (pueden sobrescribirse en config)
        if 'auto_step_max_iterations' not in self.config:
            self.config['auto_step_max_iterations'] = 15
        if 'wait_kits_after_model' not in self.config:
            self.config['wait_kits_after_model'] = False
        if 'dependency_dropdown_timeout' not in self.config:
            self.config['dependency_dropdown_timeout'] = float(_global_timeouts.get('dependency_dropdown', 8.0))
        if 'dependency_dropdown_poll_interval' not in self.config:
            self.config['dependency_dropdown_poll_interval'] = 0.2
        if 'dependency_selection_retries' not in self.config:
            self.config['dependency_selection_retries'] = 2
        # Timeouts globales accesibles para métodos internos
        self._t_page_load = int(_global_timeouts.get('page_load', 20))
        self._t_element_wait = float(_global_timeouts.get('element_wait', 8.0))
        self._t_form_submit = float(_global_timeouts.get('form_submit', 10.0))
        self._t_step_change = float(_global_timeouts.get('step_change', 5.0))
        # Registro por fila de IDs de campos seleccionados (enfocado en dropdowns)
        self.current_row_field_values = {}
        # Campos nuevos detectados durante la ejecución actual
        self._campos_nuevos_detectados = []
        # Callback para solicitar valores manuales (se inyecta desde la UI)
        self.manual_input_callback = _global_manual_input_callback

    def _scroll_element_into_view(self, element):
        """Scrolls the element into view. If focused inside an iframe (especially on Safari/macOS),
        it also scrolls the parent window so that the element is fully visible in the viewport."""
        if not element:
            return
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", element)
            time.sleep(0.1)
        except Exception:
            pass

        try:
            # Comprobar si estamos dentro de un iframe
            in_iframe = self.driver.execute_script("return window.self !== window.top;")
            if in_iframe:
                rect = self.driver.execute_script(
                    "const r = arguments[0].getBoundingClientRect(); return {top: r.top, height: r.height};",
                    element
                )
                element_top_in_iframe = rect["top"]
                element_height = rect["height"]

                # Obtener el iframe actual
                target_iframe = None
                if hasattr(self, "screenshot_manager") and self.screenshot_manager and getattr(self.screenshot_manager, "current_frame", None):
                    target_iframe = self.screenshot_manager.current_frame
                
                # Cambiar temporalmente al contenido principal
                self.driver.switch_to.default_content()

                try:
                    if not target_iframe:
                        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                        for iframe in iframes:
                            if iframe.is_displayed():
                                target_iframe = iframe
                                break
                    
                    if target_iframe:
                        # Obtener la posición del iframe en el parent
                        iframe_rect = self.driver.execute_script(
                            "const r = arguments[0].getBoundingClientRect(); return {top: r.top, yOffset: window.pageYOffset};",
                            target_iframe
                        )
                        iframe_top = iframe_rect["top"]
                        parent_y_offset = iframe_rect["yOffset"]

                        viewport_height = self.driver.execute_script("return window.innerHeight;")
                        
                        target_y_on_parent = parent_y_offset + iframe_top + element_top_in_iframe
                        # Centrar el elemento en la pantalla
                        scroll_y = target_y_on_parent - (viewport_height / 2) + (element_height / 2)
                        scroll_y = max(0, scroll_y)

                        self.driver.execute_script(f"window.scrollTo(0, {scroll_y});")
                        time.sleep(0.15)
                finally:
                    # Siempre volver al iframe
                    if target_iframe:
                        self.driver.switch_to.frame(target_iframe)
        except Exception as e:
            print(f"[DEBUG] Error en _scroll_element_into_view: {e}")

    def begin_row_tracking(self):
        """Resetea el tracking de IDs/valores para la fila actual."""
        self.current_row_field_values = {}
        self._campos_sin_mapeo_exitoso = []
        self._campos_dropdown_no_encontrados = []
        self._campos_sin_valor_asignado = []
        self._ids_din_cb_map = None  # re-elige SI/NO random por fila si hay varios valores
        self._current_step = 1
        self._ty_cta = ""
        self._link_issue = "-"
        self._link_issue_present = False
        self._datos_vs_excel = "-"
        self._datos_mismatch = False

    def _normalize_field_id(self, field_id):
        """Normaliza un field_id para uso como clave/columna."""
        if field_id is None:
            return ""
        normalized = str(field_id).strip()
        return normalized

    def _sanitize_url(self, value):
        """Limpia espacios y saltos de línea al inicio/fin de una URL."""
        if not isinstance(value, str):
            return value
        return value.strip()

    def _record_field_value(self, field_id, value):
        """Guarda el valor final aplicado a un campo ID para la fila actual."""
        normalized_id = self._normalize_field_id(field_id)
        if not normalized_id:
            return

        if isinstance(value, (list, tuple, set)):
            text_value = " | ".join(str(v).strip() for v in value if str(v).strip())
        else:
            text_value = "" if value is None else str(value).strip()

        if text_value == "":
            return

        step = getattr(self, "_current_step", 1)
        key = f"Paso{step}::{normalized_id}"
        self.current_row_field_values[key] = text_value

    # Valores de ?model= que NO son modelos (son servicios/secciones): no deben registrarse
    # como "modelo elegido". Si el form igual tiene un dropdown de modelos, ese valor gana.
    _NON_MODEL_URL_TOKENS = (
        "posventa", "postventa", "pos venta", "onstar", "servicio", "service",
        "agendamiento", "seminuevos", "seminovos", "revision", "revisao",
        "testdrive", "test drive", "cotizacion", "cotizar", "contacto", "contato",
        "oferta", "raq", "suscripcion", "renovacion", "financiamiento", "repuestos",
        "accesorios", "acessorios", "acessilab", "garantia", "posventa gral", "gral",
    )

    def _model_from_form_url(self):
        """Extrae el ?model= de la URL del FORM (iframe), no de la landing. Devuelve '' si el
        valor es un servicio/sección (no un modelo real)."""
        from urllib.parse import urlsplit, parse_qs, unquote
        for url in (getattr(self, "expected_form_url", "") or "",
                    getattr(self, "_url_form_encontrado", "") or "",
                    self.driver.current_url if self.driver else ""):
            try:
                raw = parse_qs(urlsplit(url).query).get("model", [""])[0]
                val = unquote(raw).replace("+", " ").strip()
            except Exception:
                val = ""
            if not val:
                continue
            norm = self._normalize_text(val)
            if any(tok in norm for tok in (self._normalize_text(t) for t in self._NON_MODEL_URL_TOKENS)):
                continue  # es un servicio, no un modelo
            return val
        return ""

    def _tracked_model_value(self):
        """Modelo elegido en un dropdown del form (si se registró en esta fila)."""
        val = ""
        for key, v in (self.current_row_field_values or {}).items():
            raw_id = key.split("::", 1)[1] if "::" in key else key
            if raw_id in ("models", "model") and str(v).strip():
                val = str(v).strip()
        return val

    def _effective_model_value(self):
        """Modelo efectivo para la columna 'Modelo' de resultados:
        1) el elegido en el dropdown de modelos del form (si hay), o
        2) el ?model= de la URL del form (si es un modelo real, no un servicio)."""
        return self._tracked_model_value() or self._model_from_form_url()

    @staticmethod
    def _short_fail_reason(result_text):
        """Motivo corto del fallo para el resumen rápido (modal/email). p.ej. 'form ausente'."""
        t = (result_text or "").lower()
        if "form no inserto" in t or "formulario ausente" in t:
            return "form no inserto"
        if "no coincide con el esperado" in t:
            return ("form inserto no coincide (se envió lead igualmente)"
                    if "se envió lead igualmente" in t else
                    "form inserto no coincide (lead no enviado)")
        if "formulario incorrecto" in t or "distinto al esperado" in t:
            return "form incorrecto"
        if "error visual" in t or "campos sin completar" in t:
            _det = result_text.split("Error visual", 1)[-1].lstrip(": ").strip()
            _det = _det.split("(TY Page")[0].strip(" |")
            if _det:
                return f"validación del form: {_det[:120]}"
        if "link issue typ" in t:
            return "link issue TYP"
        if "error landing" in t or "404" in t:
            return "landing 404"
        if "ty page" in t or "confirmación ty" in t or "confirmacion ty" in t or "sin confirmaci" in t:
            return "sin TYP"
        if "event_id" in t:
            return "event_id"
        if "dropdown no encontrado" in t:
            return "dropdown no encontrado"
        if "campos sin completar" in t:
            return "campos sin completar"
        if "error completando" in t or "error general" in t:
            return "error al completar"
        return "error"

    def _record_model_from_url_if_missing(self):
        """Si ningún select de Modelo quedó registrado en la fila, usar el ?model= del FORM."""
        for key in self.current_row_field_values:
            raw_id = key.split("::", 1)[1] if "::" in key else key
            if raw_id in ("models", "model"):
                return  # ya se registró un modelo (del Excel o aleatorio)
        url_model = self._model_from_form_url()
        if not url_model:
            return
        model_id = "model"
        for fc in (self.field_mapping or []):
            if isinstance(fc, dict):
                fid = fc.get("id")
                fid = fid[0] if isinstance(fid, list) else fid
                if fid in ("models", "model"):
                    model_id = fid
                    break
        self._record_field_value(model_id, url_model)
        print(f"🚗 Modelo (de la URL ?model=) = {url_model}")

    # ── Snapshot real del form antes de enviar ────────────────────────────────
    # El tracking normal guarda lo que se *intentó* escribir en cada paso. Si después el
    # form re-renderiza, un dropdown se resetea, o se vuelve a elegir un modelo en un paso
    # posterior, lo que viaja en el lead deja de coincidir con lo trackeado. Antes del
    # click en Enviar se relee el DOM y se pisa el tracking con el valor efectivo: el
    # Excel de resultados tiene que reflejar SIEMPRE lo último que quedó en el form.
    _SNAPSHOT_JS = r"""
    var out = [];
    var els = document.querySelectorAll("input, select, textarea");
    for (var i = 0; i < els.length; i++) {
        var e = els[i];
        var tag = e.tagName.toLowerCase();
        var type = (e.type || "").toLowerCase();
        if (tag === "input" && (type === "hidden" || type === "submit" ||
                                type === "button" || type === "reset" || type === "file")) continue;
        if (type === "checkbox" || type === "radio") continue;
        var id = e.id || e.getAttribute("name") || "";
        if (!id) continue;
        var value = "", text = "";
        if (tag === "select") {
            var sel = e.multiple
                ? [].filter.call(e.options, function(o){return o.selected;})
                : (e.selectedIndex >= 0 ? [e.options[e.selectedIndex]] : []);
            value = sel.map(function(o){return o.value;}).join(" | ");
            text  = sel.map(function(o){return (o.text||"").trim();}).join(" | ");
        } else {
            value = e.value || "";
            text  = value;
        }
        out.push({id: id, realId: e.id || "", name: e.getAttribute("name") || "",
                  tag: tag, type: type, value: value, text: text,
                  visible: !!(e.getClientRects().length && getComputedStyle(e).visibility !== "hidden")});
    }
    return out;
    """

    def _snapshot_form_state(self):
        """Estado REAL de todos los campos del form (un solo JS call). {id: entry}."""
        try:
            raw = self.driver.execute_script(self._SNAPSHOT_JS) or []
        except Exception as e:
            print(f"No se pudo leer el estado final del form: {e}")
            return {}
        snap = {}
        for entry in raw:
            if not isinstance(entry, dict):
                continue
            for key in (entry.get("realId"), entry.get("name"), entry.get("id")):
                key = (key or "").strip()
                if not key:
                    continue
                prev = snap.get(key)
                if prev is None:
                    snap[key] = entry
                elif not self._entry_value(prev) and self._entry_value(entry):
                    snap[key] = entry
        return snap

    def _set_input_value_js(self, el, value):
        """Setea el value con el setter nativo del prototipo (React/Angular no registran
        la asignación directa de el.value)."""
        try:
            self.driver.execute_script(
                "var e=arguments[0], v=arguments[1];"
                "var proto = e instanceof window.HTMLTextAreaElement"
                "    ? window.HTMLTextAreaElement.prototype : window.HTMLInputElement.prototype;"
                "var d = Object.getOwnPropertyDescriptor(proto, 'value');"
                "if (d && d.set) { d.set.call(e, v); } else { e.value = v; }"
                "e.dispatchEvent(new Event('input',{bubbles:true}));"
                "e.dispatchEvent(new Event('change',{bubbles:true}));",
                el, value,
            )
        except Exception:
            pass

    def _hard_clear_input(self, el):
        """Vacía un input y confirma que quedó vacío.

        `clear()` puede no vaciar nada y tampoco lanzar excepción (campos con máscara o
        manejados por React). Si eso pasa, el send_keys posterior CONCATENA y se envía
        'ApellidoApellido' o un email duplicado — el lead viaja con datos corruptos.
        """
        def _current():
            try:
                return el.get_attribute("value") or ""
            except Exception:
                return ""

        for intento in range(3):
            if not _current():
                return True
            if intento == 0:
                try:
                    el.clear()
                except Exception:
                    pass
            else:
                self._set_input_value_js(el, "")
            time.sleep(0.05)
        return not _current()

    @staticmethod
    def _entry_value(entry):
        if not entry:
            return ""
        return (entry.get("text") or "").strip() or (entry.get("value") or "").strip()

    def _sync_tracked_with_dom_before_submit(self, form_data=None):
        """Pisa current_row_field_values con el estado real del form justo antes de enviar.

        - Cada `PasoN::id` pasa a valer lo que ese campo tiene AHORA en el DOM: si un
          dropdown (modelo, fecha estimada, ...) se re-eligió más tarde, queda el último.
        - Los campos con valor que nadie trackeó se agregan como `Final::<id>`.
        - Compara contra lo pedido en el Excel y deja el detalle en self._datos_vs_excel.
        """
        snap = self._snapshot_form_state()
        if not snap:
            return

        def _resolve(ids):
            for cand in ids:
                cand = str(cand or "")
                if cand and self._entry_value(snap.get(cand)):
                    return snap.get(cand)
            for cand in ids:
                cand = str(cand or "")
                if not cand:
                    continue
                for key, entry in snap.items():
                    if key.startswith(cand + "-") and self._entry_value(entry):
                        return entry
            return None

        covered = set()
        tracked_ids = set()
        for key in list(self.current_row_field_values.keys()):
            raw_id = key.split("::", 1)[1] if "::" in key else key
            tracked_ids.add(raw_id)
            entry = _resolve([raw_id])
            if not entry:
                continue
            eid = entry.get("realId") or entry.get("name") or ""
            if eid:
                covered.add(eid)
            real_value = self._entry_value(entry)
            if not real_value:
                continue
            old_value = str(self.current_row_field_values.get(key, "")).strip()
            if self._normalize_text(old_value) != self._normalize_text(real_value):
                self.current_row_field_values[key] = real_value
                print(f"🔄 {raw_id}: el form quedó con '{real_value}' (trackeado era "
                      f"'{old_value}') — se registra el valor real")

        # Campos con valor que nadie trackeó (randoms de selects no mapeados, etc.)
        for key, entry in snap.items():
            eid = entry.get("realId") or entry.get("name") or ""
            if not eid or eid != key or eid in covered or eid in tracked_ids:
                continue
            # Solo selects: un input de texto sin trackear suele ser ruido de la landing
            # (buscador, newsletter); un select con valor sí es una elección que viaja.
            if entry.get("tag") != "select":
                continue
            value = self._entry_value(entry)
            if not value or self._is_placeholder_text(value):
                continue
            self.current_row_field_values[f"Final::{eid}"] = value
            print(f"📎 '{eid}' quedó en '{value}' sin estar trackeado — registrado como Final::{eid}")

        # ── Comparación contra lo pedido en el Excel ──────────────────────────
        # Sólo campos de elección (selects): en los de texto el propio flujo transforma
        # el dato a propósito (CPF/CNPJ/CEP regenerados en Brasil, documento saneado en
        # Perú, recorte por maxlength) y compararlos daría avisos falsos todo el tiempo.
        mismatches = []
        for fc in (self.field_mapping or []):
            if not isinstance(fc, dict):
                continue
            if fc.get("type", "text") != "select":
                continue
            fid_raw = fc.get("id")
            ids = fid_raw if isinstance(fid_raw, list) else [fid_raw]
            fname = fc.get("name", str(fid_raw))
            try:
                expected = self._resolve_field_value(form_data or {}, fc)
            except Exception:
                expected = ""
            expected = "" if expected is None else str(expected).strip()
            if not expected or self._is_placeholder_text(expected):
                continue
            entry = _resolve(ids)
            if not entry:
                continue
            real_value = self._entry_value(entry)
            if not real_value:
                mismatches.append(f"{fname}: pedido '{expected}' pero quedó vacío")
                continue
            exp_n, real_n = self._normalize_text(expected), self._normalize_text(real_value)
            if exp_n == real_n or exp_n in real_n or real_n in exp_n:
                continue
            mismatches.append(f"{fname}: pedido '{expected}' → quedó '{real_value}'")

        if mismatches:
            self._datos_vs_excel = " ; ".join(mismatches)
            self._datos_mismatch = True
            print(f"⚠️ Datos distintos a los pedidos en el Excel: {self._datos_vs_excel}")
        else:
            self._datos_vs_excel = "OK"
            self._datos_mismatch = False

    def _get_selected_text_for_select(self, select_element):
        """Obtiene el texto seleccionado actual de un select simple."""
        try:
            selected_option = Select(select_element).first_selected_option
            if selected_option:
                return (selected_option.text or "").strip()
        except Exception:
            return ""
        return ""

    def _is_option_disabled(self, option_element):
        """Determina si una option está deshabilitada de forma robusta."""
        try:
            disabled_attr = option_element.get_attribute("disabled")
            if disabled_attr is None:
                return False
            normalized = str(disabled_attr).strip().lower()
            # HTML boolean attrs suelen venir como "true", "disabled" o "".
            # Algunos frameworks pueden devolver "false" y no debe bloquearse.
            return normalized in ("", "true", "disabled")
        except Exception:
            return False

    def _get_valid_select_options(self, select_element):
        """Retorna opciones seleccionables (no placeholder y no disabled)."""
        valid_options = []
        try:
            select = Select(select_element)
        except Exception:
            return valid_options

        for opt in select.options:
            opt_text = (opt.text or "").strip()
            if not opt_text:
                continue
            if self._is_option_disabled(opt):
                continue
            if self._is_placeholder_text(opt_text):
                continue
            valid_options.append(opt)

        return valid_options

    def _resolve_data_key_from_field_config(self, field_config):
        """Obtiene el data_key efectivo de un field_config (directo o por data_index)."""
        if not isinstance(field_config, dict):
            return None

        data_key = field_config.get("data_key")
        if data_key:
            return data_key

        data_index = field_config.get("data_index")
        if isinstance(data_index, int) and 0 <= data_index < len(self.data_columns):
            return self.data_columns[data_index]

        return None

    def _build_effective_data_keys(self):
        """
        Construye el orden efectivo de data_keys según field_mapping.
        Prioriza requested_data_index y luego data_index, para alinear lectura/escritura del Excel.
        """
        ordered = []
        if isinstance(self.field_mapping, list) and self.field_mapping:
            candidates = []
            for pos, field_config in enumerate(self.field_mapping):
                if not isinstance(field_config, dict):
                    continue
                data_key = self._resolve_data_key_from_field_config(field_config)
                if not data_key:
                    continue

                requested_idx = field_config.get("requested_data_index")
                data_idx = field_config.get("data_index")

                if not isinstance(requested_idx, int) or requested_idx < 0:
                    requested_idx = data_idx if isinstance(data_idx, int) and data_idx >= 0 else pos
                if not isinstance(data_idx, int) or data_idx < 0:
                    data_idx = requested_idx

                candidates.append((requested_idx, data_idx, pos, data_key))

            candidates.sort(key=lambda item: (item[0], item[1], item[2]))
            for _, _, _, data_key in candidates:
                if data_key not in ordered:
                    ordered.append(data_key)

        if not ordered:
            ordered = list(self.data_columns)
        return ordered

    def _build_id_to_data_key_map(self):
        """Construye mapa de id de campo -> data_key según field_mapping del país."""
        id_to_data_key = {}

        for field_config in self.field_mapping or []:
            data_key = self._resolve_data_key_from_field_config(field_config)
            if not data_key:
                continue

            field_id = field_config.get("id")
            field_ids = field_id if isinstance(field_id, list) else [field_id]

            for fid in field_ids:
                normalized_id = self._normalize_field_id(fid)
                if not normalized_id:
                    continue
                if normalized_id not in id_to_data_key:
                    id_to_data_key[normalized_id] = data_key

        return id_to_data_key

    def _get_sheet_column_index_for_data_key(self, data_key):
        """Devuelve índice de columna Excel (1-based) para un data_key de entrada."""
        if not data_key:
            return None

        try:
            data_pos = self.effective_data_keys.index(data_key)
        except ValueError:
            return None

        # data_start_index es 0-based dentro de la fila (URL, Formulario, ...)
        return self.data_start_index + data_pos + 1

    @staticmethod
    def _delete_sheet_column_if_header(sheet, header_title):
        """Si la fila 1 tiene una celda con valor exacto header_title, borra esa columna."""
        target = str(header_title or "").strip()
        if not target:
            return False
        for cell in sheet[1]:
            if str(cell.value or "").strip() == target:
                sheet.delete_cols(int(cell.column))
                return True
        return False

    def write_tracked_fields_to_sheet(self, sheet, row_index):
        """Escribe en Excel columnas dinámicas por ID con los valores registrados de la fila."""
        if not self.current_row_field_values:
            return

        headers = [cell.value for cell in sheet[1] if cell.value]
        id_to_data_key = self._build_id_to_data_key_map()
        nombres_ids_dinamicos = self._cargar_nombres_ids_dinamicos()

        def _resolver_nombre_columna_dinamica(field_id):
            return nombres_ids_dinamicos.get(field_id, field_id)

        for field_id, field_value in sorted(self.current_row_field_values.items()):
            # field_id tiene formato "PasoN::raw_id" — extraer raw_id para lookup en mapping
            if "::" in field_id:
                paso_prefix, raw_id = field_id.split("::", 1)
            else:
                paso_prefix, raw_id = "", field_id

            mapped_data_key = id_to_data_key.get(raw_id)
            mapped_col_idx = self._get_sheet_column_index_for_data_key(mapped_data_key)

            # Si el ID existe en el mapping del país, columna con prefijo de paso
            if mapped_col_idx:
                resolved_label = _resolver_nombre_columna_dinamica(raw_id)
                col_name = f"{paso_prefix}::{resolved_label}" if paso_prefix else resolved_label
                if col_name not in headers:
                    headers.append(col_name)
                    sheet.cell(row=1, column=len(headers)).value = col_name
                col_idx = headers.index(col_name) + 1
                sheet.cell(row=row_index, column=col_idx).value = field_value
                continue

            resolved_label = _resolver_nombre_columna_dinamica(raw_id)
            col_name = f"{paso_prefix}::{resolved_label}" if paso_prefix else f"ID::{resolved_label}"
            if col_name not in headers:
                headers.append(col_name)
                sheet.cell(row=1, column=len(headers)).value = col_name
                print(f"Columna dinámica '{col_name}' agregada al Excel")

            col_idx = headers.index(col_name) + 1
            sheet.cell(row=row_index, column=col_idx).value = field_value

    def _cargar_nombres_ids_dinamicos(self):
        """Carga nombres de campo para IDs dinámicos aplicables al país actual."""
        nombres = {}
        pais_actual = str(self.config.get("pais") or "").strip()

        def _normalizar_paises(raw):
            if raw is None:
                return []
            if isinstance(raw, str):
                candidatos = [raw]
            elif isinstance(raw, (list, tuple, set)):
                candidatos = list(raw)
            else:
                return []
            salida = []
            for item in candidatos:
                texto = str(item).strip()
                if texto and texto not in salida:
                    salida.append(texto)
            return salida

        try:
            path = os.path.join(self.BASE_DIR, "json", "ids_dinamicos.json")
            if not os.path.exists(path):
                return nombres

            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            if isinstance(data, dict) and isinstance(data.get("entries"), list):
                for entry in data.get("entries", []):
                    if not isinstance(entry, dict):
                        continue
                    entry_id = str(entry.get("id") or "").strip()
                    if not entry_id or entry_id in nombres:
                        continue

                    paises_entry = _normalizar_paises(entry.get("paises", entry.get("countries")))
                    if paises_entry and pais_actual not in paises_entry:
                        continue

                    nombre = str(
                        entry.get("nombre_campo")
                        or entry.get("nombre")
                        or entry.get("campo")
                        or ""
                    ).strip()
                    if nombre:
                        nombres[entry_id] = nombre
            elif isinstance(data, dict):
                # Legacy: {id: valor} o {id: {nombre_campo, valor}}
                for entry_id, raw_value in data.items():
                    if entry_id in {"version", "entries"}:
                        continue
                    entry_id = str(entry_id).strip()
                    if not entry_id or entry_id in nombres:
                        continue

                    if isinstance(raw_value, dict):
                        nombre = str(
                            raw_value.get("nombre_campo")
                            or raw_value.get("nombre")
                            or raw_value.get("campo")
                            or ""
                        ).strip()
                        if nombre:
                            nombres[entry_id] = nombre
        except Exception as e:
            print(f" No se pudo cargar nombres de ids_dinamicos: {e}")

        return nombres
        
    # Valores aceptados en una columna del Excel que apunta a un checkbox (por name o id).
    CHECKBOX_SI = {"si", "sí", "yes", "true", "1", "x", "marcar", "on"}
    CHECKBOX_NO = {"no", "false", "0", "off", "desmarcar"}

    def build_checkbox_prefs(self, raw_headers, row):
        """
        Preferencias de checkbox tomadas del Excel: una columna cuyo encabezado es el
        `name` (o el `id`) del checkbox, con valor SI/NO. Ej: columna "test-drive" = "NO".

        Se resuelve después contra el DOM por name/id, así que una columna con SI/NO que
        no corresponda a ningún checkbox se ignora sola.
        """
        prefs = {}
        for idx, header in enumerate(raw_headers or []):
            key = str(header or "").strip().lower()
            if not key or idx >= len(row):
                continue
            value = str(row[idx] or "").strip().lower()
            if not value:
                continue
            if value in self.CHECKBOX_SI:
                prefs[key] = True
            elif value in self.CHECKBOX_NO:
                prefs[key] = False
        return prefs

    def _checkbox_pref_for(self, lower_name, checkbox_id):
        """Preferencia para este checkbox: primero el Excel (por name o id), si no hay
        columna, IDs dinámicos con valor SI/NO. None = sin preferencia."""
        prefs = getattr(self, "checkbox_prefs", None) or {}
        pref = prefs.get(lower_name)
        if pref is None:
            pref = prefs.get((checkbox_id or "").strip().lower())
        if pref is None:
            pref = self._checkbox_pref_dinamica(lower_name, checkbox_id)
        return pref

    def _checkbox_pref_dinamica(self, lower_name, checkbox_id):
        """SI/NO desde ids_dinamicos.json para checkboxes sin columna en el Excel.
        Si el ID tiene varios valores SI/NO cargados, elige uno al azar por fila."""
        try:
            if getattr(self, "_ids_din_cb_map", None) is None:
                _map = {}
                for _id, _vals in (self._cargar_ids_dinamicos() or {}).items():
                    _sino = [str(v).strip().lower() for v in (_vals or [])]
                    _sino = [v for v in _sino if v in self.CHECKBOX_SI or v in self.CHECKBOX_NO]
                    if not _sino:
                        continue
                    _map[_id.strip().lower()] = random.choice(_sino) in self.CHECKBOX_SI
                self._ids_din_cb_map = _map
            for _key in (lower_name, (checkbox_id or "").strip().lower()):
                if _key and _key in self._ids_din_cb_map:
                    return self._ids_din_cb_map[_key]
        except Exception:
            pass
        return None

    def _uncheck_checkbox(self, checkbox_id, name_attr):
        """Destilda un checkbox (el Excel pidió NO) y dispara los eventos del form."""
        try:
            self.driver.execute_script(
                """
                var id = arguments[0], name = arguments[1];
                var cb = (id && document.getElementById(id)) ||
                         (name && document.querySelector('input[type="checkbox"][name="' + name + '"]'));
                if (!cb || !cb.checked) return;
                cb.checked = false;
                cb.dispatchEvent(new Event('click',  {bubbles:true}));
                cb.dispatchEvent(new Event('change', {bubbles:true}));
                cb.checked = false;   // por si algún handler lo volvió a marcar
                """,
                checkbox_id, name_attr,
            )
        except Exception:
            pass

    def _tiene_campos_obligatorios(self):
        """
        ¿El formulario tiene al menos un campo obligatorio visible?

        Si no tiene ninguno, clickear "Enviar" con el form vacío NO dispara validaciones:
        lo manda de una y genera un lead basura. Es el caso del Libro de Reclamaciones de
        Perú (form-reclamos), donde todos los campos son voluntarios.
        """
        try:
            return bool(self.driver.execute_script("""
                var campos = document.querySelectorAll(
                    'input[required], select[required], textarea[required],' +
                    '[aria-required="true"]');
                for (var i = 0; i < campos.length; i++) {
                    var e = campos[i];
                    if (e.type === 'hidden' || e.disabled) continue;
                    if (e.getClientRects().length) return true;
                }
                return false;
            """))
        except Exception:
            return True   # ante la duda, comportarse como siempre

    def _revalidar_campos_llenos(self):
        """
        Hace que el formulario limpie SOLO los mensajes de error que ya no corresponden.

        No borra nada a mano: dispara en cada campo ya completo los mismos eventos que
        genera una persona escribiendo (keyup / focusout / blur) y deja que el form decida.
        Hace falta porque el "click enviar vacío" pinta las validaciones ANTES de llenar, y
        el llenado por JS setea el .value sin disparar keyup — que es justo el evento con el
        que jquery-validation borra el mensaje. Resultado: el campo quedaba válido pero con
        el cartelito rojo puesto, y la captura mostraba un falso error.

        Los campos vacíos no se tocan: ahí el error es legítimo y tiene que verse.
        """
        try:
            revalidados = self.driver.execute_script(r"""
                var revalidados = 0;
                var campos = document.querySelectorAll('input, select, textarea');
                for (var i = 0; i < campos.length; i++) {
                    var el = campos[i];
                    if (el.type === 'hidden' || el.disabled) continue;
                    if (!el.getClientRects().length) continue;              // paso oculto
                    if (!el.value || !String(el.value).trim()) continue;    // vacío: error real
                    el.dispatchEvent(new Event('keyup',    {bubbles: true}));
                    el.dispatchEvent(new Event('focusout', {bubbles: true}));
                    el.dispatchEvent(new Event('blur',     {bubbles: true}));
                    revalidados++;
                }
                // jquery-validation: su propia API para revalidar campo por campo
                try {
                    var $ = window.jQuery;
                    if ($) {
                        $('form').each(function () {
                            var v = $(this).data('validator');
                            if (!v) return;
                            $(this).find('input, select, textarea').each(function () {
                                if (this.type === 'hidden' || this.disabled) return;
                                if (!this.getClientRects().length) return;
                                if (!this.value || !String(this.value).trim()) return;
                                try { v.element(this); } catch (e) {}
                            });
                        });
                    }
                } catch (e) {}
                return revalidados;
            """)
            if revalidados:
                print(f" ✓ {revalidados} campos revalidados (limpieza de errores ya resueltos)")
        except Exception as e:
            print(f" Revalidación de campos: error no crítico — {e}")

    def extract_form_data(self, row):
        """Convierte una fila del Excel en un diccionario normalizado"""
        row_list = list(row) if row is not None else []
        start = self.data_start_index
        values = list(row_list[start:]) if len(row_list) > start else []
        
        # LOGS DE DEPURACIÓN PARA EXCEL
        print(f"[DEBUG-EXCEL] Fila completa del Excel (en crudo): {row_list}")
        print(f"[DEBUG-EXCEL] data_start_index configurada: {start}")
        print(f"[DEBUG-EXCEL] Valores extraídos (desde la columna index {start}): {values}")
        
        normalized = {}
        for idx, key in enumerate(self.effective_data_keys):
            normalized[key] = values[idx] if idx < len(values) else ""

        # Mapeo explícito por ID (sin depender de nombre/descripción de columna).
        id_values = {}
        ordered_mapping = []
        for pos, field_config in enumerate(self.field_mapping or []):
            if not isinstance(field_config, dict):
                continue
            requested_idx = field_config.get("requested_data_index")
            data_idx = field_config.get("data_index")
            if not isinstance(requested_idx, int) or requested_idx < 0:
                requested_idx = data_idx if isinstance(data_idx, int) and data_idx >= 0 else pos
            if not isinstance(data_idx, int) or data_idx < 0:
                data_idx = requested_idx
            ordered_mapping.append((requested_idx, data_idx, pos, field_config))

        ordered_mapping.sort(key=lambda item: (item[0], item[1], item[2]))
        for col_idx, (_, _, _, field_config) in enumerate(ordered_mapping):
            value = values[col_idx] if col_idx < len(values) else ""
            field_id = field_config.get("id")
            field_ids = field_id if isinstance(field_id, list) else [field_id]
            for fid in field_ids:
                normalized_id = self._normalize_field_id(fid)
                if normalized_id and normalized_id not in id_values:
                    id_values[normalized_id] = value

        normalized["__by_id"] = id_values
        print(f"[DEBUG-EXCEL] Diccionario normalizado final: {normalized}")
        return normalized

    def get_form_value(self, form_data, key):
        """Obtiene el valor de un campo respetando listas antiguas"""
        if isinstance(form_data, dict):
            return form_data.get(key, "")

        if key in self.data_columns:
            idx = self.data_columns.index(key)
            if idx < len(form_data):
                return form_data[idx]
        return ""

    def get_form_value_str(self, form_data, key):
        """Obtiene el valor del campo como cadena"""
        value = self.get_form_value(form_data, key)
        return "" if value is None else str(value)

    def _normalize_resolved_value(self, value, field_config):
        if value is None:
            return ""
            
        val_str = str(value).strip()
        # Eliminar decimal .0 si se leyó como float
        if val_str.endswith(".0"):
            val_str = val_str[:-2]
            
        # Obtener los identificadores del campo
        field_id = str(field_config.get("id") or "").lower()
        field_name = str(field_config.get("name") or "").lower()
        
        # Verificar si es un campo de teléfono o celular
        is_phone = any(x in field_id or x in field_name for x in ("phone", "tel", "celular", "movil", "telephone"))
        
        # Verificar si es un campo de documento de Uruguay/Paraguay (ci, dni, rut, documento)
        is_document = any(x in field_id or x in field_name for x in ("document", "ci", "dni", "rut", "cedula"))
        
        digits = "".join(c for c in val_str if c.isdigit())
        if not digits:
            return val_str
            
        pais = (self.config.get("pais") or "").lower()
        
        if is_phone:
            if "paraguay" in pais:
                # 10 dígitos, debe empezar con 09
                if len(digits) == 9 and digits.startswith("9"):
                    return "0" + digits
            elif "uruguay" in pais:
                # 9 dígitos, debe empezar con 09
                if len(digits) == 8 and digits.startswith("9"):
                    return "0" + digits
                    
        elif is_document:
            if "uruguay" in pais:
                # Cédula de Uruguay tiene 8 dígitos. Si tiene 7, agregarle un 0 al inicio.
                if len(digits) == 7:
                    return "0" + digits
                
        return val_str

    def _resolve_field_value(self, form_data, field_config):
        """Resuelve el valor a usar para un campo según su mapping"""
        val = ""
        if isinstance(form_data, dict):
            by_id = form_data.get("__by_id", {}) if isinstance(form_data.get("__by_id"), dict) else {}
            resolved_id = field_config.get("__resolved_id") or field_config.get("id")
            candidate_ids = resolved_id if isinstance(resolved_id, list) else [resolved_id]
            found = False
            for candidate_id in candidate_ids:
                normalized_id = self._normalize_field_id(candidate_id)
                if normalized_id and normalized_id in by_id:
                    val = by_id.get(normalized_id, "")
                    found = True
                    break

            if not found:
                if 'data_key' in field_config:
                    val = form_data.get(field_config['data_key'], "")
                else:
                    data_index = field_config.get('data_index')
                    if data_index is not None and data_index < len(self.data_columns):
                        key = self.data_columns[data_index]
                        val = form_data.get(key, "")
        else:
            data_index = field_config.get('data_index')
            if data_index is not None and data_index < len(form_data):
                val = form_data[data_index]

        return self._normalize_resolved_value(val, field_config)

    def setup_directories_and_files(self):
        """Configura carpetas y archivos con rutas relativas"""
        run_number = self._get_next_available_number()
        
        self.SCREENSHOT_DIR = os.path.join(self.RESULTADOS_DIR, f"{self.SCREENSHOT_BASENAME}{run_number}")
        self.RESULTADOS_PATH = os.path.join(self.RESULTADOS_DIR, f"{self.RESULTADOS_BASENAME}{run_number}.xlsx")
        
        os.makedirs(self.SCREENSHOT_DIR, exist_ok=True)
        
        print(f"Carpeta de screenshots: {self.SCREENSHOT_DIR}")
        print(f"Archivo de resultados: {self.RESULTADOS_PATH}")
        print(f"Archivo Excel de datos: {self.EXCEL_PATH}")
        log_runtime(f"Excel de entrada: {self.EXCEL_PATH}")
        
        # Verificar que el Excel existe
        if not os.path.exists(self.EXCEL_PATH):
            print(f"ERROR: No se encuentra el archivo Excel: {self.EXCEL_PATH}")
            print("Los archivos disponibles en data/ son:")
            for file in os.listdir(self.DATA_DIR):
                if file.endswith('.xlsx'):
                    print(f"   - {file}")
            popup_log(
                "Error Excel",
                f"No se encuentra el archivo Excel:\n{self.EXCEL_PATH}",
                level="ERROR",
            )
            raise FileNotFoundError(f"No se encuentra el archivo Excel: {self.EXCEL_PATH}")
        
        self.RUN_NUMBER = run_number
        return run_number
    
    def _get_next_available_number(self):
        """Encuentra el próximo número disponible para carpetas y archivos"""
        i = 1
        while True:
            folder_exists = os.path.exists(os.path.join(self.RESULTADOS_DIR, f"{self.SCREENSHOT_BASENAME}{i}"))
            file_exists = os.path.exists(os.path.join(self.RESULTADOS_DIR, f"{self.RESULTADOS_BASENAME}{i}.xlsx"))
            
            if not folder_exists and not file_exists:
                return i
            i += 1
            
            if i > 1000:
                return 1
    
    def initialize_browser(self):
        """Inicializa el navegador según la configuración"""
        self.driver = BrowserManager.create_browser(
            browser_type=self.config['browser'],
            viewport=self.config['viewport'],
            headless=self.config.get('headless', False),
            background=self.config.get('background', True),
        )
        self.screenshot_manager = ScreenshotManager(self.driver, self.SCREENSHOT_DIR)
        
        # Monkey patch find_element and find_elements to support name & CSS fallback for By.ID
        from selenium.webdriver.common.by import By
        original_find_element = self.driver.find_element
        original_find_elements = self.driver.find_elements

        def wrapped_find_element(by, value):
            if by == By.ID:
                try:
                    return original_find_element(by, value)
                except Exception as e:
                    # Fallback 1: Buscar por Name attribute
                    try:
                        return original_find_element(By.NAME, value)
                    except Exception:
                        pass
                    # Fallback 2: Buscar por CSS Selector
                    if any(c in value for c in ('[', '.', '#')):
                        try:
                            return original_find_element(By.CSS_SELECTOR, value)
                        except Exception:
                            pass
                    raise e
            return original_find_element(by, value)

        def wrapped_find_elements(by, value):
            if by == By.ID:
                elements = original_find_elements(by, value)
                if not elements:
                    # Fallback 1: Buscar por Name attribute
                    elements = original_find_elements(By.NAME, value)
                    # Fallback 2: Buscar por CSS Selector
                    if not elements and any(c in value for c in ('[', '.', '#')):
                        elements = original_find_elements(By.CSS_SELECTOR, value)
                return elements
            return original_find_elements(by, value)

        self.driver.find_element = wrapped_find_element
        self.driver.find_elements = wrapped_find_elements
    
    def safe_load_workbook(self, excel_path):
        """Carga el workbook de forma segura, incluso si está abierto"""
        try:
            return load_workbook(excel_path)
        except Exception as e:
            if "Permission denied" in str(e) or "used by another process" in str(e):
                print("Excel está abierto, usando modo de solo lectura...")
                try:
                    temp_path = "temp_readonly.xlsx"
                    shutil.copy2(excel_path, temp_path)
                    wb = load_workbook(temp_path)
                    os.remove(temp_path)
                    return wb
                except Exception as e2:
                    print(f"No se pudo leer el Excel: {e2}")
                    popup_log("Error Excel", f"No se pudo leer el Excel:\n{e2}", level="ERROR")
                    raise e2
            else:
                popup_log("Error Excel", f"Error al cargar workbook:\n{e}", level="ERROR")
                raise e
    
    def safe_save_workbook(self, wb, filepath):
        """Guarda el workbook de forma segura, manejando archivos abiertos"""
        try:
            wb.save(filepath)
            return True
        except Exception as e:
            if "Permission denied" in str(e) or "used by another process" in str(e):
                print(f"No se pudo guardar en {filepath} (archivo abierto)")
                print("Guardando en archivo alternativo...")
                try:
                    base, ext = os.path.splitext(filepath)
                    alt_path = f"{base}_alternativo{ext}"
                    wb.save(alt_path)
                    print(f"Resultados guardados en: {alt_path}")
                    return True
                except Exception as e2:
                    print(f"Error crítico al guardar: {e2}")
                    popup_log("Error Guardado", f"Error crítico al guardar:\n{e2}", level="ERROR")
                    return False
            else:
                popup_log("Error Guardado", f"Error al guardar workbook:\n{e}", level="ERROR")
                raise e
    
    def safe_find_element(self, by, value, max_attempts=2):
        """Busca un elemento manejando posibles stale element references"""
        attempt = 0
        while attempt < max_attempts:
            try:
                element = self.driver.find_element(by, value)
                return element
            except Exception as e:
                if "stale" in str(e).lower() and attempt < max_attempts - 1:
                    print(f"Elemento stale, reintentando... ({attempt + 1}/{max_attempts})")
                    attempt += 1
                    time.sleep(1)
                else:
                    raise e

    def _element_exists_by_id(self, element_id):
        """Chequeo rápido (sin waits) de existencia por ID, Name o Selector CSS dentro del iframe actual."""
        try:
            if bool(self.driver.execute_script("return document.getElementById(arguments[0]) !== null;", element_id)):
                return True
            if bool(self.driver.execute_script("return document.querySelector('[name=\"' + arguments[0] + '\"]') !== null;", element_id)):
                return True
            if any(x in element_id for x in ("[", ".", "#")):
                if bool(self.driver.execute_script("return document.querySelector(arguments[0]) !== null;", element_id)):
                    return True
            return False
        except Exception:
            # Fallback a Selenium (usará los métodos wrapped_find_elements)
            try:
                return len(self.driver.find_elements(By.ID, element_id)) > 0
            except Exception:
                return False

    def _pick_first_existing_id(self, ids):
        """Devuelve el primer id existente en DOM (o None)."""
        if isinstance(ids, str):
            ids = [ids]
        for fid in ids or []:
            if fid and self._element_exists_by_id(fid):
                return fid
        return None

    def _is_visible(self, by, value):
        """Verifica si un elemento está presente y visible en el iframe actual."""
        try:
            el = self.driver.find_element(by, value)
            return el.is_displayed()
        except Exception:
            return False

    def _pick_first_visible_id(self, ids):
        """Devuelve el primer id existente en DOM y visible (o None)."""
        if isinstance(ids, str):
            ids = [ids]
        for fid in ids or []:
            if fid and self._element_exists_by_id(fid):
                try:
                    el = self.driver.find_element(By.ID, fid)
                    if el.is_displayed():
                        return fid
                except Exception:
                    continue
        # Fallback visid: si ningún ID del mapping fue encontrado, probar alias estandarizado
        for fid in ids or []:
            alias = self._VISID_ID_ALIASES.get(fid if isinstance(fid, str) else "")
            if alias and self._element_exists_by_id(alias):
                try:
                    el = self.driver.find_element(By.ID, alias)
                    if el.is_displayed():
                        return alias
                except Exception:
                    continue
        return None

    def _is_next_button_element(self, element):
        """True si el elemento parece ser un botón para avanzar al siguiente paso."""
        try:
            if not element.is_displayed():
                return False

            text_parts = [
                element.text or "",
                element.get_attribute("data-dtm") or "",
                element.get_attribute("class") or "",
                element.get_attribute("aria-label") or "",
                element.get_attribute("title") or "",
            ]
            normalized = " ".join(self._normalize_text(part) for part in text_parts if part)
            next_keywords = (
                "siguiente", "seguinte", "continuar", "continuacao",
                "proximo", "next"
            )
            return any(keyword in normalized for keyword in next_keywords)
        except Exception:
            return False

    def _find_next_button(self):
        """Devuelve el primer botón visible que parezca avanzar al siguiente paso."""
        selectors = [
            "button.btn-steps-submit",   # visid multi-step
            ".button.next.pulsate.stat-button-link",
            "button[class*='next']",
            "button[data-dtm*='next']",
            "button[type='button']",
            ".next-button",
        ]

        for selector in selectors:
            try:
                btns = self.driver.find_elements(By.CSS_SELECTOR, selector)
                for btn in btns:
                    if self._is_next_button_element(btn):
                        return btn
            except Exception:
                continue

        # Barrido genérico final: cualquier button/a/role=button visible cuyo texto/atributos
        # indiquen "Siguiente/Seguinte/Continuar/Próximo/Next" (cubre markups fuera de la lista).
        try:
            for btn in self.driver.find_elements(By.CSS_SELECTOR, "button, a, [role='button']"):
                if self._is_next_button_element(btn):
                    return btn
        except Exception:
            pass
        return None

    def _has_next_button(self):
        """Detecta si hay un botón Siguiente/Seguinte/Next/Continuar visible."""
        try:
            return self._find_next_button() is not None
        except Exception:
            return False

    def _click_next_button(self, paso_label="auto"):
        """Hace clic en el botón Siguiente/Seguinte/Next/Continuar si existe."""
        try:
            btn = self._find_next_button()
            if not btn:
                print(f" No se encontró botón 'Siguiente/Seguinte' visible ({paso_label})")
                return False

            self._scroll_element_into_view(btn)
            time.sleep(0.5)
            self.driver.execute_script("arguments[0].click();", btn)
            print(f" Clic en botón de siguiente ({paso_label}) realizado")
            return True

        except Exception:
            try:
                btn = self._find_next_button()
                if btn:
                    btn.click()
                    print(f" Clic en botón de siguiente ({paso_label}) realizado con fallback")
                    return True
            except Exception:
                pass

            print(f" No se encontró botón 'Siguiente/Seguinte' visible ({paso_label})")
            return False
        except Exception as e:
            print(f" Error al hacer clic en 'Siguiente' ({paso_label}): {e}")
            return False

    def _dom_signature_visible_mapping(self):
        """Firma del DOM: ids del mapping visibles (ordenado), para detectar atascos."""
        sig = []
        for field_config in self.field_mapping or []:
            field_id = field_config.get("id")
            ids = field_id if isinstance(field_id, list) else [field_id]
            chosen = self._pick_first_visible_id(ids)
            if chosen:
                sig.append(chosen)
        return tuple(sorted(sig))

    def _finalize_model_kit_on_last_step(self, form_data):
        """Re-selecciona modelo/kit si aparecen de nuevo en el último paso (p. ej. Argentina)."""
        data = {key: self.get_form_value_str(form_data, key) for key in self.data_columns}
        modelo_valor = (data.get("model") or "").strip()
        kit_valor = (data.get("kit") or "").strip()

        # Si no hay modelo ni kit en Excel, no interferir
        if not modelo_valor and not kit_valor:
            return

        try:
            if modelo_valor and self._is_visible(By.ID, "models"):
                model_updated = False
                if self._select_current_option_matches_desired("models", modelo_valor):
                    print(" Modelo (paso final) ya coincide, se conserva")
                else:
                    print(f"Seleccionando modelo en paso final: {modelo_valor}")
                    model_updated = self.safe_select_option_if_visible("models", modelo_valor, "Modelo (paso final)")

                if model_updated:
                    try:
                        modelo_select = self.driver.find_element(By.ID, "models")
                        self.driver.execute_script(
                            """
                            var select = arguments[0];
                            select.dispatchEvent(new Event('change', { bubbles: true }));
                            select.dispatchEvent(new Event('input', { bubbles: true }));
                            if (typeof $ !== 'undefined' && $(select).selectpicker) {
                                $(select).trigger('change');
                            }
                            """,
                            modelo_select,
                        )
                    except Exception as e:
                        print(f" Aviso disparando eventos en modelo: {e}")
                    time.sleep(0.5)
        except Exception as e:
            print(f" Error con modelo en paso final: {e}")

        if not kit_valor:
            print(" Kit valor vacío en datos")
            return

        if not self._wait_for_dependent_dropdown_ready("kits[]", parent_id="models"):
            print(" No se pudo preparar kits[] en paso final dentro del timeout")
            return

        for intento in range(1, 3):
            try:
                if intento > 1:
                    print(f" Reintento {intento}/2 para seleccionar kit...")
                    time.sleep(1)
                if not self._is_visible(By.ID, "kits[]"):
                    print(" kits[] no visible en paso final")
                    return
                kit_select = self.driver.find_element(By.ID, "kits[]")
                if kit_select.get_attribute("disabled"):
                    print(" Kit dropdown aún deshabilitado")
                    continue
                if self.safe_select_option_if_visible("kits[]", kit_valor, "Kit"):
                    print(f" Kit seleccionado: {kit_valor}")
                    return
            except Exception as e:
                print(f" Error con kit (intento {intento}): {e}")
        print(" No se pudo seleccionar el kit en paso final")

    def _fill_visible_fields_from_mapping(self, form_data, dependencies):
        """
        Igual que fill_form_fields pero solo campos cuyo id resuelto está visible.
        Respeta dependencias entre selects (models→kits, region→city→dealer).
        """
        # --- NUEVO: Cargar dependencias de ids_dinamicos.json ---
        import json
        ids_dinamicos_path = os.path.join(self.BASE_DIR, "json", "ids_dinamicos.json")
        dyn_dependencies = {}
        if os.path.exists(ids_dinamicos_path):
            with open(ids_dinamicos_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                for dep in data.get("dependencies", []):
                    padre = str(dep.get("padre") or "").strip()
                    hijo = str(dep.get("hijo") or "").strip()
                    if padre and hijo:
                        dyn_dependencies[hijo] = padre

        resolved_field_configs = []
        for field_config in self.field_mapping or []:
            field_id = field_config.get("id")
            ids = field_id if isinstance(field_id, list) else [field_id]
            chosen_id = self._pick_first_visible_id(ids)
            if not chosen_id:
                print(f"⚠️ [DEBUG-FILL] Campo '{field_config.get('name')}' (IDs/Names intentados: {ids}) NO SE ENCONTRÓ en el DOM o no es visible.")
                continue
            resolved = dict(field_config)
            resolved["__resolved_id"] = chosen_id
            resolved_field_configs.append(resolved)

        processed_ids = set()
        solo_verificar_visual = bool(self.config.get("solo_verificar_visual", False) or self.config.get("no_enviar_lead", False))

        def fill_with_dependencies(field_config):
            field_id = field_config.get("__resolved_id") or field_config.get("id")
            if field_id in processed_ids:
                return
            if solo_verificar_visual and len(processed_ids) >= 2:
                return
            # Si este campo es hijo en dependencias dinámicas, completar el padre primero
            parent_id = dyn_dependencies.get(field_id)
            if parent_id:
                parent_config = next((fc for fc in resolved_field_configs if (fc.get("__resolved_id") or fc.get("id")) == parent_id), None)
                if parent_config:
                    fill_with_dependencies(parent_config)
            # Ahora completar el campo actual
            field_type = field_config.get("type", "text")
            field_name = field_config.get("name", field_id)
            field_value = self._resolve_field_value(form_data, field_config)

            # Si el campo es select y tiene dependencias normales, usar la lógica robusta
            parent_id_normal = None
            for k, v in dependencies.items():
                if v == field_id:
                    parent_id_normal = k
                    break

            if field_type == "select":
                print(f"🔹 [DEBUG-FILL] Procesando select '{field_name}' (ID: {field_id}, valor excel: '{field_value if field_value else '(vacío)'}')...")
                if parent_id_normal:
                    success = self._select_dependency_child_with_timeout(field_id, field_value, field_name, parent_id_normal)
                elif field_id == "event":
                    original_ids = field_config.get("id")
                    field_ids = original_ids if isinstance(original_ids, list) else [original_ids]
                    success = self._select_dropdown_with_fallback(field_id, field_value, field_name, fallback_ids=field_ids)
                elif field_id == "estimated-date-purchase":
                    success = self._fill_purchase_date_field(field_value, field_name)
                else:
                    success = self.safe_select_option_if_visible(field_id, field_value, field_name)
                if success:
                    processed_ids.add(field_id)
            else:
                # Text, textarea, etc. Si es dinámico, completar con valor fijo si corresponde
                ids_dinamicos = self._cargar_ids_dinamicos()
                if field_id in ids_dinamicos:
                    dynamic_candidates = self._resolve_dynamic_id_values(ids_dinamicos[field_id])
                    if dynamic_candidates:
                        fixed_value = dynamic_candidates[0]
                        try:
                            element = self.safe_find_element(By.ID, field_id)
                            if element and element.is_displayed() and element.is_enabled():
                                self.driver.execute_script(
                                    "arguments[0].focus();"
                                    "arguments[0].value = arguments[1];"
                                    "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));"
                                    "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
                                    "arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));",
                                    element,
                                    fixed_value,
                                )
                                print(f"📌 Campo '{field_id}' completado con ID dinámico: {fixed_value}")
                                self._record_field_value(field_id, fixed_value)
                                processed_ids.add(field_id)
                                return
                        except Exception as e:
                            print(f" Error completando campo dinámico '{field_id}': {e}")
                # Si no es dinámico, usar el valor normal.
                # Acepta type="text", type="textarea" en el config, o detección automática
                # por tag del DOM (input + textarea se tratan igual).
                # Brasil: generar CPF/CNPJ/CEP solo si el Excel no trae un valor válido
                _fid_lower = field_id.lower()
                _is_brasil = str(self.config.get("pais", "")).lower() in ("brasil", "brazil", "br")
                if _is_brasil and any(x in _fid_lower for x in ("cpf", "cnpj", "cep", "zip", "postal")):
                    _excel_empty = field_value in (None, "")
                    if isinstance(field_value, (int, float)):
                        _raw_str = str(int(field_value))
                    else:
                        _raw_str = str(field_value or "").strip()
                    _digits = "".join(c for c in _raw_str if c.isdigit())
                    _min_len = 14 if "cnpj" in _fid_lower else (8 if any(x in _fid_lower for x in ("cep", "zip", "postal")) else 11)
                    # Excel numérico come el cero inicial: CPF de 10 → pad a 11, CNPJ de 13 → pad a 14, CEP de 7 → pad a 8
                    if _digits and len(_digits) == _min_len - 1:
                        _digits = _digits.zfill(_min_len)
                    if _digits and len(_digits) >= _min_len:
                        field_value = _digits
                    else:
                        generated = self._generate_brazil_document(field_id)
                        if generated:
                            field_value = generated

                # Perú: sanitizar número de documento según tipo seleccionado.
                # OJO con el id: en los forms visid / gm_front el 'ci' del mapping se resuelve
                # al alias 'document' (_VISID_ID_ALIASES) ANTES de llegar acá, así que
                # comparar sólo contra "ci" dejaba estos forms sin sanitizar y el lead se
                # rechazaba por longitud cuando el tipo elegido no era DNI.
                if field_id in ("ci", "document") and str(self.config.get("pais", "")).lower() in ("peru", "pe"):
                    try:
                        doc_type_el = self.safe_find_element(By.ID, "document-type")
                        if doc_type_el:
                            from selenium.webdriver.support.ui import Select as _Select
                            doc_type_val = _Select(doc_type_el).first_selected_option.text
                            field_value = self._sanitize_peru_document(doc_type_val, field_value)
                    except Exception:
                        pass

                if field_value:
                    try:
                        element = self.safe_find_element(By.ID, field_id)
                        if element and element.is_displayed() and element.is_enabled():
                            dom_tag = (element.tag_name or "").lower()
                            dom_type = (element.get_attribute("type") or "text").lower() if dom_tag == "input" else dom_tag
                            # Ignorar tipos que no son texto libre
                            if dom_type in ("hidden", "submit", "button", "checkbox", "radio", "file", "reset", "select"):
                                processed_ids.add(field_id)
                                return
                            try:
                                # CPF en Brasil: usar JS para evitar que la máscara del campo trague dígitos
                                _is_cpf_field = (
                                    str(self.config.get("pais", "")).lower() in ("brasil", "brazil", "br")
                                    and any(x in _fid_lower for x in ("cpf", "cnpj"))
                                )
                                if _is_cpf_field:
                                    self.driver.execute_script(
                                        "var n=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value');"
                                        "n.set.call(arguments[0], arguments[1]);"
                                        "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                                        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
                                        "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
                                        element,
                                        field_value,
                                    )
                                else:
                                    self._fill_and_dispatch(element, field_value)
                            except Exception:
                                try:
                                    self._fill_and_dispatch(element, field_value)
                                except Exception:
                                    pass
                            print(f"✅ [DEBUG-FILL] {field_name} completado ({field_id}): {field_value}")
                            self._record_field_value(field_id, field_value)
                            processed_ids.add(field_id)
                            return
                        else:
                            status = "no encontrado en el DOM" if not element else ("no visible" if not element.is_displayed() else "deshabilitado")
                            print(f"⚠️ [DEBUG-FILL] {field_name} ({field_id}) existe pero está {status}")
                    except Exception as e:
                        print(f"❌ [DEBUG-FILL] Error completando campo '{field_id}': {e}")
                else:
                    print(f"⚠️ [DEBUG-FILL] Omitiendo '{field_name}' ({field_id}) porque su valor en Excel está vacío o es None.")
            processed_ids.add(field_id)

        for field_config in resolved_field_configs:
            if solo_verificar_visual and len(processed_ids) >= 2:
                print("⚠️ [DEBUG-FILL] Llenado parcial activo: se detiene el llenado tras completar 2 campos.")
                break
            fill_with_dependencies(field_config)
        if not solo_verificar_visual:
            self._auto_fill_unmapped_dropdowns(self.field_mapping)

    # === Adobe AEM Adaptive Form (formularios "2.0") =========================
    # Lógica compartida en utils/aem_fill.py (misma fuente para desktop y LambdaTest).
    # En estos forms el ID del widget es genérico y el keyword semántico vive en el
    # <label for>, así que se localiza por label, no por ID fijo del field_mapping.

    def _is_aem_adaptive_form(self):
        """Detecta si la página actual es un Adobe AEM Adaptive Form (Guide)."""
        from utils import aem_fill
        return aem_fill.is_aem_adaptive_form(self.driver)

    def _fill_aem_by_semantic_id(self, form_data):
        """Llena un AEM Adaptive Form delegando en utils.aem_fill (fuente única)."""
        from utils import aem_fill
        pais = str(self.config.get("pais", "")).lower()
        is_brasil = pais in ("brasil", "brazil", "br")
        fd = dict(form_data) if isinstance(form_data, dict) else {}
        
        # Modo rellenado parcial
        if bool(self.config.get("solo_verificar_visual", False) or self.config.get("no_enviar_lead", False)):
            fd = {
                "firstname": fd.get("firstname") or "Test",
                "lastname": fd.get("lastname") or "User"
            }
        else:
            # Asegurar cpf/cnpj/cep desde __by_id (por si el normalizado no los trae)
            by_id = fd.get("__by_id", {}) if isinstance(fd.get("__by_id"), dict) else {}
        for _src, _dst in (("cpf", "cpf"), ("cnpj", "cnpj"), ("cep", "cep"),
                           ("zip", "cep"), ("postal", "cep"),
                           ("vin", "vin"), ("vin-code", "vin"), ("chassis", "vin")):
            if not fd.get(_dst) and by_id.get(_src):
                fd[_dst] = by_id.get(_src)
        return aem_fill.fill_aem_form(
            self.driver, fd, is_brasil,
            gen_doc=self._generate_brazil_document,
            record=self._record_field_value,
        )

    # Campos propios del Libro de Reclamaciones. Alcanza con estos dos para reconocerlo:
    # su layout no coincide con el mapping del país (Documento cae en Teléfono, Celular en
    # E-mail…) y necesita el fill directo por ID.
    _LIBRO_RECLAMACIONES_IDS = ("cc_name", "cc_telephone")

    # Fragmentos de URL que identifican un Libro de Reclamaciones, tanto en la landing como
    # en la URL del form: chevrolet.com.pe/libro-reclamaciones-virtual (inserto) y
    # gm_front/form/form-reclamos (suelto).
    _LIBRO_RECLAMACIONES_URL_HINTS = ("libro-reclamaciones", "reclamos", "reclamaciones")

    def _has_libro_reclamaciones_fields(self):
        """True si el DOM actual tiene los campos propios del Libro de Reclamaciones."""
        try:
            return all(self._element_exists_by_id(fid) for fid in self._LIBRO_RECLAMACIONES_IDS)
        except Exception:
            return False

    def _is_libro_reclamaciones_form(self, landing_url):
        """True si el form actual es un Libro de Reclamaciones.

        Se mira la URL (landing Y form: en las URLs sueltas el slug 'reclamos' viene en la
        del form, no en la landing) y, como respaldo, el DOM por los ids cc_*. El chequeo
        por DOM solo no alcanza si el form React todavía no montó cuando se evalúa.

        En estos forms NO se hace el clic de "enviar en vacío" que se usa en el resto para
        disparar las validaciones: se llena primero y recién ahí se envía.
        """
        _urls = (landing_url or "", getattr(self, "expected_form_url", "") or "")
        for _u in _urls:
            _u = _u.lower()
            if any(h in _u for h in self._LIBRO_RECLAMACIONES_URL_HINTS):
                return True
        return self._has_libro_reclamaciones_fields()

    def _fill_libro_reclamaciones_direct(self, form_data):
        """
        Fill directo para chevrolet.com.pe/libro-reclamaciones-virtual.
        IDs fijos: cc_name, cc_telephone, cc_ci, cc_email + dropdowns city/dealer aleatorios.
        """
        by_id = form_data.get("__by_id", {}) if isinstance(form_data, dict) else {}

        firstname = str(by_id.get("firstname") or by_id.get("name") or form_data.get("firstname") or "").strip()
        lastname  = str(by_id.get("lastname") or form_data.get("lastname") or "").strip()
        nombre    = f"{firstname} {lastname}".strip()
        phone     = str(by_id.get("telephone") or by_id.get("cellphone") or by_id.get("phone")
                        or form_data.get("phone") or "").strip()
        document  = str(by_id.get("ci") or by_id.get("document") or by_id.get("rut")
                        or form_data.get("document") or "").strip()
        email     = str(by_id.get("email") or form_data.get("email") or "").strip()

        def _fill_by_id(field_id, value):
            if not value:
                return False
            try:
                els = self.driver.find_elements(By.ID, field_id)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        self._scroll_element_into_view(el)
                        try:
                            el.clear()
                            el.send_keys(value)
                        except Exception:
                            self.driver.execute_script(
                                "arguments[0].value = arguments[1];"
                                "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                                el, value
                            )
                        print(f"  ✓ libro-reclamaciones: '{field_id}' = '{value}'")
                        return True
            except Exception as e:
                print(f"  ✗ libro-reclamaciones: error llenando '{field_id}': {e}")
            return False

        _fill_by_id("cc_name", nombre)
        _fill_by_id("cc_telephone", phone)
        
        if bool(self.config.get("solo_verificar_visual", False) or self.config.get("no_enviar_lead", False)):
            current_ss_number = getattr(self, 'ss_counter', 0)
            if self.screenshot_manager:
                self.screenshot_manager.take_form_screenshot(current_ss_number, "completado", full_page=True)
            return f"form_completado_{current_ss_number}.png"

        _fill_by_id("cc_ci", document)
        _fill_by_id("cc_email", email)

        # Resto de campos del reclamo: no están en el mapping del país y el form los valida
        # por JS (en el DOM figuran como required=false), así que hay que llenarlos acá.
        _fill_by_id("cc_address", "Av. Javier Prado Este 1234")
        _fill_by_id("cc_amount", "1000")
        _fill_by_id("cc_details", "Detalle de prueba automatizada del formulario.")
        _fill_by_id("cc_details_claim", "Detalle de prueba automatizada del formulario.")
        _fill_by_id("cc_order_claim", "Pedido de prueba automatizada.")

        # Ciudad aleatoria, luego esperar y seleccionar Concesionario aleatorio
        self.safe_select_option_if_visible("city", "", "Ciudad")
        time.sleep(0.5)
        if self._wait_for_dependent_dropdown_ready("dealer", parent_id="city"):
            self.safe_select_option_if_visible("dealer", "", "Concesionario")

        # Radios obligatorios (mayor de edad / bien contratado / reclamo-queja) y el
        # checkbox de términos del pie.
        try:
            self._handle_terms_checkboxes()
        except Exception as _e:
            print(f"  ⚠ libro-reclamaciones: radios/checkboxes — {_e}")

        current_ss_number = getattr(self, 'ss_counter', 0)
        if self.screenshot_manager:
            # full_page: el Libro de Reclamaciones es larguísimo y sin esto la captura salía
            # recortada al viewport (solo Pedido/VIN/Ciudad/Enviar, sin los datos personales).
            self._revalidar_campos_llenos()
            self.screenshot_manager.take_form_screenshot(current_ss_number, "completado",
                                                         full_page=True)
            print("Captura 2/3: Formulario completado")
        return f"form_completado_{current_ss_number}.png"

    def fill_form_fields_auto_step(self, form_data):
        """
        Motor unificado: completa solo campos visibles, avanza con Siguiente mientras exista,
        luego modelo/kit final, términos y captura. Funciona en 1 paso o multipaso sin configuración extra.
        """
        current_ss_number = self.ss_counter
        solo_verificar_visual = bool(self.config.get("solo_verificar_visual", False) or self.config.get("no_enviar_lead", False))

        dependencies = {
            "models": "kits[]",
            "model": "kits[]",
            "region": "city",
            "city": "dealer",
        }
        max_iter = int(self.config.get("auto_step_max_iterations", 15))

        # Guardar form_data para que _auto_fill_unmapped_dropdowns pueda usarlo
        self._current_form_data = form_data or {}

        # Para forms sin iframe (gm_front y similares React/SPA): esperar a que el
        # framework monte los componentes antes de escanear el DOM.
        _is_standalone = not bool(getattr(self, "expected_form_url", ""))
        if _is_standalone:
            try:
                from selenium.webdriver.support.ui import WebDriverWait as _WDW
                _WDW(self.driver, 8).until(
                    lambda d: bool(
                        d.find_elements(By.CSS_SELECTOR,
                            "input:not([type='hidden']):not([type='submit']):not([type='button']), select"
                        )
                    )
                )
            except Exception:
                time.sleep(2)

        # Detectar y mapear campos nuevos antes de comenzar
        try:
            self._campos_nuevos_detectados = []
            self._discover_and_report_unmapped_fields()
        except Exception as _disc_err:
            print(f"Auto-discovery: error no crítico — {_disc_err}")

        # Formularios "2.0" (Adobe AEM Adaptive Form): IDs con panel volátil → se
        # llena por keyword semántico del id/label en vez de ID fijo del mapping.
        self._is_aem = self._is_aem_adaptive_form()
        if self._is_aem:
            print("🧩 [DEBUG] Formulario AEM Adaptive detectado → llenado por keyword semántico")
            # Cerrar cookies antes de llenar (también en estos forms 2.0)
            try:
                self.handle_cookie_popups()
            except Exception:
                pass

        for iteration in range(max_iter):
            paso = iteration + 1
            # Captura del paso VACÍO (para pasos posteriores; el paso 1 ya lo capturó el flujo
            # principal como form_vacio antes del clic vacío).
            if iteration > 0 and self.screenshot_manager:
                try:
                    self.screenshot_manager.take_form_screenshot(
                        current_ss_number, f"vacio_paso{paso}", full_page=True)
                    print(f"Captura form_vacio_paso{paso} tomada")
                except Exception:
                    pass
            # Clic preliminar para forzar validación y activar mensajes de error de este paso
            try:
                if solo_verificar_visual:
                    raise Exception("Modo 'No enviar lead' activo — omitiendo click preliminar.")
                # Buscar el botón Siguiente o Enviar visible en este paso.
                # Si el CTA "Siguiente" YA está visible, se clickea en vacío para arrojar los
                # mensajes de error antes de llenar. Si NO está visible (CTA gated tras un campo
                # obligatorio, ej. Chevrolet Exonerados: sin modelo no hay botón), boton_accion
                # queda None → se llena primero, el botón aparece y se avanza más abajo.
                boton_accion = self._find_next_button()
                es_submit = False
                if not boton_accion:
                    boton_accion, _ = self._resolve_submit_button(wait_seconds=1)
                    es_submit = boton_accion is not None

                # Si el botón es ENVIAR y el form no tiene ningún campo obligatorio, el clic
                # preliminar NO dispara validaciones: manda el formulario vacío y genera un
                # lead basura (ej. el Libro de Reclamaciones de Perú, donde todos los campos
                # son voluntarios). En ese caso no se clickea: se llena y recién ahí se envía.
                if es_submit and not self._tiene_campos_obligatorios():
                    print("⚠️ [DEBUG] El form no tiene campos obligatorios — se omite el clic "
                          "preliminar sobre Enviar (lo enviaría vacío).")
                    boton_accion = None

                if boton_accion and boton_accion.is_displayed() and boton_accion.is_enabled():
                    print("⚡ [DEBUG] Clic preliminar sobre botón de acción para forzar validaciones antes de rellenar...")
                    try:
                        boton_accion.click()
                    except Exception:
                        self.driver.execute_script("arguments[0].click();", boton_accion)
                    time.sleep(1.0)  # Esperar a que se pinten/activen los mensajes de error

                    # Captura de errores de este paso. En el paso 1 de un form de un solo paso ya
                    # la tomó el bloque de "click enviar vacío" (_errores_ss_taken).
                    if self.screenshot_manager and not (iteration == 0 and getattr(self, "_errores_ss_taken", False)):
                        _sufijo = "errores" if (iteration == 0 and not self._has_next_button()) else f"errores_paso{paso}"
                        self.screenshot_manager.take_form_screenshot(
                            current_ss_number, _sufijo, full_page=True)
                        self._errores_ss_taken = True
                        print(f"Captura form_{_sufijo} tomada")
            except Exception as _e:
                print(f"⚠️ No se pudo realizar clic preliminar de validación: {_e}")

            if getattr(self, "_is_aem", False):
                self._fill_aem_by_semantic_id(form_data)
            else:
                self._fill_visible_fields_from_mapping(form_data, dependencies)
                self._record_model_from_url_if_missing()

            if solo_verificar_visual:
                print("⚠️ [DEBUG] Llenado parcial activo: deteniendo tras primer paso.")
                break

            if not self._has_next_button():
                break

            # Form multi-paso: captura del paso ya completado antes de pasar al siguiente
            if self.screenshot_manager:
                self.screenshot_manager.take_form_screenshot(
                    current_ss_number, f"completado_paso{paso}", full_page=True)
                print(f"Captura form_completado_paso{paso} tomada")

            sig_before = self._dom_signature_visible_mapping()
            if not self._click_next_button("auto"):
                print(" auto_step: no hay botón Siguiente; finalizando bucle de pasos")
                break

            self._current_step += 1
            time.sleep(1.5)
            stuck = True
            for _ in range(4):
                if self._dom_signature_visible_mapping() != sig_before:
                    stuck = False
                    break
                time.sleep(0.75)
            if stuck:
                msg = (
                    "auto_step: el formulario no avanzó tras clic en Siguiente "
                    "(DOM sin cambios visibles). Revisar selectores o validaciones."
                )
                print(f" ERROR: {msg}")
                self._log(f"STUCK en iteracion {iteration}: {msg}")
                raise RuntimeError(msg)
        else:
            raise RuntimeError(f"auto_step: se alcanzó el máximo de iteraciones ({max_iter})")

        if solo_verificar_visual:
            self.reposition_to_form(self.expected_form_url)
            _last_step = int(getattr(self, "_current_step", 1) or 1)
            _stage_fin = f"completado_paso{_last_step}" if _last_step > 1 else "completado"
            form_completado_name = self.screenshot_manager.fname("form", _stage_fin, current_ss_number)
            self.screenshot_manager.take_form_screenshot(current_ss_number, _stage_fin, full_page=True)
            print(f"Captura 2/3: Formulario completado (Verificación visual - parcial)")
            return form_completado_name

        self._finalize_model_kit_on_last_step(form_data)
        # Rellenar requeridos visibles sin dato (ej. Rua/Número/Data de Nascimento) con
        # valores sintéticos, para no quedar con campos obligatorios vacíos al enviar.
        try:
            self._fill_unmapped_required_fields()
        except Exception as _e:
            print(f" Auto-relleno de requeridos: error no crítico — {_e}")
        self._handle_terms_checkboxes()
        # Antes de la captura: limpiar los errores que quedaron pintados en campos que ya
        # están completos (si no, la captura muestra falsos errores).
        self._revalidar_campos_llenos()

        self.reposition_to_form(self.expected_form_url)
        # Si el form fue multipaso, el 'completado' final es el ÚLTIMO paso: nombrarlo con su
        # número (completado_paso3) para que sea consistente con completado_paso1/paso2 y no
        # quede ambiguo. Para forms de 1 paso queda 'completado' a secas.
        _last_step = int(getattr(self, "_current_step", 1) or 1)
        _stage_fin = f"completado_paso{_last_step}" if _last_step > 1 else "completado"
        form_completado_name = self.screenshot_manager.fname("form", _stage_fin, current_ss_number)
        self.screenshot_manager.take_form_screenshot(current_ss_number, _stage_fin, full_page=True)
        print(f"Captura 2/3: Formulario completado ({_stage_fin})")
        return form_completado_name

    def fill_form_fields(self, form_data):
        """Llena el formulario con detección automática de pasos (un solo flujo)."""
        return self.fill_form_fields_auto_step(form_data)

    def _is_placeholder_text(self, option_text):
        """Detecta textos de placeholder en español/portugués (y variantes comunes)."""
        normalized = self._normalize_text(option_text)
        if not normalized:
            return True

        placeholder_keywords = (
            "seleccione", "selecciona", "seleccionar", "seleccion",
            "selecione", "selecionar", "selecao", "selecao",
            "escolha", "escolher", "escolhe",
            "elija", "elegir", "opcao", "opcoes", "opcion", "opciones",
            "porfavor", "favor", "obrigatorio", "required",
            "select", "choose", "please"
        )

        return any(keyword in normalized for keyword in placeholder_keywords)

    def _select_has_valid_selected_option(self, select_id):
        """Indica si un select ya tiene una opción elegida que no parece placeholder."""
        try:
            select_element = self.safe_find_element(By.ID, select_id)
            if not select_element:
                return False
            select = Select(select_element)
            selected_option = select.first_selected_option
            if not selected_option:
                return False

            selected_text = (selected_option.text or "").strip()
            if not selected_text:
                return False

            return not self._is_placeholder_text(selected_text)
        except Exception:
            return False

    def _select_current_option_matches_desired(self, select_id, desired_value):
        """True si el select ya tiene una opción válida que coincide con el valor pedido."""
        desired = (desired_value or "").strip()
        if not desired:
            return self._select_has_valid_selected_option(select_id)

        try:
            select_element = self.safe_find_element(By.ID, select_id)
            if not select_element:
                return False

            selected_option = Select(select_element).first_selected_option
            selected_text = (selected_option.text or "").strip()
            if not selected_text or self._is_placeholder_text(selected_text):
                return False

            selected_norm = self._normalize_text(selected_text)
            desired_norm = self._normalize_text(desired)
            if not selected_norm or not desired_norm:
                return False

            return selected_norm == desired_norm or selected_norm in desired_norm or desired_norm in selected_norm
        except Exception:
            return False

    def _get_dependency_wait_settings(self):
        """Obtiene timeout/poll/retries para dependencias entre dropdowns."""
        try:
            timeout = float(self.config.get("dependency_dropdown_timeout", 8.0))
        except Exception:
            timeout = 8.0
        try:
            poll_interval = float(self.config.get("dependency_dropdown_poll_interval", 0.2))
        except Exception:
            poll_interval = 0.2
        try:
            retries = int(self.config.get("dependency_selection_retries", 2))
        except Exception:
            retries = 2

        timeout = max(1.0, timeout)
        poll_interval = max(0.05, poll_interval)
        retries = max(1, retries)
        return timeout, poll_interval, retries

    def _wait_for_dependent_dropdown_ready(self, child_select_id, parent_id=None):
        """Espera a que el dropdown dependiente esté visible/habilitado y con opciones válidas."""
        timeout, poll_interval, _ = self._get_dependency_wait_settings()
        deadline = time.time() + timeout

        while time.time() < deadline:
            try:
                child_element = self.safe_find_element(By.ID, child_select_id)
                if not child_element or not child_element.is_displayed():
                    time.sleep(poll_interval)
                    continue

                tag_name = (child_element.tag_name or "").lower()
                if tag_name != "select":
                    return True

                if child_element.get_attribute("disabled") or not child_element.is_enabled():
                    time.sleep(poll_interval)
                    continue

                if self._get_valid_select_options(child_element):
                    return True
            except Exception:
                pass

            time.sleep(poll_interval)

        if parent_id:
            print(
                f" Timeout esperando dependiente '{child_select_id}' después de '{parent_id}' "
                f"({timeout:.1f}s)"
            )
        else:
            print(f" Timeout esperando dependiente '{child_select_id}' ({timeout:.1f}s)")
        return False

    def _select_dependency_child_with_timeout(self, child_select_id, child_value, child_field_name, parent_id):
        """Selecciona dropdown hijo con espera activa y reintentos."""
        _, _, retries = self._get_dependency_wait_settings()

        child_ready = self._wait_for_dependent_dropdown_ready(child_select_id, parent_id=parent_id)
        if not child_ready:
            raise ValueError(f"No se encontraron opciones en el dropdown '{child_field_name}' al depender de '{parent_id}'")

        for attempt in range(1, retries + 1):
            if self._select_has_valid_selected_option(child_select_id):
                print(f" {child_field_name} ya tiene valor válido, se conserva")
                return True

            selected = self.safe_select_option_if_visible(child_select_id, child_value, child_field_name)
            if selected and self._select_has_valid_selected_option(child_select_id):
                return True

            if attempt < retries:
                time.sleep(0.4)

        return False

    def _has_visible_required_field(self):
        """Indica si todavía queda visible algún campo relevante del formulario actual."""
        candidate_ids = []

        for field_config in self.field_mapping or []:
            field_id = field_config.get("id")
            if isinstance(field_id, list):
                candidate_ids.extend(fid for fid in field_id if fid)
            elif field_id:
                candidate_ids.append(field_id)

        required_fields = ((self.country_fields or {}).get("required_fields") or [])
        for field_id in required_fields:
            if field_id:
                candidate_ids.append(field_id)

        seen = set()
        for field_id in candidate_ids:
            normalized_id = self._normalize_field_id(field_id)
            if not normalized_id or normalized_id in seen:
                continue
            seen.add(normalized_id)

            if self._element_exists_by_id(normalized_id) and self._is_visible(By.ID, normalized_id):
                return True

        return False
    
    def safe_select_option_if_visible(self, select_id, option_text, field_name):
        """Selecciona una opción solo si el dropdown está visible - CON AUTO-SELECCIÓN ALEATORIA SI ESTÁ VACÍO"""
        if option_text is None:
            option_text = ""
        else:
            option_text = str(option_text)
        
        # Determinar si el campo está vacío (también si el valor es un placeholder)
        is_empty = not option_text or option_text.strip() == "" or self._is_placeholder_text(option_text)
            
        try:
            # Verificar primero si el elemento existe
            try:
                select_element = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.ID, select_id))
                )
            except TimeoutException:
                print(f" {field_name} (ID: {select_id}) no encontrado en el DOM después de 5s")
                return False
            
            # Verificar si está visible
            if not select_element.is_displayed():
                print(f" {field_name} (ID: {select_id}) existe pero NO está visible")
                return False

            # Forzar habilitación si está deshabilitado
            if select_element.get_attribute("disabled") or not select_element.is_enabled():
                try:
                    WebDriverWait(self.driver, 4).until(
                        EC.element_to_be_clickable((By.ID, select_id))
                    )
                    select_element = self.safe_find_element(By.ID, select_id)
                except TimeoutException:
                    try:
                        self.driver.execute_script("arguments[0].removeAttribute('disabled'); arguments[0].disabled = false;", select_element)
                        print(f" {field_name} habilitado forzosamente")
                    except Exception:
                        print(f"{field_name} está deshabilitado, omitiendo...")
                        return False

            # Asegurar que el select tenga opciones disponibles (solo si no está vacío)
            if not is_empty:
                try:
                    def select_has_options(_):
                        try:
                            select_el = self.safe_find_element(By.ID, select_id)
                            return len(Select(select_el).options) > 0
                        except Exception:
                            return False

                    WebDriverWait(self.driver, 4).until(select_has_options)
                except Exception:
                    pass

            select = Select(select_element)
            is_multiple = select_element.get_attribute("multiple") is not None
            
            # Detectar si es un select múltiple (como kits[])
            if is_multiple:
                print(f" {field_name} es multi-select, usando lógica especial")
                return self._select_multiple_options(select_element, option_text, field_name)

            if is_empty:
                try:
                    selected_option = select.first_selected_option
                except Exception:
                    selected_option = None

                if selected_option:
                    selected_text = (selected_option.text or "").strip()
                    if selected_text and not self._is_placeholder_text(selected_text):
                        print(f" {field_name} ya tiene valor válido ('{selected_text}'), se conserva")
                        self._record_field_value(select_id, selected_text)
                        return True

            # SI EL CAMPO ESTÁ VACÍO: Seleccionar opción ALEATORIA
            if is_empty:
                valid_options = []

                # Esperar opciones válidas para dropdowns que cargan dinámicamente.
                try:
                    def has_valid_options(_):
                        current_select = self.safe_find_element(By.ID, select_id)
                        if not current_select:
                            return False
                        options_now = self._get_valid_select_options(current_select)
                        return options_now if options_now else False

                    valid_options = WebDriverWait(self.driver, 8).until(has_valid_options)
                    select_element = self.safe_find_element(By.ID, select_id) or select_element
                    select = Select(select_element)
                except TimeoutException:
                    # Último intento sin wait para reportar correctamente.
                    valid_options = self._get_valid_select_options(select_element)
                except Exception:
                    valid_options = self._get_valid_select_options(select_element)
                
                if valid_options:
                    random_option = random.choice(valid_options)
                    select.select_by_index(select.options.index(random_option))
                    # Disparar evento change para actualizar dropdowns dependientes
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
                        "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));",
                        select_element
                    )
                    print(f"🎲 {field_name} - Auto-seleccionado (aleatorio): {random_option.text}")
                    self._record_field_value(select_id, random_option.text)
                    return True
                else:
                    print(f" {field_name} - No hay opciones válidas para auto-selección")
                    raise ValueError(f"No se encontraron opciones en el dropdown '{field_name}'")

            # SI HAY VALOR EN EXCEL: match controlado (NO usar select_by_visible_text,
            # que ante un fallo exacto hace un fallback difuso propio de Selenium y agarra
            # otra option — ej. "1 mes" terminaba eligiendo "2 meses" porque contiene "mes").
            try:
                norm_desired = self._normalize_text(option_text)
                desired_plain = option_text.strip()
                opts = list(select.options)
                chosen_idx = None
                # 1) exacto tal cual (texto visible idéntico)
                for i, o in enumerate(opts):
                    if not self._is_placeholder_text(o.text) and (o.text or "").strip() == desired_plain:
                        chosen_idx = i
                        break
                # 2) exacto normalizado (ignora nbsp/acentos/espacios)
                if chosen_idx is None and norm_desired:
                    for i, o in enumerate(opts):
                        if not self._is_placeholder_text(o.text) and self._normalize_text(o.text) == norm_desired:
                            chosen_idx = i
                            break
                # 3) contiene normalizado (último recurso tolerante)
                if chosen_idx is None and norm_desired:
                    for i, o in enumerate(opts):
                        if not self._is_placeholder_text(o.text) and norm_desired in self._normalize_text(o.text):
                            chosen_idx = i
                            break

                if chosen_idx is not None:
                    select.select_by_index(chosen_idx)
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
                        "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));",
                        select_element
                    )
                    selected_text = (opts[chosen_idx].text or option_text).strip()
                    print(f" {field_name} seleccionado: {selected_text}")
                    self._record_field_value(select_id, selected_text)
                    return True

                print(f"❌ [DEBUG-FILL] No se encontró opción para {field_name}: '{option_text}'.")
                options_now = self._get_valid_select_options(select_element)
                try:
                    options = [o.text.strip() for o in select.options if o.text.strip()]
                    print(f"   💡 Opciones disponibles en dropdown '{field_name}': {options}")
                except Exception:
                    pass
                if not options_now:
                    raise ValueError(f"No se encontraron opciones en el dropdown '{field_name}'")
                if not hasattr(self, "_campos_dropdown_no_encontrados"):
                    self._campos_dropdown_no_encontrados = []
                self._campos_dropdown_no_encontrados.append(f"{field_name}: '{option_text}'")
                return False
            except ValueError:
                raise
            except Exception as e:
                print(f" Error resolviendo opción de {field_name}: {e}")
                return False
                
        except Exception as e:
            print(f" Error en {field_name}: {e}")
            return False

    def _select_purchase_date_radio(self, desired_value):
        """Selecciona una opción de radio para 'estimated-date-purchase' si existe.

        Retorna True si se seleccionó, False si falló con radios presentes y
        None si no hay radios (para que el flujo continue con el select/input).
        """
        try:
            radios = self.driver.find_elements(By.CSS_SELECTOR, "input[type='radio'][name='estimated-date-purchase']")
            radios = [radio for radio in radios if radio.is_displayed()]
            if not radios:
                return None

            normalized_value = (desired_value or "").strip().lower()
            
            # Si el Excel no pide un valor específico y ya hay un radio seleccionado, conservarlo
            already_selected = next((r for r in radios if r.is_selected()), None)
            if already_selected and not normalized_value:
                desc = already_selected.get_attribute("title") or already_selected.get_attribute("value") or "opción"
                print(f" Fecha estimada (radio) ya seleccionada ({desc}), se conserva")
                return True

            target_radio = None

            if normalized_value:
                # Normalizar (sin acentos, sin espacios/nbsp) para que "1 mes" del Excel
                # matchee un label "1&nbsp;mes" del form. Match exacto primero, luego contiene.
                norm_desired = self._normalize_text(desired_value)

                def _radio_candidates(radio):
                    texts = [
                        radio.get_attribute("value"),
                        radio.get_attribute("title"),
                        radio.get_attribute("aria-label"),
                    ]
                    rid = radio.get_attribute("id")
                    if rid:
                        try:
                            texts.append(self.driver.find_element(
                                By.CSS_SELECTOR, f"label[for='{rid}']").text)
                        except Exception:
                            pass
                    try:
                        texts.append(radio.find_element(By.XPATH, "./ancestor::label[1]").text)
                    except Exception:
                        pass
                    return [self._normalize_text(t) for t in texts if t]

                radios_norm = [(radio, _radio_candidates(radio)) for radio in radios]
                for radio, cands in radios_norm:  # 1) match exacto normalizado
                    if norm_desired and any(c == norm_desired for c in cands):
                        target_radio = radio
                        break
                if not target_radio:  # 2) fallback: contiene
                    for radio, cands in radios_norm:
                        if norm_desired and any(norm_desired in c for c in cands):
                            target_radio = radio
                            break

                if target_radio:
                    print(f" Fecha estimada (radio) seleccionada: {desired_value}")
                else:
                    print(f" No se encontró radio que coincida con '{desired_value}', seleccionando aleatorio...")

            if not target_radio:
                enabled_radios = [radio for radio in radios if radio.is_enabled()]
                if not enabled_radios:
                    print(" No hay radios habilitados para fecha estimada")
                    return False
                target_radio = random.choice(enabled_radios)
                descripcion = target_radio.get_attribute("title") or target_radio.get_attribute("value") or "opción"
                print(f"🎲 Fecha estimada (radio) auto-seleccionada: {descripcion}")
            else:
                descripcion = target_radio.get_attribute("title") or target_radio.get_attribute("value") or (desired_value or "opción")

            self._scroll_element_into_view(target_radio)
            try:
                target_radio.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", target_radio)

            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
                "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));",
                target_radio
            )
            self._record_field_value("estimated-date-purchase", descripcion)
            return True
        except Exception as e:
            print(f" Error seleccionando fecha estimada (radios): {e}")
            return False

    def _fill_purchase_date_field(self, value, field_name="Fecha estimada"):
        """Gestiona el campo de fecha estimada con radios, select o input."""
        print(f" Intentando llenar {field_name} con valor: '{value}' (vacío: {not value})")
        
        try:
            # Intentar con radios primero
            radio_result = self._select_purchase_date_radio(value)
            if radio_result is True:
                print(f" {field_name} completado exitosamente con radio")
                return True
            if radio_result is False:
                print(f" {field_name} falló con radios")
                return False
            
            # Si no hay radios (None), intentar con select
            print(f" No hay radios para {field_name}, intentando select...")
            success = self.safe_select_option_if_visible("estimated-date-purchase", value, field_name)
            if success:
                print(f" {field_name} completado exitosamente con select")
                return True
            
            # Si el select falló, intentar como input de texto
            print(f" No se pudo seleccionar {field_name} como select, intentando como input...")
            element = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.ID, "estimated-date-purchase"))
            )
            element.clear()
            if value:
                element.send_keys(value)
                print(f" {field_name} completado como texto: {value}")
            else:
                print(f" {field_name} es input de texto pero valor está vacío")
            return True
        except Exception as e:
            print(f" Error con campo {field_name}: {e}")
            return False
    
    def handle_gm_cookie_popup(self):
        """Maneja específicamente el popup de cookies de General Motors/Chevrolet"""
        try:
            selectors = [
                "gb-legal-notification",
                ".js-close-icon",
                ".silent-consent", 
                ".close-btn",
                "gb-legal-notification .close-btn",
            ]
            
            for selector in selectors:
                try:
                    if "gb-legal-notification" in selector:
                        popups = self.driver.find_elements(By.TAG_NAME, "gb-legal-notification")
                        if popups:
                            close_buttons = self.driver.find_elements(By.CSS_SELECTOR, ".close-btn.js-close-icon.silent-consent")
                            if close_buttons:
                                close_button = close_buttons[0]
                                if close_button.is_displayed():
                                    self.driver.execute_script("arguments[0].click();", close_button)
                                    time.sleep(1)
                                    return True
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        for element in elements:
                            if element.is_displayed():
                                self.driver.execute_script("arguments[0].click();", element)
                                time.sleep(1)
                                return True
                except Exception as e:
                    continue

            return False
            
        except Exception as e:
            print(f"Error manejando popup GM: {e}")
            return False
    
    def handle_cookie_popups(self):
        """Maneja popups de cookies y legales que puedan aparecer"""
        if self.handle_gm_cookie_popup():
            return True
        
        cookie_selectors = [
            "button[onclick*='cookie']",
            "button[class*='cookie-accept']",
            "button[id*='cookie-accept']",
            ".cookie-accept",
            "#cookie-accept",
        ]
        
        for selector in cookie_selectors:
            try:
                elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    if element.is_displayed():
                        self.driver.execute_script("arguments[0].click();", element)
                        time.sleep(1)
                        return True
            except Exception as e:
                continue

        return False
    
    def find_and_position_to_form(self, expected_form_url):
        """Encuentra el iframe del formulario y posiciona el viewport sobre él.
        Usa 2 intentos: el primero espera activamente el iframe específico; el segundo
        hace pre-scroll para activar lazy-loading antes de reintentar."""
        expected_form_url = self._sanitize_url(expected_form_url)
        if not isinstance(expected_form_url, str):
            expected_form_url = str(expected_form_url or "").strip()
        else:
            expected_form_url = expected_form_url.strip()
        if not expected_form_url:
            return None

        _url = expected_form_url

        def _find_gm_now(d):
            """Devuelve el iframe GM (o el que matchea expected) si YA está en el DOM, si no None."""
            try:
                iframes = d.find_elements(By.TAG_NAME, "iframe")
            except Exception:
                return None
            cand, _es_gm = self._pick_gm_iframe(iframes, expected_url=_url)
            # Solo aceptar en la espera si es GM o matchea el esperado; el fallback "primer
            # iframe visible" se resuelve más tarde en el loop principal (marca mismatch).
            if cand is not None and (_es_gm or (_url and _url.lower() in self._iframe_src_of(cand).lower())):
                return cand
            return None

        max_attempts = 2
        for attempt in range(max_attempts):
            try:
                # Esperar el iframe GM del formulario (no cualquier iframe). Espera corta el
                # primer intento para ser rápido; si no aparece, pre-scroll (footer) y reintento.
                _wait_s = 6 if attempt == 0 else 8
                try:
                    target_iframe = WebDriverWait(self.driver, _wait_s).until(_find_gm_now)
                except TimeoutException:
                    target_iframe = None

                if not target_iframe:
                    if attempt < max_attempts - 1:
                        try:
                            self.pre_scroll_for_dynamic_content()
                        except Exception:
                            pass
                        time.sleep(3)
                    continue

                # Posicionar viewport sobre el iframe
                location = target_iframe.location
                viewport_height = self.driver.execute_script("return window.innerHeight")
                element_y = location['y']
                if element_y > viewport_height:
                    scroll_amount = element_y - (viewport_height * 0.3)
                    self.driver.execute_script(f"window.scrollTo(0, {scroll_amount});")
                else:
                    scroll_amount = element_y - 100
                    if scroll_amount > 0:
                        self.driver.execute_script(f"window.scrollTo(0, {scroll_amount});")

                time.sleep(1)
                # Registrar URLs en screenshot_manager para el banner
                _found_src = target_iframe.get_attribute("src") or ""
                if self.screenshot_manager:
                    self.screenshot_manager.url_form_esperado   = expected_form_url
                    self.screenshot_manager.url_form_encontrado = _found_src
                self._url_form_encontrado = _found_src
                return target_iframe

            except Exception as e:
                print(f"find_and_position_to_form intento {attempt + 1}: {e}")
                if attempt < max_attempts - 1:
                    time.sleep(1)

        return None
    
    def wait_for_form_ready_in_iframe(self, timeout=3):
        """Espera a que el formulario dentro del iframe esté completamente listo"""
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            time.sleep(0.3)
        except Exception as e:
            print(f"Error en wait_for_form_ready_in_iframe: {e}")
    
   
    def reposition_to_form(self, expected_form_url):
        """Vuelve a posicionarse sobre el formulario para asegurar capturas completas"""
        expected_form_url = self._sanitize_url(expected_form_url)
        if not isinstance(expected_form_url, str):
            expected_form_url = str(expected_form_url or "").strip()
        else:
            expected_form_url = expected_form_url.strip()
        try:
            self.driver.switch_to.default_content()
            if not expected_form_url:
                return True

            iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
            # Priorizar SIEMPRE el iframe GM (gm_forms/gm_front/gm_admin) que matchee el esperado.
            target_iframe, _es_gm = self._pick_gm_iframe(iframes, expected_url=expected_form_url)

            if target_iframe is not None:
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", target_iframe)
                time.sleep(1)
                self.driver.switch_to.frame(target_iframe)
                if self.screenshot_manager:
                    self.screenshot_manager.current_frame = target_iframe
                return _es_gm or (expected_form_url.lower() in self._iframe_src_of(target_iframe).lower())
            else:
                self.driver.switch_to.frame(self.driver.find_elements(By.TAG_NAME, "iframe")[0])
                return False

        except Exception as e:
            print(f"Error al reposicionar: {e}")
            try:
                self.driver.switch_to.frame(self.driver.find_elements(By.TAG_NAME, "iframe")[0])
            except:
                pass
            return False

    def _reload_form_iframe(self, expected_form_url):
        """Recarga SOLO el iframe del formulario (no toda la landing) y deja el driver
        posicionado y con el contexto cambiado adentro del iframe recargado.
        Evita tener que scrollear toda la landing de nuevo en cada reintento.
        Devuelve True si logró recargar y reposicionar sobre un iframe GM."""
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass
        try:
            iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
            target_iframe, es_gm = self._pick_gm_iframe(iframes, expected_url=expected_form_url)
            if target_iframe is None:
                return False
            # Recargar el documento del iframe. Preferimos re-setear el src (fuerza recarga
            # limpia); si no hay src usable, entramos y hacemos location.reload().
            src = self._iframe_src_of(target_iframe)
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", target_iframe)
            time.sleep(0.3)
            if src and src.startswith("http"):
                self.driver.execute_script(
                    "arguments[0].src = arguments[0].src;", target_iframe)
            else:
                self.driver.switch_to.frame(target_iframe)
                self.driver.execute_script("location.reload();")
                self.driver.switch_to.default_content()
            print("↻ Recargado solo el iframe del formulario (sin recargar la landing)")
            time.sleep(2)
            # Reposicionar sobre el iframe (puede ser un elemento nuevo tras la recarga).
            ok = self.reposition_to_form(expected_form_url)
            self.wait_for_form_ready_in_iframe()
            return bool(ok or es_gm)
        except Exception as e:
            print(f"No se pudo recargar solo el iframe: {e}")
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass
            return False

    def pre_scroll_for_dynamic_content(self):
        """Realiza pre-scroll disparando eventos de scroll para activar IntersectionObserver y lazy-loading."""
        total_height = self.driver.execute_script("return document.body.parentNode.scrollHeight")
        viewport_height = self.driver.execute_script("return window.innerHeight")
        scroll_step = max(viewport_height * 0.8, 800)

        headless = self.config.get('headless', False)
        step_wait = 0.5 if headless else 0.15
        end_wait = 1 if headless else 0.3

        _SCROLL_JS = (
            "window.scrollTo(0, arguments[0]);"
            "window.dispatchEvent(new Event('scroll', {bubbles:true,cancelable:false}));"
            "document.dispatchEvent(new Event('scroll', {bubbles:true}));"
        )

        current_position = 0
        while current_position < total_height:
            self.driver.execute_script(_SCROLL_JS, current_position)
            time.sleep(step_wait)
            current_position += scroll_step

        self.driver.execute_script(_SCROLL_JS, 999999)
        time.sleep(end_wait)

        self.driver.execute_script(_SCROLL_JS, 0)
        time.sleep(end_wait)

    def _maybe_click_raq_cta(self, *urls):
        """Brasil RAQ (solicitar-contato / raq / raq-revamp): tiene una pantalla previa con un CTA
        (#contact-by-form) que hay que clickear para llegar al formulario. Aplica tanto si el RAQ
        viene dentro del iframe de una landing como si la URL del form se corre suelta.
        raq-eletricos NO tiene esa pantalla."""
        blob = " ".join((u or "").lower() for u in urls)
        if str(self.config.get("pais", "")).lower() not in ("brasil", "brazil", "br"):
            return
        if "raq" not in blob or "eletricos" in blob:
            return
        if not ("solicitar-contato" in blob or "gm_forms/raq" in blob or "raq-revamp" in blob):
            return
        try:
            _btn = WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((By.ID, "contact-by-form"))
            )
            self._scroll_element_into_view(_btn)
            try:
                _btn.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", _btn)
            print("Click en #contact-by-form (RAQ Brasil)")
            time.sleep(1.5)
        except Exception as _e:
            print(f"#contact-by-form no encontrado: {_e}")

    def _slug_for(self, landing_url, form_url):
        """Slug corto y legible para nombrar las capturas de un form (siempre el mismo para el
        mismo form). Toma el último segmento significativo de la URL del form (o de la landing),
        p.ej. '.../gm_front/form/flotas-pesados?model=...' → 'flotas-pesados'."""
        from urllib.parse import urlsplit
        for url in (form_url or "", landing_url or ""):
            try:
                path = urlsplit(str(url)).path
            except Exception:
                path = ""
            segs = [s for s in path.split("/") if s and s.lower() not in ("form", "forms", "gm_front", "gm_forms", "gm_admin")]
            if segs:
                slug = segs[-1]
                # Limpiar a algo apto para nombre de archivo.
                slug = "".join(c if (c.isalnum() or c in "-_") else "-" for c in slug).strip("-_")
                slug = slug.lower()[:30].strip("-_")
                if slug:
                    return slug
        return ""

    def _detect_landing_error(self, landing_url):
        """Detecta si la landing devolvió 404 / página no encontrada. Heurística por título y
        encabezados visibles (Selenium no expone el status HTTP directamente). Devuelve un
        texto con el motivo, o '' si la landing cargó bien."""
        patterns = (
            "404", "not found", "page not found",
            "no encontrada", "no encontrado", "no existe",
            "não encontrada", "nao encontrada", "não encontrado", "nao encontrado",
        )
        try:
            title_raw = self.driver.title or ""
        except Exception:
            title_raw = ""
        title = title_raw.lower()
        for p in patterns:
            if p in title:
                return f"Landing 404 / no encontrada (título: {title_raw[:80]!r})"
        try:
            heads = self.driver.execute_script(
                "return Array.prototype.slice.call(document.querySelectorAll('h1,h2'),0,6)"
                ".map(function(e){return (e.innerText||'');}).join(' | ').toLowerCase();"
            ) or ""
        except Exception:
            heads = ""
        for p in patterns:
            if p in heads:
                return f"Landing 404 / no encontrada (encabezado contiene '{p}')"
        return ""

    def process_landing_page(self, landing_url, ss_counter, take_screenshot=True):
        """Procesa la página de destino inicial"""
        try:
            self.driver.switch_to.default_content()
        except Exception:
            pass
        landing_url = self._sanitize_url(landing_url)
        self.driver.get(landing_url)

        # Detección de landing rota (404 / página no encontrada): se marca para el Excel.
        try:
            _issue = self._detect_landing_error(landing_url)
            if _issue:
                self._landing_issue = _issue
                print(f"⚠ Landing con problema: {_issue}")
        except Exception:
            pass

        # Tras navegar, el sitio suele re-activar la ventana → re-aplicar 'sin foco' (background).
        try:
            from browser_manager import reapply_background_no_activate
            reapply_background_no_activate(self.driver)
        except Exception:
            pass

        self.handle_cookie_popups()

        try:
            # "interactive" alcanza: con page_load_strategy="eager" el DOM ya está armado ahí;
            # "complete" puede tardar mucho más porque incluye trackers/pixels de terceros de fondo.
            WebDriverWait(self.driver, 8).until(
                lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
            )
        except:
            pass

        extra_wait = 1.5 if self.config.get('headless', False) else 0.4
        time.sleep(extra_wait)

        self.pre_scroll_for_dynamic_content()

        # En headless el JS necesita más tiempo para procesar los eventos de scroll e inyectar el iframe
        if self.config.get('headless', False):
            time.sleep(3)

        # gm_front (React SPA): esperar a que los componentes monten tras el scroll
        if "gm_front" in (landing_url or "").lower():
            time.sleep(3)

        self.handle_cookie_popups()

        # Chevrolet Brasil: la landing es un selector de canal; hay que entrar al formulario primero
        # (RAQ revamp usa el mismo #contact-by-form pero DENTRO del iframe del form, no acá en el
        # top-level — ver el bloque en find_and_position_to_form, después de cambiar de contexto.)
        if "chevrolet.com.br/solicitar-contato" in landing_url:
            try:
                btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "contact-by-form"))
                )
                self.driver.execute_script("arguments[0].click();", btn)
                WebDriverWait(self.driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                time.sleep(2)
                self.handle_cookie_popups()
                print("Chevrolet BR: click en #contact-by-form OK")
            except Exception as e:
                print(f"Chevrolet BR: no se encontró #contact-by-form — {e}")

        form_inserto_name = self.screenshot_manager.fname("landing", "inicial", ss_counter) if self.screenshot_manager else f"landing_inicial_{ss_counter}.png"
        if take_screenshot:
            self.screenshot_manager.url_landing = landing_url
            self.screenshot_manager.take_landing_screenshot(ss_counter, "inicial")
            print("Captura 1/3: Formulario inserto en landing")

        return form_inserto_name

    def _submit_button_css_selectors(self):
        cid = self._GUIDE_SUBMIT_CONTAINER_ID
        wpre = self._GUIDE_SUBMIT_WIDGET_ID_PREFIX
        return (
            f'div[id="{cid}"] button[type="submit"]',
            f'[id^="{cid}"] button[type="submit"]',
            f'button#{wpre}widget',
            f'button[id^="{wpre}"][id$="_widget"]',
            'button[name="guideContainer-rootPanel-guidebutton___jqName"]',
            'button[type="submit"][aria-label="Enviar"]',
            "button.submit-button.stat-button-link",
            "button.btn-visid-submit.stat-button-link",
            "button[type='button'][class*='submit']",
            "button[class*='submit']",
            "button[id*='guidebutton']",
            "button[name*='guidebutton']",
            "[id*='guidebutton'] button",
            "input[type='submit']",
        )

    def _refill_brasil_doc_sendkeys(self, form_data) -> bool:
        """
        Re-ingresa los campos CPF/CNPJ/CEP carácter a carácter via send_keys,
        usando el MISMO valor del Excel. Solo para Brasil.
        Retorna True si re-llenó al menos un campo.
        """
        pais = str(self.config.get("pais", "")).lower()
        if pais not in ("brasil", "brazil", "br"):
            return False

        doc_keywords = ("cpf", "cnpj", "cep", "zip", "postal", "document", "ci")
        rellenados = 0
        for fc in (self.field_mapping or []):
            fid_raw = fc.get("id", "")
            fids    = fid_raw if isinstance(fid_raw, list) else [fid_raw]
            for fid in fids:
                if not fid or not any(k in str(fid).lower() for k in doc_keywords):
                    continue
                valor = self.get_form_value_str(form_data, fc.get("name", fid))
                if not valor:
                    continue
                # Normalizar: extraer dígitos y recuperar cero inicial comido por Excel numérico
                _fid_k = str(fid).lower()
                _min_k = 14 if "cnpj" in _fid_k else (8 if any(x in _fid_k for x in ("cep", "zip", "postal")) else 11)
                _digs = "".join(c for c in valor if c.isdigit())
                if _digs and len(_digs) == _min_k - 1:
                    _digs = _digs.zfill(_min_k)
                if _digs:
                    valor = _digs
                try:
                    els = self.driver.find_elements(By.ID, fid)
                    if not els or not els[0].is_displayed():
                        continue
                    el = els[0]
                    if not self._hard_clear_input(el):
                        print(f"  ⚠ No se pudo vaciar '{fid}' — se omite el re-ingreso "
                              f"para no duplicar el valor")
                        continue
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", el
                    )
                    time.sleep(0.05)
                    for char in str(valor):
                        el.send_keys(char)
                        time.sleep(0.005)
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
                        "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
                        el,
                    )
                    print(f"  ↺ Re-ingresado '{fid}' via send_keys (mismo valor): '{valor}'")
                    rellenados += 1
                except Exception as e:
                    print(f"  ✗ Error send_keys retry '{fid}': {e}")
        return rellenados > 0

    def _describir_errores_visuales(self) -> str:
        """Devuelve todos los errores visuales con campo asociado: 'campo: mensaje, ...'"""
        try:
            errores = self.driver.find_elements(
                By.XPATH,
                "//*[contains(@class,'error') or contains(@style,'red') or contains(@class,'invalid')]"
            )
            errores_visibles = [e for e in errores if e.is_displayed() and e.text.strip()]
        except Exception:
            return ""
        if not errores_visibles:
            return ""
        partes = []
        vistos = set()
        for err_el in errores_visibles:
            texto = err_el.text.strip()
            if texto in vistos:
                continue
            vistos.add(texto)
            campo = ""
            try:
                campo = self.driver.execute_script("""
                    var el = arguments[0];
                    var parent = el.parentElement;
                    for (var i = 0; i < 4 && parent; i++) {
                        var inp = parent.querySelector('input,select,textarea');
                        if (inp) { return inp.id || inp.name || inp.getAttribute('placeholder') || ''; }
                        var lbl = parent.querySelector('label');
                        if (lbl && lbl.textContent.trim()) { return lbl.textContent.trim(); }
                        parent = parent.parentElement;
                    }
                    return '';
                """, err_el) or ""
            except Exception:
                pass
            partes.append(f'{campo}: "{texto}"' if campo else f'"{texto}"')
        return ", ".join(partes)

    def _resolve_submit_button(self, wait_seconds=3):
        """
        Devuelve (botón, selector_usado) o (None, None).
        Usa un único WebDriverWait con todos los CSS combinados (O(1) en vez de O(N×t)).
        """
        # 1. Un solo wait con todos los CSS selectors en paralelo via coma
        combined_css = ", ".join(self._submit_button_css_selectors())
        try:
            btn = WebDriverWait(self.driver, wait_seconds).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, combined_css))
            )
            return btn, "css:combined"
        except Exception:
            pass

        # 2. XPath por texto (ES/PT) — cubre botones sin clase 'submit' pero con texto de envío.
        #    Se hace en minúsculas via translate() para no depender de mayúsculas/acentos.
        _low = ("translate(normalize-space(.),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÃÕÂÊÔÇ','abcdefghijklmnopqrstuvwxyzaeiouaoaeoc')")
        _send_words = ("enviar", "submit", "solicitar", "finalizar", "confirmar",
                       "quero", "cadastrar", "registrar", "receber")
        _txt_cond = " or ".join(f"contains({_low}, '{w}')" for w in _send_words)
        xpath_combined = (
            "//button[contains(@class,'submit') and not(contains(@class,'next'))]"
            f" | //button[({_txt_cond}) and not(contains(@class,'next')) and not(contains(@class,'cookie'))]"
            f" | //button[@type='submit']"
            f" | //input[@type='submit']"
            f" | //*[@role='button'][{_txt_cond}]"
            f" | //a[contains(@class,'submit')][{_txt_cond}]"
        )
        try:
            btn = WebDriverWait(self.driver, wait_seconds).until(
                EC.element_to_be_clickable((By.XPATH, xpath_combined))
            )
            return btn, "xpath:texto"
        except Exception:
            pass

        # 3. Fallback genérico: barrer TODOS los botones/inputs submit visibles dentro de un <form>
        #    y quedarse con el que no sea 'Siguiente' ni cierre/cookie. Esto ataca el bug de
        #    "no encuentra el botón que sí está" cuando el markup no matchea los selectores fijos.
        try:
            candidatos = self.driver.find_elements(
                By.CSS_SELECTOR,
                "form button, form input[type='submit'], form [role='button']")
            if not candidatos:
                candidatos = self.driver.find_elements(
                    By.CSS_SELECTOR, "button, input[type='submit'], [role='button']")
            for el in candidatos:
                try:
                    if not (el.is_displayed() and el.is_enabled()):
                        continue
                    if self._is_next_button_element(el):
                        continue
                    blob = self._normalize_text(
                        " ".join(filter(None, [
                            el.text or "",
                            el.get_attribute("class") or "",
                            el.get_attribute("aria-label") or "",
                            el.get_attribute("id") or "",
                            el.get_attribute("type") or "",
                        ])))
                    if any(bad in blob for bad in ("cookie", "cerrar", "close", "aceptar cookies", "fechar")):
                        continue
                    # Es un submit real si type=submit, o su texto/atributos sugieren envío.
                    _is_type_submit = (el.get_attribute("type") or "").lower() == "submit"
                    if _is_type_submit or any(w in blob for w in _send_words):
                        return el, "scan:form"
                except Exception:
                    continue
        except Exception:
            pass
        return None, None

    def fill_fields_present(self, form_data, field_mapping=None):
        """
        Llena campos usando mapping pero SOLO los presentes en el DOM actual.
        Útil para formularios multi-paso: no toma screenshots ni marca términos.
        """
        mapping = field_mapping or self.field_mapping or []

        resolved_field_configs = []
        for field_config in mapping:
            field_id = field_config.get("id")
            ids = field_id if isinstance(field_id, list) else [field_id]
            chosen_id = self._pick_first_existing_id(ids)
            if not chosen_id:
                continue
            resolved = dict(field_config)
            resolved["__resolved_id"] = chosen_id
            resolved_field_configs.append(resolved)

        for field_config in resolved_field_configs:
            field_type = field_config.get("type", "text")
            field_id = field_config.get("__resolved_id") or field_config.get("id")
            field_name = field_config.get("name", field_id)
            field_value = self._resolve_field_value(form_data, field_config)

            if field_type != "select" and not field_value:
                continue

            if field_type == "select":
                original_ids = field_config.get("id")
                fallback_ids = original_ids if isinstance(original_ids, list) else [original_ids]

                if field_id == "event":
                    self._select_dropdown_with_fallback(field_id, field_value, field_name, fallback_ids=fallback_ids)
                elif field_id == "estimated-date-purchase":
                    self._fill_purchase_date_field(field_value, field_name)
                else:
                    self.safe_select_option_if_visible(field_id, field_value, field_name)
                continue

            # text / textarea — ambos se tratan igual (input y textarea por ID)
            try:
                element = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, field_id))
                )
                try:
                    element.clear()
                    element.send_keys(field_value)
                except Exception:
                    # Fallback JS para textareas que no responden a send_keys
                    self.driver.execute_script(
                        "arguments[0].focus();"
                        "arguments[0].value = arguments[1];"
                        "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));"
                        "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
                        "arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));",
                        element,
                        field_value,
                    )
                print(f"{field_name} completado ({field_id}): {field_value}")
                self._record_field_value(field_id, field_value)
            except Exception:
                # No bloquear el flujo por un campo puntual
                continue

        self._auto_fill_unmapped_dropdowns(mapping)

    # --- Helpers reutilizables para formularios de 3 pasos ---
    def _fill_optional_model_and_kit(self, data, model_id="models", kit_id="kits[]", model_key="model", kit_key="kit"):
        """Reaplica la selección de modelo y kit si vuelven a mostrarse en pasos posteriores"""
        if not self._is_visible(By.ID, model_id):
            print(f" Modelo select (ID: {model_id}) no visible en paso 3")
            return

        modelo_valor = data.get(model_key, "") if isinstance(data, dict) else ""
        if not modelo_valor:
            print(" Modelo vacío o no disponible en los datos (paso 3)")
            return

        try:
            self.safe_select_option_if_visible(model_id, modelo_valor, "Modelo (paso 3)")
            time.sleep(1)

            kit_valor = data.get(kit_key, "") if isinstance(data, dict) else ""
            
            if kit_valor and self._is_visible(By.ID, kit_id):
                print(f" Intentando seleccionar kit: {kit_valor}")
                time.sleep(0.5)
                if not self.safe_select_option_if_visible(kit_id, kit_valor, "Kit"):
                    print(f" No se pudo seleccionar kit con valor: {kit_valor}")
            elif not kit_valor:
                print(f" Kit valor está vacío")
            elif not self._is_visible(By.ID, kit_id):
                print(f" Kit select (ID: {kit_id}) no está visible")
        except Exception as e:
            print(f" Error completando modelo/kit en paso 3: {e}")

    def _fill_text_fields_with_fallback_ids(self, fields):
        """Completa campos de texto aceptando múltiples IDs alternativos por campo"""
        for field in fields:
            name = field.get('name', 'Campo')
            value = field.get('value', '')
            ids = field.get('ids') or []

            if isinstance(ids, str):
                ids = [ids]

            if not value:
                print(f" Saltando {name}: valor vacío")
                continue

            completado = False
            for field_id in ids:
                if not self._is_visible(By.ID, field_id):
                    continue

                try:
                    element = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.ID, field_id))
                    )
                    element.clear()
                    element.send_keys(value)
                    print(f" {name} completado ({field_id}): {value}")
                    completado = True
                    break
                except Exception as e:
                    print(f" Error completando {name} con id '{field_id}': {e}")

            if not completado:
                print(f" No se pudo completar {name}: elemento no disponible")

    def _fill_mileage_field(self, value, field_ids=None, field_name="Kilometraje"):
        """Completa el campo de kilometraje aceptando inputs o selects (con auto-selección si vacío)"""
        is_empty = not value or str(value).strip() == ""
        
        ids = field_ids or ['estimated-mileage', 'mileage']

        for field_id in ids:
            if not self._is_visible(By.ID, field_id):
                continue

            try:
                element = self.driver.find_element(By.ID, field_id)
                tag_name = element.tag_name.lower()

                if tag_name == 'select':
                    # Para select: intentar auto-selección aunque esté vacío
                    if self.safe_select_option_if_visible(field_id, value, field_name):
                        return True
                else:
                    # Para input text: solo completar si hay valor
                    if is_empty:
                        print(f" Saltando {field_name} (input text): valor vacío")
                        return False
                    
                    element = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.ID, field_id))
                    )
                    element.clear()
                    element.send_keys(value)
                    print(f" {field_name} completado ({field_id}): {value}")
                    return True
            except Exception as e:
                print(f" Error completando {field_name} con id '{field_id}': {e}")

        if is_empty:
            print(f" {field_name}: no se encontró select para auto-selección")
        else:
            print(f" No se pudo completar {field_name}: elemento no disponible")
        return False

    def _select_dropdown_with_fallback(self, select_id, desired_value, field_name, allow_text_fallback=True, fallback_ids=None):
        """Selecciona opciones tolerando diferencias de texto, con auto-selección aleatoria si está vacío"""
        # Determinar si el campo está vacío
        is_empty = not desired_value or str(desired_value).strip() == ""
        
        fallback_ids = fallback_ids or [select_id]

        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, select_id))
            )
        except Exception as e:
            print(f" {field_name} con id='{select_id}' no disponible: {e}")
            if allow_text_fallback and not is_empty:
                return self._fill_text_field_direct(fallback_ids, desired_value, field_name)
            return False

        tag_name = ''
        try:
            tag_name = element.tag_name.lower()
        except Exception:
            tag_name = ''

        if tag_name != 'select':
            print(f" {field_name} con id='{select_id}' no es dropdown, intentando como texto...")
            if allow_text_fallback and not is_empty:
                return self._fill_text_field_direct(fallback_ids, desired_value, field_name)
            return False

        try:
            self._scroll_element_into_view(element)
            time.sleep(0.3)
        except Exception:
            pass

        options = element.find_elements(By.TAG_NAME, "option")
        if not options:
            print(f" {field_name} no tiene opciones disponibles")
            raise ValueError(f"No se encontraron opciones en el dropdown '{field_name}'")

        if is_empty:
            try:
                selected_option = Select(element).first_selected_option
            except Exception:
                selected_option = None

            if selected_option:
                selected_text = (selected_option.text or "").strip()
                if selected_text and not self._is_placeholder_text(selected_text):
                    print(f" {field_name} ya tiene valor válido ('{selected_text}'), se conserva")
                    return True

        # SI EL CAMPO ESTÁ VACÍO: Auto-seleccionar opción ALEATORIA
        if is_empty:
            valid_options = []
            for opt in options:
                opt_text = opt.text.strip()
                is_placeholder = self._is_placeholder_text(opt_text)
                if opt_text and not opt.get_attribute("disabled") and not is_placeholder:
                    valid_options.append(opt)
            
            if valid_options:
                random_option = random.choice(valid_options)
                option_value = random_option.get_attribute("value") or random_option.text
                option_text = random_option.text.strip()
                try:
                    Select(element).select_by_value(option_value)
                    # Disparar evento change para actualizar dropdowns dependientes
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));"
                        "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));",
                        element
                    )
                    print(f"🎲 {field_name} - Auto-seleccionado (aleatorio): {option_text}")
                    return True
                except:
                    pass
            print(f" {field_name} - No hay opciones válidas para auto-selección")
            raise ValueError(f"No se encontraron opciones en el dropdown '{field_name}'")

        # SI HAY VALOR: Intentar usar el método optimizado primero
        try:
            if self.safe_select_option_if_visible(select_id, desired_value, field_name):
                return True
        except Exception as e:
            if isinstance(e, ValueError) and "No se encontraron opciones" in str(e):
                raise

        # Búsqueda avanzada con normalización
        normalized_target = self._normalize_text(desired_value)
        fallback_option = None

        for opt in options:
            opt_text = opt.text.strip()
            opt_value = opt.get_attribute("value") or ""

            if opt_text and desired_value.strip().lower() == opt_text.lower():
                fallback_option = opt
                break

            if normalized_target:
                if normalized_target == self._normalize_text(opt_text) or normalized_target == self._normalize_text(opt_value):
                    fallback_option = opt
                    break

        if not fallback_option and normalized_target:
            for opt in options:
                opt_text = opt.text.strip()
                opt_value = opt.get_attribute("value") or ""

                if normalized_target in self._normalize_text(opt_text) or normalized_target in self._normalize_text(opt_value):
                    fallback_option = opt
                    break

        if not fallback_option:
            print(f" Sin coincidencias para {field_name} con valor '{desired_value}'")
            if allow_text_fallback:
                return self._fill_text_field_direct(fallback_ids, desired_value, field_name)
            return False

        option_value = fallback_option.get_attribute("value") or fallback_option.text
        option_text = fallback_option.text.strip()

        try:
            Select(element).select_by_value(option_value)
            print(f" {field_name} seleccionado (match flexible): {option_text or option_value}")
            return True
        except Exception:
            try:
                if option_text:
                    Select(element).select_by_visible_text(option_text)
                    print(f" {field_name} seleccionado (por texto): {option_text}")
                    return True
            except Exception:
                pass

        try:
            self.driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                element,
                option_value
            )
            print(f" {field_name} asignado vía JS: {option_text or option_value}")
            return True
        except Exception as e:
            print(f" Error completando {field_name} con id='{select_id}': {e}")
            if allow_text_fallback:
                return self._fill_text_field_direct(fallback_ids, desired_value, field_name)
            return False

    def _fill_text_field_direct(self, field_ids, value, field_name):
        """Intenta rellenar un campo de texto probando múltiples IDs"""
        for field_id in field_ids:
            try:
                element = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, field_id))
                )
                element.clear()
                element.send_keys(value)
                print(f" {field_name} completado como texto ({field_id}): {value}")
                return True
            except Exception as e:
                print(f" No se pudo completar {field_name} con id '{field_id}' como texto: {e}")
        print(f" No se pudo completar {field_name} como texto")
        return False

    # Tabla de inferencia: patrones de texto → data_key conocido
    _INFER_DATAKEY_MAP = {
        "firstname":  ("nombre", "name", "first", "fname", "nome", "given"),
        "lastname":   ("apellido", "lastname", "surname", "lname", "sobrenome", "family"),
        "email":      ("email", "correo", "mail", "e-mail", "emailaddress"),
        "phone":      ("telefono", "celular", "phone", "mobile", "cel", "tel", "fono", "movil",
                       "whatsapp", "celphone", "telefone"),
        "document":   ("documento", "doc", "cedula", "cedula", "dni", "rut", "ci", "cpf",
                       "rfc", "identif", "passport", "pasaporte"),
        "model":      ("modelo", "model", "vehiculo", "veiculo", "auto", "car"),
    }

    @staticmethod
    def _normalize_for_infer(text):
        """Minúsculas, sin tildes, sin espacios/guiones."""
        import unicodedata
        nfkd = unicodedata.normalize("NFKD", (text or "").lower())
        ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
        return ascii_str.replace(" ", "").replace("-", "").replace("_", "")

    def _fetch_4devs(self, acao, extra_params=None, timeout=8):
        """Llama a la API de 4devs y retorna el texto de respuesta."""
        import urllib.request, urllib.parse
        data = {"acao": acao}
        if extra_params:
            data.update(extra_params)
        body = urllib.parse.urlencode(data).encode()
        req = urllib.request.Request(
            "https://www.4devs.com.br/ferramentas_online.php",
            data=body,
            headers={"Content-Type": "application/x-www-form-urlencoded",
                     "User-Agent": "Mozilla/5.0"},
        )
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read().decode("utf-8").strip()

    def _generate_valid_cpf(self):
        """Fallback: genera CPF con dígitos verificadores válidos."""
        while True:
            d = [random.randint(0, 9) for _ in range(9)]
            if len(set(d)) == 1:
                continue
            r1 = sum(v * (10 - i) for i, v in enumerate(d)) % 11
            c1 = 0 if r1 < 2 else 11 - r1
            d2 = d + [c1]
            r2 = sum(v * (11 - i) for i, v in enumerate(d2)) % 11
            c2 = 0 if r2 < 2 else 11 - r2
            return "".join(map(str, d + [c1, c2]))

    def _generate_valid_cnpj(self):
        """Fallback: genera CNPJ con dígitos verificadores válidos (con puntuación)."""
        b = [random.randint(0, 9) for _ in range(8)] + [0, 0, 0, 1]
        w1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        r1 = sum(v * w for v, w in zip(b, w1)) % 11
        c1 = 0 if r1 < 2 else 11 - r1
        w2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        r2 = sum(v * w for v, w in zip(b + [c1], w2)) % 11
        c2 = 0 if r2 < 2 else 11 - r2
        digits = "".join(map(str, b + [c1, c2]))
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"

    _VALID_CEPS = [
        "01310100", "04538133", "20040020", "30112010",
        "40015970", "60060100", "80010010", "90010280",
        "69010060", "74805100",
    ]

    def _generate_brazil_document(self, field_id):
        """Genera CPF (11 díg), CNPJ (14 díg) o CEP (8 díg) via API 4devs con fallback local."""
        import re as _re
        fid = field_id.lower()
        is_cep  = "cep" in fid or "zip" in fid or "postal" in fid
        is_cnpj = "cnpj" in fid

        if is_cep:
            html = self._fetch_4devs("gerar_cep", {
                "estado": "", "cidade": "São Paulo",
                "bairro": "", "tipo_cep": "residencial",
            })
            m = _re.search(r'(\d{5}-\d{3})', html)
            if m:
                val = m.group(1).replace("-", "")
                if len(val) >= 8:
                    return val
            return random.choice(self._VALID_CEPS)

        if is_cnpj:
            raw = self._fetch_4devs("gerar_cnpj", {"pontuacao": "S"})
            digits = "".join(c for c in (raw or "") if c.isdigit())
            if len(digits) >= 14:
                return raw
            return self._generate_valid_cnpj()

        # CPF — solo dígitos, mínimo 11
        raw = self._fetch_4devs("gerar_cpf", {"pontuacao": "N"})
        digits = "".join(c for c in (raw or "") if c.isdigit())
        if len(digits) >= 11:
            return digits
        return self._generate_valid_cpf()

    def _sanitize_peru_document(self, doc_type_value, raw_value):
        """Corrige el número de documento según las reglas de Perú."""
        dt = (doc_type_value or "").lower()
        if "dni" in dt:
            required_len, no_leading_zero = 8, True
        elif "ruc" in dt:
            required_len, no_leading_zero = 11, False
        elif "pasaporte" in dt:
            # El pasaporte es ALFANUMÉRICO: se conservan las letras del valor original y se
            # completa con caracteres alfanuméricos, no con dígitos.
            alnum = "".join(c for c in str(raw_value or "") if c.isalnum()).upper()
            _pool = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
            while len(alnum) < 12:
                alnum += random.choice(_pool)
            return alnum[:12]
        elif "carn" in dt or "extran" in dt:
            required_len, no_leading_zero = 12, False
        else:
            return raw_value

        digits = "".join(c for c in str(raw_value or "") if c.isdigit())
        while len(digits) < required_len:
            digits += str(random.randint(0, 9))
        digits = digits[:required_len]
        if no_leading_zero and digits[0] == "0":
            digits = str(random.randint(1, 9)) + digits[1:]
        return digits

    def _generate_synthetic_value(self, element):
        """Genera un valor aleatorio para un campo requerido sin datos, según sus atributos HTML."""
        import re as _re
        tag = (element.tag_name or "").lower()
        input_type = (element.get_attribute("type") or "text").lower() if tag == "input" else "textarea"

        # Brasil: detectar CPF/CNPJ/CEP por ID, name o placeholder
        if str(self.config.get("pais", "")).lower() in ("brasil", "brazil", "br"):
            _hints = " ".join(filter(None, [
                element.get_attribute("id") or "",
                element.get_attribute("name") or "",
                element.get_attribute("placeholder") or "",
                element.get_attribute("aria-label") or "",
            ])).lower()
            if "cep" in _hints or "zip" in _hints or "postal" in _hints:
                return self._generate_brazil_document("cep")
            if "cnpj" in _hints:
                return self._generate_brazil_document("cnpj")
            if "cpf" in _hints:
                return self._generate_brazil_document("cpf")
        maxlength = element.get_attribute("maxlength")
        minlength = element.get_attribute("minlength")
        pattern   = element.get_attribute("pattern") or ""
        inputmode = (element.get_attribute("inputmode") or "").lower()

        # Hints generales (id/name/placeholder/aria-label) para inferir tipo semántico
        _all_hints = " ".join(filter(None, [
            element.get_attribute("id") or "",
            element.get_attribute("name") or "",
            element.get_attribute("placeholder") or "",
            element.get_attribute("aria-label") or "",
        ])).lower()

        # Fecha (nacimiento / data de nascimento / fecha): generar fecha válida de adulto
        _date_hints = ("nascimento", "nacimiento", "birth", "dob", "aniversario",
                       "fecha", "data-de", "data_de", "dataname", "cumple")
        if input_type == "date" or any(h in _all_hints for h in _date_hints):
            y = random.randint(1975, 2003)
            m = random.randint(1, 12)
            d = random.randint(1, 28)
            if input_type == "date":
                return f"{y:04d}-{m:02d}-{d:02d}"
            return f"{d:02d}/{m:02d}/{y:04d}"

        max_len = min(int(maxlength) if maxlength and str(maxlength).isdigit() else 20, 30)
        min_len = max(int(minlength) if minlength and str(minlength).isdigit() else 1, 1)
        target  = max(min_len, min(max_len, 10))

        if input_type == "email":
            _fd = getattr(self, "_current_form_data", {}) or {}
            _first = "".join(c for c in str(_fd.get("firstname") or "auto") if c.isalpha())[:8]
            _last  = "".join(c for c in str(_fd.get("lastname")  or "")      if c.isalpha())[:8]
            _pais_map = {
                "argentina": "ar", "bolivia": "bo", "brasil": "br", "brazil": "br",
                "chile": "cl", "colombia": "co", "ecuador": "ec",
                "paraguay": "py", "peru": "pe", "uruguay": "uy",
            }
            _pais_raw = str(self.config.get("pais") or "").lower().strip()
            _pais_abbr = _pais_map.get(_pais_raw, _pais_raw[:2]) if _pais_raw else "xx"
            _nums = str(random.randint(10, 99))
            _local = f"{_first}{_last}_{_pais_abbr}{_nums}" if _last else f"{_first}_{_pais_abbr}{_nums}"
            return f"{_local}@mrm.com"

        is_numeric = (
            input_type in ("number", "tel")
            or inputmode in ("numeric", "tel", "decimal")
            or bool(pattern and _re.search(r'^\^?\[?0-9', pattern))
            or any(h in _all_hints for h in ("numero", "número", "number", "quantidade", "cantidad"))
        )
        if is_numeric:
            _n = max(1, min(target, 4))  # números cortos (ej. número de casa)
            return "".join(str(random.randint(0, 9)) for _ in range(_n))

        # textarea / texto libre: alfanumérico
        chars = "abcdefghijklmnopqrstuvwxyz0123456789"
        return "".join(random.choice(chars) for _ in range(target))

    def _is_field_required(self, element):
        """True si el campo es requerido (attr required/aria-required o su label trae '*')."""
        try:
            return bool(self.driver.execute_script(
                "var el=arguments[0];"
                "if(el.required || el.getAttribute('aria-required')==='true') return true;"
                "var id=el.id;"
                "if(id){var l=document.querySelector('label[for=\"'+(window.CSS&&CSS.escape?CSS.escape(id):id)+'\"]');"
                "  if(l && (l.textContent||'').indexOf('*')>=0) return true;}"
                "var p=el.closest('label'); if(p && (p.textContent||'').indexOf('*')>=0) return true;"
                "return false;",
                element,
            ))
        except Exception:
            return bool(element.get_attribute("required") or element.get_attribute("aria-required") == "true")

    def _fill_unmapped_required_fields(self):
        """Rellena con valores sintéticos los campos REQUERIDOS visibles que quedaron vacíos
        (no mapeados o sin dato en el Excel). Cubre campos extra de ciertos forms
        (ej. Rua, Número, Data de Nascimento) sin tener que mapearlos uno por uno."""
        try:
            els = self.driver.find_elements(
                By.XPATH,
                "//input[not(@type='hidden')][not(@type='submit')][not(@type='button')]"
                "[not(@type='checkbox')][not(@type='radio')][not(@type='file')][not(@type='reset')]"
                " | //textarea",
            )
        except Exception:
            return 0
        filled = 0
        for el in els:
            try:
                if not el.is_displayed() or not el.is_enabled():
                    continue
                if el.get_attribute("disabled") is not None:
                    continue
                # OJO: no saltar readonly — los datepickers de fecha (ej. Data de Nascimento)
                # suelen ser readonly y aun así requeridos; se setean por JS igual.
                if (el.get_attribute("value") or "").strip():
                    continue  # ya tiene dato (mapeado, autocompletado por CEP, etc.)
                if not self._is_field_required(el):
                    continue
                value = self._generate_synthetic_value(el)
                if not value:
                    continue
                self._scroll_element_into_view(el)
                self._fill_and_dispatch(el, value)
                fid = el.get_attribute("id") or el.get_attribute("name") or ""
                print(f"🧩 Requerido sin dato completado con valor sintético: {fid} = {value}")
                self._record_field_value(fid, value)
                filled += 1
            except StaleElementReferenceException:
                continue
            except Exception as e:
                print(f" Requerido sintético: error — {e}")
        if filled:
            print(f"🧩 {filled} campo(s) requerido(s) sin dato completado(s) con valor sintético")
        return filled

    def _infer_data_key(self, field_id, field_name, field_placeholder=""):
        """
        Intenta deducir el data_key del campo a partir de su id/nombre/placeholder.
        Retorna un string con el data_key inferido, o None si no hay match.
        """
        tokens = [
            self._normalize_for_infer(field_id),
            self._normalize_for_infer(field_name),
            self._normalize_for_infer(field_placeholder),
        ]
        for data_key, keywords in self._INFER_DATAKEY_MAP.items():
            for token in tokens:
                if token and any(kw in token or token in kw for kw in keywords):
                    return data_key
        return None

    def _discover_and_report_unmapped_fields(self):
        """
        Escanea TODOS los campos visibles del formulario activo.
        Los que no están en el mapping se agregan como auto_discovered=True,
        se persisten en fixed_field_mappings.json y se reportan en json/nuevos_campos_<pais>.json.
        Llena los requeridos en el siguiente ciclo; los opcionales se mapean pero no se llenan.
        """
        mapped_ids = self._get_mapped_select_ids()
        
        # Ampliar mapped_ids para incluir todos sus alias de visid y evitar falsas detecciones
        expanded_mapped_ids = set(mapped_ids)
        for original_id in mapped_ids:
            alias = self._VISID_ID_ALIASES.get(original_id)
            if alias:
                expanded_mapped_ids.add(alias)
            for k, v in self._VISID_ID_ALIASES.items():
                if v == original_id:
                    expanded_mapped_ids.add(k)
        
        nuevos = []

        scan_targets = [
            ("//select", "select"),
            (
                "//input[not(@type='hidden')][not(@type='submit')]"
                "[not(@type='button')][not(@type='checkbox')][not(@type='radio')]"
                "[not(@type='file')][not(@type='reset')]",
                "text",
            ),
            ("//textarea", "textarea"),
        ]

        for xpath, ftype in scan_targets:
            try:
                elements = self.driver.find_elements(By.XPATH, xpath)
            except Exception:
                continue
            for el in elements:
                try:
                    fid = el.get_attribute("id") or el.get_attribute("name")
                    if not fid or fid in expanded_mapped_ids:
                        continue
                    if not el.is_displayed():
                        continue
                    required = bool(
                        el.get_attribute("required")
                        or el.get_attribute("aria-required") == "true"
                    )
                    label = (
                        el.get_attribute("aria-label")
                        or el.get_attribute("placeholder")
                        or el.get_attribute("name")
                        or fid
                    )
                    placeholder = el.get_attribute("placeholder") or ""
                    campo = {
                        "id": fid,
                        "type": ftype,
                        "name": label,
                        "required": required,
                        "data_index": None,
                        "auto_discovered": True,
                        "optional": not required,
                    }
                    # Para campos de texto requeridos, intentar inferir data_key
                    # Esto permite que se rellenen en la misma ejecución
                    if required and ftype in ("text", "textarea"):
                        inferred = self._infer_data_key(fid, label, placeholder)
                        if inferred:
                            campo["data_key"] = inferred
                            print(f"🔎 Campo requerido '{fid}' → inferido como '{inferred}'")
                    nuevos.append(campo)
                    mapped_ids.add(fid)
                except Exception:
                    continue

        if nuevos:
            self._campos_nuevos_detectados.extend(nuevos)
            self._persistir_campos_nuevos(nuevos)

        return nuevos

    def _persistir_campos_nuevos(self, nuevos):
        """Agrega los campos nuevos al mapping del país y genera el reporte JSON."""
        if not nuevos:
            return

        pais = str(self.config.get("pais") or "").strip()
        if not pais:
            return

        # Agregar al mapping activo en memoria (para que _auto_fill los reconozca)
        ids_ya_en_mapping = self._get_mapped_select_ids()
        for campo in nuevos:
            if campo["id"] not in ids_ya_en_mapping:
                self.field_mapping.append(campo)

        # Persistir en fixed_field_mappings.json
        try:
            from utils.fixed_field_mapping_store import (
                save_country_fixed_field_mapping,
                load_effective_country_form_config,
            )
            effective = load_effective_country_form_config(pais)
            existing_mapping = list(effective.get("field_mapping") or [])
            existing_ids = set()
            for entry in existing_mapping:
                fid = entry.get("id")
                if isinstance(fid, list):
                    existing_ids.update(fid)
                elif fid:
                    existing_ids.add(fid)
            campos_a_agregar = [c for c in nuevos if c["id"] not in existing_ids]
            if campos_a_agregar:
                merged = existing_mapping + campos_a_agregar
                save_country_fixed_field_mapping(pais, merged)
                print(f"Auto-mapeo: {len(campos_a_agregar)} campo(s) nuevo(s) guardado(s) para {pais}")
        except Exception as e:
            print(f"No se pudo persistir campos nuevos en fixed_field_mappings: {e}")

        # Agregar campos nuevos al field_validation_rules_{pais}.json para que aparezcan en la UI de validación
        try:
            _pais_key_vr = pais.lower().replace(" ", "_")
            _vr_path = os.path.join(self.BASE_DIR, "json", f"field_validation_rules_{_pais_key_vr}.json")
            if os.path.exists(_vr_path):
                with open(_vr_path, "r", encoding="utf-8") as _f:
                    _vr_data = json.load(_f)
                _vr_fields = _vr_data.get("fields") or {}
                _existing_element_ids = {v.get("element_id") for v in _vr_fields.values() if v.get("element_id")}
                _added_vr = False
                for _c in nuevos:
                    if _c["id"] in _existing_element_ids:
                        continue
                    _key = _c["id"].upper()
                    _vr_fields[_key] = {
                        "descripcion": _c.get("name") or _c["id"],
                        "campo": _c["id"],
                        "element_id": _c["id"],
                        "regex_full": "",
                        "regex_char": "",
                        "test_text": "",
                        "dropdown": _c.get("type") == "select",
                        "dropdown_error_message": "",
                        "dependencies": [],
                        "paises": [pais],
                        "teclado_mobile": False,
                        "rules": {},
                        "error_messages": {},
                        "error_message_patterns": [],
                        "error_config": {},
                        "error_priority": [],
                    }
                    _added_vr = True
                if _added_vr:
                    _vr_data["fields"] = _vr_fields
                    _tmp_vr = _vr_path + ".tmp"
                    with open(_tmp_vr, "w", encoding="utf-8") as _f:
                        json.dump(_vr_data, _f, ensure_ascii=False, indent=2)
                    os.replace(_tmp_vr, _vr_path)
                    print(f"Campos nuevos agregados a {os.path.basename(_vr_path)}")
        except Exception as _e_vr:
            print(f"No se pudo actualizar field_validation_rules: {_e_vr}")

        # Generar reporte json/nuevos_campos_<pais>.json
        try:
            from datetime import datetime as _dt
            pais_key = pais.lower().replace(" ", "_")
            reporte_path = os.path.join(self.BASE_DIR, "json", f"nuevos_campos_{pais_key}.json")
            existing_report = {}
            if os.path.exists(reporte_path):
                try:
                    with open(reporte_path, "r", encoding="utf-8") as _f:
                        existing_report = json.load(_f)
                except Exception:
                    pass
            campos_previos = existing_report.get("campos_nuevos", [])
            ids_previos = {c.get("id") for c in campos_previos}
            campos_realmente_nuevos = [c for c in nuevos if c["id"] not in ids_previos]
            if campos_realmente_nuevos:
                reporte = {
                    "pais": pais,
                    "ultima_deteccion": _dt.now().isoformat(timespec="seconds"),
                    "campos_nuevos": campos_previos + [
                        {
                            "id": c["id"],
                            "type": c["type"],
                            "required": c["required"],
                            "label": c.get("name", c["id"]),
                        }
                        for c in campos_realmente_nuevos
                    ],
                }
                tmp = reporte_path + ".tmp"
                with open(tmp, "w", encoding="utf-8") as _f:
                    json.dump(reporte, _f, ensure_ascii=False, indent=2)
                os.replace(tmp, reporte_path)
                print(f"Reporte de campos nuevos guardado: {os.path.basename(reporte_path)}")
        except Exception as e:
            print(f"No se pudo escribir reporte de campos nuevos: {e}")

    def _get_mapped_select_ids(self, field_mapping=None):
        """Obtiene IDs de campos que tienen un valor de datos asignado (data_index o data_key).
        Los campos auto-descubiertos sin asignación NO se incluyen para que _auto_fill los rellene."""
        mapped_ids = set()
        mapping = field_mapping if field_mapping is not None else self.field_mapping or []
        for field_config in mapping:
            has_data = (
                field_config.get("data_index") is not None
                or field_config.get("data_key")
            )
            if not has_data:
                continue
            field_id = field_config.get("id")
            if isinstance(field_id, list):
                for fid in field_id:
                    if fid:
                        mapped_ids.add(fid)
            elif field_id:
                mapped_ids.add(field_id)
        return mapped_ids

    def _cargar_ids_dinamicos(self):
        """Carga IDs dinámicos aplicables al país actual: globales + específicos de país."""
        ids_por_id = {}
        pais_actual = str(self.config.get("pais") or "").strip()

        def _agregar_valores(entry_id, raw_value):
            valores = self._resolve_dynamic_id_values(raw_value)
            if not valores:
                return
            if entry_id not in ids_por_id:
                ids_por_id[entry_id] = []
            for val in valores:
                if val not in ids_por_id[entry_id]:
                    ids_por_id[entry_id].append(val)

        def _normalizar_paises(raw):
            if raw is None:
                return []
            if isinstance(raw, str):
                candidatos = [raw]
            elif isinstance(raw, (list, tuple, set)):
                candidatos = list(raw)
            else:
                return []
            salida = []
            for item in candidatos:
                texto = str(item).strip()
                if texto and texto not in salida:
                    salida.append(texto)
            return salida

        try:
            path = os.path.join(self.BASE_DIR, "json", "ids_dinamicos.json")
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if isinstance(data, dict) and isinstance(data.get("entries"), list):
                    for entry in data.get("entries", []):
                        if not isinstance(entry, dict):
                            continue
                        entry_id = str(entry.get("id") or "").strip()
                        if not entry_id:
                            continue

                        paises_entry = _normalizar_paises(entry.get("paises", entry.get("countries")))
                        if paises_entry and pais_actual not in paises_entry:
                            continue

                        raw_value = (
                            entry.get("valor")
                            if "valor" in entry
                            else entry.get("valores", entry.get("value", entry.get("values")))
                        )
                        _agregar_valores(entry_id, raw_value)
                elif isinstance(data, dict):
                    # Legacy: {id: valor}
                    for entry_id, raw_value in data.items():
                        if entry_id in {"version", "entries"}:
                            continue
                        entry_id = str(entry_id).strip()
                        if not entry_id:
                            continue
                        _agregar_valores(entry_id, raw_value)
        except Exception as e:
            print(f" No se pudo cargar ids_dinamicos: {e}")
        return ids_por_id

    def _resolve_dynamic_id_values(self, raw_value):
        """Normaliza un valor dinámico a una lista de candidatos no vacíos."""
        if raw_value is None:
            return []
        if isinstance(raw_value, dict):
            raw_value = (
                raw_value.get("valor")
                if "valor" in raw_value
                else raw_value.get("valores", raw_value.get("value", raw_value.get("values")))
            )
            if raw_value is None:
                return []
        if isinstance(raw_value, (list, tuple, set)):
            return [str(item).strip() for item in raw_value if str(item).strip()]
        text = str(raw_value).strip()
        if not text:
            return []
        if "|" in text:
            return [item.strip() for item in text.split("|") if item.strip()]
        return [text]

    @staticmethod
    def _sanitize_field_value(value):
        """Reemplaza 'test' (palabra completa, cualquier capitalización) por 'prueba'."""
        import re
        return re.sub(r'\btest\b', 'prueba', str(value), flags=re.IGNORECASE)

    def _fill_and_dispatch(self, element, value):
        """Rellena un input/textarea con value y dispara los eventos necesarios para frameworks JS."""
        value = self._sanitize_field_value(value)
        self.driver.execute_script(
            "var el=arguments[0],v=arguments[1];"
            "el.focus();"
            "try{"
            "  var proto = el.tagName === 'TEXTAREA' ? window.HTMLTextAreaElement.prototype : window.HTMLInputElement.prototype;"
            "  Object.getOwnPropertyDescriptor(proto,'value').set.call(el,v);"
            "}catch(e){el.value=v;}"
            "el.dispatchEvent(new Event('input',{bubbles:true}));"
            "el.dispatchEvent(new Event('change',{bubbles:true}));"
            "el.dispatchEvent(new Event('blur',{bubbles:true}));",
            element, value,
        )

    def _auto_fill_unmapped_dropdowns(self, field_mapping=None):
        """Completa campos no mapeados: selects aleatorios y valores fijos para IDs dinámicos."""
        mapped_ids = self._get_mapped_select_ids(field_mapping)
        # Ids ya llenados/registrados por el mapping en ESTA fila. Cubre el caso en que el
        # id real del form difiere del de la config (ej. 'estimated-day' vs
        # 'estimated-date-purchase'): sin esto, el select ya llenado con el valor del Excel
        # se re-selecciona acá con la 1ª opción y se pisa el dato.
        already_filled_ids = set()
        for _k in getattr(self, "current_row_field_values", {}) or {}:
            raw = _k.split("::", 1)[1] if "::" in _k else _k
            already_filled_ids.add(raw)
        ids_dinamicos = self._cargar_ids_dinamicos()
        try:
            select_elements = self.driver.find_elements(By.XPATH, "//select")
        except Exception as e:
            print(f" Error buscando dropdowns no mapeados: {e}")
            return False
        filled_any = False

        for select_element in select_elements:
            select_id = None
            try:
                select_id = select_element.get_attribute("id") or select_element.get_attribute("name")
                if not select_id:
                    continue
                # Solo saltear si tiene id Y ese id ya está en el mapping (campos por name nunca están en mapped_ids)
                elem_real_id = select_element.get_attribute("id")
                if elem_real_id and elem_real_id in mapped_ids:
                    continue
                # Saltear si el mapping ya llenó este select en esta fila (id real ≠ id de config)
                if (elem_real_id and elem_real_id in already_filled_ids) or (select_id in already_filled_ids):
                    continue
                if not select_element.is_enabled():
                    continue
                if select_element.get_attribute("multiple") is not None:
                    continue

                # Para selects no mapeados siempre intentamos seleccionar.
                # No confiamos en first_selected_option como señal de "ya lleno" porque
                # algunos forms Angular/React pre-seleccionan la primera opción en DOM pero
                # requieren un evento activo del usuario para validar el campo.

                select = Select(select_element)
                # Construir lista de opciones válidas (no placeholder, no disabled, con value)
                valid_options = []
                for opt in select.options:
                    opt_text = opt.text.strip()
                    opt_val = (opt.get_attribute("value") or "").strip()
                    if not opt_text or opt.get_attribute("disabled"):
                        continue
                    if self._is_placeholder_text(opt_text):
                        continue
                    valid_options.append(opt)

                # Fallback: si el filtro de placeholder fue demasiado estricto,
                # tomar opciones con value no vacío y que NO sean placeholder
                if not valid_options:
                    for opt in select.options:
                        opt_text = opt.text.strip()
                        opt_val = (opt.get_attribute("value") or "").strip()
                        if opt_val and opt_val != "0" and not self._is_placeholder_text(opt_text):
                            valid_options.append(opt)

                if not valid_options:
                    print(f"⚠️ Select '{select_id}': sin opciones seleccionables, se omite")
                    continue

                # Usar valor fijo si el ID está en IDs dinámicos
                if select_id in ids_dinamicos:
                    fixed_candidates = self._resolve_dynamic_id_values(ids_dinamicos[select_id])
                    matched_options = []
                    for candidate in fixed_candidates:
                        matched = next(
                            (o for o in valid_options if o.text.strip() == candidate or o.get_attribute("value") == candidate),
                            None
                        )
                        if matched:
                            matched_options.append((matched, candidate))

                    if matched_options:
                        chosen_option, matched_value = random.choice(matched_options)
                        print(f"📌 Select no mapeado '{select_id}' valor fijo: {matched_value}")
                    else:
                        chosen_option = random.choice(valid_options)
                        print(f"⚠️ '{select_id}': ningún valor dinámico coincide, usando aleatorio")
                else:
                    chosen_option = random.choice(valid_options)

                option_value = chosen_option.get_attribute("value") or chosen_option.text
                option_text  = chosen_option.text.strip()

                # 1) Selenium nativo — cambia el DOM value
                try:
                    select.select_by_value(option_value)
                except Exception:
                    try:
                        select.select_by_visible_text(option_text)
                    except Exception:
                        pass

                # 2) send_keys con el texto — genera eventos de teclado REALES que
                #    Angular/React procesan dentro de su propio zone (no son sintéticos)
                try:
                    from selenium.webdriver.common.keys import Keys
                    select_element.send_keys(option_text)
                except Exception:
                    pass

                # 3) Eventos JS como capa adicional (focus → change → input → blur)
                self.driver.execute_script(
                    "var el=arguments[0],val=arguments[1];"
                    "try{Object.getOwnPropertyDescriptor(window.HTMLSelectElement.prototype,'value').set.call(el,val);}catch(e){}"
                    "el.dispatchEvent(new Event('focus',{bubbles:true}));"
                    "el.dispatchEvent(new Event('change',{bubbles:true,cancelable:true}));"
                    "el.dispatchEvent(new Event('input',{bubbles:true,cancelable:true}));"
                    "el.dispatchEvent(new Event('blur',{bubbles:true}));",
                    select_element, option_value
                )
                time.sleep(0.3)
                print(f"🎲 Select no mapeado '{select_id}' → '{option_text}'")
                self._record_field_value(select_id, option_text)
                filled_any = True
            except Exception as e:
                print(f" _auto_fill_unmapped_dropdowns id='{select_id}' error: {e}")
                continue

        # Aplicar IDs dinámicos en inputs y textareas no mapeados.
        try:
            text_elements = self.driver.find_elements(
                By.XPATH,
                "//input[not(@type='hidden')][not(@type='submit')][not(@type='button')]"
                "[not(@type='checkbox')][not(@type='radio')][not(@type='file')][not(@type='reset')]"
                " | //textarea"
            )
        except Exception as e:
            print(f" Error buscando inputs/textarea no mapeados: {e}")
            text_elements = []

        form_data = getattr(self, "_current_form_data", {}) or {}

        for element in text_elements:
            field_id = None
            try:
                field_id = element.get_attribute("id") or element.get_attribute("name")
                if not field_id or field_id in mapped_ids:
                    continue
                if not element.is_displayed() or not element.is_enabled():
                    continue

                tag_name = (element.tag_name or "").lower()
                input_type = (element.get_attribute("type") or "text").lower() if tag_name == "input" else "textarea"
                if input_type in {"hidden", "submit", "button", "checkbox", "radio", "file", "reset"}:
                    continue

                current_value = (element.get_attribute("value") or "").strip()
                if current_value:
                    continue

                fill_value = None
                source_label = ""

                # Prioridad 1: IDs dinámicos (comportamiento existente)
                if field_id in ids_dinamicos:
                    dynamic_candidates = self._resolve_dynamic_id_values(ids_dinamicos[field_id])
                    if dynamic_candidates:
                        fill_value = random.choice(dynamic_candidates)
                        source_label = "ID dinámico"

                # Prioridad 2: inferir data_key y tomar valor de form_data
                if not fill_value and form_data:
                    field_name = (
                        element.get_attribute("aria-label")
                        or element.get_attribute("placeholder")
                        or element.get_attribute("name")
                        or field_id
                    )
                    placeholder = element.get_attribute("placeholder") or ""
                    inferred_key = self._infer_data_key(field_id, field_name, placeholder)
                    if inferred_key and form_data.get(inferred_key):
                        fill_value = str(form_data[inferred_key])
                        source_label = f"inferido→{inferred_key}"

                # Prioridad 3: probe numérico → alfa para inputs requeridos sin valor
                if not fill_value:
                    is_req = bool(
                        element.get_attribute("required")
                        or element.get_attribute("aria-required") == "true"
                    )
                    if not is_req:
                        # Campo no mapeado sin valor asignado: queda vacío. Se registra
                        # para avisar en el Excel/mail que conviene setearlo en IDs Dinámicos.
                        if not hasattr(self, "_campos_sin_valor_asignado"):
                            self._campos_sin_valor_asignado = []
                        if field_id not in self._campos_sin_valor_asignado:
                            self._campos_sin_valor_asignado.append(field_id)
                        continue

                    # Decidir orden de prueba: si type/inputmode sugiere numérico, probar num primero
                    _inputmode = (element.get_attribute("inputmode") or "").lower()
                    _likely_num = input_type in ("number", "tel") or _inputmode in ("numeric", "tel", "decimal")
                    _probes = ["12345678", "Carlos"] if _likely_num else ["Carlos", "12345678"]

                    _probe_success = False
                    for _probe in _probes:
                        self._fill_and_dispatch(element, _probe)
                        time.sleep(0.25)
                        _cls = element.get_attribute("class") or ""
                        _has_error = any(c in _cls for c in ("error", "invalid", "is-invalid", "ng-invalid"))
                        if not _has_error:
                            fill_value = _probe
                            source_label = "auto-probe"
                            _probe_success = True
                            break

                    if not _probe_success:
                        print(f"⚠️ Campo no mapeado '{field_id}': no se pudo auto-completar")
                        if not hasattr(self, "_campos_sin_mapeo_exitoso"):
                            self._campos_sin_mapeo_exitoso = []
                        self._campos_sin_mapeo_exitoso.append(field_id)
                        continue

                if fill_value and source_label != "auto-probe":
                    self._fill_and_dispatch(element, fill_value)
                print(f"📌 Campo no mapeado '{field_id}' completado [{source_label}]: {fill_value}")
                self._record_field_value(field_id, fill_value)
                filled_any = True
                time.sleep(0.1)
            except Exception as e:
                print(f" _auto_fill_unmapped_dropdowns field='{field_id}' error: {e}")
                continue

        # Manejo especial: kits[] es un select múltiple Bootstrap Select
        # El <select> real está oculto (display:none), hay que interactuar con el botón del widget
        try:
            kits_elements = self.driver.find_elements(By.ID, "kits[]")
            if kits_elements:
                kits_el = kits_elements[0]
                # NO chequeamos is_displayed() porque Bootstrap Select oculta el <select> real

                # Esperar hasta 8s a que se habilite Y carguen opciones vía AJAX
                valid_kits = []
                for _ in range(8):
                    if kits_el.get_attribute("disabled"):
                        time.sleep(1)
                        continue
                    valid_kits = self._get_valid_select_options(kits_el)
                    if valid_kits:
                        break
                    time.sleep(1)

                if valid_kits:
                    fixed_kit_candidates = self._resolve_dynamic_id_values(ids_dinamicos.get("kits[]"))
                    matched_kits = []
                    for candidate in fixed_kit_candidates:
                        matched_kit = next(
                            (
                                opt for opt in valid_kits
                                if opt.text.strip() == str(candidate) or (opt.get_attribute("value") or "") == str(candidate)
                            ),
                            None,
                        )
                        if matched_kit:
                            matched_kits.append((matched_kit, candidate))

                    if matched_kits:
                        chosen, matched_kit_value = random.choice(matched_kits)
                    else:
                        chosen = random.choice(valid_kits)
                        matched_kit_value = None
                    chosen_val = chosen.get_attribute("value") or chosen.text.strip()
                    chosen_text = chosen.text.strip()
                    if matched_kit_value is not None:
                        print(f"📌 kits[] valor fijo: {matched_kit_value}")
                    elif fixed_kit_candidates:
                        print("⚠️ 'kits[]': ningún valor dinámico coincide, usando aleatorio")

                    # Usar la API de jQuery selectpicker con los brackets correctamente escapados
                    selected = self.driver.execute_script("""
                        var val = arguments[0];
                        try {
                            var jq = (typeof window.jQuery !== 'undefined') ? window.jQuery : (typeof $ !== 'undefined' ? $ : null);
                            if (!jq) { return null; }
                            var $sel = jq('#kits\\\\[\\\\]');
                            if ($sel.length && typeof $sel.selectpicker === 'function') {
                                $sel.selectpicker('val', [val]);
                                $sel.trigger('changed.bs.select');
                                $sel.trigger('change');
                                $sel.selectpicker('refresh');
                                return 'selectpicker';
                            }
                        } catch(e) {}
                        // Fallback: manipulación directa del DOM
                        var sel = document.getElementById('kits[]');
                        if (sel) {
                            for (var i = 0; i < sel.options.length; i++) {
                                if (sel.options[i].value === val || sel.options[i].text.trim() === val) {
                                    sel.options[i].selected = true;
                                    sel.dispatchEvent(new Event('change', { bubbles: true }));
                                    sel.dispatchEvent(new Event('input', { bubbles: true }));
                                    return 'dom';
                                }
                            }
                        }
                        return null;
                    """, chosen_val)

                    if selected:
                        print(f"🎲 kits[] auto-seleccionado via {selected}: {chosen_text}")
                        self._record_field_value("kits[]", chosen_text)
                        filled_any = True
                    else:
                        # Último fallback: clic real en el botón del dropdown y luego en el item
                        try:
                            kit_btn = self.driver.find_element(By.CSS_SELECTOR, "button[data-id='kits[]']")
                            if kit_btn.is_displayed() and kit_btn.is_enabled():
                                kit_btn.click()
                                time.sleep(0.5)
                                # Buscar el li con el texto del kit en el dropdown abierto
                                li_items = self.driver.find_elements(
                                    By.CSS_SELECTOR, ".dropdown.bootstrap-select.kits .dropdown-menu li:not(.disabled) a span.text"
                                )
                                for li in li_items:
                                    if li.text.strip() == chosen_text:
                                        li.click()
                                        print(f"🎲 kits[] auto-seleccionado via clic real: {chosen_text}")
                                        self._record_field_value("kits[]", chosen_text)
                                        filled_any = True
                                        break
                                else:
                                    # cerrar el dropdown sin seleccionar
                                    kit_btn.click()
                        except Exception as e_click:
                            print(f" kits[] clic fallback error: {e_click}")

                elif kits_el.get_attribute("disabled"):
                    print(" kits[] sigue deshabilitado tras espera, se omite")
                else:
                    print(" kits[] habilitado pero sin opciones válidas tras espera")
        except Exception as e:
            print(f" _auto_fill_unmapped_dropdowns kits[] error: {e}")

        if not filled_any:
            print(" No se encontraron campos no mapeados para auto-completar")
        return filled_any

    def _normalize_text(self, value):
        """Normaliza texto eliminando acentos, espacios y mayúsculas"""
        if value is None:
            return ""

        text = str(value).strip().lower()
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(ch for ch in text if not unicodedata.combining(ch))
        text = text.replace('-', ' ').replace('_', ' ')
        text = ''.join(ch for ch in text if ch.isalnum())
        return text

    def _handle_terms_checkboxes(self):
        """Marca radios y checkboxes requeridos respetando el orden de términos"""
        try:
            radios_marked = self._mark_preferred_radios()
        except Exception as e:
            print(f" Error marcando radios: {e}")
            radios_marked = 0

        try:
            checkboxes_marked = self._mark_required_checkboxes()
        except Exception as e:
            print(f" Error marcando checkboxes: {e}")
            checkboxes_marked = 0

        try:
            guide_cb = self._mark_guide_checkbox_widgets()
            checkboxes_marked += guide_cb
        except Exception as e:
            print(f" Error marcando checkboxes Guide (AEM): {e}")

        total = radios_marked + checkboxes_marked
        print(f" Radios marcados: {radios_marked}, Checkboxes marcados: {checkboxes_marked}")
        return total > 0

    def _mark_preferred_radios(self):
        groups = self._collect_radio_groups()
        marked = 0

        for group in groups:
            try:
                if self._ensure_radio_selected(group):
                    marked += 1
            except Exception as e:
                print(f" No se pudo procesar radios '{group['name']}': {e}")

        if groups:
            print(f" Radios marcados: {marked}/{len(groups)}")
        return marked

    def _collect_radio_groups(self):
        try:
            radios = self.driver.find_elements(By.CSS_SELECTOR, "input[type='radio']")
        except Exception:
            return []

        groups = {}
        for idx, radio in enumerate(radios):
            try:
                original_name_attr = (radio.get_attribute("name") or "").strip()
                has_real_name = bool(original_name_attr)
                if not original_name_attr:
                    original_name_attr = f"radio_group_{idx}"

                group_key = original_name_attr.lower()
                group = groups.setdefault(
                    group_key,
                    {
                        "name": group_key,
                        "original_name": original_name_attr,
                        "has_real_name": has_real_name,
                        "options": [],
                        "first_index": idx,
                    },
                )

                group["options"].append(
                    {
                        "value": (radio.get_attribute("value") or "").strip(),
                        "label": (radio.get_attribute("title") or radio.get_attribute("data-dtm") or radio.get_attribute("data-label") or radio.get_attribute("id") or radio.get_attribute("value") or f"opcion_{idx}"),
                        "index": idx,
                        "element": radio,
                    }
                )
            except StaleElementReferenceException:
                continue

        ordered_groups = sorted(groups.values(), key=lambda g: g["first_index"])
        return ordered_groups

    def _ensure_radio_selected(self, group):
        target_option = self._choose_radio_option(group)
        if not target_option:
            return False

        label = f"radio {group['original_name']}={target_option['value']}"

        for attempt in range(3):
            radio = self._locate_radio_candidate(group, target_option)
            if not radio:
                time.sleep(0.2)
                continue

            try:
                self._prepare_radio_for_interaction(radio)

                try:
                    state_info = (
                        radio.is_enabled(),
                        radio.is_displayed(),
                        radio.is_selected(),
                    )
                    print(f" Radio intento {attempt + 1} {label}: enabled={state_info[0]}, visible={state_info[1]}, seleccionado={state_info[2]}")
                except Exception:
                    pass

                if radio.is_selected():
                    print(f"ℹ {label} ya estaba seleccionado")
                    return True

                if self._set_radio_checked_via_js(radio):
                    radio = self._locate_radio_candidate(group, target_option) or radio
                    if radio.is_selected():
                        print(f" {label} seleccionado vía JS")
                        return True

                click_target = self._find_input_click_target(radio)
                if click_target and self._click_element_stable(click_target):
                    time.sleep(0.2)
                    radio = self._locate_radio_candidate(group, target_option) or radio
                    if radio.is_selected():
                        print(f" {label} seleccionado por click")
                        return True
                    else:
                        print(f" {label} no cambió tras click, reintentando...")

            except StaleElementReferenceException:
                time.sleep(0.2)
                continue
            except Exception as e:
                print(f" Error seleccionando {label}: {e}")
                break

        radio = self._locate_radio_candidate(group, target_option)
        return radio.is_selected() if radio else False

    def _choose_radio_option(self, group):
        options = group.get("options", [])
        if not options:
            return None

        for option in options:
            try:
                element = option.get("element")
                if element and element.is_selected():
                    return option
            except StaleElementReferenceException:
                continue

        priority_values = {
            "si": 10,
            "sí": 10,
            "yes": 9,
            "true": 9,
            "1": 9,
            "renovar": 8,
            "suscribir": 6,
            "no": 2,
        }

        name_key = group.get("name", "")
        specific_priority = {
            "have_interest": ["renovar", "suscribir"],
            "have-chevrolet": ["si", "sí", "no"],
            "client": ["si", "sí"],
            # Libro de Reclamaciones: elegir "menor de edad" abre un bloque extra obligatorio
            # (responsable legal) que el lead no tiene cómo completar.
            "cc-younger-status": ["soy mayor de edad"],
        }

        preferred_values = specific_priority.get(name_key, [])

        def option_priority(option):
            value = option.get("value", "").strip().lower()
            if preferred_values and value in preferred_values:
                return 100 - preferred_values.index(value)
            return priority_values.get(value, 1)

        sorted_options = sorted(options, key=lambda opt: (-option_priority(opt), opt.get("index", 0)))
        return sorted_options[0]

    def _locate_radio_candidate(self, group, target_option):
        name_original = group.get("original_name")
        has_real_name = group.get("has_real_name", True)
        value_target = (target_option.get("value") or "").strip()

        selectors = []

        if name_original and has_real_name:
            selectors.append((By.CSS_SELECTOR, f"input[type='radio'][name=\"{name_original}\"]"))

        selectors.append((By.CSS_SELECTOR, "input[type='radio']"))

        for by, selector in selectors:
            try:
                radios = self.driver.find_elements(by, selector)
            except Exception:
                continue

            for element in radios:
                try:
                    name_attr = (element.get_attribute("name") or "").strip()
                    if has_real_name and name_original and name_attr != name_original:
                        continue

                    if value_target:
                        value_attr = (element.get_attribute("value") or "").strip()
                        if value_attr != value_target:
                            continue

                    return element
                except StaleElementReferenceException:
                    continue

        return None

    def _prepare_radio_for_interaction(self, radio):
        try:
            self.driver.execute_script(
                """
                    const rb = arguments[0];
                    if (rb.hasAttribute('disabled')) {
                        rb.removeAttribute('disabled');
                    }
                    rb.disabled = false;
                    rb.setAttribute('aria-disabled', 'false');
                    if (rb.tabIndex === -1) {
                        rb.tabIndex = 0;
                    }
                """,
                radio,
            )
        except Exception:
            pass

    def _set_radio_checked_via_js(self, radio):
        try:
            result = self.driver.execute_script(
                """
                    const rb = arguments[0];
                    if (rb.checked) {
                        return true;
                    }
                    rb.focus && rb.focus();
                    rb.checked = true;
                    rb.dispatchEvent(new Event('input', { bubbles: true }));
                    rb.dispatchEvent(new Event('change', { bubbles: true }));
                    return rb.checked === true;
                """,
                radio,
            )
            if result:
                print("ℹ Radio marcado mediante JS directo")
            return bool(result)
        except Exception:
            return False

    def _checkbox_in_dom(self, checkbox):
        """True si el checkbox está en el DOM con layout, aunque tenga opacity:0."""
        try:
            return self.driver.execute_script(
                """
                    const el = arguments[0];
                    if (!el || !el.isConnected) return false;
                    const style = getComputedStyle(el);
                    if (style.display === 'none') return false;
                    if (style.visibility === 'hidden') return false;
                    // opacity:0 es ok — checkboxes custom usan esto; se hace click en label padre
                    return el.offsetWidth > 0 || el.offsetHeight > 0 || el.getClientRects().length > 0;
                """,
                checkbox,
            )
        except Exception:
            return False

    def _mark_required_checkboxes(self):
        known_names = {
            "terms",
            "terms-and-conditions",    # visid standard
            "terms-contact",
            "terms_contact",
            "termscontact",
            "terms-platform",
            "terms_platform",
            "termsplatform",
            "accept-terms",
            "accept_terms",
            "privacy",
            "privacy_policy",
        }

        priority_map = {
            "terms": 3,
            "terms-and-conditions": 3, # visid — misma prioridad que "terms"
            "terms-platform": 2,
            "terms_platform": 2,
            "termsplatform": 2,
            "terms-contact": 1,
            "terms_contact": 1,
            "termscontact": 1,
        }

        candidates = []

        try:
            raw_checkboxes = self.driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
        except Exception:
            raw_checkboxes = []

        for index, checkbox in enumerate(raw_checkboxes):
            try:
                name_attr = (checkbox.get_attribute("name") or "").strip()
                lower_name = name_attr.lower()
                checkbox_id = (checkbox.get_attribute("id") or "").strip()
                required_attr = checkbox.get_attribute("required")
                data_dtm = (checkbox.get_attribute("data-dtm") or "").strip()
                value_attr = (checkbox.get_attribute("value") or "").strip()

                is_known = lower_name in known_names
                is_html_required = bool(required_attr)

                # El Excel manda: una columna con el name/id del checkbox y valor SI/NO
                pref = self._checkbox_pref_for(lower_name, checkbox_id)
                if pref is False:
                    self._uncheck_checkbox(checkbox_id, name_attr)
                    print(f"  ⊘ {name_attr or checkbox_id} desmarcado (Excel = NO)")
                    continue

                if not is_known and not is_html_required and pref is not True:
                    # No es términos ni required HTML — incluir solo si en DOM con layout + no marcado
                    # Usamos _checkbox_in_dom en vez de is_displayed() para no excluir inputs con opacity:0
                    try:
                        if not self._checkbox_in_dom(checkbox) or not checkbox.is_enabled():
                            continue
                        if checkbox.is_selected():
                            continue
                    except StaleElementReferenceException:
                        continue
                elif not lower_name and not is_html_required and pref is not True:
                    continue  # sin nombre y sin required: saltar (comportamiento original)

                priority = priority_map.get(lower_name, 0)
                display = name_attr or checkbox_id or data_dtm or "checkbox"

                candidates.append({
                    "name": lower_name,
                    "name_original": name_attr,
                    "id": checkbox_id,
                    "value": value_attr,
                    "data_dtm": data_dtm,
                    "priority": priority,
                    "order": index,
                    "label": display,
                })
            except StaleElementReferenceException:
                continue

        if not candidates:
            print(" No se encontraron checkboxes requeridos")
            return 0

        candidates.sort(key=lambda item: (item["priority"], item["order"]))

        marked = 0
        for candidate in candidates:
            if self._ensure_checkbox_selected(candidate):
                marked += 1
            else:
                print(f" No se pudo marcar {candidate.get('label', 'checkbox')}")

        print(f" Checkboxes marcados: {marked}/{len(candidates)}")
        return marked

    def _click_aem_guide_checkbox_by_input(self, checkbox_input):
        """AEM Guide suele reaccionar al click en .guideCheckBoxItem, no solo al input."""
        try:
            if checkbox_input:
                self._scroll_element_into_view(checkbox_input)
            return bool(
                self.driver.execute_script(
                    """
                    const inp = arguments[0];
                    if (!inp) return false;
                    const item = inp.closest('.guideCheckBoxItem');
                    if (item) {
                        item.click();
                    } else {
                        inp.click();
                    }
                    return inp.checked === true || inp.getAttribute('aria-checked') === 'true';
                    """,
                    checkbox_input,
                )
            )
        except Exception:
            return False

    def _mark_guide_checkbox_widgets(self):
        """Marca checkboxes de términos Adobe AEM Guide (contenedor ___guide-item + input ___N_widget)."""
        cid = self._GUIDE_CHECKBOX_CONTAINER_ID
        inp_pre = self._GUIDE_CHECKBOX_INPUT_ID_PREFIX
        selectors = (
            f'div[id="{cid}"] input[type="checkbox"]',
            f'[id^="{cid}"] input[type="checkbox"]',
            f'input[type="checkbox"][id^="{inp_pre}"][id$="_widget"]',
        )
        marked = 0
        seen = set()
        for css in selectors:
            try:
                elems = self.driver.find_elements(By.CSS_SELECTOR, css)
            except Exception:
                continue
            for cb in elems:
                try:
                    eid = (cb.get_attribute("id") or "").strip()
                    key = eid or str(cb)
                    if key in seen:
                        continue
                    if not cb.is_displayed():
                        continue
                    seen.add(key)
                    self._scroll_element_into_view(cb)
                    time.sleep(0.15)
                    # Sin name_* falsos: _locate_checkbox_candidate filtraría por name != guide_checkbox
                    candidate = {
                        "name": "",
                        "name_original": "",
                        "id": eid,
                        "value": "",
                        "data_dtm": "",
                        "priority": 100,
                        "order": 0,
                        "label": f"guide_checkbox:{eid or 'sin-id'}",
                    }
                    if self._ensure_checkbox_selected(candidate):
                        marked += 1
                        continue
                    el = self.safe_find_element(By.ID, eid) if eid else None
                    if el and self._click_aem_guide_checkbox_by_input(el):
                        time.sleep(0.25)
                        el2 = self.safe_find_element(By.ID, eid) if eid else None
                        if el2 and el2.is_selected():
                            marked += 1
                            print(f" Guide checkbox (click AEM): {eid}")
                except StaleElementReferenceException:
                    continue
                except Exception as e:
                    print(f" Guide checkbox: {e}")
        if marked:
            print(f" Checkboxes Guide (AEM) marcados: {marked}")
        return marked

    def _ensure_checkbox_selected(self, candidate):
        label = candidate.get("label", "checkbox")

        for attempt in range(4):
            checkbox = self._locate_checkbox_candidate(candidate)
            if not checkbox:
                time.sleep(0.2)
                continue

            try:
                self._prepare_checkbox_for_interaction(checkbox)

                try:
                    state_info = (
                        checkbox.is_enabled(),
                        checkbox.is_displayed(),
                        checkbox.is_selected(),
                    )
                    print(f" Intento {attempt + 1} para {label}: enabled={state_info[0]}, visible={state_info[1]}, seleccionado={state_info[2]}")
                except Exception:
                    pass

                if checkbox.is_selected():
                    print(f"ℹ {label} ya estaba marcado")
                    return True

                if self._set_checkbox_checked_via_js(checkbox):
                    checkbox = self._locate_checkbox_candidate(candidate) or checkbox
                    if checkbox.is_selected():
                        print(f" {label} marcado vía JS")
                        return True

                click_target = self._find_input_click_target(checkbox)
                if click_target and self._click_element_stable(click_target):
                    time.sleep(0.2)
                    checkbox = self._locate_checkbox_candidate(candidate) or checkbox
                    if checkbox.is_selected():
                        print(f" {label} marcado por click")
                        return True
                    else:
                        print(f" {label} no cambió tras click, reintentando...")

            except StaleElementReferenceException:
                time.sleep(0.2)
                continue
            except Exception as e:
                print(f" Error marcando {label}: {e}")
                break

        checkbox = self._locate_checkbox_candidate(candidate)
        return checkbox.is_selected() if checkbox else False

    def _prepare_checkbox_for_interaction(self, checkbox):
        try:
            self.driver.execute_script(
                """
                    const cb = arguments[0];
                    if (cb.hasAttribute('disabled')) {
                        cb.removeAttribute('disabled');
                    }
                    cb.disabled = false;
                    cb.setAttribute('aria-disabled', 'false');
                    cb.classList.remove('is-invalid');
                    if (cb.tabIndex === -1) {
                        cb.tabIndex = 0;
                    }
                """,
                checkbox,
            )
        except Exception:
            pass

    def _locate_checkbox_candidate(self, candidate):
        selectors = []
        checkbox_id = candidate.get("id")
        name_original = candidate.get("name_original")
        lower_name = candidate.get("name")
        data_dtm = candidate.get("data_dtm")

        if checkbox_id:
            selectors.append((By.ID, checkbox_id))

        if name_original:
            selectors.append((By.CSS_SELECTOR, f"input[type='checkbox'][name=\"{name_original}\"]"))

        if data_dtm:
            selectors.append((By.CSS_SELECTOR, f"input[type='checkbox'][data-dtm=\"{data_dtm}\"]"))

        selectors.append((By.CSS_SELECTOR, "input[type='checkbox']"))

        for by, selector in selectors:
            try:
                elements = self.driver.find_elements(by, selector)
            except Exception:
                continue

            for element in elements:
                try:
                    if checkbox_id and (element.get_attribute("id") or "").strip() != checkbox_id:
                        continue

                    if name_original and (element.get_attribute("name") or "").strip() != name_original:
                        continue

                    if lower_name and (element.get_attribute("name") or "").strip().lower() != lower_name:
                        continue

                    if data_dtm and (element.get_attribute("data-dtm") or "").strip() != data_dtm:
                        continue

                    value_attr = (element.get_attribute("value") or "").strip()
                    candidate_value = candidate.get("value", "")
                    # Solo validar value si ambos tienen valores no vacíos
                    if candidate_value and value_attr and candidate_value != value_attr:
                        continue

                    return element
                except StaleElementReferenceException:
                    continue

        return None

    def _set_checkbox_checked_via_js(self, checkbox):
        try:
            result = self.driver.execute_script(
                """
                    const cb = arguments[0];
                    if (cb.checked) {
                        return true;
                    }
                    cb.focus && cb.focus();
                    // Disparar click en el label padre primero (frameworks React/Angular)
                    const parentLabel = cb.closest('label');
                    if (parentLabel) {
                        parentLabel.click();
                        if (cb.checked) return true;
                    }
                    // Fallback: click directo sobre el input
                    cb.click();
                    if (cb.checked) return true;
                    // Fallback: forzar via propiedad + eventos
                    cb.checked = true;
                    cb.setAttribute('checked', 'checked');
                    cb.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
                    cb.dispatchEvent(new Event('input', { bubbles: true }));
                    cb.dispatchEvent(new Event('change', { bubbles: true }));
                    return cb.checked === true;
                """,
                checkbox,
            )
            if result:
                print("ℹ Checkbox marcado mediante JS directo")
            return bool(result)
        except Exception:
            return False

    def _find_input_click_target(self, input_element):
        try:
            target = self.driver.execute_script(
                """
                    const elem = arguments[0];
                    // Prioridad 1: label padre (más confiable para checkboxes custom)
                    const parentLabel = elem.closest('label');
                    if (parentLabel) {
                        return parentLabel;
                    }
                    // Prioridad 2: label externo con for="id"
                    if (elem.id) {
                        const forLabel = document.querySelector(`label[for="${elem.id}"]`);
                        if (forLabel) {
                            return forLabel;
                        }
                    }
                    // Prioridad 3: hermano siguiente si no es input
                    if (elem.nextElementSibling && elem.nextElementSibling.tagName && elem.nextElementSibling.tagName.toLowerCase() !== 'input') {
                        return elem.nextElementSibling;
                    }
                    return elem;
                """,
                input_element,
            )
            return target
        except Exception:
            return input_element

    def _click_element_stable(self, element):
        if element is None:
            return False

        if not hasattr(element, "click"):
            return False

        try:
            self._scroll_element_into_view(element)
        except Exception:
            pass

        for _ in range(2):
            try:
                element.click()
                return True
            except Exception:
                try:
                    self.driver.execute_script("arguments[0].click();", element)
                    return True
                except Exception:
                    try:
                        self.driver.execute_script(
                            """
                                const el = arguments[0];
                                if (typeof PointerEvent !== 'undefined') {
                                    el.dispatchEvent(new PointerEvent('pointerdown', { bubbles: true }));
                                    el.dispatchEvent(new PointerEvent('pointerup', { bubbles: true }));
                                }
                                el.dispatchEvent(new MouseEvent('mousedown', { bubbles: true }));
                                el.dispatchEvent(new MouseEvent('click', { bubbles: true }));
                                el.dispatchEvent(new MouseEvent('mouseup', { bubbles: true }));
                            """,
                            element,
                        )
                        return True
                    except Exception:
                        pass
                    time.sleep(0.2)

        return False
    
    def _select_multiple_options(self, select_element, option_text, field_name):
        """Selecciona múltiples opciones en un select múltiple usando JavaScript"""
        try:
            # Dividir valores si vienen separados por coma, punto y coma, etc.
            values_to_select = [v.strip() for v in re.split(r'[,;|]', option_text) if v.strip()]
            if not values_to_select:
                values_to_select = [option_text.strip()]
            
            print(f" Intentando seleccionar en {field_name}: {values_to_select}")
            
            selected_count = 0
            for value in values_to_select:
                # Usar JavaScript para seleccionar sin desmarcar otras opciones
                result = self.driver.execute_script("""
                    const select = arguments[0];
                    const searchText = arguments[1].toLowerCase();
                    let matched = false;
                    
                    for (let i = 0; i < select.options.length; i++) {
                        const opt = select.options[i];
                        const optText = opt.text.trim().toLowerCase();
                        const optValue = (opt.value || '').toLowerCase();
                        
                        // Buscar coincidencia exacta o parcial
                        if (optText === searchText || optValue === searchText || 
                            optText.includes(searchText) || optValue.includes(searchText)) {
                            opt.selected = true;
                            matched = true;
                            console.log('Seleccionado:', opt.text);
                            break;
                        }
                    }
                    
                    if (matched) {
                        select.dispatchEvent(new Event('input', { bubbles: true }));
                        select.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                    
                    return matched;
                """, select_element, value)
                
                if result:
                    selected_count += 1
                    print(f"   Seleccionado: {value}")
                else:
                    print(f"   No se encontró: {value}")
            
            if selected_count > 0:
                print(f" {field_name}: {selected_count}/{len(values_to_select)} opciones seleccionadas")
                try:
                    selected_labels = [opt.text.strip() for opt in Select(select_element).all_selected_options if opt.text.strip()]
                except Exception:
                    selected_labels = values_to_select
                select_id = select_element.get_attribute("id") or field_name
                self._record_field_value(select_id, selected_labels)
                return True
            else:
                print(f" {field_name}: No se pudo seleccionar ninguna opción")
                return False
                
        except Exception as e:
            print(f" Error seleccionando múltiples opciones en {field_name}: {e}")
            return False
    
    def _get_field_label(self, element):
        """Intenta obtener la etiqueta visible de un campo del formulario."""
        try:
            field_id = element.get_attribute("id")
            if field_id:
                labels = self.driver.find_elements(By.XPATH, f"//label[@for='{field_id}']")
                if labels and labels[0].text.strip():
                    return labels[0].text.strip()
            aria = element.get_attribute("aria-label") or ""
            if aria.strip():
                return aria.strip()
            placeholder = element.get_attribute("placeholder") or ""
            if placeholder.strip():
                return placeholder.strip()
            name = element.get_attribute("name") or ""
            return name.strip()
        except Exception:
            return ""

    def _find_unfilled_visible_fields(self):
        """Devuelve lista de campos visibles que parecen vacíos y no están en field_mapping."""
        mapped_ids = set()
        for entry in self.field_mapping:
            fid = entry.get("id")
            if isinstance(fid, list):
                mapped_ids.update(str(f) for f in fid if f)
            elif fid:
                mapped_ids.add(str(fid))

        unfilled = []
        try:
            elements = self.driver.find_elements(
                By.XPATH,
                "//input[not(@type='hidden')][not(@type='submit')][not(@type='button')]"
                "[not(@type='checkbox')][not(@type='radio')][not(@type='file')][not(@type='reset')]"
                " | //select | //textarea"
            )
            for el in elements:
                try:
                    if not el.is_displayed() or not el.is_enabled():
                        continue
                    field_id = (el.get_attribute("id") or el.get_attribute("name") or "").strip()
                    if not field_id or field_id in mapped_ids:
                        continue
                    tag = el.tag_name.lower()
                    if tag == "select":
                        try:
                            selected_val = (Select(el).first_selected_option.get_attribute("value") or "").strip()
                            if selected_val in ("", "0", "-1"):
                                unfilled.append({"id": field_id, "label": self._get_field_label(el) or field_id, "element": el, "type": "select"})
                        except Exception:
                            pass
                    else:
                        val = (el.get_attribute("value") or "").strip()
                        if not val:
                            unfilled.append({"id": field_id, "label": self._get_field_label(el) or field_id, "element": el, "type": "text"})
                except Exception:
                    pass
        except Exception:
            pass
        return unfilled

    def _handle_unmapped_fields_after_failure(self):
        """Detecta campos no mapeados tras fallo de envío y solicita valores al usuario."""
        if not self.manual_input_callback:
            return False

        unfilled = self._find_unfilled_visible_fields()
        if not unfilled:
            print("⚠️ El formulario falló pero no se detectaron campos visibles sin mapear.")
            return False

        filled_any = False
        for info in unfilled:
            field_id = info["id"]
            label = info["label"]
            element = info["element"]
            ftype = info["type"]

            print(f"⚠️ Campo no mapeado detectado: '{label}' (id: {field_id})")
            try:
                value = self.manual_input_callback(field_id, label)
            except Exception:
                value = None

            if not value:
                print(f"  → Omitido por el usuario.")
                continue

            try:
                if ftype == "select":
                    sel = Select(element)
                    try:
                        sel.select_by_visible_text(value)
                    except Exception:
                        sel.select_by_value(value)
                else:
                    element.clear()
                    element.send_keys(value)
                print(f"  → Rellenado con: '{value}'")
                filled_any = True
            except Exception as e:
                print(f"  → No se pudo rellenar '{label}': {e}")

        return filled_any

    def _detect_event_id_error(self):
        """Detecta el cartel de error 'Lo siento / Ocurrió un inconveniente' (falta de event_id)."""
        try:
            page_text = self.driver.find_element(By.TAG_NAME, "body").text.lower()
        except Exception:
            page_text = ""
        if any(p in page_text for p in self._EVENT_ID_PATTERNS):
            return True
        try:
            popup_els = self.driver.find_elements(
                By.CSS_SELECTOR,
                "div[class*='modal'], div[class*='popup'], div[role='dialog'], "
                "div[role='alert'], div[class*='alert'], div[class*='error-msg'], "
                "div[class*='message'], p[class*='error']"
            )
            for el in popup_els:
                try:
                    if el.is_displayed():
                        et = el.text.lower()
                        if any(p in et for p in self._EVENT_ID_PATTERNS):
                            return True
                except Exception:
                    continue
        except Exception:
            pass
        return False

    def _apply_row_color(self, sheet, row_index, is_ok, start_col=None,
                         form_coincide_col=None, form_coincide_ok=None):
        """Colorea SOLO las columnas de resultado (desde start_col) según si el lead
        se envió OK (verde) o con error (rojo). Los datos de entrada quedan sin color."""
        GREEN_ROW  = PatternFill(fill_type="solid", fgColor="C6EFCE")
        RED_ROW    = PatternFill(fill_type="solid", fgColor="FFC7CE")
        GREEN_CELL = PatternFill(fill_type="solid", fgColor="375623")
        RED_CELL   = PatternFill(fill_type="solid", fgColor="C00000")
        WHITE_FONT = Font(color="FFFFFF", bold=True)

        row_fill = GREEN_ROW if is_ok else RED_ROW
        for col in range(start_col or 1, sheet.max_column + 1):
            sheet.cell(row=row_index, column=col).fill = row_fill

        if form_coincide_col and form_coincide_ok is not None:
            cell = sheet.cell(row=row_index, column=form_coincide_col)
            cell.fill = GREEN_CELL if form_coincide_ok else RED_CELL
            cell.font = WHITE_FONT

    def submit_and_verify_form(self, current_ss_number, expected_form_url):
        """Envía el formulario y verifica resultado"""
        expected_form_url = self._sanitize_url(expected_form_url)
        if bool(self.config.get("solo_verificar_visual", False) or self.config.get("no_enviar_lead", False)):
            print("🚫 [DEBUG] Modo 'No enviar lead' activo. Omitiendo envío final.")
            return "PASS (Verificación visual)", "Verificación visual"
        try:
            boton_enviar, used_sel = self._resolve_submit_button(wait_seconds=3)
            if not boton_enviar:
                raise Exception("No se encontró el botón de envío con ningún selector")
            print(f"Botón enviar encontrado con selector: {used_sel}")

            # Foto del estado REAL del form justo antes de enviar: es lo que viaja en el
            # lead y lo que hay que poder comparar después contra la base de datos.
            try:
                self._sync_tracked_with_dom_before_submit(getattr(self, "_current_form_data", {}))
            except Exception as _se:
                print(f"Aviso: no se pudo sincronizar el tracking con el DOM: {_se}")

            self._scroll_element_into_view(boton_enviar)
            time.sleep(0.3)

            try:
                boton_enviar.click()
            except:
                self.driver.execute_script("arguments[0].click();", boton_enviar)

            print("Clic en enviar (completado) realizado. Verificando resultado...")

            # Detección de TY para forms 2.0 (/tools/forms): mensaje de agradecimiento.
            # OJO: no basta con buscar texto — el encabezado del form ("...e entraremos
            # em contato") daría falso positivo. Sólo hay TY si el form (botón Enviar) ya
            # NO está visible (fue reemplazado por la confirmación) Y hay una frase fuerte.
            def _tiene_thankyou_texto_2_0(d):
                try:
                    _cur = (d.current_url or "").lower()
                except Exception:
                    _cur = ""
                if "/tools/" not in _cur:
                    return False
                # Si el botón Enviar sigue visible, el form NO se envió → no es TY.
                try:
                    _submit_visible = d.execute_script("""
                        var els = document.querySelectorAll("button, input[type='submit'], input[type='button']");
                        for (var i=0;i<els.length;i++){
                            var e=els[i];
                            var t=((e.innerText||e.value||"")+"").trim().toLowerCase();
                            if(e.offsetParent!==null && (t==='enviar' || t.indexOf('enviar')===0)){ return true; }
                        }
                        return false;
                    """)
                    if _submit_visible:
                        return False
                except Exception:
                    pass
                try:
                    _txt = d.execute_script(
                        "return (document.body && document.body.innerText) ? document.body.innerText : '';"
                    ) or ""
                except Exception:
                    _txt = ""
                _txt = _txt.lower()
                # Frases que sólo aparecen en la confirmación (no en el encabezado del form).
                _markers = (
                    "obrigado", "obrigada",
                    "recebemos sua solicit", "recebemos seu contato",
                    "sua solicitação foi recebida", "solicitação enviada",
                    "recebido com sucesso", "recebida com sucesso",
                    "enviado com sucesso", "cadastro realizado",
                    "dados enviados", "mensagem enviada",
                    "gracias por", "thank you for",
                )
                return any(m in _txt for m in _markers)

            # Verificación positiva: esperar div#thank-you con display:block (dentro del iframe)
            def _ty_visible(d):
                # Plataforma clásica: div#thank-you display:block
                try:
                    el = d.find_element(By.CSS_SELECTOR, "div#thank-you")
                    style = el.get_attribute("style") or ""
                    if "display: block" in style or "display:block" in style:
                        return True
                except Exception:
                    pass
                # Bolivia: iframe navega a nueva URL cuyo documento tiene div.rp-wrapper
                try:
                    d.find_element(By.CSS_SELECTOR, "div.rp-wrapper")
                    return True
                except Exception:
                    pass
                # Forms 2.0 (URL con /tools/forms): la TY NO es div#thank-you sino un
                # mensaje de agradecimiento (Gracias / Obrigado / Thank you / etc.).
                try:
                    if _tiene_thankyou_texto_2_0(d):
                        return True
                except Exception:
                    pass
                return False

            # El banner de error de event_id ("Lo siento, ocurrió un inconveniente…") es
            # EFÍMERO: en algunos forms se borra a los ~5s. Si sólo se lo busca después de
            # agotar los 15s de espera de la TY ya no está, y el envío fallido se reportaba
            # como "Formulario sigue visible" en vez de como error de servidor. Por eso se
            # chequea durante la espera y se corta apenas aparece.
            _event_id_seen = [False]

            def _ty_or_event_id_error(d):
                if _ty_visible(d):
                    return True
                try:
                    if self._detect_event_id_error():
                        _event_id_seen[0] = True
                        return True
                except Exception:
                    pass
                return False

            try:
                WebDriverWait(self.driver, 15).until(_ty_or_event_id_error)
                if _event_id_seen[0]:
                    raise TimeoutException("event_id error detectado durante la espera de TY")
                result_text = "Lead enviado correctamente"
                print("TY div detectado. Esperando carga completa...")
                time.sleep(1)

            except TimeoutException:
                if _event_id_seen[0]:
                    result_text = "ERROR_EVENT_ID: Error de servidor al envío - formulario recargado automáticamente"
                    print(f" Error event_id detectado en página: {result_text}")
                    return result_text, None

                # T3/AEM: capturar SIEMPRE el estado post-submit aunque no haya TY, para
                # que se vea qué pasó (form más largo → full-page).
                if getattr(self, "_is_aem", False) and self.screenshot_manager:
                    try:
                        self.driver.switch_to.default_content()
                        self.screenshot_manager.take_form_screenshot(current_ss_number, "post_submit", full_page=True)
                        print("Captura post_submit (sin TY) tomada")
                    except Exception as _pe:
                        print(f"Error capturando post_submit: {_pe}")

                # Chequeo específico: error de event_id del servidor ("Lo siento / Ocurrió un inconveniente")
                if self._detect_event_id_error():
                    result_text = "ERROR_EVENT_ID: Error de servidor al envío - formulario recargado automáticamente"
                    print(f" Error event_id detectado en página: {result_text}")
                    return result_text, None

                # Fallback: checks negativos
                desc_errores = self._describir_errores_visuales()
                if desc_errores:
                    result_text = f"Error visual detectado: {desc_errores} (TY Page no detectada)"
                    print(result_text)
                    return result_text, None
                if self._has_visible_required_field():
                    result_text = "Formulario sigue visible (TY Page no detectada)"
                    print(result_text)
                    return result_text, None
                # Sin errores y sin campos visibles pero sin TY — señal para retry
                print("TY Page no detectada (sin errores). Señalando para retry...")
                return "Enviado sin confirmación TY Page", None

            # TY con CTA: buscar link/CTA en la TY, verificarlo con click real y dejar
            # captura de evidencia en la carpeta de screenshots del browser. (gm_forms /
            # gm_front — no aplica a forms 2.0/AEM, que tienen su propio flujo.)
            try:
                _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                if _root not in sys.path:
                    sys.path.insert(0, _root)
                from utils.ty_cta import investigate_ty_cta, format_ty_cta, format_link_issue
                _cta_info = investigate_ty_cta(
                    self.driver, log=print,
                    evidence_dir=self.SCREENSHOT_DIR, take_screenshot=True,
                )
                self._ty_cta = format_ty_cta(_cta_info)
                self._link_issue = format_link_issue(_cta_info)
                self._link_issue_present = bool(_cta_info.get("has_weird"))
            except Exception as _cta_e:
                print(f"Error investigando CTA en TY: {_cta_e}")

            # TY div detectado — capturar la TY ACOTADA al área del form (muestra "Solicitud
            # enviada"), NO toda la landing (para eso están landing_inicial y landing_final).
            self.driver.switch_to.default_content()
            print("Salido del contexto iframe")
            self.reposition_to_form(expected_form_url)
            ty_page_name = self.screenshot_manager.fname("form", "typage", current_ss_number)
            self.screenshot_manager.take_form_screenshot(current_ss_number, "typage", full_page=True)
            print("Captura 3/3: TY Page (área del formulario)")
            return result_text, ty_page_name

        except Exception as e:
            result_text = f"Error al enviar: {e}"
            print(result_text)
            return result_text, None
    
    def _log(self, msg):
        """Log a consola. No escribe archivo: el portable no debe dejar debug_run.log al usuario."""
        try:
            print(msg)
        except Exception:
            pass

    def run(self, progress_callback=None):
        """Ejecuta el proceso completo - MÉTODO UNIFICADO PARA TODOS LOS PAÍSES"""
        try:
            # Configuración inicial
            self.setup_directories_and_files()
            self._log(f"run() iniciado — {self.config.get('pais','?')}")
            self.initialize_browser()
            
            print(f"\nINICIANDO EJECUCIÓN #{self.RUN_NUMBER} - {self.config['pais'].upper()}")
            print(f"Screenshots: {self.SCREENSHOT_DIR}")
            print(f"Resultados: {self.RESULTADOS_PATH}")
            
            # Cargar Excel
            wb = self.safe_load_workbook(self.EXCEL_PATH)
            sheet = wb.active

            if self._delete_sheet_column_if_header(sheet, "Mensajes Error"):
                print("Columna 'Mensajes Error' eliminada del libro (ya no se usa)")
            
            # Verificar y crear columnas necesarias
            headers = [cell.value for cell in sheet[1] if cell.value]
            required_columns = ["Resultado", "Formulario Inserto", "Formulario Completado", "TY Page",
                                "TYP con CTA", "LINK ISSUE TYP",
                                "Form URL esperada", "Form URL encontrada", "Form coincide",
                                "Datos vs Excel", "Motivo",
                                "Estado URL landing", "Estado URL form"]

            for col_name in required_columns:
                if col_name not in headers:
                    headers.append(col_name)
                    sheet.cell(row=1, column=len(headers)).value = col_name
                    print(f"Columna '{col_name}' agregada al Excel")

            self.safe_save_workbook(wb, self.RESULTADOS_PATH)

            # Obtener índices de columnas
            headers = [cell.value for cell in sheet[1] if cell.value]
            result_col = headers.index("Resultado") + 1
            form_inserto_col = headers.index("Formulario Inserto") + 1
            form_completado_col = headers.index("Formulario Completado") + 1
            ty_page_col = headers.index("TY Page") + 1
            ty_cta_col = headers.index("TYP con CTA") + 1 if "TYP con CTA" in headers else None
            link_issue_col = headers.index("LINK ISSUE TYP") + 1 if "LINK ISSUE TYP" in headers else None
            form_url_esperada_col   = headers.index("Form URL esperada") + 1
            form_url_encontrada_col = headers.index("Form URL encontrada") + 1
            form_coincide_col = headers.index("Form coincide") + 1
            datos_vs_excel_col = headers.index("Datos vs Excel") + 1 if "Datos vs Excel" in headers else None
            motivo_col = headers.index("Motivo") + 1 if "Motivo" in headers else None
            estado_url_landing_col = (headers.index("Estado URL landing") + 1
                                      if "Estado URL landing" in headers else None)
            estado_url_form_col = (headers.index("Estado URL form") + 1
                                   if "Estado URL form" in headers else None)

            ss_counter = 1
            _total_leads = max(0, sheet.max_row - 1)
            _done_leads = 0
            # Resumen autoritativo por formulario (lo lee la UI para el modal / email).
            self.run_summary = {"ok": 0, "fail": 0, "total": _total_leads, "fail_rows": []}

            # Encabezados SIN filtrar los vacíos: los índices tienen que alinear con la fila
            raw_headers = [cell.value for cell in sheet[1]]

            # Procesar cada fila
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                self.begin_row_tracking()
                # Columnas SI/NO que apuntan a un checkbox por name/id (ver build_checkbox_prefs)
                self.checkbox_prefs = self.build_checkbox_prefs(raw_headers, row)
                if self.checkbox_prefs:
                    print(f"  ☑ Preferencias de checkbox (Excel): {self.checkbox_prefs}")
                if len(row) < 2:
                    print(f"Saltando fila {i}: no tiene suficientes columnas")
                    continue
                    
                landing_url, expected_form_url = row[:2]
                landing_url = self._sanitize_url(landing_url)
                expected_form_url = self._sanitize_url(expected_form_url)
                if not isinstance(landing_url, str):
                    landing_url = str(landing_url or "").strip()
                if not isinstance(expected_form_url, str):
                    expected_form_url = str(expected_form_url or "").strip()
                else:
                    expected_form_url = expected_form_url.strip()

                if not landing_url or landing_url.strip() == "":
                    if not expected_form_url or expected_form_url.strip() == "":
                        print(f"Saltando fila {i}: no hay URL de landing ni de formulario")
                        continue
                    # Form standalone: navegar directo a la URL del form, sin iframe
                    print(f"Fila {i}: sin landing URL — navegando directo al form: {expected_form_url}")
                    landing_url = expected_form_url
                    expected_form_url = ""

                use_iframe = bool(expected_form_url)
                
                form_data = self.extract_form_data(row)
                result_text = ""
                form_url_mismatch = False
                self.expected_form_url = expected_form_url
                self.ss_counter = ss_counter

                form_inserto_name = ""
                form_completado_name = ""
                ty_page_name = ""
                self._url_form_encontrado = ""
                self._errores_ss_taken = False
                self._landing_issue = ""   # 404 / error de carga de la landing (se llena en process_landing_page)

                # Estado HTTP real de las URLs: Selenium no expone el status code, así que
                # si la landing da 404 o el form está caído (503), el lead falla sin decir
                # por qué. Se consulta aparte para que el Excel muestre la causa.
                self._estado_url_landing = "-"
                self._estado_url_form = "-"
                self._url_status_problema = ""
                try:
                    from utils.url_status import check_url_status, format_status_pair
                    _st_landing = check_url_status(landing_url)
                    _st_form = check_url_status(expected_form_url) if expected_form_url else {}
                    self._estado_url_landing = _st_landing.get("label", "-")
                    self._estado_url_form = _st_form.get("label", "-") if _st_form else "-"
                    self._url_status_problema = format_status_pair(_st_landing, _st_form)
                    print(f"🌐 Estado URL landing: {self._estado_url_landing}")
                    if expected_form_url:
                        print(f"🌐 Estado URL form:    {self._estado_url_form}")
                except Exception as _use:
                    print(f"Aviso: no se pudo verificar el estado de las URLs: {_use}")
                # Resetear URLs y frame en screenshot_manager al inicio de cada fila
                if self.screenshot_manager:
                    self.screenshot_manager.url_form_esperado   = expected_form_url
                    self.screenshot_manager.url_form_encontrado = ""
                    # Slug del form para los nombres de captura (siempre el mismo dentro de la fila).
                    self.screenshot_manager.form_slug = self._slug_for(landing_url, expected_form_url)
                    self.screenshot_manager.current_frame       = None

                print(f"\n Procesando fila {i}: {landing_url}")

                try:
                    # 1. Procesar landing page. Si vamos a pausar para login manual (solo en la
                    #    primera fila), la captura de landing se toma DESPUÉS del login.
                    _pausar = (i == 2 and bool(self.config.get('pausar_autenticacion', False)))
                    form_inserto_name = self.process_landing_page(landing_url, ss_counter,
                                                                  take_screenshot=not _pausar)

                    # 1b. Pausa opcional para autenticación manual (SSO / MFA / credenciales)
                    if _pausar:
                        from tkinter import messagebox
                        self._log("[INFO] Pausando ejecución para autenticación manual...")
                        res = messagebox.askokcancel(
                            "Ocioso — Autenticación Manual",
                            "Se pausó la ejecución antes del primer formulario para que puedas iniciar sesión.\n\n"
                            "1. Completá la autenticación / login en la ventana del navegador.\n"
                            "2. Asegurate de quedar en la página del formulario.\n"
                            "3. 'Aceptar' continúa con el llenado automático; 'Cancelar' detiene el proceso.",
                        )
                        if not res:
                            raise InterruptedError("Ejecución cancelada por el usuario durante la pausa de autenticación.")
                        self._log("[INFO] Resumiendo ejecución...")
                        # Ya autenticado: por defecto mandar el navegador fuera de pantalla para que
                        # el resto del llenado corra en segundo plano y no le robe el foco al usuario.
                        # Si el usuario activó "Preview: ver navegador durante todo el envío", se deja
                        # visible el navegador el resto de la corrida.
                        if not self.config.get('preview_visible_browser', False):
                            try:
                                self.driver.set_window_position(10000, 0)
                            except Exception:
                                pass
                        try:
                            self.driver.switch_to.default_content()
                        except Exception:
                            pass
                        # El pre-scroll/lazy-load anterior corrió sobre la pantalla de login, no
                        # sobre la landing: hay que reprocesarla ya autenticado o el iframe del
                        # formulario nunca llega a estar en el DOM.
                        form_inserto_name = self.process_landing_page(landing_url, ss_counter,
                                                                      take_screenshot=True)

                    # 2. Buscar iframe (solo si hay URL esperada en columna B); si no, formulario embebido en la página
                    if use_iframe:
                        target_iframe = self.find_and_position_to_form(expected_form_url)
                        if not target_iframe:
                            print("Formulario esperado no encontrado, buscando iframe GM disponible...")
                            try:
                                _cand, _es_gm = self._pick_gm_iframe(
                                    self.driver.find_elements(By.TAG_NAME, "iframe"),
                                    expected_url=expected_form_url)
                            except Exception:
                                _cand, _es_gm = None, False
                            if _cand is not None:
                                target_iframe = _cand
                                # Si es un iframe GM (aunque su src no matchee 1:1 el Excel) no lo
                                # tratamos como mismatch duro: es el form correcto de todos modos.
                                if not _es_gm:
                                    form_url_mismatch = True
                            else:
                                form_url_mismatch = True
                        if target_iframe:
                            if form_url_mismatch:
                                print("Usando iframe disponible (URL de formulario no coincide con Excel)")
                            else:
                                print("Formulario correcto encontrado y posicionado")

                            _iframe_src = ""
                            try:
                                _iframe_src = target_iframe.get_attribute("src") or ""
                            except Exception:
                                pass

                            self.driver.switch_to.frame(target_iframe)
                            if self.screenshot_manager:
                                self.screenshot_manager.current_frame = target_iframe

                            if not _iframe_src or _iframe_src.strip() == "" or _iframe_src.startswith("about:"):
                                try:
                                    _iframe_src = self.driver.execute_script("return window.location.href;") or ""
                                except Exception:
                                    pass

                            if _iframe_src and not _iframe_src.startswith("http"):
                                try:
                                    from urllib.parse import urljoin
                                    self.driver.switch_to.parent_frame()
                                    _landing_url = self.driver.current_url
                                    if self.screenshot_manager and self.screenshot_manager.current_frame:
                                        self.driver.switch_to.frame(self.screenshot_manager.current_frame)
                                    else:
                                        self.driver.switch_to.frame(target_iframe)
                                    _iframe_src = urljoin(_landing_url, _iframe_src)
                                except Exception:
                                    pass

                            if self.screenshot_manager:
                                self.screenshot_manager.url_form_encontrado = _iframe_src
                            self._url_form_encontrado = _iframe_src
                            print("Cambiado al contexto del iframe")
                            self._maybe_click_raq_cta(landing_url, _iframe_src)
                        else:
                            form_url_mismatch = True
                            self.driver.switch_to.default_content()
                            print("No se encontró ningún iframe, intentando con documento principal")
                    else:
                        self.driver.switch_to.default_content()
                        print("Columna B vacía: formulario embebido en documento principal (sin iframe)")
                        # Form suelto (sin landing): el RAQ de Brasil igual arranca en la pantalla del CTA
                        self._maybe_click_raq_cta(landing_url)

                    # 3. Esperar a que el formulario esté listo
                    self.wait_for_form_ready_in_iframe()

                    # 3a. Captura del FORMULARIO VACÍO (área del form, antes de tocar nada).
                    if self.screenshot_manager:
                        try:
                            self.screenshot_manager.take_form_screenshot(ss_counter, "vacio", full_page=True)
                            print("Captura form_vacio tomada (formulario vacío inicial)")
                        except Exception as _e:
                            print(f"  ⚠ Captura form_vacio: {_e}")

                    # Detección T3/AEM temprana: los forms 2.0 son más largos → capturas full-page
                    self._is_aem = self._is_aem_adaptive_form()

                    # 3b. Click enviar vacío + captura errores (todos excepto libro-reclamaciones).
                    # El Libro de Reclamaciones se detecta por URL (landing o form) o por DOM,
                    # pero el llenado directo por ids cc_* sólo aplica si esos campos existen
                    # de verdad: si una URL trae 'reclamos' pero el form es otro, se llena con
                    # el motor normal en vez de quedar vacío.
                    _is_libro_reclamaciones = self._is_libro_reclamaciones_form(landing_url)
                    if _is_libro_reclamaciones and not self._has_libro_reclamaciones_fields():
                        print("URL de reclamos pero sin campos cc_*: se llena con el motor normal.")
                        _is_libro_reclamaciones = False
                    if not _is_libro_reclamaciones:
                        try:
                            _btn_empty, _ = self._resolve_submit_button(wait_seconds=2)
                            # Solo se clickea en vacío un ENVIAR real. Un CTA tipo "Siguiente"
                            # (ej. btn-steps-submit) matchea class*='submit' pero NO debe
                            # clickearse vacío: puede estar ausente hasta elegir un campo
                            # obligatorio o avanzar de paso indebidamente. Se llena primero.
                            if _btn_empty and self._is_next_button_element(_btn_empty):
                                _btn_empty = None
                            if _btn_empty:
                                self._scroll_element_into_view(_btn_empty)
                                time.sleep(0.3)
                                try:
                                    _btn_empty.click()
                                except Exception:
                                    self.driver.execute_script("arguments[0].click();", _btn_empty)
                                time.sleep(0.5)  # esperar a que JS muestre los errores de validación
                                if self.screenshot_manager:
                                    self.screenshot_manager.take_form_screenshot(ss_counter, "errores", full_page=True)
                                    self._errores_ss_taken = True
                                    print("Captura form_errores tomada (formulario vacío)")
                        except Exception as _e:
                            print(f"  ⚠ Click enviar vacío: {_e}")

                    # 4. Llenar campos del formulario
                    _original_field_mapping = None

                    try:
                        if _is_libro_reclamaciones:
                            print("URL libro-reclamaciones detectada: llenando nombre/teléfono/email directamente.")
                            form_completado_name = self._fill_libro_reclamaciones_direct(form_data)
                        else:
                            self._log(f"fill_form_fields() iniciando — fila {i}")
                            form_completado_name = self.fill_form_fields(form_data)
                            self._log(f"fill_form_fields() OK — fila {i}")

                        # 5. Enviar y verificar formulario
                        result_text, ty_page_name = self.submit_and_verify_form(ss_counter, expected_form_url)

                        # 5b. Fix específico Brasil: re-ingresar doc via send_keys antes del retry genérico
                        if ty_page_name is None and "Error visual" in (result_text or ""):
                            if self._refill_brasil_doc_sendkeys(form_data):
                                print("  ↺ Re-ingreso doc Brasil via send_keys, reintentando envío...")
                                result_text, ty_page_name = self.submit_and_verify_form(ss_counter, expected_form_url)

                        # 5b2. Retry event_id: formulario recarga solo → rellenar sin recargar landing
                        _event_id_retry_done = False
                        if ty_page_name is None and isinstance(result_text, str) and result_text.startswith("ERROR_EVENT_ID:"):
                            _event_id_retry_done = True
                            print(f"  ↺ Error event_id (intento 1). Esperando recarga automática del formulario...")
                            try:
                                time.sleep(4)
                                if use_iframe:
                                    _ev_iframe = self.find_and_position_to_form(expected_form_url)
                                    if _ev_iframe:
                                        self.driver.switch_to.frame(_ev_iframe)
                                    else:
                                        self.driver.switch_to.default_content()
                                self.wait_for_form_ready_in_iframe()
                                if _is_libro_reclamaciones:
                                    form_completado_name = self._fill_libro_reclamaciones_direct(form_data)
                                else:
                                    form_completado_name = self.fill_form_fields(form_data)
                                if self.screenshot_manager:
                                    self.screenshot_manager.take_form_screenshot(ss_counter, "completado_intento2", full_page=True)
                                    print("Captura formulario completo (intento 2 tras error event_id)")
                                result_text_2, ty_page_name_2 = self.submit_and_verify_form(ss_counter, expected_form_url)
                                if ty_page_name_2 is not None:
                                    result_text = result_text_2
                                    ty_page_name = ty_page_name_2
                                elif isinstance(result_text_2, str) and result_text_2.startswith("ERROR_EVENT_ID:"):
                                    result_text = "Error event_id: formulario no pudo enviarse por falta de event_id (2 intentos fallidos)"
                                    ty_page_name = None
                                    try:
                                        self.driver.switch_to.default_content()
                                        self.reposition_to_form(expected_form_url)
                                        ty_page_name = self.screenshot_manager.fname("landing", "typage", ss_counter)
                                        self.screenshot_manager.take_landing_screenshot(ss_counter, "typage")
                                    except Exception as _cap_e:
                                        print(f"Error capturando evidencia event_id: {_cap_e}")
                                else:
                                    result_text = (result_text_2 or "Fallo") + " (intento 2 tras error event_id)"
                                    ty_page_name = ty_page_name_2
                            except Exception as _ev_e:
                                print(f"  Error en retry event_id: {_ev_e}")
                                result_text = f"Error event_id: fallo en reintento — {_ev_e}"

                        # 5c-1. Retry rápido: recargar SOLO el iframe del formulario (no toda la
                        #       landing), rellenar de nuevo y reenviar, quedando posicionado en el form.
                        if ty_page_name is None and not _event_id_retry_done and use_iframe:
                            print(f"  ↺ Intento 2/2 (rápido): recargando solo el iframe (motivo: {result_text!r})...")
                            try:
                                if self._reload_form_iframe(expected_form_url):
                                    if _is_libro_reclamaciones:
                                        form_completado_name = self._fill_libro_reclamaciones_direct(form_data)
                                    else:
                                        form_completado_name = self.fill_form_fields(form_data)
                                    result_text, ty_page_name = self.submit_and_verify_form(ss_counter, expected_form_url)
                            except Exception as _rl_e:
                                print(f"  Retry por recarga de iframe falló, se recargará la landing: {_rl_e}")

                        # 5c-2. Retry genérico (2º fallback): recargar toda la landing + rellenar + reenviar
                        if ty_page_name is None and not _event_id_retry_done:
                            print(f"  ↺ Intento 2/2: recargando landing y rellenando de nuevo (motivo: {result_text!r})...")
                            try:
                                self.driver.switch_to.default_content()
                            except Exception:
                                pass
                            form_inserto_name = self.process_landing_page(landing_url, ss_counter)
                            if use_iframe:
                                _retry_iframe = self.find_and_position_to_form(expected_form_url)
                                if not _retry_iframe and form_url_mismatch:
                                    try:
                                        _retry_iframe, _ = self._pick_gm_iframe(
                                            self.driver.find_elements(By.TAG_NAME, "iframe"),
                                            expected_url=expected_form_url)
                                    except Exception:
                                        _retry_iframe = None
                                if _retry_iframe:
                                    self.driver.switch_to.frame(_retry_iframe)
                                else:
                                    self.driver.switch_to.default_content()
                            self.wait_for_form_ready_in_iframe()
                            if _is_libro_reclamaciones:
                                form_completado_name = self._fill_libro_reclamaciones_direct(form_data)
                            else:
                                form_completado_name = self.fill_form_fields(form_data)
                            result_text, ty_page_name = self.submit_and_verify_form(ss_counter, expected_form_url)
                            if ty_page_name is None:
                                result_text = (result_text or "Fallo") + " (2 intentos fallidos)"
                                try:
                                    self.driver.switch_to.default_content()
                                    self.reposition_to_form(expected_form_url)
                                    ty_page_name = f"landing_typage_{ss_counter}.png"
                                    self.screenshot_manager.take_landing_screenshot(ss_counter, "typage")
                                    print("Captura evidencia tras 2 intentos fallidos")
                                except Exception as _cap_e:
                                    print(f"Error capturando evidencia: {_cap_e}")

                    except Exception as e:
                        # Errores durante llenado o envío (iframe o documento principal)
                        _exc_info = f"[{type(e).__name__}] {e}"
                        print(f"Error en fill/submit: {_exc_info}")
                        self._log(f"EXCEPTION fila {i}: {_exc_info}")
                        try:
                            desc_errores = self._describir_errores_visuales()
                            if desc_errores:
                                result_text = f" Error completando {_exc_info}: {desc_errores}"
                            else:
                                result_text = f" Error completando: {_exc_info}"
                        except Exception:
                            result_text = f" Error completando: {_exc_info}"
                    finally:
                        if _original_field_mapping is not None:
                            self.field_mapping = _original_field_mapping

                    # 6. Volver al contexto principal
                    self.driver.switch_to.default_content()
                    print("Vuelto al contexto principal")

                    # 6b. Captura de LANDING COMPLETA post-envío (aparezca o no la thank-you page).
                    if self.screenshot_manager:
                        try:
                            self.screenshot_manager.current_frame = None
                            self.screenshot_manager.take_landing_screenshot(ss_counter, "final")
                            print("Captura landing_final tomada (post-envío)")
                        except Exception as _e:
                            print(f"  ⚠ Captura landing_final: {_e}")

                except Exception as e:
                    result_text = f" Error general: {e}"
                    print(result_text)
                    try:
                        self.driver.switch_to.default_content()
                    except:
                        pass
                
                # 7. Guardar resultados en Excel
                # Limpiar prefijo interno ERROR_EVENT_ID si quedó sin procesar
                if isinstance(result_text, str) and result_text.startswith("ERROR_EVENT_ID:"):
                    result_text = result_text[len("ERROR_EVENT_ID:"):].strip() or "Error event_id (sin detalles)"

                # Determinar si el lead se envió OK antes de agregar prefijos
                _lead_ok = (result_text == "Lead enviado correctamente")

                # Landing rota (404): cualquier problema de la landing es fallo del form.
                _landing_issue = getattr(self, "_landing_issue", "")
                if _landing_issue:
                    _prev = (result_text or "").strip()
                    result_text = f"[Error Landing] {_landing_issue}"
                    if _prev and not _prev.lower().startswith("lead enviado"):
                        result_text += f" || {_prev}"
                    _lead_ok = False

                # Estado del form inserto en TRES niveles (no dos), y si el lead viajó
                # igual pese al form equivocado (eso va en naranja, no en rojo).
                self._form_inserto_estado = "inserto_ok"
                self._form_mismatch_enviado = False
                if form_url_mismatch:
                    _found_for_msg = getattr(self, "_url_form_encontrado", "") or ""
                    self._form_mismatch_enviado = bool(_lead_ok)
                    self._form_inserto_estado = "inserto_otro" if _found_for_msg else "no_inserto"
                    _lead_str = "se envió lead igualmente" if _lead_ok else "lead no enviado"
                    # Preservar el error de envío original (si lo hubo) para mapear TODOS los errores.
                    _orig_err = "" if _lead_ok else (result_text or "").strip()
                    if _found_for_msg:
                        # SÍ hay un formulario inserto (iframe con src), pero es distinto al esperado.
                        # "Form no inserto" se reserva para cuando NO hay ningún form.
                        _mismatch_msg = (
                            f"[Error Form] Form inserto NO coincide con el esperado, {_lead_str} — "
                            f"URL esperada: {expected_form_url or '?'} | "
                            f"URL encontrada: {_found_for_msg}"
                        )
                    else:
                        # No hay NINGÚN form GM en la landing (iframe sin src / inexistente).
                        _mismatch_msg = (
                            f"[Error Form] FORM NO INSERTO (iframe sin src / sin form en la landing) — "
                            f"URL esperada: {expected_form_url or '?'} | "
                            f"URL encontrada: ninguno | {_lead_str}"
                        )
                    # Mapear TODOS los errores: el del form + el de envío (si el form llenó pero falló).
                    if _orig_err and not _orig_err.lower().startswith("lead enviado"):
                        result_text = f"{_mismatch_msg} || Error de envío: {_orig_err}"
                    else:
                        result_text = _mismatch_msg
                    _lead_ok = False
                _sin_mapeo = getattr(self, "_campos_sin_mapeo_exitoso", [])
                if _sin_mapeo:
                    result_text = (result_text or "") + (
                        f" | Campos sin completar (sin valor asignado): {', '.join(_sin_mapeo)}"
                        f" — asignales un valor en ⚙ IDs Dinámicos"
                    )
                    # No se pudieron completar todos los campos → el lead NO cuenta como PASS
                    _lead_ok = False
                _sin_valor = getattr(self, "_campos_sin_valor_asignado", [])
                if _sin_valor:
                    result_text = (result_text or "") + (
                        f" | Aviso — campos opcionales vacíos (sin valor asignado): {', '.join(_sin_valor)}"
                        f" — podés asignarles un valor en ⚙ IDs Dinámicos"
                    )
                _no_encontrados = getattr(self, "_campos_dropdown_no_encontrados", [])
                if _no_encontrados:
                    result_text = (result_text or "") + f" | Dropdown no encontrado: {'; '.join(_no_encontrados)}"
                    # Un dropdown que no se pudo completar es un fallo de campo → NO es PASS.
                    _lead_ok = False
                # Issue en el CTA/link de la TY page → columna LINK ISSUE en rojo y el lead falla.
                if getattr(self, "_link_issue_present", False):
                    _li = getattr(self, "_link_issue", "") or ""
                    result_text = (result_text or "") + f" | LINK ISSUE TYP: {_li}".rstrip()
                    _lead_ok = False
                sheet.cell(row=i, column=result_col).value = result_text
                # OJO: NO se tocan las columnas de entrada (Modelo/Nombre/etc. las setea el
                # usuario). El modelo realmente elegido queda en las columnas de tracking
                # (PasoN::models / PasoN::model), para poder comparar pedido vs completado.
                sheet.cell(row=i, column=form_inserto_col).value = form_inserto_name
                sheet.cell(row=i, column=form_completado_col).value = form_completado_name if form_completado_name else "-"
                sheet.cell(row=i, column=ty_page_col).value = ty_page_name if ty_page_name else "-"
                if ty_cta_col:
                    sheet.cell(row=i, column=ty_cta_col).value = getattr(self, "_ty_cta", "") or "-"
                if link_issue_col:
                    sheet.cell(row=i, column=link_issue_col).value = getattr(self, "_link_issue", "-") or "-"
                if expected_form_url:
                    sheet.cell(row=i, column=form_url_esperada_col).value = expected_form_url
                _found_url = getattr(self, "_url_form_encontrado", "")
                sheet.cell(row=i, column=form_url_encontrada_col).value = _found_url
                _form_coincide_val = ""
                _form_coincide_ok = None
                if expected_form_url:
                    def _norm(u):
                        if not u: return ""
                        u = u.strip().split("?")[0].split("#")[0]
                        if u.endswith("/"): u = u[:-1]
                        return u.lower()
                    _form_coincide_ok = (_norm(_found_url) == _norm(expected_form_url))
                    _mismatch_enviado = getattr(self, "_form_mismatch_enviado", False)
                    # Si la landing nunca cargó (404 / sin respuesta / redirect), el form
                    # jamás se buscó: decir "no coincide" sería engañoso.
                    _lt = str(getattr(self, "_estado_url_landing", "-") or "-").upper()
                    _landing_falla = bool(
                        _lt not in ("-", "") and
                        (not _lt.startswith("200") or "REDIRIGE" in _lt)
                    )
                    if _form_coincide_ok:
                        _form_coincide_val = "PASS"
                    elif _landing_falla:
                        _form_coincide_val = (
                            f"N/D — la landing no cargó "
                            f"({getattr(self, '_estado_url_landing', '-')})")
                    elif _mismatch_enviado:
                        _form_coincide_val = "FAIL — form inserto no coincide, se envió lead igualmente"
                    elif getattr(self, "_form_inserto_estado", "") == "no_inserto":
                        _form_coincide_val = "FAIL — form NO inserto (iframe sin src)"
                    else:
                        _form_coincide_val = "FAIL — form inserto no coincide con el esperado"
                sheet.cell(row=i, column=form_coincide_col).value = _form_coincide_val
                if datos_vs_excel_col:
                    sheet.cell(row=i, column=datos_vs_excel_col).value = (
                        getattr(self, "_datos_vs_excel", "-") or "-")

                # Motivo del fallo: por qué no salió el lead, sin leer el Resultado entero
                _motivo = "-"
                if not _lead_ok:
                    _motivo = self._short_fail_reason(result_text)
                    _url_prob = getattr(self, "_url_status_problema", "") or ""
                    if _url_prob:
                        # Una landing 404 explica cualquier "form no encontrado": va primero
                        _motivo = f"{_url_prob} | {_motivo}"
                if motivo_col:
                    sheet.cell(row=i, column=motivo_col).value = _motivo
                if estado_url_landing_col:
                    sheet.cell(row=i, column=estado_url_landing_col).value = (
                        getattr(self, "_estado_url_landing", "-") or "-")
                if estado_url_form_col:
                    sheet.cell(row=i, column=estado_url_form_col).value = (
                        getattr(self, "_estado_url_form", "-") or "-")

                self.write_tracked_fields_to_sheet(sheet, i)

                # Colorear SOLO las columnas de resultado: verde si lead OK, rojo si error
                self._apply_row_color(sheet, i, _lead_ok, start_col=result_col,
                                      form_coincide_col=form_coincide_col, form_coincide_ok=_form_coincide_ok)

                # Form distinto al esperado PERO el lead viajó: ámbar, no rojo — hay dato en
                # la base, sólo que entró por otro formulario.
                if getattr(self, "_form_mismatch_enviado", False):
                    for _c in (result_col, form_coincide_col, motivo_col):
                        if not _c:
                            continue
                        _cell = sheet.cell(row=i, column=_c)
                        _cell.fill = PatternFill(fill_type="solid", fgColor="FFC000")
                        _cell.font = Font(color="000000", bold=True)

                # Estado de las URLs: verde 200, ámbar si redirige, rojo si 404/503/sin respuesta
                for _c in (estado_url_landing_col, estado_url_form_col):
                    if not _c:
                        continue
                    _cell = sheet.cell(row=i, column=_c)
                    _txt = str(_cell.value or "").upper()
                    if _txt.startswith("200"):
                        _fg = "C6EFCE"
                    elif "REDIRIGE" in _txt:
                        _fg = "FFC000"
                    elif _txt in ("-", ""):
                        continue
                    else:
                        _fg = "FFC7CE"
                    _cell.fill = PatternFill(fill_type="solid", fgColor=_fg)

                # Enviado OK pero con datos distintos a los pedidos en el Excel: la fila no
                # puede quedar verde a secas — se marca en ámbar el detalle de la diferencia.
                if datos_vs_excel_col and getattr(self, "_datos_mismatch", False):
                    _dv = sheet.cell(row=i, column=datos_vs_excel_col)
                    _dv.fill = PatternFill(fill_type="solid", fgColor="FFC000")
                    _dv.font = Font(color="000000", bold=True)

                # LINK ISSUE TYP: rojo si hay link raro, aunque el lead se haya enviado OK
                if link_issue_col and getattr(self, "_link_issue_present", False):
                    _cell = sheet.cell(row=i, column=link_issue_col)
                    _cell.fill = PatternFill(fill_type="solid", fgColor="C00000")
                    _cell.font = Font(color="FFFFFF", bold=True)

                self.safe_save_workbook(wb, self.RESULTADOS_PATH)
                # Acumular resumen por formulario (autoritativo, mismo criterio que el color).
                try:
                    if _lead_ok:
                        self.run_summary["ok"] += 1
                    else:
                        self.run_summary["fail"] += 1
                        # fila + motivo corto (para el resumen rápido del modal / email).
                        self.run_summary["fail_rows"].append(
                            {"row": i, "reason": self._short_fail_reason(result_text)})
                except Exception:
                    pass
                _done_leads += 1
                if progress_callback:
                    try:
                        progress_callback(_done_leads, _total_leads)
                    except Exception:
                        pass
                ss_counter += 1

            # Layout final: columnas de RESULTADO primero, datos de entrada al final. Se
            # hace recién acá porque durante la corrida se leen los datos del lead de la
            # misma hoja (URL en A, Formulario en B, campos desde data_start_index) y
            # reordenar antes desalinearía esa lectura.
            try:
                from utils.excel_layout import reordenar_archivo
                reordenar_archivo(self.RESULTADOS_PATH, log=print)
            except Exception as _re:
                print(f"Aviso: no se pudo reordenar el Excel: {_re}")

            print(f"\nProceso finalizado. Ejecución #{self.RUN_NUMBER} completada")
            print(f"Screenshots guardados en: {self.SCREENSHOT_DIR}")
            print(f"Resultados guardados en: {self.RESULTADOS_PATH}")
            print(f"Se capturaron {ss_counter-1} sets completos de screenshots")

            # Limpiar temporales de captura que hayan quedado sueltos (no deben viajar al email).
            try:
                import glob as _g
                for _t in _g.glob(os.path.join(self.SCREENSHOT_DIR, "temp_*.png")):
                    try:
                        os.remove(_t)
                    except Exception:
                        pass
            except Exception:
                pass

            # Garantizar que TODAS las capturas pesen < 50 MB (recomprime solo si se pasa).
            try:
                from screenshot_manager import enforce_screenshot_budget
                enforce_screenshot_budget(self.SCREENSHOT_DIR, max_mb=48)
            except Exception as _eb:
                print(f"Aviso: no se pudo aplicar el presupuesto de capturas: {_eb}")

        except Exception as e:
            print(f"Error crítico: {e}")
        finally:
            self.cleanup()
    def detect_form_type(self):
        """Detecta si es formulario de 1 paso o 3 pasos"""
        try:
            if self._find_next_button():
                return 3
            
            # Si no encuentra botones de siguiente, es de 1 paso
            return 1
            
        except Exception as e:
            print(f"Error detectando tipo de formulario: {e}")
            return 1  # Por defecto asumir 1 paso
    
    def cleanup(self):
        """Limpia recursos"""
        if self.driver:
            self.driver.quit()