import os
import time
import requests
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from selenium.webdriver.common.by import By

# Import base classes dynamically to ensure project compatibility
import sys
_base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for p in (_base, os.path.join(_base, "core")):
    if p not in sys.path:
        sys.path.insert(0, p)

from core.generic_country_base import GenericCountryBase
from utils.data_generator import generar_fila_datos

# Igual que en run.py: hace que Python use la lista de certificados de confianza de Windows.
# Sin esto, detrás de un proxy corporativo que inspecciona TLS (Netskope/Zscaler) la
# verificación de certificados fallaría y el chequeo marcaría FAIL a URLs sanas. Se repite
# acá porque este módulo es el que hace las llamadas HTTP y puede importarse suelto.
try:
    import truststore
    truststore.inject_into_ssl()
except Exception:
    pass

class MassiveFormFiller(GenericCountryBase):
    """
    Subclase especializada de GenericCountryBase para el chequeo masivo sin envíos.
    Configura de forma dura el modo rellenado parcial / no enviar lead,
    y escribe las capturas en la ruta específica solicitada.
    """
    def __init__(self, country_name, browser="chrome", headless=True, background=True):
        super().__init__(country_name, browser=browser, headless=headless, background=background)
        self.config["solo_verificar_visual"] = True
        self.config["no_enviar_lead"] = True

    def setup_directories_and_files(self, custom_dir=None):
        """Sobrescribir para guardar las capturas en la subcarpeta resultados/resultado_urlsinsertas/Capturas"""
        if custom_dir:
            self.SCREENSHOT_DIR = os.path.normpath(custom_dir)
        else:
            self.SCREENSHOT_DIR = os.path.normpath(os.path.join(self.BASE_DIR, "resultados", "resultado_urlsinsertas", "Capturas"))
        os.makedirs(self.SCREENSHOT_DIR, exist_ok=True)
        self.RUN_NUMBER = 1
        return 1

    def fill_minimum_fields_massive(self):
        """
        Completa al menos un campo de texto (Nombre/Apellido) con caracteres aleatorios (mínimo 4)
        y un dropdown (Select o personalizado) seleccionando una opción aleatoria no placeholder.
        """
        import random
        nombres_db = ["Carlos", "Daniel", "Mariana", "Alejandro", "Sofia", "Mateo", "Camila", "Nicolas", "Valentina", "Sebastian"]
        apellidos_db = ["Gomez", "Rodriguez", "Lopez", "Fernandez", "Martinez", "Perez", "Gonzalez", "Sanchez", "Alvarez", "Romero"]
        
        # Generar nombre y apellido aleatorios de mínimo 4 caracteres
        num_names = random.randint(1, 2)
        first_val = " ".join(random.choices(nombres_db, k=num_names))
        while len(first_val) < 4:
            first_val = " ".join(random.choices(nombres_db, k=num_names))
            
        last_val = " ".join(random.choices(apellidos_db, k=random.randint(1, 2)))
        while len(last_val) < 4:
            last_val = " ".join(random.choices(apellidos_db, k=random.randint(1, 2)))

        # 1. Rellenar Campo de Texto (Nombre o Apellido)
        inputs = self.driver.find_elements(By.TAG_NAME, "input")
        name_inputs = []
        lastname_inputs = []
        generic_text_inputs = []
        
        for inp in inputs:
            try:
                if not inp.is_displayed() or not inp.is_enabled():
                    continue
                itype = str(inp.get_attribute("type") or "").lower()
                if itype not in ["text", "", "email", "tel"]:
                    continue
                    
                name_attr = str(inp.get_attribute("name") or "").lower()
                id_attr = str(inp.get_attribute("id") or "").lower()
                placeholder = str(inp.get_attribute("placeholder") or "").lower()
                class_attr = str(inp.get_attribute("class") or "").lower()
                
                if any(x in name_attr or x in id_attr or x in placeholder or x in class_attr for x in ["firstname", "first_name", "nombre", "name"]):
                    if "last" not in name_attr and "last" not in id_attr and "apellido" not in name_attr and "apellido" not in id_attr:
                        name_inputs.append(inp)
                        continue
                
                if any(x in name_attr or x in id_attr or x in placeholder or x in class_attr for x in ["lastname", "last_name", "apellido"]):
                    lastname_inputs.append(inp)
                    continue
                    
                generic_text_inputs.append(inp)
            except Exception:
                pass

        filled_text = False
        target_inp = None
        target_val = first_val
        
        if name_inputs:
            target_inp = name_inputs[0]
            target_val = first_val
        elif lastname_inputs:
            target_inp = lastname_inputs[0]
            target_val = last_val
        elif generic_text_inputs:
            target_inp = generic_text_inputs[0]
            target_val = first_val

        if target_inp:
            try:
                target_inp.clear()
                target_inp.send_keys(target_val)
                filled_text = True
            except Exception as e:
                print(f"Error rellenando campo de texto masivo: {e}")

        # Rellenar también el segundo campo de texto si existe
        if filled_text:
            try:
                if target_inp in name_inputs and lastname_inputs:
                    lastname_inputs[0].clear()
                    lastname_inputs[0].send_keys(last_val)
                elif target_inp in lastname_inputs and name_inputs:
                    name_inputs[0].clear()
                    name_inputs[0].send_keys(first_val)
            except Exception:
                pass

        # 2. Seleccionar Dropdown (Select o custom)
        selects = self.driver.find_elements(By.TAG_NAME, "select")
        filled_dropdown = False
        
        for sel in selects:
            try:
                if not sel.is_displayed() or not sel.is_enabled():
                    continue
                from selenium.webdriver.support.ui import Select
                select_obj = Select(sel)
                valid_options = []
                for opt in select_obj.options:
                    otxt = (opt.text or "").strip()
                    oval = (opt.get_attribute("value") or "").strip()
                    otxt_lower = otxt.lower()
                    
                    if any(p in otxt_lower for p in ["selecciona", "seleccione", "selecione", "select", "elige", "elegir", "choose", "--", "placeholder"]):
                        continue
                    if not otxt or not oval:
                        continue
                    valid_options.append(opt)
                    
                if valid_options:
                    chosen_opt = random.choice(valid_options)
                    select_obj.select_by_visible_text(chosen_opt.text)
                    filled_dropdown = True
                    break
            except Exception as e:
                print(f"Error dropdown nativo: {e}")

        # Fallback para custom dropdowns (AEM / custom divs)
        if not filled_dropdown:
            try:
                dropdown_triggers = self.driver.find_elements(By.XPATH, "//*[contains(@class, 'select') or contains(@class, 'dropdown') or @role='combobox' or @role='listbox']")
                for trigger in dropdown_triggers:
                    if not trigger.is_displayed() or not trigger.is_enabled():
                        continue
                    try:
                        trigger.click()
                        time.sleep(0.5)
                        options = self.driver.find_elements(By.XPATH, "//li[not(contains(translate(text(), 'SELECI ONAREG', 'selecionareg'), 'selecciona')) and not(contains(translate(text(), 'SELECI ONAREG', 'selecionareg'), 'seleccione'))] | //*[ @role='option' and not(contains(translate(text(), 'SELECI ONAREG', 'selecionareg'), 'selecciona'))]")
                        valid_custom_opts = [o for o in options if o.is_displayed() and (o.text or "").strip()]
                        if valid_custom_opts:
                            chosen_opt = random.choice(valid_custom_opts)
                            chosen_opt.click()
                            filled_dropdown = True
                            break
                    except Exception:
                        pass
            except Exception as e:
                print(f"Error custom dropdown fallback: {e}")
        return (filled_text and filled_dropdown)

    _UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
           "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    def _http_status_str(self, url, report_redirects=False):
        """Status HTTP de una URL para el reporte. Devuelve (texto, url_final).

        La verificación TLS va ACTIVADA a propósito. Antes se llamaba con verify=False y
        se silenciaba el InsecureRequestWarning: un certificado vencido, mal emitido o un
        intermediario que interceptara el tráfico devolvían "200 PASS" igual. En una
        herramienta cuyo trabajo es justamente validar que las URLs estén sanas, un
        certificado roto tiene que salir como FAIL, no taparse.

        En oficinas con proxy que inspecciona TLS (Netskope/Zscaler) el certificado lo
        firma el proxy: Windows ya confía en él, y truststore hace que Python use esa misma
        lista, así que no genera falsos FAIL (ver la inyección arriba y en run.py).
        """
        try:
            r = requests.get(url, headers={"User-Agent": self._UA},
                             timeout=8, allow_redirects=True)
        except requests.exceptions.SSLError as e:
            # Certificado inválido/vencido/no confiable: es un hallazgo real del chequeo.
            return f"SSL - ❌ FAIL Certificado TLS inválido o no confiable: {str(e)[:160]}", url
        except requests.exceptions.Timeout:
            return "Timeout - ❌ FAIL La URL no respondió en 8s", url
        except requests.exceptions.RequestException as e:
            return f"Error - ❌ FAIL {type(e).__name__}: {str(e)[:160]}", url
        except Exception as e:
            return f"Error - ❌ FAIL HTTP Code: {type(e).__name__}", url

        if report_redirects and r.history:
            code = r.history[0].status_code
            return (f"{code} - 🔀 FAIL Link redirects. Code: {code}, "
                    f"URL Original: {url} redirects to: {r.url}"), r.url
        if r.status_code == 200:
            return "200 - ✅ PASS Valid link. HTTP Code: 200", r.url
        return f"{r.status_code} - ❌ FAIL HTTP Code: {r.status_code}", r.url

    def check_single_url(self, landing_url, expected_form_url, row_num, tomar_capturas=True):
        """
        Navega a la URL y ejecuta la lógica de verificación (HTTP Status, Iframe Match, Llenado parcial)
        """
        if getattr(self, "stop_event", None) and self.stop_event.is_set():
            raise Exception("Cancelado por el usuario")
        landing_url = self._sanitize_url(landing_url)
        expected_form_url = self._sanitize_url(expected_form_url)

        # URL suelta (form standalone, sin landing que lo embeba). Pasa cuando la fila trae
        # solo URL SECURE, o cuando las dos columnas son la misma URL — típico de los forms
        # nuevos tipo .../gm_frontend/chevrolet/t3/<pais>/form/<slug>, que son una página
        # entera y no un iframe. Se abre el form directo y no se busca ningún iframe: es la
        # misma convención que ya usa Envío de Leads (columna URL vacía = form standalone).
        def _norm(u):
            return (u or "").strip().rstrip("/").lower()

        standalone = (not landing_url) or (_norm(landing_url) == _norm(expected_form_url))
        if standalone:
            landing_url = landing_url or expected_form_url
            expected_form_url = ""

        self.expected_form_url = expected_form_url
        self.ss_counter = row_num
        self._url_form_encontrado = ""
        self._landing_issue = ""
        self._errores_ss_taken = False
        
        if self.screenshot_manager:
            self.screenshot_manager.url_form_esperado = expected_form_url
            self.screenshot_manager.url_form_encontrado = ""
            self.screenshot_manager.form_slug = self._slug_for(landing_url, expected_form_url)
            self.screenshot_manager.current_frame = None

        # 1. Obtener HTTP Status Code y Redirecciones del landing page
        landing_status_str, final_url_found = self._http_status_str(landing_url, report_redirects=True)

        # 1b. Obtener HTTP Status del expected secure URL (Secure_Excel)
        if expected_form_url:
            secure_status_str, _ = self._http_status_str(expected_form_url)
        else:
            secure_status_str = "N/A"

        if getattr(self, "stop_event", None) and self.stop_event.is_set():
            raise Exception("Cancelado por el usuario")
        # 2. Navegar con Selenium
        ss_landing_path = ""
        try:
            ss_landing_path = self.process_landing_page(landing_url, row_num, take_screenshot=tomar_capturas)
        except Exception as e:
            return {
                "ok": False,
                "error": f"Error navegación: {str(e)}",
                "current_url": final_url_found,
                "landing_status": landing_status_str,
                "iframe_correct": "NO",
                "form_url_found": "",
                "form_coincide": "❌ FAIL",
                "displayed": "❌ FAIL",
                "secure_status": secure_status_str,
                "forms_count": 0,
                "ss_landing": ss_landing_path or "",
                "ss_completado": ""
            }

        # Validar si base_form_filler detectó 404 en el título/encabezados
        if self._landing_issue:
            return {
                "ok": False,
                "error": self._landing_issue,
                "current_url": self.driver.current_url,
                "landing_status": f"404 - ❌ FAIL {self._landing_issue}",
                "iframe_correct": "NO",
                "form_url_found": "",
                "form_coincide": "❌ FAIL",
                "displayed": "❌ FAIL",
                "secure_status": secure_status_str,
                "forms_count": 0,
                "ss_landing": ss_landing_path,
                "ss_completado": ""
            }

        # 3. Detectar iframe e inyectar URL de form
        use_iframe = bool(expected_form_url)
        form_url_mismatch = False
        target_iframe = None
        iframe_src = ""
        
        # Contar iframes disponibles (solo los que correspondan a formularios GM).
        # Muchas landings cargan el iframe del form en diferido: si se mira el DOM apenas
        # termina de navegar, el form todavía no está y se marcaba FAIL un form que sí existe.
        # Por eso se espera hasta FORM_WAIT_SECONDS a que aparezca alguno; si ya está, sigue
        # de largo sin esperar nada.
        FORM_WAIT_SECONDS = 5

        def _buscar_iframes_gm():
            try:
                todos = self.driver.find_elements(By.TAG_NAME, "iframe")
            except Exception:
                return [], []
            # Mismo criterio que Envio de Leads (BaseFormFiller.GM_FORM_URL_MARKERS), para no
            # tener dos listas de marcadores que se desincronicen.
            gm = [fr for fr in todos if self._is_gm_form_src(self._iframe_src_of(fr))]
            return todos, gm

        iframes_found, gm_iframes = _buscar_iframes_gm()
        if use_iframe:
            # Solo tiene sentido esperar si se espera un iframe: en un form standalone no hay
            # ninguno y la espera serían 5 segundos perdidos por fila.
            _deadline = time.time() + FORM_WAIT_SECONDS
            while not gm_iframes and time.time() < _deadline:
                if getattr(self, "stop_event", None) and self.stop_event.is_set():
                    break
                time.sleep(0.5)
                iframes_found, gm_iframes = _buscar_iframes_gm()

        # En un form standalone la página ES el formulario: cuenta como 1, no como 0.
        forms_count = 1 if standalone else len(gm_iframes)

        if use_iframe:
            try:
                target_iframe = self.find_and_position_to_form(expected_form_url)
                if not target_iframe:
                    # Intento de fallback
                    try:
                        _cand, _es_gm = self._pick_gm_iframe(iframes_found, expected_url=expected_form_url)
                    except Exception:
                        _cand, _es_gm = None, False
                        
                    if _cand is not None:
                        target_iframe = _cand
                        if not _es_gm:
                            form_url_mismatch = True
                    else:
                        form_url_mismatch = True
                
                if target_iframe:
                    iframe_src = target_iframe.get_attribute("src") or ""
                    self.driver.switch_to.frame(target_iframe)
                    if self.screenshot_manager:
                        self.screenshot_manager.current_frame = target_iframe
            except Exception as e:
                return {
                    "ok": False,
                    "error": f"Error con iframe: {str(e)}",
                    "current_url": self.driver.current_url,
                    "landing_status": landing_status_str,
                    "iframe_correct": "NO",
                    "form_url_found": "",
                    "form_coincide": "❌ FAIL",
                    "displayed": "❌ FAIL",
                    "secure_status": secure_status_str,
                    "forms_count": forms_count,
                    "ss_landing": ss_landing_path,
                    "ss_completado": ""
                }
        else:
            iframe_src = self.driver.current_url

        iframe_correct_str = "SÍ" if (target_iframe and not form_url_mismatch) else ("NO" if use_iframe else "N/A (Standalone)")
        
        # Validar coincidencia estricta de URL (Excel == Inserted)
        form_coincide_str = "✅ PASS"
        if use_iframe:
            if not target_iframe or form_url_mismatch:
                form_coincide_str = f"❌ FAIL Mismatch: expected {expected_form_url} but found {iframe_src}"
        
        if use_iframe and not target_iframe:
            return {
                "ok": False,
                "error": "Iframe del formulario no encontrado en la landing.",
                "current_url": self.driver.current_url,
                "landing_status": landing_status_str,
                "iframe_correct": "NO",
                "form_url_found": "",
                "form_coincide": form_coincide_str,
                "displayed": "❌ FAIL",
                "secure_status": secure_status_str,
                "forms_count": forms_count,
                "ss_landing": ss_landing_path,
                "ss_completado": ""
            }

        if getattr(self, "stop_event", None) and self.stop_event.is_set():
            raise Exception("Cancelado por el usuario")
        # 4. Rellenado parcial (Nombre/Apellido y Dropdown)
        ss_completado_path = ""
        displayed_str = "❌ FAIL"
        fill_success = False
        try:
            self.wait_for_form_ready_in_iframe()
            if getattr(self, "stop_event", None) and self.stop_event.is_set():
                raise Exception("Cancelado por el usuario")
            # Ejecutar rellenado mínimo personalizado (dos campos interactivos)
            fill_success = self.fill_minimum_fields_massive()
            if fill_success:
                displayed_str = "✅ PASS"
            else:
                displayed_str = "❌ FAIL"
            time.sleep(0.5)
            
            # Tomar captura después del rellenado si tomar_capturas está activo
            if tomar_capturas and self.screenshot_manager:
                ss_completado_path = self.screenshot_manager.take_form_screenshot(row_num, "completado", full_page=True)
        except Exception as e:
            displayed_str = "❌ FAIL"
            return {
                "ok": False,
                "error": f"Error completando campos: {str(e)}",
                "current_url": self.driver.current_url,
                "landing_status": landing_status_str,
                "iframe_correct": iframe_correct_str,
                "form_url_found": iframe_src,
                "form_coincide": form_coincide_str,
                "displayed": displayed_str,
                "secure_status": secure_status_str,
                "forms_count": forms_count,
                "ss_landing": ss_landing_path,
                "ss_completado": ""
            }

        return {
            "ok": True if fill_success else False,
            "error": "" if fill_success else "No se pudo rellenar los dos campos interactivos (nombre y dropdown)",
            "current_url": self.driver.current_url,
            "landing_status": landing_status_str,
            "iframe_correct": iframe_correct_str,
            "form_url_found": iframe_src,
            "form_coincide": form_coincide_str,
            "displayed": displayed_str,
            "secure_status": secure_status_str,
            "forms_count": forms_count,
            "ss_landing": ss_landing_path,
            "ss_completado": ss_completado_path
        }


def load_passed_results_from_wb(wb, custom_cols):
    """Devuelve {(url_live, url_secure): resultados} de las filas que pasaron LIMPIAS.

    Una fila cuenta como pasada solo si TODAS las columnas de resultado tienen valor y
    ninguna dice FAIL ni SKIPPED. Así, con "Rerun fails previos" tildado, se vuelve a correr
    cualquier fila que haya fallado en cualquiera de las columnas (antes solo se miraban
    Excel==Inserted?, Displayed? y LandingStatus, y un FAIL en statusInserted o en
    Secure_Inserted quedaba sin reintentar).
    """
    passed_results = {}
    for s_name in wb.sheetnames:
        sheet = wb[s_name]
        header_row_idx = None
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_vals = [str(x).strip().upper() for x in row if x is not None]
            if (custom_cols["segmento"].upper() in row_vals and
                custom_cols["estado"].upper() in row_vals and
                custom_cols["url_live"].upper() in row_vals and
                custom_cols["url_secure"].upper() in row_vals):
                header_row_idx = r_idx
                break
        if not header_row_idx:
            continue
            
        header_row = [str(cell.value).strip().upper() if cell.value is not None else "" for cell in sheet[header_row_idx]]
        try:
            col_seg = header_row.index(custom_cols["segmento"].upper()) + 1
            col_est = header_row.index(custom_cols["estado"].upper()) + 1
            col_url_live = header_row.index(custom_cols["url_live"].upper()) + 1
            col_url_secure = header_row.index(custom_cols["url_secure"].upper()) + 1
            
            col_current = header_row.index("CURRENTURL") + 1
            col_landing = header_row.index("LANDINGSTATUS") + 1
            col_secure_ins = header_row.index("SECURE_INSERTED") + 1
            col_excel_ins = header_row.index("EXCEL==INSERTED?") + 1
            col_disp = header_row.index("DISPLAYED?") + 1
            col_status_ins = header_row.index("STATUSINSERTED") + 1
            col_forms = header_row.index("FORMSCOUNT") + 1
            
            col_comments = None
            for col_idx in range(1, len(header_row) + 1):
                val = header_row[col_idx - 1]
                if val == "COMENTARIOS" or val == "COMENTARIO":
                    col_comments = col_idx
                    break
        except ValueError:
            continue
            
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            url_live = str(sheet.cell(row=r_idx, column=col_url_live).value or "").strip()
            url_secure = str(sheet.cell(row=r_idx, column=col_url_secure).value or "").strip()
            if not url_live and not url_secure:
                continue
                
            cols_resultado = (col_current, col_landing, col_secure_ins, col_excel_ins,
                              col_disp, col_status_ins, col_forms)
            valores = [str(sheet.cell(row=r_idx, column=c).value or "").strip() for c in cols_resultado]

            # Sin resultados previos (fila nunca corrida) -> hay que correrla.
            # Con FAIL o SKIPPED en cualquier columna -> hay que volver a correrla.
            is_ok = (
                all(v != "" for v in valores)
                and not any("FAIL" in v.upper() or "SKIPPED" in v.upper() for v in valores)
            )
            if is_ok:
                key = (url_live.lower(), url_secure.lower())
                passed_results[key] = {
                    "currentUrl": sheet.cell(row=r_idx, column=col_current).value,
                    "LandingStatus": sheet.cell(row=r_idx, column=col_landing).value,
                    "Secure_Inserted": sheet.cell(row=r_idx, column=col_secure_ins).value,
                    "Excel==Inserted?": sheet.cell(row=r_idx, column=col_excel_ins).value,
                    "Displayed?": sheet.cell(row=r_idx, column=col_disp).value,
                    "statusInserted": sheet.cell(row=r_idx, column=col_status_ins).value,
                    "formsCount": sheet.cell(row=r_idx, column=col_forms).value,
                    "Comentarios": sheet.cell(row=r_idx, column=col_comments).value if col_comments else ""
                }
    return passed_results


# Columnas de resultado que se insertan al principio de cada hoja del Excel matriz.
# El orden importa: se escriben en bloque contiguo arrancando en la columna 1.
RESULT_HEADERS = [
    "currentUrl",
    "LandingStatus",
    "Secure_Inserted",
    "Excel==Inserted?",
    "Displayed?",
    "statusInserted",
    "formsCount",
]

RESUMEN_SHEET_NAME = "RESUMEN REVISIÓN"


def _sheet_headers(sheet, header_row_idx):
    return [str(c.value).strip().upper() if c.value is not None else "" for c in sheet[header_row_idx]]


def _find_result_block(sheet, header_row_idx):
    """Columna (1-indexed) donde arranca el bloque de resultados si ya existe, o None."""
    headers = _sheet_headers(sheet, header_row_idx)
    wanted = [h.upper() for h in RESULT_HEADERS]
    for start in range(0, max(1, len(headers) - len(wanted) + 1)):
        if headers[start:start + len(wanted)] == wanted:
            return start + 1
    return None


def _find_comments_col(headers):
    for idx, val in enumerate(headers, start=1):
        if val in ("COMENTARIOS", "COMENTARIO"):
            return idx
    return None


def run_massive_check(excel_path, custom_cols, selected_markets, borrar_comentarios=False, tomar_capturas=True, solo_fails=False, browser="chrome", headless=True, progress_callback=None, stop_event=None, paralelo=False, max_workers=3):
    """
    Función controladora del chequeo masivo.

    El Excel matriz que se recibe NO se modifica: se trabaja sobre una copia en memoria y de
    cada corrida sale un único archivo, en resultados/resultado_urlsinsertas/:

        Resultados_Revision_Masiva_<matriz>_<timestamp>.xlsx

    que es el mismo Excel matriz pero ya editado con los resultados: en cada hoja (una por
    mercado) las columnas de resultado quedan al principio y a continuación el resto de las
    columnas originales. No se generan archivos por país: todo queda en ese único Excel.

    paralelo=True revisa varios mercados a la vez, cada uno con su propio navegador
    (hasta max_workers en simultáneo). La escritura del Excel siempre es de a uno al final,
    así que el paralelo no puede pisar resultados.
    """
    import time
    start_time = time.time()

    # 1. Carpeta destino (capturas y respaldos)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dest_dir = os.path.join(base_dir, "resultados", "resultado_urlsinsertas")
    os.makedirs(dest_dir, exist_ok=True)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    if progress_callback:
        progress_callback("GENERAL", "Cargando archivo Excel...", 0)

    # Cargar Excel matriz con openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        return False, f"Error cargando Excel: {e}"

    # Cargar nombres de hojas
    sheet_names = list(wb.sheetnames)

    # Mapeo de hojas a países
    sheet_to_country = {
        "COLOMBIA": "Colombia",
        "ECUADOR": "Ecuador",
        "ARGENTINA": "Argentina",
        "CHILE": "Chile",
        "PERÚ": "Peru",
        "PERU": "Peru",
        "PARAGUAY": "Paraguay",
        "URUGUAY": "Uruguay",
        "BOLIVIA": "Bolivia",
        "BRASIL": "Brasil",
        "BRAZIL": "Brasil"
    }

    # Definir estilos premium con colores verde/rojo
    fill_pass = PatternFill(fill_type="solid", fgColor="C6EFCE") # verde claro
    font_pass = Font(color="006100", bold=True)

    fill_fail = PatternFill(fill_type="solid", fgColor="FFC7CE") # rojo claro
    font_fail = Font(color="9C0006", bold=True)

    fill_skip = PatternFill(fill_type="solid", fgColor="EAECEE") # gris claro
    font_bold = Font(bold=True)

    # Helper para colorear celdas basado en contenido
    def color_cell_by_content(cell):
        val_str = str(cell.value or "")
        if "✅ PASS" in val_str or "PASS" in val_str or "SÍ" == val_str:
            cell.fill = fill_pass
            cell.font = font_pass
        elif "❌ FAIL" in val_str or "FAIL" in val_str or "NO" == val_str or "Mismatch" in val_str:
            cell.fill = fill_fail
            cell.font = font_fail

    # Contadores generales
    total_processed = 0
    total_skipped = 0
    total_failed = 0
    total_passed = 0
    hojas_procesadas = []
    stats_por_mercado = {}

    # Resultados previos (solo_fails). El matriz nunca se edita, así que la corrida anterior
    # hay que buscarla en el último reporte generado en resultados/resultado_urlsinsertas/.
    passed_results = {}
    if solo_fails:
        try:
            previos = [os.path.join(dest_dir, f) for f in os.listdir(dest_dir)
                       if f.startswith("Resultados_Revision_Masiva_") and f.endswith(".xlsx")
                       and not f.startswith("~$")]
            if previos:
                previos.sort(key=os.path.getmtime, reverse=True)
                wb_previo = openpyxl.load_workbook(previos[0], data_only=True)
                passed_results.update(load_passed_results_from_wb(wb_previo, custom_cols))
                print(f"Rerun fails: reusando resultados PASS de {os.path.basename(previos[0])}")
            else:
                print("Rerun fails: no hay reportes previos, se revisa todo.")
        except Exception as e:
            print(f"Rerun fails: no se pudieron leer resultados previos ({e}). Se revisa todo.")

    # Normalizar mercados seleccionados a mayúsculas
    selected_markets_upper = [m.upper() for m in selected_markets]

    # ==================================================================================
    # La revision se hace en TRES fases para que el modo paralelo sea seguro:
    #
    #   1. PLANIFICAR  (un solo hilo): se lee el Excel, se insertan las columnas de
    #      resultado y se arma la lista de filas a revisar de cada mercado.
    #   2. REVISAR     (secuencial o en paralelo): cada mercado abre SU navegador y
    #      devuelve los resultados en memoria. Nadie toca el workbook aca.
    #   3. ESCRIBIR    (un solo hilo): se vuelcan los resultados al workbook y se guarda.
    #
    # openpyxl no es thread-safe y ademas todos los mercados escriben en el MISMO libro:
    # si los hilos escribieran en vivo, el ultimo en guardar pisaria a los demas. Separando
    # la parte lenta (los navegadores) de la escritura, el paralelo queda sin riesgo.
    # ==================================================================================

    planes = []
    for s_name in sheet_names:
        if stop_event and stop_event.is_set():
            break

        if s_name == RESUMEN_SHEET_NAME:
            continue

        # El nombre de la hoja no siempre coincide con el del mercado en la UI: la hoja de
        # Perú se llama "PERÚ" (con tilde) y la UI manda "Peru". Se compara contra el país
        # normalizado por sheet_to_country además del nombre crudo, si no Perú (y Brazil)
        # quedaban afuera de la revisión sin ningún aviso.
        country_name = sheet_to_country.get(s_name.upper())
        if (s_name.upper() not in selected_markets_upper
                and (country_name or "").upper() not in selected_markets_upper):
            # Omitir mercados no seleccionados
            continue

        if not country_name:
            continue

        sheet = wb[s_name]

        # Buscar la fila cabecera
        header_row_idx = None
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_vals = [str(x).strip().upper() for x in row if x is not None]
            if (custom_cols["segmento"].upper() in row_vals and
                custom_cols["estado"].upper() in row_vals and
                custom_cols["url_live"].upper() in row_vals and
                custom_cols["url_secure"].upper() in row_vals):
                header_row_idx = r_idx
                break

        if not header_row_idx:
            print(f"No se encontró la fila cabecera en hoja {s_name}")
            continue

        # --- Bloque de resultados al principio de la hoja ---
        # Si ya existe (corrida anterior) se reutiliza y se sobrescribe; si no, se inserta.
        # Así el Excel no se llena de columnas nuevas en cada ejecución.
        if _find_result_block(sheet, header_row_idx) is None:
            sheet.insert_cols(1, amount=len(RESULT_HEADERS))
            for off, title in enumerate(RESULT_HEADERS):
                c = sheet.cell(row=header_row_idx, column=1 + off)
                c.value = title
                c.font = font_bold

        headers = _sheet_headers(sheet, header_row_idx)
        res_start = _find_result_block(sheet, header_row_idx)
        new_col_indices = {name: res_start + off for off, name in enumerate(RESULT_HEADERS)}

        # Índices de las columnas originales (ya corridas a la derecha por el insert)
        col_indices = {}
        try:
            col_indices["segmento"] = headers.index(custom_cols["segmento"].upper()) + 1
            col_indices["estado"] = headers.index(custom_cols["estado"].upper()) + 1
            col_indices["url_live"] = headers.index(custom_cols["url_live"].upper()) + 1
            col_indices["url_secure"] = headers.index(custom_cols["url_secure"].upper()) + 1
        except ValueError as e:
            print(f"Error indexando columnas en la hoja {s_name}: {e}")
            continue

        # Columna de Comentarios: se usa la del Excel si existe, si no se crea al final.
        comentarios_col_idx = _find_comments_col(headers)
        if not comentarios_col_idx:
            comentarios_col_idx = len(headers) + 1
            c = sheet.cell(row=header_row_idx, column=comentarios_col_idx)
            c.value = "Comentarios"
            c.font = font_bold
        new_col_indices["Comentarios"] = comentarios_col_idx

        # Carpeta de capturas del mercado
        country_capturas_dir = None
        if tomar_capturas:
            country_capturas_dir = os.path.join(dest_dir, f"resultadoMasivo_{country_name}", "Capturas")
            os.makedirs(country_capturas_dir, exist_ok=True)

        # Filas a revisar: se leen ACA, en un solo hilo, para que los workers no toquen
        # el workbook.
        filas = []
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            url_live = str(sheet.cell(row=r_idx, column=col_indices["url_live"]).value or "").strip()
            url_secure = str(sheet.cell(row=r_idx, column=col_indices["url_secure"]).value or "").strip()
            if not url_live and not url_secure:
                continue
            filas.append({
                "r_idx": r_idx,
                "segmento": str(sheet.cell(row=r_idx, column=col_indices["segmento"]).value or "").strip(),
                "estado": str(sheet.cell(row=r_idx, column=col_indices["estado"]).value or "").strip(),
                "url_live": url_live,
                "url_secure": url_secure,
                "comentario": sheet.cell(row=r_idx, column=comentarios_col_idx).value or "",
            })

        if not filas:
            continue

        planes.append({
            "s_name": s_name,
            "country_name": country_name,
            "new_col_indices": new_col_indices,
            "comentarios_col_idx": comentarios_col_idx,
            "capturas_dir": country_capturas_dir,
            "filas": filas,
        })

    # ---------------------------------------------------------------- FASE 2: revisar
    def _revisar_mercado(plan):
        """Revisa todas las filas de un mercado con su propio navegador.
        Devuelve {r_idx: {"valores": {...}, "estado": "pass"|"fail"|"skip"}}.
        No toca el workbook: solo devuelve datos."""
        country_name = plan["country_name"]
        filas = plan["filas"]
        total_sheet = len(filas)
        procesadas = 0
        salida = {}
        filler = None
        try:
            for fila in filas:
                if stop_event and stop_event.is_set():
                    break

                r_idx = fila["r_idx"]
                segmento = fila["segmento"]
                url_live = fila["url_live"]
                url_secure = fila["url_secure"]
                is_off = "OFF" in fila["estado"].upper() or "❌" in fila["estado"]

                # solo_fails: si esta URL ya pasó limpia en la corrida anterior, se reusa
                key = (url_live.lower(), url_secure.lower())
                if solo_fails and key in passed_results and not is_off:
                    procesadas += 1
                    prev = passed_results[key]
                    salida[r_idx] = {
                        "valores": {k: prev.get(k) for k in RESULT_HEADERS},
                        "comentario_previo": prev.get("Comentarios") or "",
                        "estado": "pass",
                    }
                    if progress_callback:
                        pct = int((procesadas / total_sheet) * 100)
                        progress_callback(country_name, f"Revisando row {r_idx}: {segmento} (Prev PASS)", pct)
                    continue

                if is_off:
                    procesadas += 1
                    salida[r_idx] = {
                        "valores": {
                            "currentUrl": "N/A",
                            "LandingStatus": "OFF - Omitido",
                            "Secure_Inserted": "N/A",
                            "Excel==Inserted?": "SKIPPED",
                            "Displayed?": "SKIPPED",
                            "statusInserted": "N/A",
                            "formsCount": 0,
                        },
                        "estado": "skip",
                    }
                    if progress_callback:
                        pct = int((procesadas / total_sheet) * 100)
                        progress_callback(country_name, f"{procesadas}/{total_sheet} (OFF)", pct)
                    continue

                # ON o prioridad: hay que abrir el navegador
                if not filler:
                    if progress_callback:
                        progress_callback(country_name, "Abriendo navegador...", 0)
                    filler = MassiveFormFiller(country_name, browser=browser, headless=headless, background=True)
                    filler.stop_event = stop_event
                    filler.setup_directories_and_files(custom_dir=plan["capturas_dir"])
                    filler.initialize_browser()
                    if not tomar_capturas:
                        filler.screenshot_manager = None

                procesadas += 1
                if progress_callback:
                    pct = int((procesadas / total_sheet) * 100)
                    progress_callback(country_name, f"Revisando row {r_idx}: {segmento}", pct)

                res = filler.check_single_url(url_live, url_secure, r_idx, tomar_capturas=tomar_capturas)
                salida[r_idx] = {
                    "valores": {
                        "currentUrl": res["current_url"],
                        "LandingStatus": res["landing_status"],
                        "Secure_Inserted": res["form_url_found"],
                        "Excel==Inserted?": res["form_coincide"],
                        "Displayed?": res["displayed"],
                        "statusInserted": res["secure_status"],
                        "formsCount": res["forms_count"],
                    },
                    "estado": "pass" if res["ok"] else "fail",
                }
        except Exception as e:
            print(f"Error procesando la hoja {plan['s_name']}: {e}")
        finally:
            if filler:
                try:
                    filler.driver.quit()
                except Exception:
                    pass
        return plan["s_name"], salida

    resultados_por_hoja = {}
    if paralelo and len(planes) > 1:
        # Un navegador por mercado a la vez. Se acota la cantidad: cada worker levanta un
        # Chrome completo y abrir nueve de golpe deja la maquina inservible.
        import concurrent.futures
        n_workers = max(1, min(int(max_workers or 3), len(planes)))
        print(f"Revisión Masiva en paralelo: {len(planes)} mercado(s), {n_workers} navegador(es) a la vez.")
        if progress_callback:
            progress_callback("GENERAL", f"Paralelo: {n_workers} navegadores simultáneos...", 0)
        with concurrent.futures.ThreadPoolExecutor(max_workers=n_workers) as pool:
            for s_name, salida in pool.map(_revisar_mercado, planes):
                resultados_por_hoja[s_name] = salida
    else:
        for plan in planes:
            if stop_event and stop_event.is_set():
                break
            s_name, salida = _revisar_mercado(plan)
            resultados_por_hoja[s_name] = salida

    # ---------------------------------------------------------------- FASE 3: escribir
    for plan in planes:
        s_name = plan["s_name"]
        salida = resultados_por_hoja.get(s_name) or {}
        if not salida:
            continue
        sheet = wb[s_name]
        new_col_indices = plan["new_col_indices"]
        comentarios_col_idx = plan["comentarios_col_idx"]
        hojas_procesadas.append(s_name)
        m_proc = m_pass = m_fail = m_skip = 0

        for fila in plan["filas"]:
            r_idx = fila["r_idx"]
            r = salida.get(r_idx)
            if not r:
                continue  # fila no revisada (corrida detenida antes de llegar)

            for col_key, val in r["valores"].items():
                cell = sheet.cell(row=r_idx, column=new_col_indices[col_key])
                cell.value = val
                if r["estado"] == "skip":
                    cell.fill = fill_skip
                else:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font()
                    color_cell_by_content(cell)

            if borrar_comentarios:
                sheet.cell(row=r_idx, column=comentarios_col_idx).value = ""
            elif r.get("comentario_previo") and not fila["comentario"]:
                sheet.cell(row=r_idx, column=comentarios_col_idx).value = r["comentario_previo"]

            if r["estado"] == "skip":
                total_skipped += 1
                m_skip += 1
            elif r["estado"] == "pass":
                total_processed += 1
                total_passed += 1
                m_proc += 1
                m_pass += 1
            else:
                total_processed += 1
                total_failed += 1
                m_proc += 1
                m_fail += 1

        stats_por_mercado[s_name] = (m_proc, m_pass, m_fail, m_skip)

    # Calcular tiempo transcurrido antes de guardar para incluirlo en el Excel
    elapsed = time.time() - start_time
    hours = int(elapsed // 3600)
    minutes = int((elapsed % 3600) // 60)
    seconds = int(elapsed % 60)
    if hours > 0:
        time_str = f"{hours}h {minutes}m {seconds}s"
    elif minutes > 0:
        time_str = f"{minutes}m {seconds}s"
    else:
        time_str = f"{seconds}s"

    if not hojas_procesadas:
        return False, ("No se encontró ninguna hoja para los mercados seleccionados "
                       f"({', '.join(selected_markets)}). Hojas del Excel: {', '.join(sheet_names)}.")

    # El reporte es de ESTA corrida: se dejan solo las hojas de los mercados que se
    # revisaron. Las de los mercados no seleccionados viajaban sin columnas de resultado y
    # hacian parecer que se habian revisado y no habian dado nada. El Excel matriz original
    # queda intacto igual, asi que no se pierde nada.
    for _s in list(wb.sheetnames):
        if _s not in hojas_procesadas and _s != RESUMEN_SHEET_NAME:
            try:
                del wb[_s]
            except Exception:
                pass

    # Hoja de resumen (se regenera en cada corrida para no acumular filas sueltas)
    summary_font = Font(bold=True, color="7D4E9F")
    summary_fill = PatternFill(start_color="F2E6FF", end_color="F2E6FF", fill_type="solid")
    try:
        if RESUMEN_SHEET_NAME in wb.sheetnames:
            del wb[RESUMEN_SHEET_NAME]
        ws_res = wb.create_sheet(title=RESUMEN_SHEET_NAME)
        ws_res.cell(row=1, column=1).value = "Última revisión masiva"
        ws_res.cell(row=1, column=2).value = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws_res.cell(row=2, column=1).value = "Duración del test:"
        ws_res.cell(row=2, column=2).value = time_str
        ws_res.cell(row=3, column=1).value = "Resumen:"
        ws_res.cell(row=3, column=2).value = (f"Procesados: {total_processed} | PASS: {total_passed} | "
                                              f"FAIL: {total_failed} | Omitidos: {total_skipped}")
        for r in (1, 2, 3):
            for c in (1, 2):
                ws_res.cell(row=r, column=c).font = summary_font
                ws_res.cell(row=r, column=c).fill = summary_fill

        ws_res.cell(row=5, column=1).value = "Mercado"
        ws_res.cell(row=5, column=2).value = "Procesados"
        ws_res.cell(row=5, column=3).value = "PASS"
        ws_res.cell(row=5, column=4).value = "FAIL"
        ws_res.cell(row=5, column=5).value = "Omitidos"
        for c in range(1, 6):
            ws_res.cell(row=5, column=c).font = font_bold
        for i, (m_name, (m_proc, m_pass, m_fail, m_skip)) in enumerate(stats_por_mercado.items(), start=6):
            ws_res.cell(row=i, column=1).value = m_name
            ws_res.cell(row=i, column=2).value = m_proc
            ws_res.cell(row=i, column=3).value = m_pass
            ws_res.cell(row=i, column=4).value = m_fail
            ws_res.cell(row=i, column=5).value = m_skip
        ws_res.column_dimensions["A"].width = 26
        ws_res.column_dimensions["B"].width = 24
    except Exception as e:
        print(f"No se pudo escribir la hoja de resumen: {e}")

    # Único archivo de salida: copia del matriz ya editada con los resultados.
    # El Excel matriz original queda intacto.
    nombre_base = os.path.splitext(os.path.basename(excel_path))[0]
    reporte_path = os.path.join(dest_dir, f"Resultados_Revision_Masiva_{nombre_base}_{timestamp}.xlsx")
    try:
        wb.save(reporte_path)
    except Exception as e:
        print(f"Error guardando el reporte de resultados: {e}")
        return False, f"Error al guardar reporte: {e}"

    msg = (f"Revisión finalizada en {time_str}. Procesados: {total_processed} "
           f"(PASS: {total_passed}, FAIL: {total_failed}), Omitidos: {total_skipped}.")
    return True, {
        "msg": msg,
        "excel_path": reporte_path,
        "reporte_path": reporte_path,
        "total_processed": total_processed,
        "total_passed": total_passed,
        "total_failed": total_failed,
        "total_skipped": total_skipped,
        "elapsed_time": time_str
    }



def sincronizar_reporte_con_matriz(master_path, results_path, custom_cols):
    """
    Sincroniza y compara el reporte de resultados (results_path) con la matriz origen (master_path).
    Soporta la adición de nuevas filas, actualización de estados (ON/OFF), actualización de URLs (live/secure),
    y detección de eliminados. Utiliza matching exacto de URLs y fallback por Segmento.
    """
    try:
        wb_master = openpyxl.load_workbook(master_path, data_only=True)
        wb_results = openpyxl.load_workbook(results_path)
    except Exception as e:
        return False, f"Error cargando los archivos Excel: {e}"

    stats = {
        "added": 0,
        "updated_state": 0,
        "deleted": 0
    }
    
    font_bold = Font(bold=True)
    timestamp = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    changes_log = []

    # Mapear nombres de columnas de salida y sus posiciones en el reporte
    new_headers = [
        custom_cols["segmento"],
        custom_cols["estado"],
        custom_cols["url_live"],
        custom_cols["url_secure"],
        "currentUrl",
        "LandingStatus",
        "Secure_Inserted",
        "Excel==Inserted?",
        "Displayed?",
        "statusInserted",
        "formsCount",
        "Comentarios"
    ]

    # Iterar por cada hoja en la matriz maestra
    for s_name in wb_master.sheetnames:
        sheet_master = wb_master[s_name]
        
        # 1. Buscar cabecera en maestro
        header_row_idx_master = None
        for r_idx, row in enumerate(sheet_master.iter_rows(values_only=True), start=1):
            row_vals = [str(x).strip().upper() for x in row if x is not None]
            if (custom_cols["segmento"].upper() in row_vals and
                custom_cols["estado"].upper() in row_vals and
                custom_cols["url_live"].upper() in row_vals and
                custom_cols["url_secure"].upper() in row_vals):
                header_row_idx_master = r_idx
                break
        
        if not header_row_idx_master:
            continue

        # Obtener índices de columna de maestro
        header_row_master = [str(cell.value).strip().upper() if cell.value is not None else "" for cell in sheet_master[header_row_idx_master]]
        
        try:
            col_seg_m = header_row_master.index(custom_cols["segmento"].upper()) + 1
            col_est_m = header_row_master.index(custom_cols["estado"].upper()) + 1
            col_url_live_m = header_row_master.index(custom_cols["url_live"].upper()) + 1
            col_url_secure_m = header_row_master.index(custom_cols["url_secure"].upper()) + 1
        except ValueError:
            continue

        comentarios_col_idx_m = None
        for col_idx in range(1, len(header_row_master) + 1):
            val = header_row_master[col_idx - 1]
            if val == "COMENTARIOS" or val == "COMENTARIO":
                comentarios_col_idx_m = col_idx
                break

        # 2. Obtener o crear la hoja correspondiente en resultados
        if s_name not in wb_results.sheetnames:
            sheet_results = wb_results.create_sheet(title=s_name)
            for c_idx, title in enumerate(new_headers, start=1):
                cell = sheet_results.cell(row=1, column=c_idx)
                cell.value = title
                cell.font = font_bold
            header_row_idx_res = 1
        else:
            sheet_results = wb_results[s_name]
            header_row_idx_res = None
            for r_idx, row in enumerate(sheet_results.iter_rows(values_only=True), start=1):
                row_vals = [str(x).strip().upper() for x in row if x is not None]
                if (custom_cols["segmento"].upper() in row_vals and
                    custom_cols["estado"].upper() in row_vals and
                    custom_cols["url_live"].upper() in row_vals and
                    custom_cols["url_secure"].upper() in row_vals):
                    header_row_idx_res = r_idx
                    break
            
            if not header_row_idx_res:
                # Si la hoja existe pero está rota o vacía, la recreamos
                sheet_results.delete_rows(1, sheet_results.max_row + 1)
                for c_idx, title in enumerate(new_headers, start=1):
                    cell = sheet_results.cell(row=1, column=c_idx)
                    cell.value = title
                    cell.font = font_bold
                header_row_idx_res = 1

        # 3. Leer registros de maestro
        master_rows = []
        for r_idx in range(header_row_idx_master + 1, sheet_master.max_row + 1):
            url_live = str(sheet_master.cell(row=r_idx, column=col_url_live_m).value or "").strip()
            url_secure = str(sheet_master.cell(row=r_idx, column=col_url_secure_m).value or "").strip()
            if not url_live and not url_secure:
                continue
                
            segmento = str(sheet_master.cell(row=r_idx, column=col_seg_m).value or "").strip()
            estado = str(sheet_master.cell(row=r_idx, column=col_est_m).value or "").strip()
            comentario = ""
            if comentarios_col_idx_m:
                comentario = str(sheet_master.cell(row=r_idx, column=comentarios_col_idx_m).value or "").strip()

            master_rows.append({
                "segmento": segmento,
                "estado": estado,
                "url_live": url_live,
                "url_secure": url_secure,
                "comentario": comentario,
                "matched": False
            })

        # 4. Leer registros de resultados
        results_rows = []
        header_row_res = [str(cell.value).strip().upper() if cell.value is not None else "" for cell in sheet_results[header_row_idx_res]]
        
        try:
            col_seg_r = header_row_res.index(custom_cols["segmento"].upper()) + 1
            col_est_r = header_row_res.index(custom_cols["estado"].upper()) + 1
            col_url_live_r = header_row_res.index(custom_cols["url_live"].upper()) + 1
            col_url_secure_r = header_row_res.index(custom_cols["url_secure"].upper()) + 1
        except ValueError:
            # Si no coincide exactamente con el formato esperado, no podemos mapear
            continue

        col_comentarios_r = 12 # Columna Comentarios estándar
        
        for r_idx in range(header_row_idx_res + 1, sheet_results.max_row + 1):
            url_live = str(sheet_results.cell(row=r_idx, column=col_url_live_r).value or "").strip()
            url_secure = str(sheet_results.cell(row=r_idx, column=col_url_secure_r).value or "").strip()
            if not url_live and not url_secure:
                continue
                
            segmento = str(sheet_results.cell(row=r_idx, column=col_seg_r).value or "").strip()
            estado = str(sheet_results.cell(row=r_idx, column=col_est_r).value or "").strip()
            comentario = str(sheet_results.cell(row=r_idx, column=col_comentarios_r).value or "").strip()

            results_rows.append({
                "row_idx": r_idx,
                "segmento": segmento,
                "estado": estado,
                "url_live": url_live,
                "url_secure": url_secure,
                "comentario": comentario,
                "matched": False
            })

        # 5. Sincronizar - Paso A: Matching exacto por (url_live, url_secure)
        for m_row in master_rows:
            for r_row in results_rows:
                if r_row["matched"]:
                    continue
                if (m_row["url_live"].lower() == r_row["url_live"].lower() and 
                    m_row["url_secure"].lower() == r_row["url_secure"].lower()):
                    
                    # Match exacto encontrado
                    m_row["matched"] = True
                    r_row["matched"] = True
                    r_idx = r_row["row_idx"]
                    
                    # Sincronizar Segmento si cambió
                    if r_row["segmento"] != m_row["segmento"]:
                        sheet_results.cell(row=r_idx, column=col_seg_r).value = m_row["segmento"]

                    # Sincronizar Estado si cambió
                    if r_row["estado"].upper() != m_row["estado"].upper():
                        sheet_results.cell(row=r_idx, column=col_est_r).value = m_row["estado"]
                        
                        # Si cambió a ON, limpiar celdas de prueba para obligar al rerun
                        is_new_on = "OFF" not in m_row["estado"].upper() and "❌" not in m_row["estado"]
                        if is_new_on:
                            for col_pos in range(5, 12):
                                cell = sheet_results.cell(row=r_idx, column=col_pos)
                                cell.value = None
                                cell.fill = PatternFill(fill_type=None)
                                cell.font = Font(color="000000", bold=False)

                        # Registrar comentario
                        prev_comm = str(sheet_results.cell(row=r_idx, column=col_comentarios_r).value or "").strip()
                        new_comm = f"[Actualización {timestamp}]: Estado cambiado de {r_row['estado']} a {m_row['estado']}."
                        if prev_comm:
                            sheet_results.cell(row=r_idx, column=col_comentarios_r).value = f"{prev_comm} | {new_comm}"
                        else:
                            sheet_results.cell(row=r_idx, column=col_comentarios_r).value = new_comm
                        
                        stats["updated_state"] += 1
                        changes_log.append(f"[ESTADO] {r_row['segmento']}: {r_row['estado']} → {m_row['estado']}")
                    break

        # 6. Sincronizar - Paso B: Matching por Segmento (para detectar cambios de URL en la misma fila/segmento)
        for m_row in master_rows:
            if m_row["matched"]:
                continue
            for r_row in results_rows:
                if r_row["matched"]:
                    continue
                if m_row["segmento"].lower() == r_row["segmento"].lower():
                    # Match por segmento encontrado (las URLs cambiaron)
                    m_row["matched"] = True
                    r_row["matched"] = True
                    r_idx = r_row["row_idx"]
                    
                    # Actualizar URLs
                    sheet_results.cell(row=r_idx, column=col_url_live_r).value = m_row["url_live"]
                    sheet_results.cell(row=r_idx, column=col_url_secure_r).value = m_row["url_secure"]
                    
                    # Sincronizar Estado si cambió
                    sheet_results.cell(row=r_idx, column=col_est_r).value = m_row["estado"]
                    
                    # Como cambiaron las URLs, limpiar SIEMPRE las columnas de resultados anteriores
                    for col_pos in range(5, 12):
                        cell = sheet_results.cell(row=r_idx, column=col_pos)
                        cell.value = None
                        cell.fill = PatternFill(fill_type=None)
                        cell.font = Font(color="000000", bold=False)

                    # Registrar comentario de actualización
                    prev_comm = str(sheet_results.cell(row=r_idx, column=col_comentarios_r).value or "").strip()
                    new_comm = f"[Actualización {timestamp}]: URLs y datos actualizados (Live/Secure modificadas)."
                    if prev_comm:
                        sheet_results.cell(row=r_idx, column=col_comentarios_r).value = f"{prev_comm} | {new_comm}"
                    else:
                        sheet_results.cell(row=r_idx, column=col_comentarios_r).value = new_comm
                    
                    stats["updated_state"] += 1
                    changes_log.append(f"[URLS] {m_row['segmento']}: URLs Live/Secure actualizadas")
                    break

        # 7. Procesar eliminados (registros en resultados que no se emparejaron)
        for r_row in results_rows:
            if not r_row["matched"]:
                r_idx = r_row["row_idx"]
                old_est = r_row["estado"]
                if "ELIMINADO" not in old_est.upper() and "OFF" not in old_est.upper():
                    sheet_results.cell(row=r_idx, column=col_est_r).value = "❌ OFF (Eliminado de origen)"
                    
                    # Registrar comentario
                    prev_comm = str(sheet_results.cell(row=r_idx, column=col_comentarios_r).value or "").strip()
                    new_comm = f"[Actualización {timestamp}]: Eliminado de la matriz origen."
                    if prev_comm:
                        sheet_results.cell(row=r_idx, column=col_comentarios_r).value = f"{prev_comm} | {new_comm}"
                    else:
                        sheet_results.cell(row=r_idx, column=col_comentarios_r).value = new_comm
                    
                    stats["deleted"] += 1
                    changes_log.append(f"[ELIMINADO] {r_row['segmento']}: Eliminado de la matriz origen")

        # 8. Agregar nuevas filas (registros en maestro que no se emparejaron)
        for m_row in master_rows:
            if not m_row["matched"]:
                next_row = sheet_results.max_row + 1
                
                # Escribir columnas de entrada
                sheet_results.cell(row=next_row, column=1).value = m_row["segmento"]
                sheet_results.cell(row=next_row, column=2).value = m_row["estado"]
                sheet_results.cell(row=next_row, column=3).value = m_row["url_live"]
                sheet_results.cell(row=next_row, column=4).value = m_row["url_secure"]
                
                # Inicializar columnas de test
                for col_pos in range(5, 12):
                    sheet_results.cell(row=next_row, column=col_pos).value = ""
                    
                # Comentario
                new_comm = f"[Actualización {timestamp}]: Agregado por sincronización."
                if m_row["comentario"]:
                    sheet_results.cell(row=next_row, column=12).value = f"{m_row['comentario']} | {new_comm}"
                else:
                    sheet_results.cell(row=next_row, column=12).value = new_comm
                
                stats["added"] += 1
                changes_log.append(f"[NUEVO] {m_row['segmento']}: {m_row['url_live']}")

    # Guardar el reporte actualizado
    try:
        wb_results.save(results_path)
    except Exception as e:
        return False, f"Error guardando los resultados actualizados: {e}"

    # Generar Excel de changelog en la misma carpeta del reporte
    changelog_path = ""
    if changes_log:
        try:
            ts_file = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            results_dir = os.path.dirname(results_path)
            changelog_path = os.path.join(results_dir, f"Changelog_Sincronizacion_{ts_file}.xlsx")
            
            wb_log = openpyxl.Workbook()
            ws_log = wb_log.active
            ws_log.title = "Changelog"
            
            # Headers
            log_headers = ["Tipo", "Hoja", "Segmento", "Detalle", "Fecha/Hora"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="7D4E9F", end_color="7D4E9F", fill_type="solid")
            for c_idx, h in enumerate(log_headers, 1):
                cell = ws_log.cell(row=1, column=c_idx)
                cell.value = h
                cell.font = header_font
                cell.fill = header_fill
            
            # Filas de cambios
            add_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            upd_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            del_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for log_idx, entry in enumerate(changes_log, 2):
                if entry.startswith("[NUEVO]"):
                    tipo = "Agregado"
                    fill = add_fill
                elif entry.startswith("[ESTADO]"):
                    tipo = "Estado Actualizado"
                    fill = upd_fill
                elif entry.startswith("[URLS]"):
                    tipo = "URLs Actualizadas"
                    fill = upd_fill
                elif entry.startswith("[ELIMINADO]"):
                    tipo = "Eliminado de Origen"
                    fill = del_fill
                else:
                    tipo = "Cambio"
                    fill = upd_fill
                
                # Parse: [TYPE] segmento: detalle
                parts = entry.split("] ", 1)
                detail = parts[1] if len(parts) > 1 else entry
                seg_parts = detail.split(": ", 1)
                segmento = seg_parts[0] if seg_parts else ""
                detalle = seg_parts[1] if len(seg_parts) > 1 else detail
                
                ws_log.cell(row=log_idx, column=1).value = tipo
                ws_log.cell(row=log_idx, column=1).fill = fill
                ws_log.cell(row=log_idx, column=2).value = ""  # Hoja (se puede expandir)
                ws_log.cell(row=log_idx, column=3).value = segmento
                ws_log.cell(row=log_idx, column=4).value = detalle
                ws_log.cell(row=log_idx, column=5).value = timestamp
            
            # Ajustar anchos
            ws_log.column_dimensions["A"].width = 20
            ws_log.column_dimensions["B"].width = 15
            ws_log.column_dimensions["C"].width = 30
            ws_log.column_dimensions["D"].width = 60
            ws_log.column_dimensions["E"].width = 22
            
            wb_log.save(changelog_path)
        except Exception as e:
            print(f"Error generando changelog: {e}")
            changelog_path = ""

    stats["changelog_path"] = changelog_path
    stats["changes_log"] = changes_log
    return True, stats
